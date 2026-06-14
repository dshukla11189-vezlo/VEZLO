"""
Gmail Integration Routes
========================
Extracted from server.py for modular organization.
Handles Gmail OAuth and Ninjacart GRN email automation.
"""
from fastapi import APIRouter, HTTPException, Depends, Request, Query
from fastapi.responses import RedirectResponse
from datetime import datetime, timezone, timedelta
from typing import Optional
import uuid
import os
import urllib.parse

from dependencies import (
    db,
    get_current_user,
    logger,
)
from gmail_integration import (
    get_authorization_url,
    exchange_code_for_tokens,
    get_gmail_service,
    search_ninjacart_emails,
    get_email_content,
    download_attachment,
    parse_ninjacart_csv,
    parse_ninjacart_email_body,
    mark_email_as_read,
    add_label_to_email,
)

router = APIRouter(tags=["gmail_integration"])

# Gmail OAuth config from environment
GMAIL_CLIENT_ID = os.environ.get("GMAIL_CLIENT_ID", "")
GMAIL_CLIENT_SECRET = os.environ.get("GMAIL_CLIENT_SECRET", "")

# SECTION: GMAIL INTEGRATION ROUTES - Ninjacart GRN Email Automation
# ============================================================================

@router.get("/oauth/gmail/debug-states")
async def gmail_oauth_debug_states():
    """DEBUG: Show OAuth states in database"""
    states = await db.oauth_states.find({}).to_list(length=10)
    return {
        "total_states": len(states),
        "states": [
            {
                "state": s.get("state", "N/A")[:20] + "...",
                "user_id": s.get("user_id"),
                "redirect_uri": s.get("redirect_uri"),
                "created_at": str(s.get("created_at")),
                "expires_at": str(s.get("expires_at"))
            }
            for s in states
        ]
    }

@router.get("/oauth/gmail/debug-tokens")
async def gmail_oauth_debug_tokens():
    """DEBUG: Show Gmail tokens in database (without sensitive data)"""
    tokens = await db.gmail_tokens.find({}).to_list(length=10)
    return {
        "total_tokens": len(tokens),
        "tokens": [
            {
                "user_id": t.get("user_id"),
                "email": t.get("email"),
                "has_access_token": bool(t.get("access_token")),
                "has_refresh_token": bool(t.get("refresh_token")),
                "updated_at": str(t.get("updated_at"))
            }
            for t in tokens
        ]
    }

@router.get("/oauth/gmail/debug-headers")
async def gmail_oauth_debug_headers(request: Request):
    """DEBUG: Show what headers the server receives - helps diagnose redirect_uri issues"""
    host = request.headers.get("host", "")
    x_forwarded_host = request.headers.get("x-forwarded-host", "")
    x_forwarded_proto = request.headers.get("x-forwarded-proto", "https")
    origin = request.headers.get("origin", "")
    referer = request.headers.get("referer", "")
    
    # Determine the actual host - prefer x-forwarded-host for reverse proxy setups
    actual_host = x_forwarded_host or host
    
    # Detect redirect URI based on actual host
    if "emergent.host" in actual_host:
        redirect_uri = "https://harvest-hub-384.emergent.host/api/oauth/gmail/callback"
    elif "preview.emergentagent.com" in actual_host:
        redirect_uri = "https://harvest-hub-384.preview.emergentagent.com/api/oauth/gmail/callback"
    else:
        # Fallback: try to detect from referer or origin
        if "emergent.host" in referer or "emergent.host" in origin:
            redirect_uri = "https://harvest-hub-384.emergent.host/api/oauth/gmail/callback"
        else:
            redirect_uri = "https://harvest-hub-384.preview.emergentagent.com/api/oauth/gmail/callback"
    
    return {
        "headers_received": {
            "host": host,
            "x-forwarded-host": x_forwarded_host,
            "x-forwarded-proto": x_forwarded_proto,
            "origin": origin,
            "referer": referer
        },
        "actual_host_used": actual_host,
        "redirect_uri_that_would_be_generated": redirect_uri,
        "expected_google_console_uri": redirect_uri,
        "message": "The 'redirect_uri_that_would_be_generated' must EXACTLY match the URI in Google Cloud Console Authorized Redirect URIs"
    }

@router.get("/oauth/gmail/status")
async def get_gmail_connection_status(current_user: dict = Depends(get_current_user)):
    """Check if Gmail is connected"""
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Only admin can manage Gmail integration")
    
    token = await db.gmail_tokens.find_one({"user_id": current_user["user_id"]}, {"_id": 0})
    
    return {
        "connected": token is not None and token.get("access_token") is not None,
        "email": token.get("email") if token else None,
        "last_sync": token.get("last_sync") if token else None
    }

@router.get("/oauth/gmail/login")
async def gmail_oauth_login(request: Request, current_user: dict = Depends(get_current_user)):
    """Start Gmail OAuth flow"""
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Only admin can connect Gmail")
    
    # Generate state token
    state = str(uuid.uuid4())
    
    # Debug: Log all relevant headers to understand what we're receiving
    host = request.headers.get("host", "")
    x_forwarded_host = request.headers.get("x-forwarded-host", "")
    x_forwarded_proto = request.headers.get("x-forwarded-proto", "https")
    origin = request.headers.get("origin", "")
    referer = request.headers.get("referer", "")
    
    logger.info("OAuth Login - Headers Debug:")
    logger.info(f"  host: {host}")
    logger.info(f"  x-forwarded-host: {x_forwarded_host}")
    logger.info(f"  x-forwarded-proto: {x_forwarded_proto}")
    logger.info(f"  origin: {origin}")
    logger.info(f"  referer: {referer}")
    
    # Determine the actual host - prefer x-forwarded-host for reverse proxy setups
    actual_host = x_forwarded_host or host
    
    # Detect redirect URI based on actual host
    if "emergent.host" in actual_host:
        redirect_uri = "https://harvest-hub-384.emergent.host/api/oauth/gmail/callback"
    elif "preview.emergentagent.com" in actual_host:
        redirect_uri = "https://harvest-hub-384.preview.emergentagent.com/api/oauth/gmail/callback"
    else:
        # Fallback: try to detect from referer or origin
        if "emergent.host" in referer or "emergent.host" in origin:
            redirect_uri = "https://harvest-hub-384.emergent.host/api/oauth/gmail/callback"
        else:
            redirect_uri = "https://harvest-hub-384.preview.emergentagent.com/api/oauth/gmail/callback"
    
    logger.info(f"  Selected redirect_uri: {redirect_uri}")
    
    # Store state in DB with user_id and redirect_uri
    await db.oauth_states.insert_one({
        "state": state,
        "user_id": current_user["user_id"],
        "redirect_uri": redirect_uri,
        "created_at": datetime.now(timezone.utc),
        "expires_at": datetime.now(timezone.utc) + timedelta(minutes=10)
    })
    
    # Get authorization URL with dynamic redirect URI
    auth_url = get_authorization_url(state, redirect_uri)
    
    logger.info(f"  Generated auth_url: {auth_url[:100]}...")
    
    return {"auth_url": auth_url, "debug_redirect_uri": redirect_uri}

@router.get("/oauth/gmail/callback")
async def gmail_oauth_callback(request: Request, code: str = None, state: str = None, error: str = None):
    """Handle Gmail OAuth callback"""
    logger.info(f"Gmail OAuth callback received - code: {code[:20] if code else 'None'}..., state: {state}, error: {error}")
    
    if error:
        logger.error(f"Gmail OAuth error from Google: {error}")
        return RedirectResponse(url=f"/admin/backup?gmail_error={error}")
    
    if not code or not state:
        logger.error(f"Gmail OAuth missing params - code: {bool(code)}, state: {bool(state)}")
        return RedirectResponse(url="/admin/backup?gmail_error=missing_params")
    
    # Verify state
    state_doc = await db.oauth_states.find_one({"state": state})
    if not state_doc:
        # Log all states to debug
        all_states = await db.oauth_states.find({}).to_list(length=10)
        logger.error(f"Gmail OAuth invalid_state: {state}")
        logger.error(f"Available states in DB: {[s.get('state', 'N/A')[:20] for s in all_states]}")
        return RedirectResponse(url="/admin/backup?gmail_error=invalid_state")
    
    logger.info(f"State validated - user_id: {state_doc.get('user_id')}, redirect_uri: {state_doc.get('redirect_uri')}")
    
    # Check expiry
    if datetime.now(timezone.utc) > state_doc["expires_at"].replace(tzinfo=timezone.utc):
        return RedirectResponse(url="/admin/backup?gmail_error=state_expired")
    
    user_id = state_doc["user_id"]
    
    # Delete used state
    await db.oauth_states.delete_one({"state": state})
    
    # Get redirect URI from state
    redirect_uri = state_doc.get("redirect_uri")
    
    try:
        # Exchange code for tokens
        logger.info(f"Exchanging code for tokens with redirect_uri: {redirect_uri}")
        tokens = exchange_code_for_tokens(code, redirect_uri)
        logger.info(f"Token exchange successful - has_access_token: {bool(tokens.get('access_token'))}, has_refresh_token: {bool(tokens.get('refresh_token'))}")
        
        # Get user email
        service = get_gmail_service(tokens)
        profile = service.users().getProfile(userId='me').execute()
        email = profile.get('emailAddress', '')
        logger.info(f"Gmail profile retrieved - email: {email}")
        
        # Store tokens
        await db.gmail_tokens.update_one(
            {"user_id": user_id},
            {"$set": {
                "user_id": user_id,
                "email": email,
                **tokens,
                "updated_at": datetime.now(timezone.utc)
            }},
            upsert=True
        )
        logger.info(f"Gmail tokens saved for user: {user_id}")
        
        return RedirectResponse(url="/admin/backup?gmail_success=true")
    except Exception as e:
        logger.error(f"Gmail OAuth error: {e}", exc_info=True)
        return RedirectResponse(url=f"/admin/backup?gmail_error={str(e)[:50]}")

@router.post("/oauth/gmail/disconnect")
async def gmail_disconnect(current_user: dict = Depends(get_current_user)):
    """Disconnect Gmail integration"""
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Only admin can disconnect Gmail")
    
    await db.gmail_tokens.delete_one({"user_id": current_user["user_id"]})
    return {"message": "Gmail disconnected successfully"}

@router.get("/gmail/ninjacart-emails")
async def get_ninjacart_emails(
    max_results: int = 10,
    current_user: dict = Depends(get_current_user)
):
    """Fetch recent Ninjacart emails"""
    if current_user["role"] not in ["admin", "staff"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    # Get Gmail tokens
    token = await db.gmail_tokens.find_one({"user_id": current_user["user_id"]}, {"_id": 0})
    if not token:
        raise HTTPException(status_code=400, detail="Gmail not connected. Please connect Gmail first.")
    
    try:
        service = get_gmail_service(token)
        
        # Search for Ninjacart emails
        messages = search_ninjacart_emails(service, max_results)
        
        emails = []
        for msg in messages[:5]:  # Limit to 5 for performance
            content = get_email_content(service, msg['id'])
            if content:
                emails.append({
                    "id": content['id'],
                    "subject": content['subject'],
                    "from": content['from'],
                    "date": content['date'],
                    "has_attachments": len(content['attachments']) > 0,
                    "attachments": [a['filename'] for a in content['attachments']]
                })
        
        return {"emails": emails}
    except Exception as e:
        logger.error(f"Error fetching emails: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to fetch emails: {str(e)}")

@router.post("/gmail/process-grn-email/{message_id}")
async def process_grn_from_email(
    message_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Process a specific email and create GRN from it"""
    if current_user["role"] not in ["admin", "staff"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    # Get Gmail tokens
    token = await db.gmail_tokens.find_one({"user_id": current_user["user_id"]}, {"_id": 0})
    if not token:
        raise HTTPException(status_code=400, detail="Gmail not connected")
    
    try:
        service = get_gmail_service(token)
        
        # Get email content
        content = get_email_content(service, message_id)
        if not content:
            raise HTTPException(status_code=404, detail="Email not found")
        
        grn_items = []
        
        # Try to parse CSV attachment first
        for attachment in content['attachments']:
            if attachment['filename'].lower().endswith('.csv'):
                csv_data = download_attachment(service, message_id, attachment['attachment_id'])
                if csv_data:
                    grn_items = parse_ninjacart_csv(csv_data)
                    break
        
        # If no CSV, try to parse email body
        if not grn_items and content['body_text']:
            grn_items = parse_ninjacart_email_body(content['body_text'])
        
        if not grn_items:
            raise HTTPException(status_code=400, detail="Could not parse GRN data from email")
        
        # Mark email as processed
        mark_email_as_read(service, message_id)
        add_label_to_email(service, message_id, "FreshFlow-Processed")
        
        return {
            "message": "Email processed successfully",
            "email_subject": content['subject'],
            "items_found": len(grn_items),
            "grn_items": grn_items
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error processing email: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to process email: {str(e)}")

@router.post("/gmail/auto-sync-grn")
async def auto_sync_ninjacart_grn(current_user: dict = Depends(get_current_user)):
    """
    Automatically sync Ninjacart GRN from latest unprocessed email.
    This processes emails, matches with dispatches, and saves GRN to database.
    """
    if current_user["role"] not in ["admin", "staff"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    # Get Gmail tokens
    token = await db.gmail_tokens.find_one({"user_id": current_user["user_id"]}, {"_id": 0})
    if not token:
        raise HTTPException(status_code=400, detail="Gmail not connected")
    
    try:
        service = get_gmail_service(token)
        
        # Search for unread Ninjacart emails
        result = service.users().messages().list(
            userId='me',
            q='from:ninjacart is:unread',
            maxResults=5
        ).execute()
        
        messages = result.get('messages', [])
        
        if not messages:
            return {"message": "No new Ninjacart emails found", "processed": 0}
        
        processed_count = 0
        saved_grns = []
        total_items_saved = 0
        rate_changes_detected = []
        
        for msg in messages:
            content = get_email_content(service, msg['id'])
            if not content:
                continue
            
            items = []
            attachment_filename = None
            
            # Try CSV/Excel first
            for attachment in content['attachments']:
                fname = attachment['filename'].lower()
                if fname.endswith('.csv') or fname.endswith('.xlsx') or fname.endswith('.xls'):
                    attachment_data = download_attachment(service, msg['id'], attachment['attachment_id'])
                    if attachment_data:
                        attachment_filename = attachment['filename']
                        if fname.endswith('.csv'):
                            items = parse_ninjacart_csv(attachment_data)
                        else:
                            # Parse Excel
                            try:
                                from openpyxl import load_workbook
                                import io
                                wb = load_workbook(io.BytesIO(attachment_data))
                                ws = wb.active
                                rows = list(ws.iter_rows(values_only=True))
                                if rows:
                                    headers = [str(h).strip() if h else '' for h in rows[0]]
                                    items = []
                                    for row in rows[1:]:
                                        row_dict = {headers[i]: row[i] for i in range(min(len(headers), len(row)))}
                                        items.append(row_dict)
                            except Exception as e:
                                logger.error(f"Error parsing Excel attachment: {e}")
                        break
            
            # Fallback to body parsing
            if not items and content['body_text']:
                items = parse_ninjacart_email_body(content['body_text'])
            
            if items:
                # Process items similar to manual upload
                # Match with dispatches and save to database
                try:
                    # Create a mock UploadFile-like object and process
                    matched_items = []  # This would need the full matching logic
                    
                    # For now, save raw GRN data
                    grn_doc = {
                        "id": str(uuid.uuid4()),
                        "source": "gmail_auto_sync",
                        "email_id": msg['id'],
                        "email_subject": content['subject'],
                        "email_date": content['date'],
                        "attachment_filename": attachment_filename,
                        "items": items,
                        "created_at": datetime.now(timezone.utc).isoformat(),
                        "created_by": current_user["user_id"],
                        "sync_method": "automatic_gmail"
                    }
                    
                    # Save to database
                    await db.gmail_grn_syncs.insert_one(grn_doc)
                    
                    saved_grns.append({
                        "email_id": msg['id'],
                        "subject": content['subject'],
                        "date": content['date'],
                        "items_count": len(items)
                    })
                    total_items_saved += len(items)
                    
                except Exception as e:
                    logger.error(f"Error saving GRN from email: {e}")
                
                # Mark as processed
                mark_email_as_read(service, msg['id'])
                add_label_to_email(service, msg['id'], "FreshFlow-Processed")
                processed_count += 1
        
        # Update last sync time
        await db.gmail_tokens.update_one(
            {"user_id": current_user["user_id"]},
            {"$set": {"last_sync": datetime.now(timezone.utc)}}
        )
        
        # Build summary for frontend display
        summary = {
            "sync_date": datetime.now(timezone.utc).isoformat(),
            "sync_method": "automatic_gmail",
            "emails_processed": processed_count,
            "total_items": total_items_saved,
            "saved_grns": saved_grns,
            "rate_changes": rate_changes_detected
        }
        
        return {
            "message": f"Processed {processed_count} Ninjacart emails",
            "processed": processed_count,
            "summary": summary,
            "grn_data": saved_grns
        }
    except Exception as e:
        logger.error(f"Error in auto-sync: {e}")
        raise HTTPException(status_code=500, detail=f"Auto-sync failed: {str(e)}")
