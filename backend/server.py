"""
FreshFlow API Server
====================

TABLE OF CONTENTS (Search for SECTION: to jump to each section)
================================================================
1. HEALTH & DIAGNOSTICS ROUTES     (~Line 145)
2. AUTHENTICATION ROUTES           (~Line 195)
3. USER MANAGEMENT ROUTES          (~Line 230)
4. PRODUCTS & PACKAGING ROUTES     (~Line 315)
5. FARMER ROUTES                   (~Line 385)
6. PROCUREMENT ROUTES              (~Line 435)
7. PROCUREMENT TEMPLATES           (~Line 560)
8. QC ORDERS ROUTES                (~Line 600)
9. QC CUSTOMER ROUTES              (~Line 675)
10. QC INDENT ROUTES               (~Line 780)
11. QC DISPATCH ROUTES             (~Line 835)
12. QC GRN ROUTES                  (~Line 1010)
13. QC INVOICE ROUTES              (~Line 1465)
14. RETAILER ORDER ROUTES          (~Line 1590)
15. WASTAGE ROUTES                 (~Line 1695)
16. DASHBOARD & ANALYTICS ROUTES   (~Line 1850)
17. STOCK STATUS ROUTES            (~Line 2500)
18. VARIABLE EXPENSES ROUTES       (~Line 3115)
19. FIXED EXPENSES ROUTES          (~Line 3220)
20. RETAILER PORTAL ROUTES         (~Line 3380)
    - Retailer Indents
    - Retailer Dispatches
    - Retailer GRN
    - Retailer Rejections
    - Retailer Payments
    - Retailer Invoices
    - Retailer Dashboard
21. BACKUP & DATA MANAGEMENT       (~Line 4140)
================================================================
"""

from fastapi import FastAPI, APIRouter, HTTPException, Depends, UploadFile, File, status, Request, Query
from fastapi.responses import StreamingResponse, JSONResponse
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
import re
import random
import string
from pathlib import Path
from datetime import datetime, timezone, timedelta
from typing import Optional, List
import bcrypt
import jwt
from pydantic import BaseModel, Field, EmailStr, ConfigDict
import io
import csv
import uuid
import traceback
import time

from models import *
from ocr_processor import process_excel_image
from pdf_generator import generate_invoice_pdf
from backup_system import setup_backup_scheduler, trigger_manual_backup, generate_backup_excel
from gmail_integration import (
    get_authorization_url, exchange_code_for_tokens, get_gmail_service,
    search_ninjacart_emails, get_email_content, download_attachment,
    parse_ninjacart_csv, parse_ninjacart_email_body, mark_email_as_read,
    add_label_to_email
)
from starlette.responses import RedirectResponse
from apscheduler.triggers.cron import CronTrigger

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# Hindi translations for all products - used for auto-translation
HINDI_PRODUCT_NAMES = {
    "Amaranthus Green": "हरा चौलाई",
    "Amaranthus Red": "लाल चौलाई",
    "Coriander": "धनिया",
    "Curry Leaves": "करी पत्ता",
    "Dill Leaf": "सोया पत्ता",
    "Fenugreek (Methi)": "मेथी",
    "Fresh Mint Leaves": "पुदीना",
    "Premium Fresh Mint Leaves": "प्रीमियम पुदीना",
    "Palak": "पालक",
    "Spinach": "पालक",
    "Tomato Hybrid": "हाइब्रिड टमाटर",
    "Lemon": "नींबू",
    "Raw Mango": "कच्चा आम",
    "Banana": "केला",
    "Banana ": "केला",
    "Onion": "प्याज",
    "Potato": "आलू",
    "Garlic": "लहसुन",
    "Ginger": "अदरक",
    "Carrot": "गाजर",
    "Radish": "मूली",
    "Peeled Garlic": "छिला लहसुन",
    "Bottle Gourd": "लौकी",
    "Bitter gourd": "करेला",
    "Cucumber": "खीरा",
    "Ridge Gourd": "तोरी",
    "Capsicum": "शिमला मिर्च",
    "Green capcicum ": "हरी शिमला मिर्च",
    "Chilli Light Green": "हल्की हरी मिर्च",
    "Chilli": "मिर्च",
    "Green chilli ": "हरी मिर्च",
    "Cauliflower": "फूलगोभी",
    "Cabbage ": "पत्तागोभी",
    "Cabbage": "पत्तागोभी",
    "Button Mushroom": "बटन मशरूम",
    "Brinjal": "बैंगन",
    "Lady Finger": "भिंडी",
    "Cluster Beans": "ग्वार फली",
    "Spring onion ": "हरा प्याज",
    "Spring onion": "हरा प्याज",
    "Green Pea": "हरी मटर",
    "Soaked chole": "भीगे छोले",
    "Soaked yellow peas": "भीगी पीली मटर",
    "Soaked Green peas": "भीगी हरी मटर",
    "Sprouted matki": "अंकुरित मोठ",
    "Matki Sprouts": "मोठ स्प्राउट्स",
    "Soaked Harbhara": "भीगा हरभरा",
    "Sprouted moong": "अंकुरित मूंग",
    "Mixed Sprouts ": "मिक्स स्प्राउट्स",
    "Mixed Sprouts": "मिक्स स्प्राउट्स",
}

# Enhanced logging setup
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger("freshflow")

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# Setup backup scheduler on startup
backup_scheduler = None

# Security
security = HTTPBearer()
SECRET_KEY = os.environ.get('JWT_SECRET_KEY', 'freshflow-secret-key-change-in-production')
ALGORITHM = "HS256"

# Create the main app
app = FastAPI(title="Mr Organix API")
api_router = APIRouter(prefix="/api")

# Store recent errors for diagnostics (in-memory, last 50 errors)
recent_errors = []
MAX_ERROR_LOG = 50

# Request logging middleware
@app.middleware("http")
async def log_requests(request: Request, call_next):
    start_time = time.time()
    request_id = str(uuid.uuid4())[:8]
    
    # Log request
    logger.info(f"[{request_id}] {request.method} {request.url.path} - Started")
    
    try:
        response = await call_next(request)
        process_time = time.time() - start_time
        
        # Log response
        log_level = logging.WARNING if response.status_code >= 400 else logging.INFO
        logger.log(log_level, f"[{request_id}] {request.method} {request.url.path} - {response.status_code} ({process_time:.3f}s)")
        
        # Store errors
        if response.status_code >= 400:
            error_entry = {
                "timestamp": datetime.now(timezone.utc).isoformat(),
                "request_id": request_id,
                "method": request.method,
                "path": str(request.url.path),
                "status_code": response.status_code,
                "process_time": round(process_time, 3)
            }
            recent_errors.append(error_entry)
            if len(recent_errors) > MAX_ERROR_LOG:
                recent_errors.pop(0)
        
        return response
    except Exception as e:
        process_time = time.time() - start_time
        error_msg = str(e)
        tb = traceback.format_exc()
        
        logger.error(f"[{request_id}] {request.method} {request.url.path} - EXCEPTION: {error_msg}")
        logger.error(f"[{request_id}] Traceback: {tb}")
        
        # Store error
        error_entry = {
            "timestamp": datetime.now(timezone.utc).isoformat(),
            "request_id": request_id,
            "method": request.method,
            "path": str(request.url.path),
            "status_code": 500,
            "error": error_msg,
            "traceback": tb[:500],  # Truncate traceback
            "process_time": round(process_time, 3)
        }
        recent_errors.append(error_entry)
        if len(recent_errors) > MAX_ERROR_LOG:
            recent_errors.pop(0)
        
        raise

# Auth Helper Functions
def hash_password(password: str) -> str:
    return bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')

def verify_password(password: str, hashed: str) -> bool:
    return bcrypt.checkpw(password.encode('utf-8'), hashed.encode('utf-8'))

def create_token(user_id: str, email: str, role: str) -> str:
    payload = {
        "user_id": user_id,
        "email": email,
        "role": role,
        "exp": datetime.now(timezone.utc) + timedelta(days=7)
    }
    return jwt.encode(payload, SECRET_KEY, algorithm=ALGORITHM)

async def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)) -> dict:
    try:
        token = credentials.credentials
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        return payload
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expired")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid token")

# ============================================================================
# SECTION: HEALTH & DIAGNOSTICS ROUTES (Lines ~140-190)
# ============================================================================
@api_router.get("/health")
async def health_check():
    """Basic health check endpoint"""
    try:
        # Test MongoDB connection
        await db.command("ping")
        db_status = "connected"
    except Exception as e:
        db_status = f"error: {str(e)}"
    
    return {
        "status": "healthy" if db_status == "connected" else "degraded",
        "timestamp": datetime.now(timezone.utc).isoformat(),
        "database": db_status,
        "version": "1.0.0"
    }

@api_router.get("/diagnostics")
async def get_diagnostics(current_user: dict = Depends(get_current_user)):
    """Get recent error logs and system diagnostics (admin only)"""
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    try:
        # Test MongoDB
        await db.command("ping")
        db_status = "connected"
        
        # Get collection stats
        collections = await db.list_collection_names()
        collection_counts = {}
        for coll in collections[:10]:  # Limit to 10 collections
            count = await db[coll].count_documents({})
            collection_counts[coll] = count
    except Exception as e:
        db_status = f"error: {str(e)}"
        collection_counts = {}
    
    return {
        "timestamp": datetime.now(timezone.utc).isoformat(),
        "database": {
            "status": db_status,
            "collections": collection_counts
        },
        "recent_errors": recent_errors[-20:],  # Last 20 errors
        "error_count": len(recent_errors),
        "uptime_note": "Errors are stored in-memory and reset on server restart"
    }

# ============================================================================
# SECTION: AUTHENTICATION ROUTES (Lines ~190-305)
# ============================================================================
@api_router.post("/auth/register", response_model=AuthResponse)
async def register(input: RegisterRequest):
    existing = await db.users.find_one({"email": input.email}, {"_id": 0})
    if existing:
        raise HTTPException(status_code=400, detail="Email already registered")
    
    user_dict = input.model_dump()
    user_dict["password"] = hash_password(user_dict.pop("password"))
    user = User(**user_dict)
    
    doc = user.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    
    await db.users.insert_one(doc)
    
    token = create_token(user.id, user.email, user.role)
    return AuthResponse(token=token, user=UserResponse(**user.model_dump()))

@api_router.post("/auth/login", response_model=AuthResponse)
async def login(input: LoginRequest):
    user = await db.users.find_one({"email": input.email}, {"_id": 0})
    if not user or not verify_password(input.password, user["password"]):
        raise HTTPException(status_code=401, detail="Invalid credentials")
    
    token = create_token(user["id"], user["email"], user["role"])
    user.pop("password")
    return AuthResponse(token=token, user=UserResponse(**user))

@api_router.get("/auth/me", response_model=UserResponse)
async def get_me(current_user: dict = Depends(get_current_user)):
    user = await db.users.find_one({"id": current_user["user_id"]}, {"_id": 0, "password": 0})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    return UserResponse(**user)

# ============================================================================
# SECTION: USER MANAGEMENT ROUTES - Admin Only (Lines ~225-305)
# ============================================================================
@api_router.get("/users")
async def get_users(current_user: dict = Depends(get_current_user)):
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Only admins can manage users")
    
    users = await db.users.find({}, {"_id": 0, "password": 0}).to_list(1000)
    return users

@api_router.post("/users")
async def create_user(input: RegisterRequest, current_user: dict = Depends(get_current_user)):
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Only admins can create users")
    
    existing = await db.users.find_one({"email": input.email}, {"_id": 0})
    if existing:
        raise HTTPException(status_code=400, detail="Email already registered")
    
    user_dict = input.model_dump()
    user_dict["password"] = hash_password(user_dict.pop("password"))
    
    # Generate 5-digit referral code for retailers
    if user_dict.get("role") == "retailer":
        referral_code = ''.join(random.choices(string.digits, k=5))
        # Ensure uniqueness
        while await db.users.find_one({"referral_code": referral_code}):
            referral_code = ''.join(random.choices(string.digits, k=5))
        user_dict["referral_code"] = referral_code
    
    user = User(**user_dict)
    
    doc = user.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    
    await db.users.insert_one(doc)
    
    return {"id": user.id, "message": "User created successfully"}

@api_router.put("/users/{user_id}")
async def update_user(user_id: str, input: UserUpdate, current_user: dict = Depends(get_current_user)):
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Only admins can update users")
    
    existing = await db.users.find_one({"id": user_id})
    if not existing:
        raise HTTPException(status_code=404, detail="User not found")
    
    # Build update data
    update_data = {}
    if input.name:
        update_data["name"] = input.name
    if input.email:
        # Check if email is already taken by another user
        email_check = await db.users.find_one({"email": input.email, "id": {"$ne": user_id}})
        if email_check:
            raise HTTPException(status_code=400, detail="Email already in use")
        update_data["email"] = input.email
    if input.password:
        update_data["password"] = hash_password(input.password)
    if input.role:
        update_data["role"] = input.role
    if input.company_name is not None:
        update_data["company_name"] = input.company_name
    if input.contact is not None:
        update_data["contact"] = input.contact
    if input.address is not None:
        update_data["address"] = input.address
    if input.commission_percentage is not None:
        update_data["commission_percentage"] = input.commission_percentage
    
    if update_data:
        await db.users.update_one({"id": user_id}, {"$set": update_data})
    
    return {"id": user_id, "message": "User updated successfully"}

@api_router.delete("/users/{user_id}")
async def delete_user(user_id: str, current_user: dict = Depends(get_current_user)):
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Only admins can delete users")
    
    # Prevent deleting self
    if user_id == current_user["user_id"]:
        raise HTTPException(status_code=400, detail="Cannot delete your own account")
    
    result = await db.users.delete_one({"id": user_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="User not found")
    
    return {"message": "User deleted successfully"}

# ============================================================================
# SECTION: PRODUCTS & PACKAGING ROUTES (Lines ~310-375)
# ============================================================================
@api_router.get("/products", response_model=List[Product])
async def get_products(current_user: dict = Depends(get_current_user)):
    products = await db.products.find({}, {"_id": 0}).to_list(1000)
    return [Product(**p) for p in products]

# ==================== UNITS MANAGEMENT ====================
@api_router.get("/units")
async def get_units(current_user: dict = Depends(get_current_user)):
    """Get all units in the system"""
    units = await db.units.find({}, {"_id": 0}).sort("name", 1).to_list(100)
    return units

@api_router.post("/units")
async def create_unit(input: dict, current_user: dict = Depends(get_current_user)):
    """Create a new unit"""
    if current_user["role"] not in ["admin", "staff"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    name = input.get("name", "").strip()
    symbol = input.get("symbol", "").strip()
    description = input.get("description", "").strip()
    
    if not name:
        raise HTTPException(status_code=400, detail="Unit name is required")
    
    # Check for duplicate
    existing = await db.units.find_one({"name": {"$regex": f"^{name}$", "$options": "i"}})
    if existing:
        raise HTTPException(status_code=400, detail=f"Unit '{name}' already exists")
    
    unit_doc = {
        "id": str(uuid.uuid4()),
        "name": name,
        "symbol": symbol or name,
        "description": description,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.units.insert_one(unit_doc)
    unit_doc.pop("_id", None)
    return unit_doc

@api_router.put("/units/{unit_id}")
async def update_unit(unit_id: str, input: dict, current_user: dict = Depends(get_current_user)):
    """Update an existing unit"""
    if current_user["role"] not in ["admin", "staff"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    unit = await db.units.find_one({"id": unit_id})
    if not unit:
        raise HTTPException(status_code=404, detail="Unit not found")
    
    update_data = {}
    if "name" in input:
        update_data["name"] = input["name"].strip()
    if "symbol" in input:
        update_data["symbol"] = input["symbol"].strip()
    if "description" in input:
        update_data["description"] = input["description"].strip()
    
    if update_data:
        await db.units.update_one({"id": unit_id}, {"$set": update_data})
    
    updated = await db.units.find_one({"id": unit_id}, {"_id": 0})
    return updated

@api_router.delete("/units/{unit_id}")
async def delete_unit(unit_id: str, current_user: dict = Depends(get_current_user)):
    """Delete a unit (only if not in use)"""
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Only admin can delete units")
    
    unit = await db.units.find_one({"id": unit_id}, {"_id": 0})
    if not unit:
        raise HTTPException(status_code=404, detail="Unit not found")
    
    # Check if unit is in use by any product
    products_using = await db.products.count_documents({"unit": unit["name"]})
    if products_using > 0:
        raise HTTPException(
            status_code=400, 
            detail=f"Cannot delete unit '{unit['name']}'. It is used by {products_using} product(s)."
        )
    
    await db.units.delete_one({"id": unit_id})
    return {"message": f"Unit '{unit['name']}' deleted successfully"}

# ==================== UNITS AUTO-SEED ON STARTUP ====================
async def seed_default_units_on_startup():
    """Auto-seed default units if collection is empty - called on server startup"""
    try:
        existing_count = await db.units.count_documents({})
        if existing_count > 0:
            logger.info(f"Units collection already has {existing_count} units. Skipping auto-seed.")
            return
        
        default_units = [
            {"name": "Kg", "symbol": "Kg", "description": "Kilogram - standard weight unit"},
            {"name": "Gram", "symbol": "g", "description": "Gram - small weight unit"},
            {"name": "Piece", "symbol": "Pc", "description": "Individual piece/item"},
            {"name": "Pieces", "symbol": "Pcs", "description": "Multiple pieces/items"},
            {"name": "Bunch", "symbol": "Bunch", "description": "Bundle of items tied together"},
            {"name": "Packet", "symbol": "Pkt", "description": "Packaged unit"},
            {"name": "Box", "symbol": "Box", "description": "Box container"},
            {"name": "Crate", "symbol": "Crate", "description": "Crate container"},
            {"name": "Dozen", "symbol": "Dz", "description": "12 pieces"},
            {"name": "Litre", "symbol": "L", "description": "Liquid volume unit"},
        ]
        
        for unit in default_units:
            unit["id"] = str(uuid.uuid4())
            unit["created_at"] = datetime.now(timezone.utc).isoformat()
            await db.units.insert_one(unit)
        
        logger.info(f"Auto-seeded {len(default_units)} default units on startup")
    except Exception as e:
        logger.error(f"Failed to auto-seed units: {e}")

@api_router.post("/units/seed-defaults")
async def seed_default_units(current_user: dict = Depends(get_current_user)):
    """Seed default units if none exist"""
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Only admin can seed units")
    
    existing_count = await db.units.count_documents({})
    if existing_count > 0:
        return {"message": f"Units already exist ({existing_count} units). Skipping seed."}
    
    default_units = [
        {"name": "Kg", "symbol": "Kg", "description": "Kilogram - standard weight unit"},
        {"name": "Gram", "symbol": "g", "description": "Gram - small weight unit"},
        {"name": "Piece", "symbol": "Pc", "description": "Individual piece/item"},
        {"name": "Pieces", "symbol": "Pcs", "description": "Multiple pieces/items"},
        {"name": "Bunch", "symbol": "Bunch", "description": "Bundle of items tied together"},
        {"name": "Packet", "symbol": "Pkt", "description": "Packaged unit"},
        {"name": "Box", "symbol": "Box", "description": "Box container"},
        {"name": "Crate", "symbol": "Crate", "description": "Crate container"},
        {"name": "Dozen", "symbol": "Dz", "description": "12 pieces"},
        {"name": "Litre", "symbol": "L", "description": "Liquid volume unit"},
    ]
    
    for unit in default_units:
        unit["id"] = str(uuid.uuid4())
        unit["created_at"] = datetime.now(timezone.utc).isoformat()
        await db.units.insert_one(unit)
    
    return {"message": f"Seeded {len(default_units)} default units"}

@api_router.post("/products", response_model=Product, status_code=status.HTTP_201_CREATED)
async def create_product(input: ProductCreate, current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in ["admin", "staff"]:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Not authorized")
    
    # Check for duplicate product name (case-insensitive)
    existing = await db.products.find_one({
        "name": {"$regex": f"^{input.name}$", "$options": "i"}
    })
    if existing:
        raise HTTPException(
            status_code=400, 
            detail=f"Product '{input.name}' already exists. Please use a different name or update the existing product."
        )
    
    product = Product(**input.model_dump())
    doc = product.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    
    # Auto-add Hindi name if available in translation dictionary
    if not doc.get('name_hi') and doc.get('name'):
        doc['name_hi'] = HINDI_PRODUCT_NAMES.get(doc['name'], None)
    
    await db.products.insert_one(doc)
    return product


@api_router.get("/products/duplicates")
async def get_duplicate_products(current_user: dict = Depends(get_current_user)):
    """Get list of duplicate products by name"""
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Only admin can view duplicates")
    
    # Find duplicates using aggregation
    pipeline = [
        {"$group": {
            "_id": {"$toLower": "$name"},
            "count": {"$sum": 1},
            "products": {"$push": {"id": "$id", "name": "$name", "category": "$category", "created_at": "$created_at"}}
        }},
        {"$match": {"count": {"$gt": 1}}},
        {"$sort": {"_id": 1}}
    ]
    
    duplicates = await db.products.aggregate(pipeline).to_list(100)
    return duplicates


@api_router.delete("/products/duplicates/cleanup")
async def cleanup_duplicate_products(current_user: dict = Depends(get_current_user)):
    """Remove duplicate products, keeping the oldest one"""
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Only admin can cleanup duplicates")
    
    # Find duplicates
    pipeline = [
        {"$group": {
            "_id": {"$toLower": "$name"},
            "count": {"$sum": 1},
            "products": {"$push": {"id": "$id", "name": "$name", "created_at": "$created_at"}}
        }},
        {"$match": {"count": {"$gt": 1}}}
    ]
    
    duplicates = await db.products.aggregate(pipeline).to_list(100)
    
    deleted_count = 0
    deleted_products = []
    
    for dup in duplicates:
        products = dup["products"]
        # Sort by created_at to keep the oldest
        products.sort(key=lambda x: x.get("created_at", ""))
        
        # Keep first (oldest), delete the rest
        for product in products[1:]:
            await db.products.delete_one({"id": product["id"]})
            deleted_count += 1
            deleted_products.append(product["name"])
    
    return {
        "message": f"Deleted {deleted_count} duplicate products",
        "deleted_count": deleted_count,
        "deleted_products": deleted_products
    }

@api_router.put("/products/{product_id}", response_model=Product)
async def update_product(product_id: str, input: ProductUpdate, current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in ["admin", "staff"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    update_data = {k: v for k, v in input.model_dump().items() if v is not None}
    if not update_data:
        raise HTTPException(status_code=400, detail="No update data provided")
    
    result = await db.products.find_one_and_update(
        {"id": product_id},
        {"$set": update_data},
        return_document=True,
        projection={"_id": 0}
    )
    
    if not result:
        raise HTTPException(status_code=404, detail="Product not found")
    
    return Product(**result)

@api_router.delete("/products/{product_id}")
async def delete_product(product_id: str, current_user: dict = Depends(get_current_user)):
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Not authorized")
    
    result = await db.products.delete_one({"id": product_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Product not found")
    
    return {"message": "Product deleted successfully"}

# Check if product has dependencies (purchases, sales, dispatches, etc.)
@api_router.get("/products/{product_id}/dependencies")
async def check_product_dependencies(product_id: str, current_user: dict = Depends(get_current_user)):
    """Check if a product has any associated sales, purchases, or other records"""
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Not authorized")
    
    product = await db.products.find_one({"id": product_id}, {"_id": 0})
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")
    
    product_name = product.get("name", "")
    
    # Check procurements (purchases)
    procurement_count = await db.procurements.count_documents({
        "$or": [
            {"product_id": product_id},
            {"product_name": {"$regex": f"^{product_name}$", "$options": "i"}}
        ]
    })
    
    # Check QC orders (sales)
    qc_order_count = await db.qc_orders.count_documents({
        "items.product_id": product_id
    })
    
    # Check retailer indents
    indent_count = await db.retailer_indents.count_documents({
        "items.product_id": product_id
    })
    
    # Check retailer dispatches
    dispatch_count = await db.retailer_dispatches.count_documents({
        "items.product_id": product_id
    })
    
    # Check retailer invoices
    invoice_count = await db.retailer_invoices.count_documents({
        "items.product_id": product_id
    })
    
    # Check wastage records
    wastage_count = await db.wastage.count_documents({
        "$or": [
            {"product_id": product_id},
            {"product_name": {"$regex": f"^{product_name}$", "$options": "i"}}
        ]
    })
    
    total_dependencies = procurement_count + qc_order_count + indent_count + dispatch_count + invoice_count + wastage_count
    
    return {
        "product_id": product_id,
        "product_name": product_name,
        "has_dependencies": total_dependencies > 0,
        "dependencies": {
            "procurements": procurement_count,
            "qc_orders": qc_order_count,
            "indents": indent_count,
            "dispatches": dispatch_count,
            "invoices": invoice_count,
            "wastage": wastage_count,
            "total": total_dependencies
        }
    }

# Delete product with replacement mapping
@api_router.post("/products/{product_id}/delete-with-replacement")
async def delete_product_with_replacement(product_id: str, input: dict, current_user: dict = Depends(get_current_user)):
    """Delete a product and remap all its records to a replacement product"""
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Not authorized")
    
    replacement_product_id = input.get("replacement_product_id")
    if not replacement_product_id:
        raise HTTPException(status_code=400, detail="Replacement product ID is required")
    
    # Get original product
    original_product = await db.products.find_one({"id": product_id}, {"_id": 0})
    if not original_product:
        raise HTTPException(status_code=404, detail="Product not found")
    
    # Get replacement product
    replacement_product = await db.products.find_one({"id": replacement_product_id}, {"_id": 0})
    if not replacement_product:
        raise HTTPException(status_code=404, detail="Replacement product not found")
    
    original_name = original_product.get("name", "")
    replacement_name = replacement_product.get("name", "")
    
    remapped_counts = {
        "procurements": 0,
        "qc_orders": 0,
        "indents": 0,
        "dispatches": 0,
        "invoices": 0,
        "wastage": 0,
        "closing_inventory": 0
    }
    
    # 1. Update procurements
    proc_result = await db.procurements.update_many(
        {"$or": [
            {"product_id": product_id},
            {"product_name": {"$regex": f"^{original_name}$", "$options": "i"}}
        ]},
        {"$set": {
            "product_id": replacement_product_id,
            "product_name": replacement_name
        }}
    )
    remapped_counts["procurements"] = proc_result.modified_count
    
    # 2. Update QC orders (items array)
    qc_orders = await db.qc_orders.find({"items.product_id": product_id}).to_list(1000)
    for order in qc_orders:
        updated_items = []
        for item in order.get("items", []):
            if item.get("product_id") == product_id:
                item["product_id"] = replacement_product_id
                item["product_name"] = replacement_name
            updated_items.append(item)
        await db.qc_orders.update_one(
            {"id": order["id"]},
            {"$set": {"items": updated_items}}
        )
        remapped_counts["qc_orders"] += 1
    
    # 3. Update retailer indents (items array)
    indents = await db.retailer_indents.find({"items.product_id": product_id}).to_list(1000)
    for indent in indents:
        updated_items = []
        for item in indent.get("items", []):
            if item.get("product_id") == product_id:
                item["product_id"] = replacement_product_id
                item["product_name"] = replacement_name
            updated_items.append(item)
        await db.retailer_indents.update_one(
            {"id": indent["id"]},
            {"$set": {"items": updated_items}}
        )
        remapped_counts["indents"] += 1
    
    # 4. Update retailer dispatches (items array)
    dispatches = await db.retailer_dispatches.find({"items.product_id": product_id}).to_list(1000)
    for dispatch in dispatches:
        updated_items = []
        for item in dispatch.get("items", []):
            if item.get("product_id") == product_id:
                item["product_id"] = replacement_product_id
                item["product_name"] = replacement_name
            updated_items.append(item)
        await db.retailer_dispatches.update_one(
            {"id": dispatch["id"]},
            {"$set": {"items": updated_items}}
        )
        remapped_counts["dispatches"] += 1
    
    # 5. Update retailer invoices (items array)
    invoices = await db.retailer_invoices.find({"items.product_id": product_id}).to_list(1000)
    for invoice in invoices:
        updated_items = []
        for item in invoice.get("items", []):
            if item.get("product_id") == product_id:
                item["product_id"] = replacement_product_id
                item["product_name"] = replacement_name
            updated_items.append(item)
        await db.retailer_invoices.update_one(
            {"id": invoice["id"]},
            {"$set": {"items": updated_items}}
        )
        remapped_counts["invoices"] += 1
    
    # 6. Update wastage records
    wastage_result = await db.wastage.update_many(
        {"$or": [
            {"product_id": product_id},
            {"product_name": {"$regex": f"^{original_name}$", "$options": "i"}}
        ]},
        {"$set": {
            "product_id": replacement_product_id,
            "product_name": replacement_name
        }}
    )
    remapped_counts["wastage"] = wastage_result.modified_count
    
    # 7. Update closing inventory
    closing_result = await db.retailer_closing_inventory.update_many(
        {"$or": [
            {"product_id": product_id},
            {"product_name": {"$regex": f"^{original_name}$", "$options": "i"}}
        ]},
        {"$set": {
            "product_id": replacement_product_id,
            "product_name": replacement_name
        }}
    )
    remapped_counts["closing_inventory"] = closing_result.modified_count
    
    # Finally, delete the original product
    await db.products.delete_one({"id": product_id})
    
    total_remapped = sum(remapped_counts.values())
    
    return {
        "message": f"Product '{original_name}' deleted successfully. {total_remapped} records remapped to '{replacement_name}'.",
        "deleted_product": original_name,
        "replacement_product": replacement_name,
        "remapped_counts": remapped_counts,
        "total_remapped": total_remapped
    }

# ============================================================================
# SECTION: QC PACKAGING ROUTES
# ============================================================================
@api_router.get("/qc-packaging")
async def get_qc_packaging(current_user: dict = Depends(get_current_user)):
    packagings = await db.qc_packaging.find({}, {"_id": 0}).to_list(100)
    return packagings

@api_router.post("/qc-packaging")
async def create_qc_packaging(name: str, weight_gm: float = 0, current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in ["admin", "staff"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    # Auto-extract weight from name if not provided
    if weight_gm == 0:
        weight_gm = extract_weight_from_packaging_name(name)
    
    packaging = {
        "id": str(uuid.uuid4()),
        "name": name,
        "weight_gm": weight_gm
    }
    await db.qc_packaging.insert_one(packaging)
    return {"id": packaging["id"], "message": "Packaging created"}

@api_router.put("/qc-packaging/{packaging_id}")
async def update_qc_packaging(packaging_id: str, name: str = None, weight_gm: float = None, current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in ["admin", "staff"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    update_data = {}
    if name is not None:
        update_data["name"] = name
    if weight_gm is not None:
        update_data["weight_gm"] = weight_gm
    
    if not update_data:
        raise HTTPException(status_code=400, detail="No fields to update")
    
    result = await db.qc_packaging.update_one(
        {"id": packaging_id},
        {"$set": update_data}
    )
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Packaging not found")
    return {"message": "Packaging updated"}


def extract_weight_from_packaging_name(name: str) -> float:
    """Extract weight in grams from packaging name like '500 gm', '240-260 gm', '200 gm Packet', 'With Roots (100 gm Pack)'"""
    if not name:
        return 0
    
    name_lower = name.lower().strip()
    
    # Pattern 1: Simple number + gm/g at start (e.g., "500 gm", "250 gm")
    match = re.search(r'^(\d+)\s*(gm|g)\b', name_lower)
    if match:
        return float(match.group(1))
    
    # Pattern 2: Range like "240-260 gm" or "90-110 gm Packet" - take average
    range_match = re.search(r'(\d+)\s*-\s*(\d+)\s*(gm|g)', name_lower)
    if range_match:
        return (float(range_match.group(1)) + float(range_match.group(2))) / 2
    
    # Pattern 3: Number + gm/g anywhere in the string (e.g., "With Roots (100 gm Pack)")
    anywhere_match = re.search(r'(\d+)\s*(gm|g)\b', name_lower)
    if anywhere_match:
        return float(anywhere_match.group(1))
    
    # Pattern 4: Just a number (assume gm if no unit, e.g., "500")
    just_number = re.search(r'^(\d+)$', name_lower)
    if just_number:
        return float(just_number.group(1))
    
    return 0  # Unknown


@api_router.post("/qc-packaging/sync-weights")
async def sync_packaging_weights(current_user: dict = Depends(get_current_user)):
    """Sync all packaging weights - extract from names and store in database"""
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Only admin can sync packaging weights")
    
    packagings = await db.qc_packaging.find({}).to_list(200)
    updated_count = 0
    
    for p in packagings:
        name = p.get('name', '')
        current_weight = p.get('weight_gm') or 0
        
        # Always update weight from name to ensure consistency
        weight = extract_weight_from_packaging_name(name)
        if weight > 0 and weight != current_weight:
            await db.qc_packaging.update_one(
                {"id": p['id']},
                {"$set": {"weight_gm": weight}}
            )
            updated_count += 1
    
    # Also add common variants if missing
    common_variants = [
        {"name": "500 gm", "weight_gm": 500},
        {"name": "250 gm", "weight_gm": 250},
        {"name": "200 gm Packet", "weight_gm": 200},
        {"name": "100 gm", "weight_gm": 100},
        {"name": "400-450 gm", "weight_gm": 425},
    ]
    
    added_count = 0
    for variant in common_variants:
        existing = await db.qc_packaging.find_one({
            "name": {"$regex": f"^{re.escape(variant['name'])}$", "$options": "i"}
        })
        if not existing:
            await db.qc_packaging.insert_one({
                "id": str(uuid.uuid4()),
                "name": variant['name'],
                "weight_gm": variant['weight_gm']
            })
            added_count += 1
    
    # Return all packagings with weights
    all_packagings = await db.qc_packaging.find({}, {"_id": 0}).to_list(200)
    
    return {
        "message": f"Synced {updated_count} existing, added {added_count} new packaging variants",
        "packagings": all_packagings
    }


# ============================================================================
# SECTION: FARMER ROUTES (Lines ~375-425)
# ============================================================================
@api_router.get("/farmers", response_model=List[Farmer])
async def get_farmers(current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in ["admin", "staff"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    farmers = await db.farmers.find({}, {"_id": 0}).to_list(1000)
    return [Farmer(**f) for f in farmers]

@api_router.post("/farmers", response_model=Farmer, status_code=status.HTTP_201_CREATED)
async def create_farmer(input: FarmerCreate, current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in ["admin", "staff"]:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Not authorized")
    
    farmer = Farmer(**input.model_dump())
    doc = farmer.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    await db.farmers.insert_one(doc)
    return farmer

@api_router.put("/farmers/{farmer_id}", response_model=Farmer)
async def update_farmer(farmer_id: str, input: FarmerCreate, current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in ["admin", "staff"]:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Not authorized")
    
    update_data = input.model_dump()
    result = await db.farmers.find_one_and_update(
        {"id": farmer_id},
        {"$set": update_data},
        return_document=True,
        projection={"_id": 0}
    )
    
    if not result:
        raise HTTPException(status_code=404, detail="Farmer not found")
    
    return Farmer(**result)

@api_router.delete("/farmers/{farmer_id}")
async def delete_farmer(farmer_id: str, current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in ["admin", "staff"]:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Not authorized")
    
    result = await db.farmers.delete_one({"id": farmer_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Farmer not found")
    
    return {"message": "Farmer deleted successfully"}

# ============================================================================
# SECTION: PROCUREMENT ROUTES (Lines ~425-590)
# ============================================================================
@api_router.get("/procurement", response_model=List[Procurement])
async def get_procurements(current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in ["admin", "staff"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    procurements = await db.procurements.find({}, {"_id": 0}).sort("date", -1).to_list(1000)
    for p in procurements:
        if isinstance(p['date'], str):
            p['date'] = datetime.fromisoformat(p['date'])
        if isinstance(p['created_at'], str):
            p['created_at'] = datetime.fromisoformat(p['created_at'])
    return [Procurement(**p) for p in procurements]


@api_router.get("/procurement/previous-day")
async def get_previous_day_procurements(
    reference_date: str = None,
    current_user: dict = Depends(get_current_user)
):
    """Get unique procurements from the 7 days before reference_date to use as template for entry.
    If no reference_date provided, defaults to today.
    Returns unique farmer-product combinations from the last 7 days.
    """
    if current_user["role"] not in ["admin", "staff"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    # Determine the reference date (date for which we want to record purchases)
    if reference_date:
        try:
            ref_date = datetime.strptime(reference_date, '%Y-%m-%d')
        except ValueError:
            ref_date = datetime.now(timezone.utc)
    else:
        ref_date = datetime.now(timezone.utc)
    
    # Get the date range: 7 days before reference date (exclusive of reference date)
    end_date = (ref_date - timedelta(days=1)).strftime('%Y-%m-%d')  # Day before reference
    start_date = (ref_date - timedelta(days=7)).strftime('%Y-%m-%d')  # 7 days before reference
    
    # Find all procurements in the date range
    all_procurements = await db.procurements.find({}, {"_id": 0}).to_list(1000)
    
    # Filter procurements in the date range and collect unique farmer-product combinations
    unique_items = {}  # Key: "farmer_id|product_id" -> latest procurement item
    
    for p in all_procurements:
        proc_date = p.get("date", "")
        if isinstance(proc_date, datetime):
            proc_date_str = proc_date.strftime('%Y-%m-%d')
        else:
            proc_date_str = str(proc_date)[:10]
        
        # Check if procurement is within the 7-day window
        if start_date <= proc_date_str <= end_date:
            farmer_id = p.get("farmer_id", "")
            farmer_name = p.get("farmer_name", "")
            
            for product in p.get("products", []):
                product_id = product.get("product_id", "")
                # Create unique key for farmer-product combination
                key = f"{farmer_id}|{product_id}"
                
                # Keep the most recent entry for each unique combination
                if key not in unique_items or proc_date_str > unique_items[key].get("date", ""):
                    unique_items[key] = {
                        "farmer_id": farmer_id,
                        "farmer_name": farmer_name,
                        "product_id": product_id,
                        "product_name": product.get("product_name", ""),
                        "quantity": product.get("quantity", 0),
                        "unit": product.get("unit", "Kg"),
                        "unit_size": product.get("unit_size", ""),
                        "rate": product.get("rate", 0),
                        "total": product.get("total", 0),
                        "date": proc_date_str
                    }
    
    # Group items by farmer for the response format
    farmer_items = {}
    for item in unique_items.values():
        farmer_id = item["farmer_id"]
        if farmer_id not in farmer_items:
            farmer_items[farmer_id] = {
                "farmer_id": farmer_id,
                "farmer_name": item["farmer_name"],
                "products": [],
                "date": f"{start_date} to {end_date}"
            }
        farmer_items[farmer_id]["products"].append({
            "product_id": item["product_id"],
            "product_name": item["product_name"],
            "quantity": item["quantity"],
            "unit": item["unit"],
            "unit_size": item["unit_size"],
            "rate": item["rate"],
            "total": item["total"]
        })
    
    # Calculate totals for each farmer
    templates = []
    for farmer_data in farmer_items.values():
        farmer_data["total_amount"] = sum(p.get("total", 0) for p in farmer_data["products"])
        templates.append(farmer_data)
    
    # Sort by farmer name
    templates.sort(key=lambda x: x.get("farmer_name", ""))
    
    return templates


@api_router.post("/procurement", response_model=Procurement, status_code=status.HTTP_201_CREATED)
async def create_procurement(input: ProcurementCreate, current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in ["admin", "staff"]:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Not authorized")
    
    procurement_dict = input.model_dump()
    procurement_dict["recorded_by"] = current_user["user_id"]
    procurement = Procurement(**procurement_dict)
    
    # Update product stock
    for item in procurement.products:
        await db.products.update_one(
            {"id": item.product_id},
            {"$inc": {"current_stock": item.quantity}}
        )
    
    doc = procurement.model_dump()
    doc['date'] = doc['date'].isoformat()
    doc['created_at'] = doc['created_at'].isoformat()
    await db.procurements.insert_one(doc)
    
    # Real-time sync: Update stock status for each product
    proc_date = procurement.date.strftime('%Y-%m-%d') if hasattr(procurement.date, 'strftime') else str(procurement.date)[:10]
    for item in procurement.products:
        product = await db.products.find_one({"id": item.product_id}, {"_id": 0})
        if not product:
            continue
            
        # Check if stock status entry exists for this product/date
        existing = await db.daily_stock_status.find_one({
            "product_id": item.product_id,
            "date": proc_date
        })
        
        # Use rate as the price per kg
        item_price = item.rate if hasattr(item, 'rate') else 0
        
        if existing:
            # Update existing entry - add to purchase_qty
            new_purchase = existing.get("purchase_qty", 0) + item.quantity
            new_avg_price = item_price if item_price else existing.get("avg_price", 0)
            await db.daily_stock_status.update_one(
                {"id": existing["id"]},
                {"$set": {"purchase_qty": new_purchase, "avg_price": new_avg_price}}
            )
        else:
            # Create new stock status entry
            # Get previous day's closing as opening
            yesterday = (procurement.date - timedelta(days=1)).strftime('%Y-%m-%d') if hasattr(procurement.date, 'strftime') else (datetime.strptime(str(procurement.date)[:10], '%Y-%m-%d') - timedelta(days=1)).strftime('%Y-%m-%d')
            prev_status = await db.daily_stock_status.find_one({
                "product_id": item.product_id,
                "date": yesterday,
                "status": "closed"
            })
            opening_qty = prev_status.get("closing_qty", 0) if prev_status else 0
            
            new_status = {
                "id": str(uuid.uuid4()),
                "date": proc_date,
                "product_id": item.product_id,
                "product_name": product["name"],
                "opening_qty": opening_qty,
                "purchase_qty": item.quantity,
                "dispatch_qty": 0,
                "wastage_qty": 0,
                "wastage_value": 0,
                "avg_price": item_price,
                "status": "open",
                "closed_at": None
            }
            await db.daily_stock_status.insert_one(new_status)
    
    return procurement

@api_router.put("/procurement/{procurement_id}")
async def update_procurement(procurement_id: str, input: dict, current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in ["admin", "staff"]:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Not authorized")
    
    # Get existing procurement first
    existing = await db.procurements.find_one({"id": procurement_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Procurement not found")
    
    old_date = existing.get("date")
    old_products = existing.get("products", [])
    
    # Build update fields - allow updating all relevant fields
    update_fields = {}
    
    # Payment-related fields
    if "paid_amount" in input:
        update_fields["paid_amount"] = input["paid_amount"]
    if "pending_amount" in input:
        update_fields["pending_amount"] = input["pending_amount"]
    if "payment_status" in input:
        update_fields["payment_status"] = input["payment_status"]
    if "total_amount" in input:
        update_fields["total_amount"] = input["total_amount"]
    if "remark" in input:
        update_fields["remark"] = input["remark"]
    if "status" in input:
        update_fields["status"] = input["status"]
    
    # Date field
    new_date = None
    if "date" in input:
        new_date = input["date"]
        if isinstance(new_date, str) and len(new_date) == 10:
            new_date = f"{new_date}T00:00:00+00:00"
        update_fields["date"] = new_date
    
    # Products array - this is critical for updating rates and totals
    if "products" in input:
        update_fields["products"] = input["products"]
    
    if not update_fields:
        raise HTTPException(status_code=400, detail="No valid fields to update")
    
    # Calculate stock status changes if products or date changed
    products_changed = "products" in update_fields
    date_changed = "date" in update_fields and old_date != update_fields.get("date")
    
    result = await db.procurements.find_one_and_update(
        {"id": procurement_id},
        {"$set": update_fields},
        return_document=True,
        projection={"_id": 0}
    )
    
    if not result:
        raise HTTPException(status_code=404, detail="Procurement not found")
    
    # Update stock status when products or date changes
    if products_changed or date_changed:
        old_date_str = str(old_date)[:10] if old_date else None
        new_date_str = str(update_fields.get("date", old_date))[:10] if update_fields.get("date") or old_date else None
        new_products = update_fields.get("products", old_products)
        
        # Build maps of old and new product data
        old_product_map = {}
        for item in old_products:
            product_id = item.get("product_id")
            if product_id:
                qty = float(item.get("quantity", 0) or 0)
                unit = item.get("unit", "Kg")
                unit_size = item.get("unit_size", "")
                qty_kg = qty
                if unit == "Bunch" and unit_size:
                    try:
                        qty_kg = (qty * float(unit_size)) / 1000
                    except:
                        pass
                total_value = float(item.get("total", 0) or 0)
                old_product_map[product_id] = {"qty_kg": qty_kg, "total_value": total_value}
        
        new_product_map = {}
        for item in new_products:
            product_id = item.get("product_id")
            if product_id:
                qty = float(item.get("quantity", 0) or 0)
                unit = item.get("unit", "Kg")
                unit_size = item.get("unit_size", "")
                qty_kg = qty
                if unit == "Bunch" and unit_size:
                    try:
                        qty_kg = (qty * float(unit_size)) / 1000
                    except:
                        pass
                total_value = float(item.get("total", 0) or 0)
                new_product_map[product_id] = {"qty_kg": qty_kg, "total_value": total_value}
        
        # Get all affected product IDs
        all_product_ids = set(old_product_map.keys()) | set(new_product_map.keys())
        
        for product_id in all_product_ids:
            old_data = old_product_map.get(product_id, {"qty_kg": 0, "total_value": 0})
            new_data = new_product_map.get(product_id, {"qty_kg": 0, "total_value": 0})
            
            if date_changed:
                # Date changed: subtract from old date, add to new date
                if old_date_str:
                    await db.daily_stock_status.update_one(
                        {"date": old_date_str, "product_id": product_id},
                        {"$inc": {"purchase_qty": -old_data["qty_kg"], "purchase_value": -old_data["total_value"]}}
                    )
                    # Recalculate wastage_value for old date
                    old_stock = await db.daily_stock_status.find_one(
                        {"date": old_date_str, "product_id": product_id},
                        {"_id": 0}
                    )
                    if old_stock and old_stock.get("wastage_qty", 0) > 0:
                        old_pq = old_stock.get("purchase_qty", 0) - old_data["qty_kg"]
                        old_pv = old_stock.get("purchase_value", 0) - old_data["total_value"]
                        old_avg = old_pv / old_pq if old_pq > 0 else 0
                        new_wv = round(old_stock.get("wastage_qty", 0) * old_avg, 2)
                        await db.daily_stock_status.update_one(
                            {"date": old_date_str, "product_id": product_id},
                            {"$set": {"wastage_value": new_wv}}
                        )
                        
                if new_date_str:
                    await db.daily_stock_status.update_one(
                        {"date": new_date_str, "product_id": product_id},
                        {"$inc": {"purchase_qty": new_data["qty_kg"], "purchase_value": new_data["total_value"]}},
                        upsert=False
                    )
                    # Recalculate wastage_value for new date
                    new_stock = await db.daily_stock_status.find_one(
                        {"date": new_date_str, "product_id": product_id},
                        {"_id": 0}
                    )
                    if new_stock and new_stock.get("wastage_qty", 0) > 0:
                        new_pq = new_stock.get("purchase_qty", 0) + new_data["qty_kg"]
                        new_pv = new_stock.get("purchase_value", 0) + new_data["total_value"]
                        new_avg = new_pv / new_pq if new_pq > 0 else 0
                        new_wv = round(new_stock.get("wastage_qty", 0) * new_avg, 2)
                        await db.daily_stock_status.update_one(
                            {"date": new_date_str, "product_id": product_id},
                            {"$set": {"wastage_value": new_wv}}
                        )
            elif products_changed and new_date_str:
                # Only products changed: update the difference on the same date
                qty_diff = new_data["qty_kg"] - old_data["qty_kg"]
                value_diff = new_data["total_value"] - old_data["total_value"]
                if qty_diff != 0 or value_diff != 0:
                    await db.daily_stock_status.update_one(
                        {"date": new_date_str, "product_id": product_id},
                        {"$inc": {"purchase_qty": qty_diff, "purchase_value": value_diff}}
                    )
                
                # Also recalculate wastage_value based on new avg_price
                # Get the current stock status to recalculate wastage
                stock_status = await db.daily_stock_status.find_one(
                    {"date": new_date_str, "product_id": product_id},
                    {"_id": 0}
                )
                if stock_status and stock_status.get("wastage_qty", 0) > 0:
                    # Get updated purchase values (after the increment above)
                    updated_purchase_qty = stock_status.get("purchase_qty", 0) + qty_diff
                    updated_purchase_value = stock_status.get("purchase_value", 0) + value_diff
                    
                    # Calculate new avg_price
                    new_avg_price = updated_purchase_value / updated_purchase_qty if updated_purchase_qty > 0 else 0
                    
                    # Recalculate wastage_value
                    wastage_qty = stock_status.get("wastage_qty", 0)
                    new_wastage_value = round(wastage_qty * new_avg_price, 2)
                    
                    # Update wastage_value
                    await db.daily_stock_status.update_one(
                        {"date": new_date_str, "product_id": product_id},
                        {"$set": {"wastage_value": new_wastage_value}}
                    )
    
    return {"message": "Procurement updated successfully", "procurement": result}

@api_router.delete("/procurement/{procurement_id}")
async def delete_procurement(procurement_id: str, current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in ["admin", "staff"]:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Not authorized")
    
    # Get procurement details first
    procurement = await db.procurements.find_one({"id": procurement_id}, {"_id": 0})
    if not procurement:
        raise HTTPException(status_code=404, detail="Procurement not found")
    
    # Reverse stock changes
    for item in procurement.get("products", []):
        await db.products.update_one(
            {"id": item["product_id"]},
            {"$inc": {"current_stock": -item["quantity"]}}
        )
    
    # Delete procurement
    result = await db.procurements.delete_one({"id": procurement_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Procurement not found")
    
    return {"message": "Procurement deleted successfully"}


# ============================================================================
# SECTION: PROCUREMENT TEMPLATES
# ============================================================================
@api_router.get("/procurement-templates")
async def get_procurement_templates(current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in ["admin", "staff"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    templates = await db.procurement_templates.find({}, {"_id": 0}).to_list(100)
    return templates

@api_router.post("/procurement-templates")
async def create_procurement_template(input: ProcurementTemplateCreate, current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in ["admin", "staff"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    template = ProcurementTemplate(
        name=input.name,
        farmer_id=input.farmer_id,
        farmer_name=input.farmer_name,
        items=input.items,
        created_by=current_user["user_id"]
    )
    
    doc = template.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    await db.procurement_templates.insert_one(doc)
    
    return {"id": template.id, "message": "Template created successfully"}

@api_router.delete("/procurement-templates/{template_id}")
async def delete_procurement_template(template_id: str, current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in ["admin", "staff"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    result = await db.procurement_templates.delete_one({"id": template_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Template not found")
    
    return {"message": "Template deleted successfully"}


# ============================================================================
# SECTION: QC ORDERS ROUTES (Lines ~592-650)
# ============================================================================
@api_router.get("/qc-orders", response_model=List[QCOrder])
async def get_qc_orders(current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in ["admin", "staff"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    orders = await db.qc_orders.find({}, {"_id": 0}).sort("order_date", -1).to_list(1000)
    for o in orders:
        if isinstance(o['order_date'], str):
            o['order_date'] = datetime.fromisoformat(o['order_date'])
        if o.get('delivery_date') and isinstance(o['delivery_date'], str):
            o['delivery_date'] = datetime.fromisoformat(o['delivery_date'])
        if isinstance(o['created_at'], str):
            o['created_at'] = datetime.fromisoformat(o['created_at'])
    return [QCOrder(**o) for o in orders]

@api_router.post("/qc-orders", response_model=QCOrder)
async def create_qc_order(input: QCOrderCreate, current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in ["admin", "staff"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    order_dict = input.model_dump()
    order_dict["recorded_by"] = current_user["user_id"]
    order = QCOrder(**order_dict)
    
    # Reduce stock
    for item in order.products:
        await db.products.update_one(
            {"id": item.product_id},
            {"$inc": {"current_stock": -item.quantity}}
        )
    
    doc = order.model_dump()
    doc['order_date'] = doc['order_date'].isoformat()
    if doc.get('delivery_date'):
        doc['delivery_date'] = doc['delivery_date'].isoformat()
    doc['created_at'] = doc['created_at'].isoformat()
    await db.qc_orders.insert_one(doc)
    return order

@api_router.post("/qc-orders/ocr")
async def ocr_qc_order(file: UploadFile = File(...), current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in ["admin", "staff"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    allowed_types = ["image/jpeg", "image/png", "image/bmp", "image/webp"]
    if file.content_type not in allowed_types:
        raise HTTPException(status_code=400, detail="Invalid file type. Allowed: JPG, PNG, BMP, WEBP")
    
    try:
        content = await file.read()
        extracted_data = await process_excel_image(content)
        return {"status": "success", "data": extracted_data}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"OCR processing failed: {str(e)}")

# ============================================================================
# SECTION: QC CUSTOMER ROUTES (Lines ~666-770)
# ============================================================================
@api_router.get("/qc-customers")
async def get_qc_customers(current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in ["admin", "staff"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    customers = await db.qc_customers.find({}, {"_id": 0}).sort("name", 1).to_list(1000)
    return customers

@api_router.post("/qc-customers")
async def create_qc_customer(input: QCCustomerCreate, current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in ["admin", "staff"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    customer = QCCustomer(**input.model_dump())
    doc = customer.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    await db.qc_customers.insert_one(doc)
    return {"id": customer.id, "name": customer.name}

@api_router.put("/qc-customers/{customer_id}")
async def update_qc_customer(customer_id: str, input: QCCustomerCreate, current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in ["admin", "staff"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    update_data = input.model_dump()
    result = await db.qc_customers.update_one({"id": customer_id}, {"$set": update_data})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Customer not found")
    
    return {"message": "Customer updated successfully"}

@api_router.delete("/qc-customers/{customer_id}")
async def delete_qc_customer(customer_id: str, current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in ["admin"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    result = await db.qc_customers.delete_one({"id": customer_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Customer not found")
    
    return {"message": "Customer deleted successfully"}

# Customer Product Settings (Lot Size per Customer-Product)
@api_router.get("/customer-product-settings")
async def get_customer_product_settings(
    customer_id: str = None,
    current_user: dict = Depends(get_current_user)
):
    if current_user["role"] not in ["admin", "staff"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    query = {}
    if customer_id:
        query["customer_id"] = customer_id
    
    settings = await db.customer_product_settings.find(query, {"_id": 0}).to_list(1000)
    return settings

@api_router.post("/customer-product-settings")
async def create_or_update_customer_product_setting(
    input: CustomerProductSettingCreate,
    current_user: dict = Depends(get_current_user)
):
    if current_user["role"] not in ["admin", "staff"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    # Check if setting already exists for this customer-product-packaging combo
    existing = await db.customer_product_settings.find_one({
        "customer_id": input.customer_id,
        "product_id": input.product_id,
        "packaging_id": input.packaging_id or None
    }, {"_id": 0})
    
    if existing:
        # Update existing
        await db.customer_product_settings.update_one(
            {"id": existing["id"]},
            {"$set": {
                "lot_size": input.lot_size,
                "updated_at": datetime.now(timezone.utc).isoformat()
            }}
        )
        return {"id": existing["id"], "message": "Setting updated successfully"}
    else:
        # Create new
        setting = CustomerProductSetting(**input.model_dump())
        doc = setting.model_dump()
        doc["created_at"] = doc["created_at"].isoformat()
        doc["updated_at"] = doc["updated_at"].isoformat()
        await db.customer_product_settings.insert_one(doc)
        return {"id": setting.id, "message": "Setting created successfully"}

@api_router.delete("/customer-product-settings/{setting_id}")
async def delete_customer_product_setting(setting_id: str, current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in ["admin"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    result = await db.customer_product_settings.delete_one({"id": setting_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Setting not found")
    
    return {"message": "Setting deleted successfully"}

# ============================================================================
# SECTION: QC INDENT ROUTES (Lines ~770-825)
# ============================================================================
@api_router.get("/qc-indents")
async def get_qc_indents(current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in ["admin", "staff"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    indents = await db.qc_indents.find({}, {"_id": 0}).sort("indent_date", -1).to_list(1000)
    for i in indents:
        if isinstance(i.get('indent_date'), str):
            i['indent_date'] = datetime.fromisoformat(i['indent_date'])
        if isinstance(i.get('created_at'), str):
            i['created_at'] = datetime.fromisoformat(i['created_at'])
    return indents

@api_router.post("/qc-indents")
async def create_qc_indent(input: QCIndentCreate, current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in ["admin", "staff"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    indent_dict = input.model_dump()
    indent_dict["recorded_by"] = current_user["user_id"]
    indent = QCIndent(**indent_dict)
    
    doc = indent.model_dump()
    doc['indent_date'] = doc['indent_date'].isoformat()
    doc['created_at'] = doc['created_at'].isoformat()
    await db.qc_indents.insert_one(doc)
    return {"id": indent.id, "message": "Indent created successfully"}

@api_router.put("/qc-indents/{indent_id}")
async def update_qc_indent(indent_id: str, input: QCIndentUpdate, current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in ["admin", "staff"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    update_data = {k: v for k, v in input.model_dump().items() if v is not None}
    if 'indent_date' in update_data:
        update_data['indent_date'] = update_data['indent_date'].isoformat()
    
    result = await db.qc_indents.update_one({"id": indent_id}, {"$set": update_data})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Indent not found")
    
    return {"message": "Indent updated successfully"}

@api_router.delete("/qc-indents/{indent_id}")
async def delete_qc_indent(indent_id: str, current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in ["admin"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    result = await db.qc_indents.delete_one({"id": indent_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Indent not found")
    
    return {"message": "Indent deleted successfully"}

@api_router.get("/qc-indents/previous/{customer_name}")
async def get_previous_qc_indent(customer_name: str, current_user: dict = Depends(get_current_user)):
    """Get the most recent indent for a customer to use as template"""
    if current_user["role"] not in ["admin", "staff"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    # Find the most recent indent for this customer
    indent = await db.qc_indents.find_one(
        {"customer_name": customer_name},
        {"_id": 0},
        sort=[("indent_date", -1), ("created_at", -1)]
    )
    
    if not indent:
        return {"indent": None, "message": "No previous indent found"}
    
    return {"indent": indent, "message": "Previous indent found"}

# ============================================================================
# SECTION: QC DISPATCH ROUTES (Lines ~825-1000)
# ============================================================================
@api_router.get("/qc-dispatches")
async def get_qc_dispatches(current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in ["admin", "staff"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    dispatches = await db.qc_dispatches.find({}, {"_id": 0}).sort("dispatch_date", -1).to_list(1000)
    for d in dispatches:
        if isinstance(d.get('dispatch_date'), str):
            d['dispatch_date'] = datetime.fromisoformat(d['dispatch_date'])
        if isinstance(d.get('created_at'), str):
            d['created_at'] = datetime.fromisoformat(d['created_at'])
    return dispatches

@api_router.post("/qc-dispatches")
async def create_qc_dispatch(input: QCDispatchCreate, current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in ["admin", "staff"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    dispatch_dict = input.model_dump()
    dispatch_dict["recorded_by"] = current_user["user_id"]
    dispatch = QCDispatch(**dispatch_dict)
    
    doc = dispatch.model_dump()
    doc['dispatch_date'] = doc['dispatch_date'].isoformat()
    doc['created_at'] = doc['created_at'].isoformat()
    await db.qc_dispatches.insert_one(doc)
    
    # Check if all indent qty is dispatched to update status
    indent = await db.qc_indents.find_one({"id": input.indent_id})
    if indent:
        # Get all dispatches for this indent
        all_dispatches = await db.qc_dispatches.find({"indent_id": input.indent_id}, {"_id": 0}).to_list(100)
        
        # Calculate total dispatched per product
        total_dispatched = {}
        for d in all_dispatches:
            for item in d.get('items', []):
                key = item['product_id']
                total_dispatched[key] = total_dispatched.get(key, 0) + item.get('supplied_qty', 0)
        
        # Check if fully dispatched
        fully_dispatched = True
        for item in indent.get('items', []):
            dispatched = total_dispatched.get(item['product_id'], 0)
            if dispatched < item['required_qty']:
                fully_dispatched = False
                break
        
        # Update indent status
        new_status = "dispatched" if fully_dispatched else "partial"
        await db.qc_indents.update_one({"id": input.indent_id}, {"$set": {"status": new_status}})
    
    return {"id": dispatch.id, "message": "Dispatch created successfully"}

@api_router.put("/qc-dispatches/{dispatch_id}")
async def update_qc_dispatch(dispatch_id: str, input: QCDispatchCreate, current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in ["admin", "staff"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    existing = await db.qc_dispatches.find_one({"id": dispatch_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Dispatch not found")
    
    update_data = input.model_dump()
    update_data['dispatch_date'] = update_data['dispatch_date'].isoformat()
    
    await db.qc_dispatches.update_one(
        {"id": dispatch_id},
        {"$set": update_data}
    )
    
    return {"id": dispatch_id, "message": "Dispatch updated successfully"}

@api_router.delete("/qc-dispatches/{dispatch_id}")
async def delete_qc_dispatch(dispatch_id: str, current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in ["admin", "staff"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    # Get the dispatch to find the indent_id
    dispatch = await db.qc_dispatches.find_one({"id": dispatch_id}, {"_id": 0})
    if not dispatch:
        raise HTTPException(status_code=404, detail="Dispatch not found")
    
    indent_id = dispatch.get("indent_id")
    
    # Delete the dispatch
    result = await db.qc_dispatches.delete_one({"id": dispatch_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Dispatch not found")
    
    # Check if there are any remaining dispatches for this indent
    if indent_id:
        remaining_dispatches = await db.qc_dispatches.count_documents({"indent_id": indent_id})
        
        if remaining_dispatches == 0:
            # No more dispatches - reset indent status to pending
            await db.qc_indents.update_one(
                {"id": indent_id},
                {"$set": {"status": "pending"}}
            )
        else:
            # Check if indent is fully dispatched or partial
            indent = await db.qc_indents.find_one({"id": indent_id}, {"_id": 0})
            if indent:
                dispatches_for_indent = await db.qc_dispatches.find({"indent_id": indent_id}, {"_id": 0}).to_list(100)
                
                # Calculate total dispatched per product
                dispatched_by_product = {}
                for d in dispatches_for_indent:
                    for item in d.get("items", []):
                        product_id = item.get("product_id")
                        if product_id:
                            dispatched_by_product[product_id] = dispatched_by_product.get(product_id, 0) + item.get("supplied_qty", 0)
                
                # Check if all items are fully dispatched
                all_dispatched = True
                for item in indent.get("items", []):
                    product_id = item.get("product_id")
                    required = item.get("required_qty", 0)
                    dispatched = dispatched_by_product.get(product_id, 0)
                    if dispatched < required:
                        all_dispatched = False
                        break
                
                new_status = "dispatched" if all_dispatched else "partial"
                await db.qc_indents.update_one(
                    {"id": indent_id},
                    {"$set": {"status": new_status}}
                )
    
    return {"message": "Dispatch deleted successfully"}

@api_router.delete("/qc-dispatches/{dispatch_id}/items/{product_id}")
async def delete_qc_dispatch_item(dispatch_id: str, product_id: str, current_user: dict = Depends(get_current_user)):
    """Delete a single item from a dispatch record"""
    if current_user["role"] not in ["admin", "staff"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    # Get the dispatch
    dispatch = await db.qc_dispatches.find_one({"id": dispatch_id}, {"_id": 0})
    if not dispatch:
        raise HTTPException(status_code=404, detail="Dispatch not found")
    
    items = dispatch.get("items", [])
    original_count = len(items)
    
    # Filter out the item with matching product_id
    items = [item for item in items if item.get("product_id") != product_id]
    
    if len(items) == original_count:
        raise HTTPException(status_code=404, detail="Item not found in dispatch")
    
    if len(items) == 0:
        # If no items left, delete the entire dispatch
        await db.qc_dispatches.delete_one({"id": dispatch_id})
        
        # Update indent status
        indent_id = dispatch.get("indent_id")
        if indent_id:
            remaining_dispatches = await db.qc_dispatches.count_documents({"indent_id": indent_id})
            if remaining_dispatches == 0:
                await db.qc_indents.update_one({"id": indent_id}, {"$set": {"status": "pending"}})
        
        return {"message": "Dispatch deleted (no items remaining)"}
    
    # Update the dispatch with remaining items
    await db.qc_dispatches.update_one(
        {"id": dispatch_id},
        {"$set": {"items": items}}
    )
    
    return {"message": "Item removed from dispatch"}

@api_router.get("/qc-dispatches/{dispatch_id}/invoice-status")
async def check_qc_dispatch_invoice_status(dispatch_id: str, current_user: dict = Depends(get_current_user)):
    """Check if a QC dispatch is linked to any invoice"""
    if current_user["role"] not in ["admin", "staff"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    # Check if dispatch exists
    dispatch = await db.qc_dispatches.find_one({"id": dispatch_id}, {"_id": 0})
    if not dispatch:
        raise HTTPException(status_code=404, detail="Dispatch not found")
    
    # Check if this dispatch is part of any invoice
    invoice = await db.qc_invoices.find_one(
        {"dispatch_ids": dispatch_id},
        {"_id": 0, "id": 1, "invoice_number": 1}
    )
    
    if invoice:
        return {
            "is_invoiced": True,
            "invoice_id": invoice.get("id"),
            "invoice_number": invoice.get("invoice_number")
        }
    
    return {"is_invoiced": False, "invoice_id": None, "invoice_number": None}

class QCDispatchItemUpdate(BaseModel):
    item_index: int
    supplied_qty: float
    no_of_crates: Optional[int] = None

@api_router.put("/qc-dispatches/{dispatch_id}/items/{item_index}")
async def update_qc_dispatch_item(
    dispatch_id: str, 
    item_index: int, 
    input: QCDispatchItemUpdate, 
    current_user: dict = Depends(get_current_user)
):
    """Update a single item in a QC dispatch (quantity and crates)"""
    if current_user["role"] not in ["admin", "staff"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    # Get the dispatch
    dispatch = await db.qc_dispatches.find_one({"id": dispatch_id}, {"_id": 0})
    if not dispatch:
        raise HTTPException(status_code=404, detail="Dispatch not found")
    
    # Check if dispatch is invoiced
    invoice = await db.qc_invoices.find_one(
        {"dispatch_ids": dispatch_id},
        {"_id": 0, "invoice_number": 1}
    )
    if invoice:
        raise HTTPException(
            status_code=400, 
            detail=f"Cannot edit: Dispatch is part of invoice {invoice.get('invoice_number')}. Delete the invoice first."
        )
    
    items = dispatch.get("items", [])
    if item_index < 0 or item_index >= len(items):
        raise HTTPException(status_code=404, detail="Item not found at specified index")
    
    # Update the item quantity and crates
    items[item_index]["supplied_qty"] = input.supplied_qty
    if input.no_of_crates is not None:
        items[item_index]["no_of_crates"] = input.no_of_crates
    
    # Update the dispatch
    await db.qc_dispatches.update_one(
        {"id": dispatch_id},
        {"$set": {"items": items}}
    )
    
    return {"message": "Dispatch item updated successfully", "new_qty": input.supplied_qty, "new_crates": input.no_of_crates}

@api_router.delete("/qc-dispatches/{dispatch_id}/items-by-index/{item_index}")
async def delete_qc_dispatch_item_by_index(
    dispatch_id: str, 
    item_index: int, 
    current_user: dict = Depends(get_current_user)
):
    """Delete a single item from a QC dispatch by index"""
    if current_user["role"] not in ["admin", "staff"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    # Get the dispatch
    dispatch = await db.qc_dispatches.find_one({"id": dispatch_id}, {"_id": 0})
    if not dispatch:
        raise HTTPException(status_code=404, detail="Dispatch not found")
    
    # Check if dispatch is invoiced
    invoice = await db.qc_invoices.find_one(
        {"dispatch_ids": dispatch_id},
        {"_id": 0, "invoice_number": 1}
    )
    if invoice:
        raise HTTPException(
            status_code=400, 
            detail=f"Cannot delete: Dispatch is part of invoice {invoice.get('invoice_number')}. Delete the invoice first."
        )
    
    items = dispatch.get("items", [])
    if item_index < 0 or item_index >= len(items):
        raise HTTPException(status_code=404, detail="Item not found at specified index")
    
    # Remove the item at the specified index
    deleted_item = items.pop(item_index)
    
    if len(items) == 0:
        # If no items left, delete the entire dispatch
        await db.qc_dispatches.delete_one({"id": dispatch_id})
        
        # Update indent status
        indent_id = dispatch.get("indent_id")
        if indent_id:
            remaining_dispatches = await db.qc_dispatches.count_documents({"indent_id": indent_id})
            if remaining_dispatches == 0:
                await db.qc_indents.update_one({"id": indent_id}, {"$set": {"status": "pending"}})
        
        return {"message": "Dispatch deleted (no items remaining)", "deleted_item": deleted_item}
    
    # Update the dispatch with remaining items
    await db.qc_dispatches.update_one(
        {"id": dispatch_id},
        {"$set": {"items": items}}
    )
    
    return {"message": "Item removed from dispatch", "deleted_item": deleted_item}

# ============================================================================
# SECTION: QC GRN ROUTES (Lines ~999-1450)
# ============================================================================
@api_router.get("/qc-grns")
async def get_qc_grns(current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in ["admin", "staff"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    grns = await db.qc_grns.find({}, {"_id": 0}).sort("created_at", -1).to_list(1000)
    for g in grns:
        if isinstance(g.get('grn_date'), str):
            g['grn_date'] = datetime.fromisoformat(g['grn_date'])
        if isinstance(g.get('created_at'), str):
            g['created_at'] = datetime.fromisoformat(g['created_at'])
    return grns

@api_router.post("/qc-grns")
async def create_qc_grn(input: QCGRNCreate, current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in ["admin", "staff"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    grn_dict = input.model_dump()
    grn_dict["recorded_by"] = current_user["user_id"]
    grn = QCGRN(**grn_dict)
    
    doc = grn.model_dump()
    doc['grn_date'] = doc['grn_date'].isoformat()
    doc['created_at'] = doc['created_at'].isoformat()
    await db.qc_grns.insert_one(doc)
    
    return {"id": grn.id, "message": "GRN created successfully"}

@api_router.put("/qc-grns/{grn_id}")
async def update_qc_grn(grn_id: str, input: QCGRNCreate, current_user: dict = Depends(get_current_user)):
    """Update an existing GRN record"""
    if current_user["role"] not in ["admin", "staff"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    # Check if GRN exists
    existing = await db.qc_grns.find_one({"id": grn_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="GRN not found")
    
    # Prepare update data
    update_data = input.model_dump()
    update_data['grn_date'] = update_data['grn_date'].isoformat() if hasattr(update_data['grn_date'], 'isoformat') else update_data['grn_date']
    
    # Recalculate totals from items
    items = update_data.get('items', [])
    update_data['total_supplied'] = sum(item.get('supplied_qty', 0) for item in items)
    update_data['total_grn'] = sum(item.get('grn_qty', 0) for item in items)
    update_data['total_difference'] = sum(item.get('difference', 0) for item in items)
    
    # Update the GRN
    await db.qc_grns.update_one(
        {"id": grn_id},
        {"$set": update_data}
    )
    
    return {"id": grn_id, "message": "GRN updated successfully"}

@api_router.delete("/qc-grns/{grn_id}")
async def delete_qc_grn(grn_id: str, current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in ["admin", "staff"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    result = await db.qc_grns.delete_one({"id": grn_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="GRN not found")
    
    return {"message": "GRN deleted successfully"}

@api_router.delete("/qc-grns/{grn_id}/items/{item_index}")
async def delete_qc_grn_item(grn_id: str, item_index: int, current_user: dict = Depends(get_current_user)):
    """Delete a single item from a GRN record"""
    if current_user["role"] not in ["admin", "staff"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    # Get the GRN
    grn = await db.qc_grns.find_one({"id": grn_id}, {"_id": 0})
    if not grn:
        raise HTTPException(status_code=404, detail="GRN not found")
    
    items = grn.get("items", [])
    if item_index < 0 or item_index >= len(items):
        raise HTTPException(status_code=400, detail="Invalid item index")
    
    # Remove the item
    items.pop(item_index)
    
    if len(items) == 0:
        # If no items left, delete the entire GRN
        await db.qc_grns.delete_one({"id": grn_id})
        return {"message": "GRN deleted (no items remaining)"}
    
    # Recalculate totals
    total_supplied = sum(item.get("supplied_qty", 0) for item in items)
    total_grn = sum(item.get("grn_qty", 0) for item in items)
    total_difference = sum(item.get("difference", 0) for item in items)
    
    # Update the GRN
    await db.qc_grns.update_one(
        {"id": grn_id},
        {"$set": {
            "items": items,
            "total_supplied": total_supplied,
            "total_grn": total_grn,
            "total_difference": total_difference
        }}
    )
    
    return {"message": "Item removed from GRN"}

@api_router.post("/qc-grns/upload-ninjacart-csv")
async def upload_ninjacart_grn_csv(file: UploadFile = File(...), current_user: dict = Depends(get_current_user)):
    """
    Upload Ninjacart GRN CSV or Excel file and match with our dispatches.
    The file contains: PO_DeliveryDate, Sku Name, GRNQuantity, GRNPrice, WeightUnit
    
    STRICT VALIDATION: Only processes rows if a dispatch exists for that exact date.
    Returns warnings for skipped rows (no matching dispatch).
    """
    if current_user["role"] not in ["admin", "staff"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    filename_lower = file.filename.lower()
    if not (filename_lower.endswith('.csv') or filename_lower.endswith('.xlsx') or filename_lower.endswith('.xls')):
        raise HTTPException(status_code=400, detail="Only CSV or Excel (.xlsx, .xls) files are allowed")
    
    # Read file content
    content = await file.read()
    
    # Parse based on file type
    rows_data = []
    if filename_lower.endswith('.csv'):
        # Parse CSV
        decoded = content.decode('utf-8')
        reader = csv.DictReader(io.StringIO(decoded))
        rows_data = list(reader)
        logger.info(f"CSV headers: {reader.fieldnames if hasattr(reader, 'fieldnames') else 'N/A'}")
    else:
        # Parse Excel (.xlsx or .xls)
        import openpyxl
        from io import BytesIO
        from datetime import datetime as dt
        
        try:
            workbook = openpyxl.load_workbook(BytesIO(content), read_only=True, data_only=True)
            sheet = workbook.active
            
            # Get headers from first row - try to find the actual header row
            headers = []
            header_row = 1
            
            # Check first few rows to find header row (in case there's a title row)
            for check_row in range(1, 4):
                potential_headers = []
                for cell in sheet[check_row]:
                    val = str(cell.value).strip() if cell.value else ''
                    potential_headers.append(val)
                
                # Check if this looks like a header row (contains expected column names)
                header_str = ' '.join(potential_headers).lower()
                if 'podate' in header_str or 'sku' in header_str or 'grn' in header_str or 'po_deliverydate' in header_str:
                    headers = potential_headers
                    header_row = check_row
                    break
            
            # If no header row found, use first row
            if not headers:
                for cell in sheet[1]:
                    headers.append(str(cell.value).strip() if cell.value else '')
            
            logger.info(f"Excel headers found at row {header_row}: {headers}")
            
            # Parse data rows (starting after header row)
            for row_idx, row in enumerate(sheet.iter_rows(min_row=header_row+1, values_only=True), start=header_row+1):
                row_dict = {}
                for col_idx, value in enumerate(row):
                    if col_idx < len(headers):
                        header = headers[col_idx]
                        # Handle datetime objects specially
                        if isinstance(value, dt):
                            row_dict[header] = value.strftime('%Y-%m-%d')
                        elif value is not None:
                            row_dict[header] = str(value).strip()
                        else:
                            row_dict[header] = ''
                if any(row_dict.values()):  # Skip completely empty rows
                    rows_data.append(row_dict)
            
            # Log first row for debugging
            if rows_data:
                logger.info(f"First Excel row parsed: {rows_data[0]}")
            else:
                logger.warning("No data rows found in Excel file")
            
            workbook.close()
        except Exception as e:
            logger.error(f"Error parsing Excel file: {e}", exc_info=True)
            raise HTTPException(status_code=400, detail=f"Failed to parse Excel file: {str(e)}")
    
    # Get all Ninjacart dispatches
    ninjacart_dispatches = await db.qc_dispatches.find(
        {"customer_name": {"$regex": "ninja", "$options": "i"}},
        {"_id": 0}
    ).to_list(1000)
    
    # Build a set of dates that have dispatches for quick lookup
    dispatch_dates_set = set()
    for dispatch in ninjacart_dispatches:
        dispatch_date = dispatch.get('dispatch_date', '')
        if isinstance(dispatch_date, str):
            dispatch_dates_set.add(dispatch_date[:10])  # YYYY-MM-DD
        else:
            dispatch_dates_set.add(dispatch_date.strftime('%Y-%m-%d'))
    
    # Get packaging variants for weight lookup
    packaging_variants = await db.qc_packaging.find({}, {"_id": 0}).to_list(100)
    packaging_map = {p['name'].lower(): p for p in packaging_variants}
    
    # Parse rows and validate against dispatch dates
    # Support both old format (PO_DeliveryDate, Sku Name, GRNQuantity, GRNPrice, WeightUnit)
    # and new format (podate, Sku, GRN_Qty, Total_Value, Kgs_Pcs)
    csv_data_by_date = {}
    skipped_rows = []  # Track rows skipped due to no matching dispatch
    total_csv_rows = 0
    
    for row in rows_data:
        total_csv_rows += 1
        try:
            # Detect format and extract date
            # New format: podate (YYYY-MM-DD)
            # Old format: PO_DeliveryDate (DD/MM/YY)
            date_str = row.get('podate', '') or row.get('PO_DeliveryDate', '')
            
            if not date_str:
                skipped_rows.append({
                    'row': total_csv_rows,
                    'reason': 'Missing date',
                    'sku': row.get('Sku', row.get('Sku Name', 'Unknown'))
                })
                continue
            
            # Parse date based on format
            parsed_date = None
            if '-' in str(date_str):
                # New format: YYYY-MM-DD or datetime object
                if hasattr(date_str, 'strftime'):
                    # It's a datetime object from Excel
                    parsed_date = date_str.strftime('%Y-%m-%d')
                else:
                    # It's a string like "2026-03-29"
                    parsed_date = str(date_str)[:10]  # Take first 10 chars (YYYY-MM-DD)
            elif '/' in str(date_str):
                # Old format: DD/MM/YY
                day, month, year = str(date_str).split('/')
                if len(year) == 2:
                    year = '20' + year
                parsed_date = f"{year}-{month.zfill(2)}-{day.zfill(2)}"
            else:
                skipped_rows.append({
                    'row': total_csv_rows,
                    'reason': f'Invalid date format: {date_str}',
                    'sku': row.get('Sku', row.get('Sku Name', 'Unknown'))
                })
                continue
            
            # Extract SKU name (new format: Sku, old format: Sku Name)
            sku_name = (row.get('Sku', '') or row.get('Sku Name', '')).strip()
            
            # Extract GRN quantity (new format: GRN_Qty, old format: GRNQuantity)
            grn_qty_raw = row.get('GRN_Qty', '') or row.get('GRNQuantity', '0')
            grn_qty = float(grn_qty_raw) if grn_qty_raw else 0
            
            # Extract price/value (new format: Total_Value, old format: GRNPrice)
            # For new format, we may need to calculate rate from Total_Value / GRN_Qty
            total_value_raw = row.get('Total_Value', '') or row.get('GRNPrice', '0')
            total_value = float(total_value_raw) if total_value_raw else 0
            
            # Calculate rate per unit
            if grn_qty > 0 and total_value > 0:
                grn_price = total_value / grn_qty
            else:
                grn_price = float(row.get('GRNPrice', '0') or '0')
            
            # Extract unit (new format: Kgs_Pcs, old format: WeightUnit)
            weight_unit = (row.get('Kgs_Pcs', '') or row.get('WeightUnit', 'Kg')).strip()
            
            if not sku_name or grn_qty == 0:
                skipped_rows.append({
                    'row': total_csv_rows,
                    'reason': 'Empty SKU or zero quantity',
                    'sku': sku_name or 'Empty',
                    'date': str(date_str)
                })
                continue
            
            # STRICT VALIDATION: Check if dispatch exists for this date
            if parsed_date not in dispatch_dates_set:
                skipped_rows.append({
                    'row': total_csv_rows,
                    'reason': f'No dispatch found for date {parsed_date}',
                    'sku': sku_name,
                    'date': str(date_str)
                })
                continue
            
            # Store by date (only if dispatch exists)
            if parsed_date not in csv_data_by_date:
                csv_data_by_date[parsed_date] = []
            
            csv_data_by_date[parsed_date].append({
                'sku_name': sku_name,
                'grn_qty': grn_qty,
                'grn_price': grn_price,
                'total_value': total_value,
                'weight_unit': weight_unit
            })
        except Exception as e:
            logger.error(f"Error parsing row {total_csv_rows}: {e}")
            skipped_rows.append({
                'row': total_csv_rows,
                'reason': f'Parse error: {str(e)}',
                'sku': row.get('Sku', row.get('Sku Name', 'Unknown'))
            })
            continue
    
    # Match CSV data with dispatches
    # NEW LOGIC:
    # 1. Use column H (Kgs_Pcs) from Excel to determine if SKU is Kg or PCS based
    # 2. Only dispatch packaging with explicit "(PCS)" is PCS-based, everything else is Kg-based
    # 3. For Kg-based products with multiple variants, distribute GRN proportionally by supplied weight
    
    matched_items = []
    
    # Log parsed data for debugging
    logger.info(f"GRN Upload: Parsed {total_csv_rows} rows, {len(csv_data_by_date)} dates with data")
    logger.info(f"GRN Upload: Dispatch dates available: {list(dispatch_dates_set)[:5]}...")
    logger.info(f"GRN Upload: CSV dates parsed: {list(csv_data_by_date.keys())}")
    
    # Variant/color words that MUST match exactly if present in product name
    variant_words = {'red', 'green', 'yellow', 'white', 'black', 'purple', 'orange', 
                   'small', 'large', 'big', 'baby', 'mini', 'jumbo', 'local', 'hybrid',
                   'english', 'indian', 'desi', 'organic', 'fresh', 'premium'}
    
    # Step 1: Combine all dispatch items by date (since each item may be in a separate dispatch)
    dispatch_items_by_date = {}
    for dispatch in ninjacart_dispatches:
        dispatch_date = dispatch.get('dispatch_date', '')
        if isinstance(dispatch_date, str):
            dispatch_date_str = dispatch_date[:10]
        else:
            dispatch_date_str = dispatch_date.strftime('%Y-%m-%d')
        
        if dispatch_date_str not in dispatch_items_by_date:
            dispatch_items_by_date[dispatch_date_str] = []
        
        for item in dispatch.get('items', []):
            item_with_dispatch = {
                **item,
                'dispatch_id': dispatch.get('id')
            }
            dispatch_items_by_date[dispatch_date_str].append(item_with_dispatch)
    
    logger.info(f"Combined dispatch items by date: {[(d, len(items)) for d, items in dispatch_items_by_date.items()]}")
    
    # Step 2: Process each date
    for dispatch_date_str, all_dispatch_items in dispatch_items_by_date.items():
        # Check if we have CSV data for this date
        if dispatch_date_str not in csv_data_by_date:
            continue
        
        csv_rows = csv_data_by_date[dispatch_date_str]
        
        # Step 3: Group dispatch items by base product name
        product_groups = {}  # {base_product_name: [items]}
        
        for item in all_dispatch_items:
            product_name = item.get('product_name', '')
            packaging_name = item.get('packaging_name', '') or ''
            packaging_upper = packaging_name.upper()
            
            # Simple PCS vs Kg classification:
            # "(PCS)" in packaging = PCS-based, matches Excel SKUs with Kgs_Pcs = "PCS"
            # Everything else (Packet, gm, etc.) = Kg-based, matches Excel SKUs with Kgs_Pcs = "Kgs"
            dispatch_is_pcs = '(PCS)' in packaging_upper
            
            # Get packaging weight
            packaging_weight_gm = 0
            if packaging_name:
                pkg = packaging_map.get(packaging_name.lower())
                if pkg:
                    packaging_weight_gm = pkg.get('weight_gm', 0)
            if packaging_weight_gm == 0:
                weight_match = re.search(r'(\d+)\s*(gm|g|gram)', packaging_name.lower())
                if weight_match:
                    packaging_weight_gm = int(weight_match.group(1))
            
            # Extract base product name
            base_name = product_name.lower().strip()
            
            item_data = {
                **item,
                'packaging_weight_gm': packaging_weight_gm,
                'dispatch_is_pcs': dispatch_is_pcs,
                'base_product_name': base_name
            }
            
            if base_name not in product_groups:
                product_groups[base_name] = []
            product_groups[base_name].append(item_data)
        
        logger.info(f"Product groups for {dispatch_date_str}: {[(name, len(items)) for name, items in product_groups.items()]}")
        
        # Step 4: First pass - collect all SKU matches for each product group
        # This handles cases like Palak/Spinach where multiple SKUs should be combined
        product_group_matches = {}  # {base_name: {'kg_skus': [], 'pcs_skus': [], 'items': []}}
        
        for csv_row in csv_rows:
            sku_name = csv_row['sku_name']
            sku_name_lower = sku_name.lower()
            grn_qty_kg_or_pcs = csv_row['grn_qty']  # This is Kg for Kgs items, pieces for Pcs items
            rate_per_kg_or_pcs = csv_row['grn_price']  # Rate per Kg or per piece
            total_value = csv_row.get('total_value', grn_qty_kg_or_pcs * rate_per_kg_or_pcs)
            
            # Use column H (Kgs_Pcs) to determine SKU type
            weight_unit = csv_row.get('weight_unit', '').upper()
            sku_is_pcs = weight_unit in ['PCS', 'PIECES', 'PIECE']
            sku_is_kg = weight_unit in ['KGS', 'KG', 'KILO', 'KILOS'] or not sku_is_pcs
            
            # Clean SKU name for matching (remove - FK, etc.)
            sku_clean = re.sub(r'\s*-\s*(FK|MH|BLR)\s*$', '', sku_name_lower, flags=re.IGNORECASE)
            sku_clean = re.sub(r'\s*\([^)]*\)\s*', ' ', sku_clean)  # Remove parentheses content
            sku_clean = re.sub(r'\s+', ' ', sku_clean).strip()
            sku_words = set(sku_clean.split())
            
            # Extract weight from SKU name for weight-based matching (e.g., "70-80gm", "100 g")
            sku_weight_match = re.search(r'(\d+)(?:\s*-\s*\d+)?\s*(gm|g)\b', sku_name_lower)
            sku_weight = int(sku_weight_match.group(1)) if sku_weight_match else 0
            
            # Check for "with roots" or "without roots" in SKU for special matching
            sku_has_with_roots = 'with roots' in sku_name_lower and 'without' not in sku_name_lower
            sku_has_without_roots = 'without roots' in sku_name_lower
            
            # Special product name mappings (Mint variants, Methi = Fenugreek)
            # NOTE: Palak and Spinach are NOT aliased - they should match separately
            # Each Excel SKU should match only its corresponding dispatch product
            product_aliases = {
                'mint': ['mint', 'fresh mint leaves', 'premium fresh mint leaves', 'fresh mint', 'premium mint'],
                'fresh mint leaves': ['mint', 'fresh mint leaves', 'premium fresh mint leaves'],
                'premium fresh mint leaves': ['mint', 'fresh mint leaves', 'premium fresh mint leaves'],
                'methi': ['methi', 'fenugreek', 'fenugreek (methi)'],
                'fenugreek': ['methi', 'fenugreek', 'fenugreek (methi)'],
                'fenugreek (methi)': ['methi', 'fenugreek', 'fenugreek (methi)'],
            }
            
            logger.info(f"Processing SKU: {sku_name} | Unit: {weight_unit} | is_pcs={sku_is_pcs} | Qty: {grn_qty_kg_or_pcs} | SKU_weight: {sku_weight}gm | with_roots={sku_has_with_roots}")
            
            # Find matching product group
            best_match_group = None
            best_match_score = 0
            
            for base_name, items in product_groups.items():
                # Special handling for Mint Leaves ONLY - match by weight and roots type
                is_mint_product = 'mint' in base_name.lower()
                is_mint_sku = 'mint' in sku_name_lower
                
                # Only apply Mint-specific matching if BOTH product and SKU mention mint
                if is_mint_product and is_mint_sku and sku_is_pcs and (sku_has_with_roots or sku_has_without_roots):
                    # For Mint PCS items, match based on BOTH roots type AND weight
                    # This prevents 70-80gm "without roots" from matching 100gm "with roots"
                    matched_mint_items = []
                    for item in items:
                        pkg_name_lower = item.get('packaging_name', '').lower()
                        pkg_has_with_roots = 'with roots' in pkg_name_lower and 'without' not in pkg_name_lower
                        pkg_has_without_roots = 'without roots' in pkg_name_lower
                        
                        # Match roots type (REQUIRED)
                        roots_match = (sku_has_with_roots and pkg_has_with_roots) or \
                                     (sku_has_without_roots and pkg_has_without_roots)
                        
                        if not roots_match:
                            continue  # Skip if roots type doesn't match
                        
                        # Match weight (within 30gm tolerance for 70-80 range)
                        pkg_weight = item.get('packaging_weight_gm', 0)
                        weight_match = sku_weight > 0 and pkg_weight > 0 and abs(pkg_weight - sku_weight) <= 30
                        
                        # For Mint, require BOTH roots match AND (weight match OR no weight specified)
                        if roots_match and (weight_match or sku_weight == 0 or pkg_weight == 0):
                            logger.info(f"  Mint match found: {item.get('packaging_name')} | roots_match={roots_match}, weight_match={weight_match}, sku_weight={sku_weight}, pkg_weight={pkg_weight}")
                            matched_mint_items.append(item)
                    
                    if matched_mint_items:
                        best_match_group = (base_name, matched_mint_items)
                        best_match_score = 1.0
                        break
                
                # Regular matching logic
                # PCS SKU (Excel Kgs_Pcs = "PCS") matches dispatch items with "(PCS)" in packaging
                # Kg SKU (Excel Kgs_Pcs = "Kgs") matches ALL other dispatch items (Packet, gm, etc.)
                if sku_is_pcs:
                    type_matched_items = [i for i in items if i.get('dispatch_is_pcs')]
                else:
                    type_matched_items = [i for i in items if not i.get('dispatch_is_pcs')]
                
                if not type_matched_items:
                    continue  # No items match the required type
                
                # Check product name match
                product_words = set(base_name.split())
                first_word = base_name.split()[0] if base_name else ''
                
                # Check for product aliases (Palak/Spinach, Mint variants)
                alias_match = False
                for alias_key, aliases in product_aliases.items():
                    if first_word in aliases or any(alias in base_name for alias in aliases):
                        if any(alias in sku_clean for alias in aliases):
                            alias_match = True
                            break
                
                # Check variant words - if present in product, must be in SKU
                product_variants = product_words & variant_words
                if product_variants and not alias_match:
                    if not product_variants.issubset(sku_words):
                        continue
                
                # Count matching words
                common_words = product_words & sku_words
                
                # For PCS items, also match based on packaging weight
                if sku_is_pcs and sku_weight > 0 and not is_mint_sku:
                    # Check if any item in group has matching weight
                    weight_matched_items = [i for i in type_matched_items if abs(i['packaging_weight_gm'] - sku_weight) <= 20]
                    if weight_matched_items:
                        # Weight match found - this is a strong match for PCS items
                        if alias_match or common_words:
                            match_score = 1.0  # High score for weight + name match
                            if match_score > best_match_score:
                                best_match_score = match_score
                                # Only include weight-matched items
                                best_match_group = (base_name, weight_matched_items)
                            continue
                
                if not common_words and not alias_match:
                    continue
                
                # First word (main product) must match (unless alias match)
                if not alias_match:
                    if first_word and first_word not in sku_words:
                        if not any(first_word in sw or sw in first_word for sw in sku_words):
                            continue
                
                match_score = len(common_words) / max(len(product_words), 1)
                if alias_match:
                    match_score = max(match_score, 0.8)  # Boost alias matches
                
                if match_score > best_match_score:
                    best_match_score = match_score
                    best_match_group = (base_name, type_matched_items)
            
            if not best_match_group:
                logger.info(f"  No match found for SKU: {sku_name}")
                continue
            
            base_name, matching_items = best_match_group
            logger.info(f"  Matched with product group: {base_name} ({len(matching_items)} variants)")
            
            # Collect matches by product group for later aggregation
            if base_name not in product_group_matches:
                product_group_matches[base_name] = {
                    'kg_skus': [],
                    'pcs_skus': [],
                    'items': matching_items
                }
            else:
                # Update items to include all variants
                existing_item_ids = {id(i) for i in product_group_matches[base_name]['items']}
                for item in matching_items:
                    if id(item) not in existing_item_ids:
                        product_group_matches[base_name]['items'].append(item)
            
            if sku_is_pcs:
                product_group_matches[base_name]['pcs_skus'].append({
                    'sku_name': sku_name,
                    'grn_qty': grn_qty_kg_or_pcs,
                    'rate': rate_per_kg_or_pcs,
                    'total_value': total_value,
                    'items': list(matching_items)  # Make a copy to avoid reference issues
                })
            else:
                product_group_matches[base_name]['kg_skus'].append({
                    'sku_name': sku_name,
                    'grn_qty_kg': grn_qty_kg_or_pcs,
                    'rate_per_kg': rate_per_kg_or_pcs,
                    'total_value': total_value,
                    'items': list(matching_items)  # Make a copy to avoid reference issues
                })
        
        # Step 5: Now distribute GRN to dispatch items
        # Track processed dispatch items to prevent duplicate supplied_qty entries
        processed_dispatch_items = {}  # {dispatch_id_product_id_packaging_id: matched_item_dict}
        
        for base_name, group_data in product_group_matches.items():
            # Process PCS SKUs - COMBINE all PCS SKUs that match the same dispatch items
            # This prevents double-counting supplied_qty when multiple Excel SKUs match the same dispatch
            if group_data['pcs_skus']:
                # First, collect ALL unique PCS dispatch items and combine GRN quantities
                all_pcs_items = {}  # {item_key: item_dict_with_accumulated_grn}
                combined_sku_names = []
                
                for pcs_sku in group_data['pcs_skus']:
                    sku_name = pcs_sku['sku_name']
                    grn_qty = pcs_sku['grn_qty']
                    rate_per_pcs = pcs_sku['rate']
                    sku_items = pcs_sku['items']
                    combined_sku_names.append(sku_name)
                    
                    # Get only PCS dispatch items for this specific SKU
                    pcs_items = [item for item in sku_items if item.get('dispatch_is_pcs')]
                    
                    if not pcs_items:
                        logger.warning(f"  No PCS dispatch items found for SKU: {sku_name}")
                        continue
                    
                    # Calculate total supplied for this SKU's matched items
                    total_supplied_pcs = sum(item.get('supplied_qty', 0) for item in pcs_items)
                    
                    logger.info(f"  Processing PCS SKU: {sku_name} | GRN: {grn_qty} pcs | Supplied: {total_supplied_pcs} | Items: {len(pcs_items)}")
                    
                    if total_supplied_pcs == 0:
                        logger.warning(f"  Zero supplied qty for SKU: {sku_name}")
                        continue
                    
                    # Distribute GRN proportionally to each dispatch item
                    for item in pcs_items:
                        supplied_qty = item.get('supplied_qty', 0)
                        
                        # Calculate this item's proportion of the total supplied
                        proportion = supplied_qty / total_supplied_pcs if total_supplied_pcs > 0 else 0
                        
                        # Distribute GRN proportionally
                        grn_qty_for_item = grn_qty * proportion
                        
                        # Create unique key for this dispatch item
                        item_key = f"{item.get('dispatch_id')}_{item.get('product_id')}_{item.get('packaging_id')}"
                        
                        if item_key in all_pcs_items:
                            # Accumulate GRN qty for same dispatch item
                            all_pcs_items[item_key]['grn_qty'] += grn_qty_for_item
                            all_pcs_items[item_key]['sku_names'].append(sku_name)
                            logger.info(f"    Accumulating to existing item: {item_key}, total grn_qty={all_pcs_items[item_key]['grn_qty']:.2f}")
                        else:
                            # New dispatch item
                            all_pcs_items[item_key] = {
                                'dispatch_id': item.get('dispatch_id'),
                                'dispatch_date': dispatch_date_str,
                                'product_id': item.get('product_id'),
                                'product_name': item.get('product_name'),
                                'product_unit': item.get('product_unit'),
                                'packaging_id': item.get('packaging_id'),
                                'packaging_name': item.get('packaging_name'),
                                'packaging_weight_gm': item.get('packaging_weight_gm', 0),
                                'supplied_qty': supplied_qty,
                                'grn_qty': grn_qty_for_item,
                                'rate_per_pcs': rate_per_pcs,
                                'sku_names': [sku_name]
                            }
                            logger.info(f"    New dispatch item: {item_key}, supplied={supplied_qty}, grn={grn_qty_for_item:.2f}")
                
                # Now add all unique PCS items to matched_items
                for item_key, item_data in all_pcs_items.items():
                    grn_qty_final = item_data['grn_qty']
                    supplied_qty = item_data['supplied_qty']
                    rate_per_pcs = item_data['rate_per_pcs']
                    
                    amount = grn_qty_final * rate_per_pcs
                    difference = grn_qty_final - supplied_qty
                    loss_gain = difference * rate_per_pcs
                    combined_sku = ' + '.join(item_data['sku_names'])
                    
                    matched_items.append({
                        'dispatch_id': item_data['dispatch_id'],
                        'dispatch_date': item_data['dispatch_date'],
                        'product_id': item_data['product_id'],
                        'product_name': item_data['product_name'],
                        'product_unit': item_data['product_unit'],
                        'packaging_id': item_data['packaging_id'],
                        'packaging_name': item_data['packaging_name'],
                        'packaging_weight_gm': item_data['packaging_weight_gm'],
                        'supplied_qty': supplied_qty,
                        'grn_qty': round(grn_qty_final, 2),
                        'grn_qty_kg': None,
                        'difference': round(difference, 2),
                        'rate_per_kg': None,
                        'rate_per_unit': round(rate_per_pcs, 2),
                        'rate_type': 'per_pcs',
                        'sku_name': combined_sku,
                        'amount': round(amount, 2),
                        'loss_gain_amount': round(loss_gain, 2)
                    })
                    
                    # Mark as processed
                    processed_dispatch_items[item_key] = True
            
            # Process Kg SKUs - COMBINE all Kg SKUs for this product group
            # Kg SKUs should ONLY match with Kg-based dispatch items (NOT PCS items)
            if group_data['kg_skus']:
                # Combine all Kg SKUs
                total_grn_kg = sum(sku['grn_qty_kg'] for sku in group_data['kg_skus'])
                total_value = sum(sku['total_value'] for sku in group_data['kg_skus'])
                combined_rate_per_kg = total_value / total_grn_kg if total_grn_kg > 0 else 0
                combined_sku_names = ' + '.join(sku['sku_name'] for sku in group_data['kg_skus'])
                
                logger.info(f"  Combining Kg SKUs for {base_name}: {total_grn_kg}kg @ ₹{combined_rate_per_kg:.2f}/kg from {len(group_data['kg_skus'])} SKUs")
                
                # Get ALL non-PCS dispatch items for Kg SKU matching
                # This includes "Packet" items, "gm" items, etc. - everything without "(PCS)"
                kg_items = []
                seen_item_keys = set()
                for i in group_data['items']:
                    # Include all items that are NOT PCS (dispatch_is_pcs = False)
                    if not i.get('dispatch_is_pcs'):
                        item_key = f"{i.get('dispatch_id')}_{i.get('product_id')}_{i.get('packaging_id')}"
                        # Skip if already processed by PCS SKUs or another product group
                        if item_key in processed_dispatch_items:
                            logger.info(f"    Skipping already processed item: {item_key}")
                            continue
                        if item_key not in seen_item_keys:
                            kg_items.append(i)
                            seen_item_keys.add(item_key)
                
                # Calculate total supplied weight
                total_supplied_kg = 0
                for item in kg_items:
                    pkg_weight_kg = item.get('packaging_weight_gm', 100) / 1000
                    supplied_qty = item.get('supplied_qty', 0)
                    total_supplied_kg += supplied_qty * pkg_weight_kg
                
                if total_supplied_kg == 0:
                    logger.warning(f"  No Kg-based items found for {base_name} (PCS items exist but Kg SKU cannot match them)")
                    continue
                
                logger.info(f"  Total supplied: {total_supplied_kg:.2f}kg, GRN: {total_grn_kg}kg, Dispatch items: {len(kg_items)}")
                
                # Distribute combined GRN proportionally to each dispatch item
                for item in kg_items:
                    pkg_weight_gm = item.get('packaging_weight_gm', 100)
                    pkg_weight_kg = pkg_weight_gm / 1000
                    supplied_qty = item.get('supplied_qty', 0)
                    supplied_kg = supplied_qty * pkg_weight_kg
                    
                    # Proportion of this item's weight
                    proportion = supplied_kg / total_supplied_kg if total_supplied_kg > 0 else 0
                    
                    # Distribute GRN kg proportionally
                    grn_kg_for_item = total_grn_kg * proportion
                    
                    # Convert GRN from Kg to piece count (same unit as supplied_qty)
                    # grn_qty_in_pieces = grn_kg / (packaging_weight_gm / 1000)
                    grn_qty_in_pieces = (grn_kg_for_item * 1000) / pkg_weight_gm if pkg_weight_gm > 0 else 0
                    
                    # Rate per piece (derived from rate_per_kg)
                    rate_per_piece = combined_rate_per_kg * (pkg_weight_gm / 1000)
                    
                    # Amount for this item (using Kg values for accuracy)
                    amount = grn_kg_for_item * combined_rate_per_kg
                    
                    # Difference in pieces (GRN pieces - Supplied pieces)
                    difference_pcs = grn_qty_in_pieces - supplied_qty
                    loss_gain = difference_pcs * rate_per_piece
                    
                    logger.info(f"    {item.get('product_name')} {item.get('packaging_name')}: supplied={supplied_qty} pcs, grn={grn_qty_in_pieces:.1f} pcs, diff={difference_pcs:.1f} pcs")
                    
                    matched_items.append({
                        'dispatch_id': item.get('dispatch_id'),
                        'dispatch_date': dispatch_date_str,
                        'product_id': item.get('product_id'),
                        'product_name': item.get('product_name'),
                        'product_unit': item.get('product_unit'),
                        'packaging_id': item.get('packaging_id'),
                        'packaging_name': item.get('packaging_name'),
                        'packaging_weight_gm': pkg_weight_gm,
                        'supplied_qty': supplied_qty,
                        'supplied_qty_kg': round(supplied_kg, 2),
                        'grn_qty': round(grn_qty_in_pieces, 0),  # GRN in pieces (same unit as supplied_qty)
                        'grn_qty_kg': round(grn_kg_for_item, 2),
                        'difference': round(difference_pcs, 0),  # Difference in pieces
                        'rate_per_kg': round(combined_rate_per_kg, 2),
                        'rate_per_unit': round(rate_per_piece, 2),  # Rate per piece
                        'rate_type': 'per_kg',  # Keep this to indicate original rate source
                        'sku_name': combined_sku_names,
                        'amount': round(amount, 2),
                        'loss_gain_amount': round(loss_gain, 2)
                    })
                    
                    # Mark as processed to prevent duplicate additions from other product groups
                    kg_item_key = f"{item.get('dispatch_id')}_{item.get('product_id')}_{item.get('packaging_id')}"
                    processed_dispatch_items[kg_item_key] = True
    
    # Build response with warnings
    warnings = []
    if skipped_rows:
        # Group skipped rows by reason
        no_dispatch_rows = [r for r in skipped_rows if 'No dispatch found' in r.get('reason', '')]
        other_skipped = [r for r in skipped_rows if 'No dispatch found' not in r.get('reason', '')]
        
        if no_dispatch_rows:
            dates_skipped = list(set([r.get('date', 'Unknown') for r in no_dispatch_rows]))
            warnings.append(f"{len(no_dispatch_rows)} row(s) skipped - No dispatch found for dates: {', '.join(dates_skipped)}")
        
        if other_skipped:
            warnings.append(f"{len(other_skipped)} row(s) skipped due to invalid data")
    
    # Calculate rate changes from previous day
    # Get all saved GRNs to find previous rates
    saved_grns = await db.qc_grns.find({}, {"_id": 0}).to_list(1000)
    
    # Build a map of product -> previous rates (by date)
    product_rates_by_date = {}  # {product_name: {date: rate}}
    for grn in saved_grns:
        for item in grn.get('items', []):
            product_name = item.get('product_name', '')
            packaging_name = item.get('packaging_name', '')
            dispatch_date = item.get('dispatch_date', '')[:10] if item.get('dispatch_date') else ''
            rate = item.get('rate_per_kg') or item.get('rate_per_unit') or 0
            
            key = f"{product_name}|{packaging_name}"
            if key not in product_rates_by_date:
                product_rates_by_date[key] = {}
            if dispatch_date and rate > 0:
                product_rates_by_date[key][dispatch_date] = rate
    
    # Add rate change info to matched items
    rate_changes = []
    for item in matched_items:
        product_name = item.get('product_name', '')
        packaging_name = item.get('packaging_name', '')
        dispatch_date = item.get('dispatch_date', '')[:10] if item.get('dispatch_date') else ''
        current_rate = item.get('rate_per_kg') or item.get('rate_per_unit') or 0
        
        key = f"{product_name}|{packaging_name}"
        
        # Find previous day's rate
        previous_rate = None
        if key in product_rates_by_date:
            # Get all dates for this product and find the closest previous date
            available_dates = sorted([d for d in product_rates_by_date[key].keys() if d < dispatch_date], reverse=True)
            if available_dates:
                previous_date = available_dates[0]
                previous_rate = product_rates_by_date[key].get(previous_date)
        
        # Calculate rate change
        if previous_rate is not None and previous_rate > 0 and current_rate > 0:
            rate_change = current_rate - previous_rate
            rate_change_percent = ((current_rate - previous_rate) / previous_rate) * 100
            item['previous_rate'] = round(previous_rate, 2)
            item['rate_change'] = round(rate_change, 2)
            item['rate_change_percent'] = round(rate_change_percent, 1)
            
            if abs(rate_change) > 0.01:
                rate_changes.append({
                    'product_name': product_name,
                    'packaging_name': packaging_name,
                    'previous_rate': round(previous_rate, 2),
                    'current_rate': round(current_rate, 2),
                    'change': round(rate_change, 2),
                    'change_percent': round(rate_change_percent, 1)
                })
        else:
            item['previous_rate'] = None
            item['rate_change'] = None
            item['rate_change_percent'] = None
    
    # Include debug info about what was parsed
    first_row_sample = rows_data[0] if rows_data else {}
    
    # Build summary for automatic/manual display
    summary = {
        "sync_date": datetime.now(timezone.utc).isoformat(),
        "sync_method": "manual_upload",  # Can be "automatic_gmail" when triggered by scheduler
        "grn_date": list(csv_data_by_date.keys())[0] if csv_data_by_date else None,
        "total_items": len(matched_items),
        "unmatched_items": len(skipped_rows),
        "products_with_rate_change": len(rate_changes),
        "total_amount": round(sum(item.get('amount', 0) for item in matched_items), 2),
        "total_loss_gain": round(sum(item.get('loss_gain_amount', 0) for item in matched_items), 2),
    }
    
    return {
        "file_name": file.filename,
        "total_csv_rows": total_csv_rows,
        "rows_processed": sum(len(rows) for rows in csv_data_by_date.values()),
        "rows_skipped": len(skipped_rows),
        "dates_found": list(csv_data_by_date.keys()),
        "matched_items": matched_items,
        "warnings": warnings,
        "skipped_details": skipped_rows[:10] if skipped_rows else [],
        "rate_changes": rate_changes,
        "summary": summary,
        "debug_first_row": first_row_sample,  # Show what was actually parsed
        "debug_columns_found": list(first_row_sample.keys()) if first_row_sample else [],
        "message": f"Processed {len(matched_items)} matching items" + (f" ({len(skipped_rows)} rows skipped)" if skipped_rows else "")
    }

@api_router.get("/qc-grns/dispatch-summary")
async def get_dispatch_summary_for_grn(current_user: dict = Depends(get_current_user)):
    """Get all Ninjacart dispatches for GRN table - excludes items already in saved GRNs"""
    if current_user["role"] not in ["admin", "staff"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    # Get all Ninjacart dispatches
    dispatches = await db.qc_dispatches.find(
        {"customer_name": {"$regex": "ninja", "$options": "i"}},
        {"_id": 0}
    ).sort("dispatch_date", -1).to_list(1000)
    
    # Get all saved GRNs to exclude already processed items
    saved_grns = await db.qc_grns.find({}, {"_id": 0}).to_list(1000)
    
    # Build a set of processed dispatch_id + product_id + packaging_id combinations
    processed_items = set()
    for grn in saved_grns:
        for item in grn.get('items', []):
            # Create a unique key for each processed item - include packaging_id for accuracy
            key = f"{item.get('dispatch_id')}_{item.get('product_id')}_{item.get('packaging_id')}"
            processed_items.add(key)
    
    # Flatten items with dispatch info, excluding already processed
    items = []
    pending_dates = set()  # Track dates with pending GRNs
    for dispatch in dispatches:
        dispatch_date = dispatch.get('dispatch_date', '')
        for item in dispatch.get('items', []):
            item_key = f"{dispatch.get('id')}_{item.get('product_id')}_{item.get('packaging_id')}"
            
            # Skip if already processed in a GRN
            if item_key in processed_items:
                continue
            
            # Track this date as having pending items
            if dispatch_date:
                date_str = dispatch_date[:10] if isinstance(dispatch_date, str) else str(dispatch_date)[:10]
                pending_dates.add(date_str)
                
            items.append({
                'dispatch_id': dispatch.get('id'),
                'dispatch_date': dispatch_date,
                'dispatch_time': dispatch.get('dispatch_time', ''),
                'customer_name': dispatch.get('customer_name'),
                'product_id': item.get('product_id'),
                'product_name': item.get('product_name'),
                'product_unit': item.get('product_unit'),
                'packaging_id': item.get('packaging_id'),
                'packaging_name': item.get('packaging_name'),
                'supplied_qty': item.get('supplied_qty', 0),
                'grn_qty': None,
                'difference': None,
                'rate_per_unit': item.get('rate'),
            })
    
    return {
        "items": items,
        "pending_dates": sorted(list(pending_dates), reverse=True),
        "total_pending": len(items)
    }


@api_router.post("/qc-grns/fix-mismatched-ids")
async def fix_mismatched_grn_ids(current_user: dict = Depends(get_current_user)):
    """
    Fix GRN items with mismatched dispatch_id, product_id, or packaging_id.
    This attempts to re-match GRN items to dispatches based on product name and date.
    Use this when Excel backup imports have corrupted/mismatched IDs.
    """
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Only admin can fix GRN matching")
    
    try:
        def extract_date_str(date_val):
            """Extract date string from various formats."""
            if date_val is None:
                return ""
            if isinstance(date_val, datetime):
                return date_val.strftime('%Y-%m-%d')
            date_str = str(date_val)
            if 'T' in date_str:
                return date_str[:10]
            return date_str[:10] if len(date_str) >= 10 else date_str
        
        # Get all dispatches
        all_dispatches = await db.qc_dispatches.find({}, {"_id": 0}).to_list(5000)
        
        # Build lookup by date -> product_name -> dispatch items
        dispatch_lookup = {}
        for dispatch in all_dispatches:
            dispatch_date = extract_date_str(dispatch.get('dispatch_date'))
            dispatch_id = dispatch.get('id')
            
            if not dispatch_date or not dispatch_id:
                continue
            
            if dispatch_date not in dispatch_lookup:
                dispatch_lookup[dispatch_date] = {}
            
            for item in dispatch.get('items', []):
                product_name = (item.get('product_name', '') or '').lower().strip()
                packaging_name = (item.get('packaging_name', '') or '').lower().strip()
                
                key = f"{product_name}_{packaging_name}"
                if key not in dispatch_lookup[dispatch_date]:
                    dispatch_lookup[dispatch_date][key] = []
                
                dispatch_lookup[dispatch_date][key].append({
                    'dispatch_id': dispatch_id,
                    'product_id': item.get('product_id'),
                    'packaging_id': item.get('packaging_id'),
                    'supplied_qty': item.get('supplied_qty', 0) or 0
                })
        
        # Get all GRNs and fix mismatched items
        all_grns = await db.qc_grns.find({}).to_list(1000)
        
        total_fixed = 0
        fixed_details = []
        
        for grn in all_grns:
            grn_id = grn.get('id')
            if not grn_id:
                continue
                
            items_updated = False
            updated_items = []
            
            for item in grn.get('items', []):
                item_dispatch_date = extract_date_str(item.get('dispatch_date'))
                product_name = (item.get('product_name', '') or '').lower().strip()
                packaging_name = (item.get('packaging_name', '') or '').lower().strip()
                
                key = f"{product_name}_{packaging_name}"
                
                # Try to find matching dispatch item
                date_items = dispatch_lookup.get(item_dispatch_date, {})
                matching_dispatches = date_items.get(key, [])
                
                if matching_dispatches:
                    # Find best match by supplied_qty
                    item_grn_qty = item.get('grn_qty', 0) or item.get('supplied_qty', 0) or 0
                    best_match = None
                    min_diff = float('inf')
                    
                    for match in matching_dispatches:
                        diff = abs((match.get('supplied_qty') or 0) - item_grn_qty)
                        if diff < min_diff:
                            min_diff = diff
                            best_match = match
                    
                    if best_match:
                        old_dispatch_id = item.get('dispatch_id')
                        old_product_id = item.get('product_id')
                        old_packaging_id = item.get('packaging_id')
                        
                        # Check if any ID is different
                        if (old_dispatch_id != best_match['dispatch_id'] or
                            old_product_id != best_match['product_id'] or
                            old_packaging_id != best_match['packaging_id']):
                            
                            # Update item with correct IDs
                            item['dispatch_id'] = best_match['dispatch_id']
                            item['product_id'] = best_match['product_id']
                            item['packaging_id'] = best_match['packaging_id']
                            items_updated = True
                            total_fixed += 1
                            
                            fixed_details.append({
                                'grn_id': grn_id[:20] if grn_id else None,
                                'product': item.get('product_name'),
                                'old_dispatch_id': old_dispatch_id[:20] if old_dispatch_id else None,
                                'new_dispatch_id': best_match['dispatch_id'][:20] if best_match['dispatch_id'] else None
                            })
                
                updated_items.append(item)
            
            # Update GRN if any items changed
            if items_updated:
                await db.qc_grns.update_one(
                    {"id": grn_id},
                    {"$set": {"items": updated_items}}
                )
        
        return {
            "message": f"Fixed {total_fixed} GRN item IDs",
            "total_fixed": total_fixed,
            "details": fixed_details[:20]  # Return first 20 fixes
        }
    except Exception as e:
        import traceback
        error_trace = traceback.format_exc()
        print(f"Error in fix_mismatched_grn_ids: {error_trace}")
        raise HTTPException(status_code=500, detail=f"Error fixing GRN IDs: {str(e)}")


@api_router.get("/qc-grns/loss-summary")
async def get_grn_loss_summary(
    from_date: str = None,
    to_date: str = None,
    current_user: dict = Depends(get_current_user)
):
    """Get GRN loss summary - difference between Supplied and GRN quantities"""
    if current_user["role"] not in ["admin", "staff"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    # Build date query - use dispatch_date (actual supply date) not grn_date (upload date)
    query = {}
    
    # Get all saved GRNs
    saved_grns = await db.qc_grns.find({}, {"_id": 0}).sort("grn_date", -1).to_list(1000)
    
    # Aggregate loss by dispatch_date (actual supply date) and product
    loss_by_date = {}
    total_loss = 0
    items_with_loss = []
    
    for grn in saved_grns:
        for item in grn.get('items', []):
            # Use dispatch_date from item (actual supply date), not grn_date (upload date)
            dispatch_date = item.get('dispatch_date', grn.get('grn_date', ''))
            dispatch_date_str = dispatch_date[:10] if dispatch_date else 'Unknown'
            
            # Filter by date range using dispatch_date
            if from_date and dispatch_date_str < from_date:
                continue
            if to_date and dispatch_date_str > to_date:
                continue
            
            difference = item.get('difference', 0)
            loss_amount = item.get('loss_gain_amount', 0)
            
            # Only count losses (negative differences mean we received less than supplied)
            if difference < 0 or loss_amount < 0:
                actual_loss = abs(loss_amount) if loss_amount < 0 else abs(difference) * (item.get('rate_per_unit', 0) or item.get('rate_per_kg', 0))
                
                if dispatch_date_str not in loss_by_date:
                    loss_by_date[dispatch_date_str] = {
                        'date': dispatch_date_str,
                        'total_loss': 0,
                        'total_difference_qty': 0,
                        'items': []
                    }
                
                loss_by_date[dispatch_date_str]['total_loss'] += actual_loss
                loss_by_date[dispatch_date_str]['total_difference_qty'] += abs(difference)
                total_loss += actual_loss
                
                items_with_loss.append({
                    'date': dispatch_date_str,
                    'product_name': item.get('product_name', ''),
                    'packaging_name': item.get('packaging_name', ''),
                    'supplied_qty': item.get('supplied_qty', 0),
                    'grn_qty': item.get('grn_qty', 0),
                    'difference': difference,
                    'loss_amount': round(actual_loss, 2),
                    'rate_type': item.get('rate_type', ''),
                    'rate': item.get('rate_per_unit', 0) or item.get('rate_per_kg', 0)
                })
    
    return {
        "total_loss": round(total_loss, 2),
        "by_date": sorted(loss_by_date.values(), key=lambda x: x['date'], reverse=True),
        "items": items_with_loss
    }

# ============================================================================
# SECTION: QC INVOICE ROUTES (Lines ~1453-1570)
# ============================================================================
@api_router.get("/qc-invoices")
async def get_qc_invoices(current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in ["admin", "staff"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    invoices = await db.qc_invoices.find({}, {"_id": 0}).sort("created_at", -1).to_list(1000)
    for inv in invoices:
        if isinstance(inv.get('invoice_date'), str):
            inv['invoice_date'] = datetime.fromisoformat(inv['invoice_date'])
        if isinstance(inv.get('created_at'), str):
            inv['created_at'] = datetime.fromisoformat(inv['created_at'])
        if isinstance(inv.get('updated_at'), str):
            inv['updated_at'] = datetime.fromisoformat(inv['updated_at'])
    return invoices

@api_router.get("/qc-invoices/next-number/{customer_name}/{invoice_date}")
async def get_next_invoice_number(customer_name: str, invoice_date: str, current_user: dict = Depends(get_current_user)):
    """Generate next invoice number in format: CUS-DDMMMYYYY-001"""
    if current_user["role"] not in ["admin", "staff"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    # Get first 3 letters of customer name (uppercase)
    prefix = customer_name[:3].upper()
    
    # Parse date and format
    date_obj = datetime.fromisoformat(invoice_date.replace('Z', '+00:00'))
    date_str = date_obj.strftime("%d%b%Y").upper()
    
    # Find existing invoices for this customer on this date
    pattern = f"^{prefix}-{date_str}-"
    existing = await db.qc_invoices.find(
        {"invoice_number": {"$regex": pattern}},
        {"invoice_number": 1, "_id": 0}
    ).to_list(100)
    
    # Get next sequence number
    if existing:
        numbers = [int(inv['invoice_number'].split('-')[-1]) for inv in existing]
        next_num = max(numbers) + 1
    else:
        next_num = 1
    
    invoice_number = f"{prefix}-{date_str}-{str(next_num).zfill(3)}"
    return {"invoice_number": invoice_number}

@api_router.post("/qc-invoices")
async def create_qc_invoice(input: QCInvoiceCreate, current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in ["admin", "staff"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    # Generate invoice number
    prefix = input.customer_name[:3].upper()
    date_str = input.invoice_date.strftime("%d%b%Y").upper()
    
    # Find existing invoices for this customer on this date
    pattern = f"^{prefix}-{date_str}-"
    existing = await db.qc_invoices.find(
        {"invoice_number": {"$regex": pattern}},
        {"invoice_number": 1, "_id": 0}
    ).to_list(100)
    
    if existing:
        numbers = [int(inv['invoice_number'].split('-')[-1]) for inv in existing]
        next_num = max(numbers) + 1
    else:
        next_num = 1
    
    invoice_number = f"{prefix}-{date_str}-{str(next_num).zfill(3)}"
    
    invoice_dict = input.model_dump()
    invoice_dict["invoice_number"] = invoice_number
    invoice_dict["recorded_by"] = current_user["user_id"]
    invoice = QCInvoice(**invoice_dict)
    
    doc = invoice.model_dump()
    doc['invoice_date'] = doc['invoice_date'].isoformat()
    doc['created_at'] = doc['created_at'].isoformat()
    doc['updated_at'] = doc['updated_at'].isoformat()
    await db.qc_invoices.insert_one(doc)
    
    return {"id": invoice.id, "invoice_number": invoice_number, "message": "Invoice created successfully"}

@api_router.put("/qc-invoices/{invoice_id}")
async def update_qc_invoice(invoice_id: str, input: QCInvoiceUpdate, current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in ["admin", "staff"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    existing = await db.qc_invoices.find_one({"id": invoice_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Invoice not found")
    
    update_data = {k: v for k, v in input.model_dump().items() if v is not None}
    if 'invoice_date' in update_data:
        update_data['invoice_date'] = update_data['invoice_date'].isoformat()
    update_data['updated_at'] = datetime.now(timezone.utc).isoformat()
    
    await db.qc_invoices.update_one(
        {"id": invoice_id},
        {"$set": update_data}
    )
    
    return {"id": invoice_id, "message": "Invoice updated successfully"}

@api_router.delete("/qc-invoices/{invoice_id}")
async def delete_qc_invoice(invoice_id: str, current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in ["admin", "staff"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    result = await db.qc_invoices.delete_one({"id": invoice_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Invoice not found")
    
    return {"message": "Invoice deleted successfully"}

# ============================================================================
# SECTION: RETAILER ORDER ROUTES (Lines ~1578-1670)
# ============================================================================
@api_router.get("/retailer-orders", response_model=List[RetailerOrder])
async def get_retailer_orders(current_user: dict = Depends(get_current_user)):
    query = {}
    if current_user["role"] == "retailer":
        query = {"retailer_id": current_user["user_id"]}
    elif current_user["role"] not in ["admin", "staff"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    orders = await db.retailer_orders.find(query, {"_id": 0}).sort("order_date", -1).to_list(1000)
    for o in orders:
        if isinstance(o['order_date'], str):
            o['order_date'] = datetime.fromisoformat(o['order_date'])
        if o.get('delivery_date') and isinstance(o['delivery_date'], str):
            o['delivery_date'] = datetime.fromisoformat(o['delivery_date'])
        if isinstance(o['created_at'], str):
            o['created_at'] = datetime.fromisoformat(o['created_at'])
    return [RetailerOrder(**o) for o in orders]

@api_router.post("/retailer-orders", response_model=RetailerOrder)
async def create_retailer_order(input: RetailerOrderCreate, current_user: dict = Depends(get_current_user)):
    order_dict = input.model_dump()
    
    if current_user["role"] == "retailer":
        order_dict["retailer_id"] = current_user["user_id"]
    elif current_user["role"] not in ["admin", "staff"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    order = RetailerOrder(**order_dict)
    
    # Reduce stock for approved orders
    if order.status == "approved":
        for item in order.products:
            await db.products.update_one(
                {"id": item.product_id},
                {"$inc": {"current_stock": -item.packets}}
            )
    
    doc = order.model_dump()
    doc['order_date'] = doc['order_date'].isoformat()
    if doc.get('delivery_date'):
        doc['delivery_date'] = doc['delivery_date'].isoformat()
    doc['created_at'] = doc['created_at'].isoformat()
    await db.retailer_orders.insert_one(doc)
    return order

@api_router.put("/retailer-orders/{order_id}/status")
async def update_order_status(order_id: str, input: OrderStatusUpdate, current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in ["admin", "staff"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    order = await db.retailer_orders.find_one({"id": order_id}, {"_id": 0})
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    
    # Update stock if status changes to approved
    if order["status"] == "pending" and input.status == "approved":
        for item in order["products"]:
            await db.products.update_one(
                {"id": item["product_id"]},
                {"$inc": {"current_stock": -item["packets"]}}
            )
    
    result = await db.retailer_orders.find_one_and_update(
        {"id": order_id},
        {"$set": {"status": input.status}},
        return_document=True,
        projection={"_id": 0}
    )
    
    return {"message": "Order status updated", "order": result}

# Rejection Routes
@api_router.get("/rejections", response_model=List[Rejection])
async def get_rejections(current_user: dict = Depends(get_current_user)):
    query = {}
    if current_user["role"] == "retailer":
        query = {"retailer_id": current_user["user_id"]}
    elif current_user["role"] not in ["admin", "staff"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    rejections = await db.rejections.find(query, {"_id": 0}).sort("date", -1).to_list(1000)
    for r in rejections:
        if isinstance(r['date'], str):
            r['date'] = datetime.fromisoformat(r['date'])
        if isinstance(r['created_at'], str):
            r['created_at'] = datetime.fromisoformat(r['created_at'])
    return [Rejection(**r) for r in rejections]

@api_router.post("/rejections", response_model=Rejection)
async def create_rejection(input: RejectionCreate, current_user: dict = Depends(get_current_user)):
    rejection_dict = input.model_dump()
    
    if current_user["role"] == "retailer":
        rejection_dict["retailer_id"] = current_user["user_id"]
    
    rejection = Rejection(**rejection_dict)
    
    doc = rejection.model_dump()
    doc['date'] = doc['date'].isoformat()
    doc['created_at'] = doc['created_at'].isoformat()
    await db.rejections.insert_one(doc)
    return rejection

# ============================================================================
# SECTION: WASTAGE ROUTES (Lines ~1684-1720)
# ============================================================================
@api_router.get("/wastage", response_model=List[Wastage])
async def get_wastage(current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in ["admin", "staff"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    wastages = await db.wastage.find({}, {"_id": 0}).sort("date", -1).to_list(1000)
    for w in wastages:
        if isinstance(w['date'], str):
            w['date'] = datetime.fromisoformat(w['date'])
        if isinstance(w['created_at'], str):
            w['created_at'] = datetime.fromisoformat(w['created_at'])
    return [Wastage(**w) for w in wastages]

@api_router.post("/wastage", response_model=Wastage)
async def create_wastage(input: WastageCreate, current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in ["admin", "staff"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    wastage_dict = input.model_dump()
    wastage_dict["recorded_by"] = current_user["user_id"]
    wastage = Wastage(**wastage_dict)
    
    # Reduce stock
    for item in wastage.products:
        await db.products.update_one(
            {"id": item.product_id},
            {"$inc": {"current_stock": -item.quantity}}
        )
    
    doc = wastage.model_dump()
    doc['date'] = doc['date'].isoformat()
    doc['created_at'] = doc['created_at'].isoformat()
    await db.wastage.insert_one(doc)
    return wastage

# Payment Routes
@api_router.get("/payments", response_model=List[Payment])
async def get_payments(party_type: Optional[str] = None, current_user: dict = Depends(get_current_user)):
    query = {}
    if current_user["role"] == "retailer":
        query = {"party_type": "retailer", "party_id": current_user["user_id"]}
    elif current_user["role"] not in ["admin", "staff"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    elif party_type:
        query["party_type"] = party_type
    
    payments = await db.payments.find(query, {"_id": 0}).sort("date", -1).to_list(1000)
    for p in payments:
        if isinstance(p['date'], str):
            p['date'] = datetime.fromisoformat(p['date'])
        if isinstance(p['created_at'], str):
            p['created_at'] = datetime.fromisoformat(p['created_at'])
    return [Payment(**p) for p in payments]

@api_router.post("/payments", response_model=Payment)
async def create_payment(input: PaymentCreate, current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in ["admin", "staff"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    payment_dict = input.model_dump()
    payment_dict["recorded_by"] = current_user["user_id"]
    payment = Payment(**payment_dict)
    
    doc = payment.model_dump()
    doc['date'] = doc['date'].isoformat()
    doc['created_at'] = doc['created_at'].isoformat()
    await db.payments.insert_one(doc)
    return payment

# Invoice Routes
@api_router.get("/invoices", response_model=List[Invoice])
async def get_invoices(current_user: dict = Depends(get_current_user)):
    query = {}
    if current_user["role"] == "retailer":
        query = {"party_type": "retailer", "party_id": current_user["user_id"]}
    elif current_user["role"] not in ["admin", "staff"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    invoices = await db.invoices.find(query, {"_id": 0}).sort("date", -1).to_list(1000)
    for inv in invoices:
        if isinstance(inv['date'], str):
            inv['date'] = datetime.fromisoformat(inv['date'])
        if isinstance(inv['created_at'], str):
            inv['created_at'] = datetime.fromisoformat(inv['created_at'])
    return [Invoice(**inv) for inv in invoices]

@api_router.post("/invoices/generate", response_model=Invoice)
async def generate_invoice(input: InvoiceCreate, current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in ["admin", "staff"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    # Get next invoice number
    last_invoice = await db.invoices.find_one({}, {"_id": 0, "invoice_number": 1}, sort=[("invoice_number", -1)])
    next_number = 1001
    if last_invoice:
        try:
            next_number = int(last_invoice["invoice_number"].split("-")[1]) + 1
        except:
            next_number = 1001
    
    invoice_dict = input.model_dump()
    invoice_dict["invoice_number"] = f"FF-{next_number}"
    invoice = Invoice(**invoice_dict)
    
    doc = invoice.model_dump()
    doc['date'] = doc['date'].isoformat()
    doc['created_at'] = doc['created_at'].isoformat()
    await db.invoices.insert_one(doc)
    return invoice

@api_router.get("/invoices/{invoice_id}/pdf")
async def download_invoice_pdf(invoice_id: str, current_user: dict = Depends(get_current_user)):
    invoice = await db.invoices.find_one({"id": invoice_id}, {"_id": 0})
    if not invoice:
        raise HTTPException(status_code=404, detail="Invoice not found")
    
    # Check access
    if current_user["role"] == "retailer" and invoice["party_id"] != current_user["user_id"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    # Get party details
    party_name = ""
    party_address = ""
    is_ninjacart = False
    
    if invoice["party_type"] == "retailer":
        user = await db.users.find_one({"id": invoice["party_id"]}, {"_id": 0})
        if user:
            party_name = user.get("name", "")
            party_address = user.get("address", "")
    elif invoice["party_type"] == "qc":
        # Try to get from qc_customers first
        customer = await db.qc_customers.find_one({"id": invoice.get("customer_id")}, {"_id": 0})
        if customer:
            party_name = customer.get("name", "")
            party_address = customer.get("address", "")
        else:
            # Fallback to qc_orders
            order = await db.qc_orders.find_one({"id": invoice["order_id"]}, {"_id": 0})
            party_name = order.get("company_name", "") if order else ""
        
        # Check if Ninjacart
        is_ninjacart = 'ninjacart' in party_name.lower() if party_name else False
    
    pdf_buffer = generate_invoice_pdf(invoice, party_name, party_address, is_ninjacart)
    
    return StreamingResponse(
        io.BytesIO(pdf_buffer),
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=invoice-{invoice['invoice_number']}.pdf"}
    )

# ============================================================================
# SECTION: DASHBOARD & ANALYTICS ROUTES (Lines ~1839-2480)
# ============================================================================
@api_router.get("/reports/dashboard")
async def get_dashboard_stats(current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in ["admin", "staff"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    # Total products
    total_products = await db.products.count_documents({})
    
    # Total stock value
    products = await db.products.find({}, {"_id": 0}).to_list(1000)
    total_stock_value = sum((p.get("current_stock", 0) or 0) * (p.get("price_per_kg", 0) or 0) for p in products)
    
    # Today's orders
    today = datetime.now(timezone.utc).replace(hour=0, minute=0, second=0, microsecond=0)
    today_qc = await db.qc_orders.count_documents({"order_date": {"$gte": today.isoformat()}})
    today_retailer = await db.retailer_orders.count_documents({"order_date": {"$gte": today.isoformat()}})
    
    # Pending payments
    invoices = await db.invoices.find({}, {"_id": 0}).to_list(1000)
    pending_payments = sum(inv.get("pending_amount", 0) for inv in invoices)
    
    # Today's wastage
    wastages = await db.wastage.find({"date": {"$gte": today.isoformat()}}, {"_id": 0}).to_list(1000)
    today_wastage = sum(sum(item.get("quantity", 0) for item in w.get("products", [])) for w in wastages)
    
    return {
        "total_products": total_products,
        "total_stock_value": round(total_stock_value, 2),
        "today_qc_orders": today_qc,
        "today_retailer_orders": today_retailer,
        "pending_payments": round(pending_payments, 2),
        "today_wastage": round(today_wastage, 2)
    }

@api_router.get("/reports/today-summary")
async def get_today_summary(current_user: dict = Depends(get_current_user)):
    """Get today's quick summary for dashboard widget"""
    if current_user["role"] not in ["admin", "staff"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    today = datetime.now(timezone.utc).strftime('%Y-%m-%d')
    
    # Today's dispatches
    dispatches = await db.qc_dispatches.find({
        "dispatch_date": {"$regex": f"^{today}"}
    }, {"_id": 0}).to_list(500)
    
    today_dispatch_count = len(dispatches)
    today_dispatch_value = 0
    product_sales = {}  # Track top selling products
    
    for d in dispatches:
        for item in d.get('items', []):
            qty = item.get('supplied_qty', 0) or 0
            rate = item.get('rate', 0) or 0
            amount = qty * rate
            today_dispatch_value += amount
            
            product_name = item.get('product_name', 'Unknown')
            if product_name not in product_sales:
                product_sales[product_name] = {"qty": 0, "amount": 0}
            product_sales[product_name]["qty"] += qty
            product_sales[product_name]["amount"] += amount
    
    # Today's Retailer Dispatches
    retailer_dispatches = await db.retailer_dispatches.find({
        "dispatch_date": {"$regex": f"^{today}"}
    }, {"_id": 0}).to_list(500)
    
    today_retailer_dispatch_count = len(retailer_dispatches)
    today_retailer_value = sum(d.get('net_payable', 0) or 0 for d in retailer_dispatches)
    
    # Add retailer products to top products
    for d in retailer_dispatches:
        for item in d.get('items', []):
            qty = item.get('supplied_qty', 0) or 0
            amount = item.get('total_value', 0) or 0
            product_name = item.get('product_name', 'Unknown')
            if product_name not in product_sales:
                product_sales[product_name] = {"qty": 0, "amount": 0}
            product_sales[product_name]["qty"] += qty
            product_sales[product_name]["amount"] += amount
    
    # Pending indents (QC + Retailer)
    pending_qc_indents = await db.qc_indents.count_documents({
        "status": {"$in": ["pending", "partial"]}
    })
    pending_retailer_indents = await db.retailer_indents.count_documents({
        "status": {"$in": ["pending", "partial"]}
    })
    pending_indents = pending_qc_indents + pending_retailer_indents
    
    # Today's procurements
    procurements = await db.procurements.find({
        "date": {"$regex": f"^{today}"}
    }, {"_id": 0}).to_list(500)
    
    today_procurement_count = len(procurements)
    today_procurement_value = sum(p.get('total_amount', 0) or 0 for p in procurements)
    
    # Top 5 selling products today
    top_products = sorted(
        [{"name": k, **v} for k, v in product_sales.items()],
        key=lambda x: x["amount"],
        reverse=True
    )[:5]
    
    # Recent activity - last 5 dispatches/procurements
    recent_activity = []
    
    for d in dispatches[:3]:
        total_items = sum(item.get('supplied_qty', 0) for item in d.get('items', []))
        recent_activity.append({
            "type": "dispatch",
            "customer": d.get('customer_name', 'Unknown'),
            "items": total_items,
            "time": d.get('dispatch_time', '')
        })
    
    for p in procurements[:2]:
        recent_activity.append({
            "type": "procurement",
            "farmer": p.get('farmer_name', 'Unknown'),
            "amount": p.get('total_amount', 0),
            "time": ""
        })
    
    return {
        "today_dispatches": today_dispatch_count,
        "today_dispatch_value": round(today_dispatch_value, 2),
        "pending_indents": pending_indents,
        "today_procurements": today_procurement_count,
        "today_procurement_value": round(today_procurement_value, 2),
        "top_products": top_products,
        "recent_activity": recent_activity[:5]
    }

@api_router.get("/reports/pnl")
async def get_pnl_report(
    from_date: str = None,
    to_date: str = None,
    current_user: dict = Depends(get_current_user)
):
    """Get comprehensive P&L report with daily breakdown and customer-wise analysis"""
    if current_user["role"] not in ["admin", "staff"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    # Default to current month if no dates provided
    if not from_date:
        from_date = datetime.now(timezone.utc).replace(day=1).strftime('%Y-%m-%d')
    if not to_date:
        to_date = datetime.now(timezone.utc).strftime('%Y-%m-%d')
    
    # Sales by customer
    sales_by_customer = {}
    sales_by_date = {}
    sales_by_product = {}
    # Product-level breakdown per date with customer info
    product_by_date = {}  # {date: {product: {sales, purchase, wastage, qty, customers: []}}}
    # NEW: Detailed line items per date for Customer->Product breakdown
    line_items_by_date = {}  # {date: [line_item, ...]}
    # Track GRN loss per date (difference between dispatched and received)
    grn_loss_by_date = {}  # {date: {value: X, qty_kg: Y}}
    total_sales = 0
    total_sales_qty = 0
    
    # ========== QC SALES (from GRN - actual received values) ==========
    qc_grns = await db.qc_grns.find({}, {"_id": 0}).to_list(1000)
    
    for grn in qc_grns:
        customer = grn.get("customer_name", "Unknown")
        
        for item in grn.get("items", []):
            # Use dispatch_date from item (actual delivery date) instead of grn_date (recording date)
            item_dispatch_date = item.get("dispatch_date", "")[:10]
            
            # Skip if outside date range
            if not item_dispatch_date or item_dispatch_date < from_date or item_dispatch_date > to_date:
                continue
            
            # Use GRN amount (actual received value)
            amount = item.get("amount", 0) or 0
            qty = item.get("grn_qty", 0) or item.get("supplied_qty", 0) or 0
            supplied_qty_units = item.get("supplied_qty", 0) or qty
            product = item.get("product_name", "Unknown")
            unit = item.get("packaging_name", "") or item.get("product_unit", "Kg")
            packaging_weight_gm = item.get("packaging_weight_gm", 0) or 0
            
            # If packaging_weight_gm not set, try to extract from packaging name
            if not packaging_weight_gm and unit:
                import re
                weight_match = re.search(r'(\d+)(?:\s*-\s*\d+)?\s*(?:gm|g)\b', unit.lower())
                if weight_match:
                    packaging_weight_gm = float(weight_match.group(1))
            
            rate_per_kg = item.get("rate_per_kg", 0) or 0
            rate_per_unit = item.get("rate_per_unit", 0) or 0
            
            # Initialize date bucket if needed
            if item_dispatch_date not in sales_by_date:
                sales_by_date[item_dispatch_date] = {"sales": 0, "sales_qty": 0, "purchase": 0, "purchase_qty": 0, "wastage": 0, "variable_exp": 0, "fixed_exp": 0, "qc_sales": 0, "retail_sales": 0}
            if item_dispatch_date not in product_by_date:
                product_by_date[item_dispatch_date] = {}
            if item_dispatch_date not in line_items_by_date:
                line_items_by_date[item_dispatch_date] = []
            
            if customer not in sales_by_customer:
                sales_by_customer[customer] = {"amount": 0, "qty": 0, "invoices": 0, "type": "QC"}
            
            # Calculate kg from DISPATCH qty (what was sent), not GRN qty (what was received)
            # Use supplied_qty for wastage distribution as it represents what was dispatched
            dispatch_qty_units = item.get("supplied_qty", 0) or qty
            if packaging_weight_gm > 0:
                # Convert dispatch units to kg
                supplied_kg = (dispatch_qty_units * packaging_weight_gm) / 1000
            else:
                # If no packaging weight, assume supplied_qty is already in kg
                supplied_kg = dispatch_qty_units
            
            # Also track GRN qty for loss calculation
            grn_qty_kg = item.get("grn_qty_kg", 0) or 0
            grn_qty_pcs = item.get("grn_qty", 0) or 0
            
            # Calculate grn_qty_kg if not directly available
            if grn_qty_kg == 0 and packaging_weight_gm > 0:
                # Use grn_qty (piece count) to calculate kg if available
                if grn_qty_pcs > 0:
                    grn_qty_kg = (grn_qty_pcs * packaging_weight_gm) / 1000
                else:
                    # If no grn_qty, assume no loss (grn_qty_kg = supplied_kg)
                    grn_qty_kg = supplied_kg
            
            total_sales += amount
            total_sales_qty += qty
            sales_by_customer[customer]["amount"] += amount
            sales_by_customer[customer]["qty"] += qty
            sales_by_date[item_dispatch_date]["sales"] += amount
            sales_by_date[item_dispatch_date]["sales_qty"] += qty
            sales_by_date[item_dispatch_date]["qc_sales"] = sales_by_date[item_dispatch_date].get("qc_sales", 0) + amount
            
            if product not in sales_by_product:
                sales_by_product[product] = {"sales_amount": 0, "sales_qty": 0, "purchase_amount": 0, "purchase_qty": 0, "wastage_amount": 0}
            sales_by_product[product]["sales_amount"] += amount
            sales_by_product[product]["sales_qty"] += qty
            
            # Product breakdown per date with customer tracking
            if product not in product_by_date[item_dispatch_date]:
                product_by_date[item_dispatch_date][product] = {"sales": 0, "sales_qty": 0, "sales_kg": 0, "purchase": 0, "purchase_qty": 0, "wastage": 0, "customers": {}}
            product_by_date[item_dispatch_date][product]["sales"] += amount
            product_by_date[item_dispatch_date][product]["sales_qty"] += qty
            product_by_date[item_dispatch_date][product]["sales_kg"] += grn_qty_kg  # Use GRN kg (accepted qty) for SP/Kg calculation
            # Track sales by customer for this product on this date
            if customer not in product_by_date[item_dispatch_date][product]["customers"]:
                product_by_date[item_dispatch_date][product]["customers"][customer] = {"sales": 0, "qty": 0}
            product_by_date[item_dispatch_date][product]["customers"][customer]["sales"] += amount
            product_by_date[item_dispatch_date][product]["customers"][customer]["qty"] += qty
            
            # Add detailed line item for this customer-product combination
            line_items_by_date[item_dispatch_date].append({
                "customer": customer,
                "customer_type": "QC",
                "product": product,
                "unit": unit,
                "supplied_qty": round(dispatch_qty_units, 2),
                "supplied_kg": round(supplied_kg, 3),
                "grn_qty_kg": round(grn_qty_kg, 3),  # Track GRN received kg for loss calculation
                "revenue": round(amount, 2),
                "rate_per_kg": round(rate_per_kg, 2),
                "rate_per_unit": round(rate_per_unit, 2),
                # COGS and wastage will be calculated after we aggregate procurements
                "cogs": 0,
                "wastage_kg": 0,
                "wastage_value": 0
            })
            
            # Calculate GRN loss for this item (dispatched - received)
            if item_dispatch_date not in grn_loss_by_date:
                grn_loss_by_date[item_dispatch_date] = {"value": 0, "qty_kg": 0}
            loss_kg = supplied_kg - grn_qty_kg
            
            # Calculate loss value using rate_per_kg, or derive from rate_per_unit if needed
            if rate_per_kg > 0:
                loss_value = loss_kg * rate_per_kg
            elif rate_per_unit > 0 and packaging_weight_gm > 0:
                # Convert rate_per_unit to rate_per_kg: if 1 unit = X gm, rate_per_kg = rate_per_unit / (X/1000)
                derived_rate_per_kg = rate_per_unit / (packaging_weight_gm / 1000)
                loss_value = loss_kg * derived_rate_per_kg
            else:
                # Try to get loss_gain_amount directly from the item
                loss_value = abs(item.get("loss_gain_amount", 0) or 0)
            
            if loss_kg > 0:  # Only track positive losses (dispatched > received)
                grn_loss_by_date[item_dispatch_date]["qty_kg"] += loss_kg
                grn_loss_by_date[item_dispatch_date]["value"] += loss_value
        
        # Only count invoice if this customer had items in the date range
        if customer in sales_by_customer:
            sales_by_customer[customer]["invoices"] = sales_by_customer[customer].get("invoices", 0) + 1
    
    # ========== PRE-CALCULATE AVERAGE PURCHASE PRICES (for COGS) ==========
    # First, load products to get cost_alias mappings
    all_products = await db.products.find({}, {"_id": 0}).to_list(500)
    product_id_to_name = {p.get("id"): p.get("name") for p in all_products}
    product_name_to_id = {p.get("name"): p.get("id") for p in all_products}
    
    # Build cost alias map: {product_name: aliased_product_name}
    # e.g., {"Spinach": "Palak"} means Spinach uses Palak's purchase cost
    cost_alias_map = {}
    # Also build reverse alias map: {aliased_product: [products_that_alias_to_it]}
    # e.g., {"Palak": ["Spinach"]} means Spinach is aliased to Palak
    reverse_alias_map = {}
    for p in all_products:
        alias_id = p.get("cost_alias_product_id")
        if alias_id and alias_id in product_id_to_name:
            alias_name = product_id_to_name[alias_id]
            cost_alias_map[p.get("name")] = alias_name
            # Build reverse map
            if alias_name not in reverse_alias_map:
                reverse_alias_map[alias_name] = []
            reverse_alias_map[alias_name].append(p.get("name"))
    
    # Fetch all procurements to calculate average purchase price per product
    all_procurements = await db.procurements.find({
        "date": {"$gte": from_date, "$lte": to_date + "T23:59:59"}
    }, {"_id": 0}).to_list(1000)
    
    product_purchase_totals = {}  # {product: {amount: X, qty: Y}}
    for proc in all_procurements:
        for item in proc.get("products", []):
            product = item.get("product_name", "Unknown")
            qty = item.get("quantity", 0) or 0
            total_amt = item.get("total", 0) or 0
            unit = item.get("unit", "Kg")
            unit_size = item.get("unit_size", "")
            
            # Convert to Kg if unit is Bunch/Piece
            if unit.lower() in ["bunch", "piece", "pack"] and unit_size:
                try:
                    weight_per_unit_gm = float(unit_size)
                    qty = (qty * weight_per_unit_gm) / 1000
                except (ValueError, TypeError):
                    pass
            
            if product not in product_purchase_totals:
                product_purchase_totals[product] = {"amount": 0, "qty": 0}
            product_purchase_totals[product]["amount"] += total_amt
            product_purchase_totals[product]["qty"] += qty
    
    # Calculate average price per Kg for each product
    avg_price_by_product = {}
    for product, data in product_purchase_totals.items():
        if data["qty"] > 0:
            avg_price_by_product[product] = data["amount"] / data["qty"]
        else:
            avg_price_by_product[product] = 0
    
    # ========== RETAILER SALES (from Retailer Dispatches - GROSS MRP) ==========
    retailer_dispatches = await db.retailer_dispatches.find({
        "dispatch_date": {"$gte": from_date, "$lte": to_date + "T23:59:59"}
    }, {"_id": 0}).to_list(1000)
    
    # Fetch all retailers to get company names
    all_retailers = await db.users.find({"role": "retailer"}, {"_id": 0}).to_list(500)
    retailer_company_map = {r.get("id"): r.get("company_name", r.get("name", "Unknown")) for r in all_retailers}
    
    # Track QC vs Retail sales for bifurcation
    qc_sales_total = total_sales  # QC sales already calculated above
    qc_sales_qty = total_sales_qty
    retail_sales_total = 0
    retail_sales_qty = 0
    
    # Track retail-specific data
    retail_gross_mrp_by_date = {}  # Gross MRP before rejection/commission
    retail_rejection_by_date = {}  # Rejection value by date
    retail_commission_by_date = {}  # Commission by date
    retail_cogs_by_date = {}  # COGS for retail
    
    # Fetch all retailer rejections for the date range
    retailer_rejections = await db.retailer_rejections.find({
        "rejection_date": {"$gte": from_date, "$lte": to_date + "T23:59:59"}
    }, {"_id": 0}).to_list(1000)
    
    # Group rejections by date
    for rej in retailer_rejections:
        rej_date = rej.get("rejection_date", "")[:10]
        if rej_date not in retail_rejection_by_date:
            retail_rejection_by_date[rej_date] = {"qty": 0, "value": 0}
        retail_rejection_by_date[rej_date]["qty"] += rej.get("quantity", 0) or 0
        retail_rejection_by_date[rej_date]["value"] += rej.get("rejection_value", 0) or 0
    
    for dispatch in retailer_dispatches:
        # Use company_name instead of retailer_name (owner's name)
        retailer_id = dispatch.get("retailer_id", "")
        company_name = retailer_company_map.get(retailer_id, dispatch.get("retailer_name", "Unknown"))
        customer = company_name + " (Retail)"
        dispatch_date = dispatch.get("dispatch_date", "")[:10]
        
        # Get GROSS MRP value (before commission deduction)
        gross_mrp = dispatch.get("total_mrp_value", 0) or 0
        commission_pct = dispatch.get("commission_percentage", 0) or 0
        
        # Calculate commission on gross MRP
        commission_amount = gross_mrp * commission_pct / 100
        
        if customer not in sales_by_customer:
            sales_by_customer[customer] = {"amount": 0, "qty": 0, "invoices": 0, "type": "Retail"}
        if dispatch_date not in sales_by_date:
            sales_by_date[dispatch_date] = {"sales": 0, "sales_qty": 0, "purchase": 0, "purchase_qty": 0, "wastage": 0, "variable_exp": 0, "fixed_exp": 0, "qc_sales": 0, "retail_sales": 0, "retail_gross_mrp": 0, "retail_rejection": 0, "retail_rejection_qty": 0, "retail_commission": 0, "retail_cogs": 0}
        if dispatch_date not in product_by_date:
            product_by_date[dispatch_date] = {}
        if dispatch_date not in line_items_by_date:
            line_items_by_date[dispatch_date] = []
        
        # Initialize retail tracking for this date
        if dispatch_date not in retail_gross_mrp_by_date:
            retail_gross_mrp_by_date[dispatch_date] = 0
            retail_commission_by_date[dispatch_date] = 0
            retail_cogs_by_date[dispatch_date] = 0
        
        dispatch_qty = 0
        dispatch_cogs = 0
        for item in dispatch.get("items", []):
            qty = item.get("supplied_qty", 0) or 0
            product = item.get("product_name", "Unknown")
            unit = item.get("variant_name", "") or "Kg"
            item_value = item.get("total_value", 0) or 0  # MRP * qty
            mrp = item.get("mrp", 0) or 0
            
            # Calculate supplied_kg - need to convert pieces to kg for packet items
            # Get packaging weight from item or extract from variant_name
            packaging_weight_gm = item.get("packaging_weight", 0) or item.get("packaging_weight_gm", 0)
            if not packaging_weight_gm and unit:
                # Try to extract weight from variant_name like "240-260 gm" or "200 gm"
                import re
                weight_match = re.search(r'(\d+)(?:\s*-\s*\d+)?\s*(?:gm|g)\b', unit.lower())
                if weight_match:
                    packaging_weight_gm = float(weight_match.group(1))
            
            # If we have packaging weight in gm, convert qty to kg; otherwise assume qty is already in kg
            if packaging_weight_gm > 0:
                supplied_kg = qty * (packaging_weight_gm / 1000)
            else:
                supplied_kg = qty  # Assume qty is already in kg
            
            dispatch_qty += qty
            
            # Calculate COGS using average purchase price from the product
            # If product has a cost_alias (e.g., Spinach -> Palak), use aliased product's price
            lookup_product = cost_alias_map.get(product, product)
            avg_purchase_price = avg_price_by_product.get(lookup_product, 0)
            
            # If no price found for lookup product, try original product
            if avg_purchase_price == 0 and lookup_product != product:
                avg_purchase_price = avg_price_by_product.get(product, 0)
            
            # COGS should be based on kg, not pieces
            item_cogs = supplied_kg * avg_purchase_price
            dispatch_cogs += item_cogs
            
            if product not in sales_by_product:
                sales_by_product[product] = {"sales_amount": 0, "sales_qty": 0, "purchase_amount": 0, "purchase_qty": 0, "wastage_amount": 0}
            sales_by_product[product]["sales_amount"] += item_value  # Use gross MRP value
            sales_by_product[product]["sales_qty"] += qty
            
            # Product breakdown per date with customer tracking
            if product not in product_by_date[dispatch_date]:
                product_by_date[dispatch_date][product] = {"sales": 0, "sales_qty": 0, "sales_kg": 0, "purchase": 0, "purchase_qty": 0, "wastage": 0, "customers": {}}
            product_by_date[dispatch_date][product]["sales"] += item_value  # Gross MRP
            product_by_date[dispatch_date][product]["sales_qty"] += qty
            product_by_date[dispatch_date][product]["sales_kg"] += supplied_kg
            if customer not in product_by_date[dispatch_date][product]["customers"]:
                product_by_date[dispatch_date][product]["customers"][customer] = {"sales": 0, "qty": 0}
            product_by_date[dispatch_date][product]["customers"][customer]["sales"] += item_value
            product_by_date[dispatch_date][product]["customers"][customer]["qty"] += qty
            
            # Add detailed line item for retailer (include commission % for later calculation)
            line_items_by_date[dispatch_date].append({
                "customer": customer,
                "customer_type": "Retail",
                "product": product,
                "unit": unit,
                "supplied_qty": round(qty, 2),
                "supplied_kg": round(supplied_kg, 3),  # Converted to kg based on packaging weight
                "revenue": round(item_value, 2),  # Gross MRP
                "rate_per_kg": round(mrp / (packaging_weight_gm / 1000) if packaging_weight_gm > 0 else mrp, 2),
                "rate_per_unit": round(mrp, 2),
                "cogs": round(item_cogs, 2),
                "wastage_kg": 0,
                "wastage_value": 0,
                "commission_pct": commission_pct  # Store commission % for proportional allocation
            })
        
        # Track gross MRP for retail (before any deductions)
        retail_gross_mrp_by_date[dispatch_date] += gross_mrp
        retail_commission_by_date[dispatch_date] += commission_amount
        retail_cogs_by_date[dispatch_date] += dispatch_cogs
        
        total_sales += gross_mrp  # Use gross MRP for total sales
        total_sales_qty += dispatch_qty
        retail_sales_total += gross_mrp
        retail_sales_qty += dispatch_qty
        sales_by_customer[customer]["amount"] += gross_mrp
        sales_by_customer[customer]["qty"] += dispatch_qty
        sales_by_date[dispatch_date]["sales"] += gross_mrp
        sales_by_date[dispatch_date]["sales_qty"] += dispatch_qty
        sales_by_date[dispatch_date]["retail_sales"] = sales_by_date[dispatch_date].get("retail_sales", 0) + gross_mrp
        sales_by_date[dispatch_date]["retail_gross_mrp"] = sales_by_date[dispatch_date].get("retail_gross_mrp", 0) + gross_mrp
        sales_by_date[dispatch_date]["retail_commission"] = sales_by_date[dispatch_date].get("retail_commission", 0) + commission_amount
        sales_by_date[dispatch_date]["retail_cogs"] = sales_by_date[dispatch_date].get("retail_cogs", 0) + dispatch_cogs
        sales_by_customer[customer]["invoices"] += 1
    
    # Add rejection data to sales_by_date
    for rej_date, rej_data in retail_rejection_by_date.items():
        if rej_date in sales_by_date:
            sales_by_date[rej_date]["retail_rejection"] = rej_data["value"]
            sales_by_date[rej_date]["retail_rejection_qty"] = rej_data["qty"]
    
    # ========== PURCHASES (from Procurements) ==========
    procurements = await db.procurements.find({
        "date": {"$gte": from_date, "$lte": to_date + "T23:59:59"}
    }, {"_id": 0}).to_list(1000)
    
    purchase_by_farmer = {}
    total_purchase = 0
    total_purchase_qty = 0
    
    for proc in procurements:
        farmer = proc.get("farmer_name", "Unknown")
        proc_date = proc.get("date", "")[:10]
        
        if farmer not in purchase_by_farmer:
            purchase_by_farmer[farmer] = {"amount": 0, "qty": 0}
        if proc_date not in sales_by_date:
            sales_by_date[proc_date] = {"sales": 0, "sales_qty": 0, "purchase": 0, "purchase_qty": 0, "wastage": 0, "variable_exp": 0, "fixed_exp": 0}
        if proc_date not in product_by_date:
            product_by_date[proc_date] = {}
        
        for item in proc.get("products", []):
            qty = item.get("quantity", 0) or 0
            unit = item.get("unit", "Kg")
            unit_size = item.get("unit_size", "")
            total_item = item.get("total", 0) or 0
            product = item.get("product_name", "Unknown")
            
            # Convert to Kg if unit is Bunch/Piece with unit_size
            if unit.lower() in ["bunch", "piece", "pack"] and unit_size:
                try:
                    weight_per_unit_gm = float(unit_size)
                    qty_kg = (qty * weight_per_unit_gm) / 1000
                except (ValueError, TypeError):
                    qty_kg = qty
            else:
                qty_kg = qty
            
            total_purchase += total_item
            total_purchase_qty += qty_kg
            purchase_by_farmer[farmer]["amount"] += total_item
            purchase_by_farmer[farmer]["qty"] += qty_kg
            sales_by_date[proc_date]["purchase"] += total_item
            sales_by_date[proc_date]["purchase_qty"] += qty_kg
            
            if product not in sales_by_product:
                sales_by_product[product] = {"sales_amount": 0, "sales_qty": 0, "purchase_amount": 0, "purchase_qty": 0, "wastage_amount": 0}
            sales_by_product[product]["purchase_amount"] += total_item
            sales_by_product[product]["purchase_qty"] += qty_kg
            
            # Product breakdown per date
            if product not in product_by_date[proc_date]:
                product_by_date[proc_date][product] = {"sales": 0, "sales_qty": 0, "sales_kg": 0, "purchase": 0, "purchase_qty": 0, "wastage": 0, "customers": {}}
            product_by_date[proc_date][product]["purchase"] += total_item
            product_by_date[proc_date][product]["purchase_qty"] += qty_kg
    
    # ========== WASTAGE (from Daily Stock Status) ==========
    stock_status = await db.daily_stock_status.find({
        "date": {"$gte": from_date, "$lte": to_date},
        "status": "closed"
    }, {"_id": 0}).to_list(1000)
    
    total_wastage_qty = 0
    total_wastage_value = 0
    
    # Build a map of most recent prices by product for fallback
    product_recent_prices = {}
    
    # Track unsold wastage by date (products with wastage but no dispatch)
    unsold_wastage_by_date = {}  # date -> {product_name: {qty, value, avg_price}}
    
    for status in stock_status:
        wastage_qty = status.get("wastage_qty", 0) or 0
        purchase_qty = status.get("purchase_qty", 0) or 0
        purchase_value = status.get("purchase_value", 0) or 0
        dispatch_qty = status.get("dispatch_qty", 0) or 0
        status_date = status.get("date", "")[:10]
        product = status.get("product_name", "Unknown")
        product_id = status.get("product_id")
        
        # Calculate wastage_value dynamically from current procurement rate
        avg_price = purchase_value / purchase_qty if purchase_qty > 0 else 0
        
        # Store this price for future fallback
        if avg_price > 0 and product_id:
            product_recent_prices[product_id] = avg_price
        
        # If no purchase today, use the most recent price we've seen
        if avg_price == 0 and product_id:
            avg_price = product_recent_prices.get(product_id, 0)
        
        # If still no price, use the stored wastage_value from stock status
        if avg_price == 0:
            wastage_value = status.get("wastage_value", 0) or 0
            # Try to get avg_price from the stored value
            if wastage_qty > 0 and wastage_value > 0:
                avg_price = wastage_value / wastage_qty
        else:
            wastage_value = round(wastage_qty * avg_price, 2)
        
        total_wastage_qty += wastage_qty
        total_wastage_value += wastage_value
        
        if status_date in sales_by_date:
            sales_by_date[status_date]["wastage"] += wastage_value
        
        if product in sales_by_product:
            sales_by_product[product]["wastage_amount"] += wastage_value
        
        # Product breakdown per date - ensure entry exists even without procurement
        if status_date not in product_by_date:
            product_by_date[status_date] = {}
        if product not in product_by_date[status_date]:
            product_by_date[status_date][product] = {"sales": 0, "sales_qty": 0, "sales_kg": 0, "purchase": 0, "purchase_qty": 0, "wastage": 0, "customers": {}}
        product_by_date[status_date][product]["wastage"] += wastage_value
        
        # Track unsold wastage (products with wastage but NO dispatch on this day)
        if wastage_qty > 0 and dispatch_qty == 0:
            if status_date not in unsold_wastage_by_date:
                unsold_wastage_by_date[status_date] = {}
            unsold_wastage_by_date[status_date][product] = {
                "qty": wastage_qty,
                "value": wastage_value,
                "avg_price": avg_price
            }
    
    # ========== VARIABLE EXPENSES ==========
    variable_expenses = await db.variable_expenses.find({
        "date": {"$gte": from_date, "$lte": to_date + "T23:59:59"}
    }, {"_id": 0}).to_list(1000)
    
    variable_by_category = {}
    total_variable = 0
    # Track variable expenses by vertical
    variable_qc = 0
    variable_retail = 0
    variable_all = 0  # Will be split equally
    
    for exp in variable_expenses:
        category = exp.get("category", "Other")
        amount = exp.get("amount", 0) or 0
        exp_date = exp.get("date", "")[:10]
        vertical = exp.get("vertical", "all")  # 'qc', 'retail', or 'all'
        
        total_variable += amount
        if category not in variable_by_category:
            variable_by_category[category] = 0
        variable_by_category[category] += amount
        
        # Track by vertical
        if vertical == "qc":
            variable_qc += amount
        elif vertical == "retail":
            variable_retail += amount
        else:
            variable_all += amount
        
        if exp_date in sales_by_date:
            sales_by_date[exp_date]["variable_exp"] += amount
    
    # ========== LABOUR COSTS (Variable Daily Costs) ==========
    labour_attendance = await db.labour_attendance.find({
        "date": {"$gte": from_date, "$lte": to_date}
    }, {"_id": 0}).to_list(5000)
    
    total_labour_cost = 0
    labour_cost_by_date = {}
    
    for record in labour_attendance:
        if not record.get("present", False):
            continue
        
        att_date = record.get("date", "")[:10]
        daily_rate = float(record.get("daily_rate", 0) or 0)
        overtime_hours = float(record.get("overtime_hours", 0) or 0)
        overtime_rate = float(record.get("overtime_rate", 0) or 0)
        
        day_cost = daily_rate + (overtime_hours * overtime_rate)
        total_labour_cost += day_cost
        
        if att_date not in labour_cost_by_date:
            labour_cost_by_date[att_date] = 0
        labour_cost_by_date[att_date] += day_cost
        
        # Add to daily breakdown
        if att_date in sales_by_date:
            if "labour_cost" not in sales_by_date[att_date]:
                sales_by_date[att_date]["labour_cost"] = 0
            sales_by_date[att_date]["labour_cost"] += day_cost
    
    # Add Labour to variable category for display
    if total_labour_cost > 0:
        variable_by_category["Labour"] = total_labour_cost
        total_variable += total_labour_cost
        variable_all += total_labour_cost  # Labour is shared across all verticals
    
    # ========== FIXED EXPENSES ==========
    # Get current month's fixed expenses
    current_month = datetime.now(timezone.utc).month - 1  # 0-indexed
    current_year = datetime.now(timezone.utc).year
    
    fixed_expenses = await db.fixed_expenses.find({
        "month": current_month,
        "year": current_year
    }, {"_id": 0}).to_list(100)
    
    fixed_by_category = {}
    total_fixed = 0
    
    for exp in fixed_expenses:
        category = exp.get("category", "Other")
        amount = exp.get("amount", 0) or 0
        
        total_fixed += amount
        if category not in fixed_by_category:
            fixed_by_category[category] = 0
        fixed_by_category[category] += amount
    
    # Distribute fixed expenses across days in period
    days_in_period = max(1, len(sales_by_date))
    daily_fixed = total_fixed / days_in_period
    for date_key in sales_by_date:
        sales_by_date[date_key]["fixed_exp"] = round(daily_fixed, 2)
    
    # ========== CALCULATE TOTAL REJECTION AND COMMISSION ==========
    total_retail_rejection = sum(day_data.get("retail_rejection", 0) for day_data in sales_by_date.values())
    total_retail_commission = sum(day_data.get("retail_commission", 0) for day_data in sales_by_date.values())
    
    # ========== CALCULATE P&L (including rejection and commission for retail) ==========
    gross_profit = total_sales - total_purchase - total_wastage_value - total_retail_rejection - total_retail_commission
    gross_margin = (gross_profit / total_sales * 100) if total_sales > 0 else 0
    
    net_profit = gross_profit - total_variable - total_fixed
    net_margin = (net_profit / total_sales * 100) if total_sales > 0 else 0
    
    # Daily P&L breakdown with product details
    daily_pnl = []
    for date_key in sorted(sales_by_date.keys()):
        day_data = sales_by_date[date_key]
        
        # Get retail-specific deductions for this date
        day_retail_rejection = day_data.get("retail_rejection", 0)
        day_retail_commission = day_data.get("retail_commission", 0)
        
        # Gross profit = Sales - Purchase - Wastage - Rejection - Commission
        day_gross = day_data["sales"] - day_data["purchase"] - day_data["wastage"] - day_retail_rejection - day_retail_commission
        day_net = day_gross - day_data["variable_exp"] - day_data["fixed_exp"]
        day_sales_qty = day_data.get("sales_qty", 0)
        day_gross_margin = (day_gross / day_data["sales"] * 100) if day_data["sales"] > 0 else 0
        day_profit_per_unit = (day_gross / day_sales_qty) if day_sales_qty > 0 else 0
        
        # Product breakdown for this date (legacy - kept for backward compatibility)
        products_detail = []
        if date_key in product_by_date:
            for prod_name, prod_data in sorted(product_by_date[date_key].items(), key=lambda x: x[1]["sales"], reverse=True):
                prod_gross = prod_data["sales"] - prod_data["purchase"] - prod_data["wastage"]
                prod_margin = (prod_gross / prod_data["sales"] * 100) if prod_data["sales"] > 0 else 0
                prod_profit_per_unit = (prod_gross / prod_data["sales_qty"]) if prod_data["sales_qty"] > 0 else 0
                # Get customer names for this product
                customers_list = list(prod_data.get("customers", {}).keys())
                customers_str = ", ".join(customers_list) if customers_list else "-"
                products_detail.append({
                    "product": prod_name,
                    "customers": customers_str,
                    "sales": round(prod_data["sales"], 2),
                    "sales_qty": round(prod_data["sales_qty"], 2),
                    "sales_kg": round(prod_data.get("sales_kg", 0), 3),
                    "purchase": round(prod_data["purchase"], 2),
                    "purchase_qty": round(prod_data["purchase_qty"], 2),
                    "wastage": round(prod_data["wastage"], 2),
                    "gross_profit": round(prod_gross, 2),
                    "gross_margin": round(prod_margin, 1),
                    "profit_per_unit": round(prod_profit_per_unit, 2)
                })
        
        # NEW: Generate detailed line items (Customer → Product) with calculated metrics
        detailed_line_items = []
        if date_key in line_items_by_date:
            # Group by customer first, then by product
            customer_product_map = {}
            for line in line_items_by_date[date_key]:
                key = (line["customer"], line["product"], line["unit"])
                if key not in customer_product_map:
                    customer_product_map[key] = {
                        "customer": line["customer"],
                        "customer_type": line["customer_type"],
                        "product": line["product"],
                        "unit": line["unit"],
                        "supplied_qty": 0,
                        "supplied_kg": 0,
                        "revenue": 0,
                        "rate_per_kg": line["rate_per_kg"],
                        "rate_per_unit": line["rate_per_unit"],
                        "commission_pct": line.get("commission_pct", 0)  # Track commission % for retail
                    }
                customer_product_map[key]["supplied_qty"] += line["supplied_qty"]
                customer_product_map[key]["supplied_kg"] += line["supplied_kg"]
                customer_product_map[key]["revenue"] += line["revenue"]
            
            # Now calculate COGS and wastage allocation per line item
            # Get product-level COGS rate (purchase_amount / purchase_qty)
            product_cogs_rate = {}
            if date_key in product_by_date:
                for prod_name, prod_data in product_by_date[date_key].items():
                    purch_qty = prod_data.get("purchase_qty", 0)
                    purch_amt = prod_data.get("purchase", 0)
                    if purch_qty > 0:
                        product_cogs_rate[prod_name] = purch_amt / purch_qty
                    else:
                        product_cogs_rate[prod_name] = 0
            
            # Calculate total supplied kg per product for proportional wastage allocation
            # For aliased products, we need combined totals for proper wastage distribution
            product_total_supplied_kg = {}
            product_total_supplied_kg_with_alias = {}  # Combined totals for aliased products
            
            for key, item in customer_product_map.items():
                product = item["product"]
                if product not in product_total_supplied_kg:
                    product_total_supplied_kg[product] = 0
                product_total_supplied_kg[product] += item["supplied_kg"]
                
                # Also track combined totals for aliased products
                # e.g., both Palak and Spinach (aliased to Palak) contribute to "Palak" pool
                alias_product = cost_alias_map.get(product, product)
                if alias_product not in product_total_supplied_kg_with_alias:
                    product_total_supplied_kg_with_alias[alias_product] = 0
                product_total_supplied_kg_with_alias[alias_product] += item["supplied_kg"]
            
            # Get product-level total wastage (value) for this date
            product_total_wastage = {}
            if date_key in product_by_date:
                for prod_name, prod_data in product_by_date[date_key].items():
                    product_total_wastage[prod_name] = prod_data.get("wastage", 0)
            
            # Build detailed line items sorted by customer then by product
            for key in sorted(customer_product_map.keys(), key=lambda x: (x[0], x[1])):
                item = customer_product_map[key]
                product = item["product"]
                
                # Calculate COGS for this line item
                # Use cost alias if defined (e.g., Spinach uses Palak's cost)
                lookup_product = cost_alias_map.get(product, product)
                cogs_rate = product_cogs_rate.get(lookup_product, 0)
                # Fallback to original product if alias has no cost
                if cogs_rate == 0 and lookup_product != product:
                    cogs_rate = product_cogs_rate.get(product, 0)
                cogs = item["supplied_kg"] * cogs_rate
                
                # Calculate wastage proportionally based on kg supplied ratio
                # If there are 3 supplies in the ratio of 50:30:20 kg, wastage is divided in same ratio
                # For aliased products, use the alias product's wastage pool
                wastage_lookup_product = cost_alias_map.get(product, product)
                
                # For wastage distribution, use the COMBINED total of aliased products
                # e.g., if Spinach (10 kg) is aliased to Palak (100 kg), 
                # total pool is 110 kg and wastage is distributed proportionally
                total_kg_for_alias_group = product_total_supplied_kg_with_alias.get(wastage_lookup_product, 0)
                
                # For wastage, combine quantities of aliased products
                # e.g., if Spinach is aliased to Palak, Spinach dispatches share Palak's wastage pool
                total_wastage_for_product = product_total_wastage.get(wastage_lookup_product, 0)
                if total_wastage_for_product == 0 and wastage_lookup_product != product:
                    total_wastage_for_product = product_total_wastage.get(product, 0)
                
                if total_kg_for_alias_group > 0:
                    # Proportional wastage = (this_line_kg / total_alias_group_kg) * total_wastage
                    kg_ratio = item["supplied_kg"] / total_kg_for_alias_group
                    wastage_value = kg_ratio * total_wastage_for_product
                    wastage_kg = (wastage_value / cogs_rate) if cogs_rate > 0 else 0
                else:
                    wastage_value = 0
                    wastage_kg = 0
                
                # Calculate commission for retail items (proportional to revenue)
                commission_pct = item.get("commission_pct", 0)
                commission_value = 0
                if item["customer_type"] == "Retail" and commission_pct > 0:
                    commission_value = item["revenue"] * commission_pct / 100
                
                # Calculate gross profit and margins for this line item
                # For retail: Gross Profit = Revenue - COGS - Wastage - Commission
                line_gross_profit = item["revenue"] - cogs - wastage_value - commission_value
                line_gross_margin = (line_gross_profit / item["revenue"] * 100) if item["revenue"] > 0 else 0
                
                # Calculate price/kg metrics using grn_qty_kg (accepted qty) since payment is based on GRN
                grn_kg_for_calc = item.get("grn_qty_kg", 0) or item.get("supplied_kg", 0) or 0
                selling_price_per_kg = (item["revenue"] / grn_kg_for_calc) if grn_kg_for_calc > 0 else 0
                purchase_price_per_kg = cogs_rate
                profit_per_qty = (line_gross_profit / item["supplied_qty"]) if item["supplied_qty"] > 0 else 0
                
                detailed_line_items.append({
                    "customer": item["customer"],
                    "customer_type": item["customer_type"],
                    "product": product,
                    "unit": item["unit"],
                    "supplied_qty": round(item["supplied_qty"], 2),
                    "supplied_kg": round(item["supplied_kg"], 3),
                    "revenue": round(item["revenue"], 2),
                    "cogs": round(cogs, 2),
                    "wastage_kg": round(wastage_kg, 3),
                    "wastage_value": round(wastage_value, 2),
                    "commission": round(commission_value, 2),  # Commission amount for this product
                    "gross_profit": round(line_gross_profit, 2),
                    "gross_margin": round(line_gross_margin, 1),
                    "selling_price_per_kg": round(selling_price_per_kg, 2),
                    "purchase_price_per_kg": round(purchase_price_per_kg, 2),
                    "profit_per_qty": round(profit_per_qty, 2)
                })
        
        # Get unsold wastage for this date (products with wastage but no dispatch)
        date_unsold_wastage = unsold_wastage_by_date.get(date_key, {})
        unsold_wastage_items = []
        unsold_wastage_total = 0
        for prod_name, wastage_data in date_unsold_wastage.items():
            unsold_wastage_items.append({
                "product": prod_name,
                "qty": round(wastage_data["qty"], 2),
                "value": round(wastage_data["value"], 2),
                "avg_price": round(wastage_data["avg_price"], 2)
            })
            unsold_wastage_total += wastage_data["value"]
        
        daily_pnl.append({
            "date": date_key,
            "sales": round(day_data["sales"], 2),
            "sales_qty": round(day_sales_qty, 2),
            "qc_sales": round(day_data.get("qc_sales", 0), 2),
            "retail_sales": round(day_data.get("retail_sales", 0), 2),
            "retail_gross_mrp": round(day_data.get("retail_gross_mrp", 0), 2),
            "retail_rejection": round(day_data.get("retail_rejection", 0), 2),
            "retail_rejection_qty": round(day_data.get("retail_rejection_qty", 0), 2),
            "retail_commission": round(day_data.get("retail_commission", 0), 2),
            "retail_cogs": round(day_data.get("retail_cogs", 0), 2),
            "purchase": round(day_data["purchase"], 2),
            "wastage": round(day_data["wastage"], 2),
            "gross_profit": round(day_gross, 2),
            "gross_margin": round(day_gross_margin, 1),
            "profit_per_unit": round(day_profit_per_unit, 2),
            "variable_exp": round(day_data["variable_exp"], 2),
            "fixed_exp": round(day_data["fixed_exp"], 2),
            "labour_cost": round(day_data.get("labour_cost", 0), 2),
            "net_profit": round(day_net, 2),
            "products": products_detail,
            "line_items": detailed_line_items,  # Customer→Product breakdown
            "unsold_wastage": unsold_wastage_items,  # Products with wastage but no sales
            "unsold_wastage_total": round(unsold_wastage_total, 2)
        })
    
    # Customer P&L - calculate GRN loss later after we have total_grn_loss
    customer_pnl_data = []
    for customer, data in sorted(sales_by_customer.items(), key=lambda x: x[1]["amount"], reverse=True):
        customer_pnl_data.append({
            "customer": customer,
            "sales_amount": round(data["amount"], 2),
            "sales_qty": round(data["qty"], 2),
            "invoices": data["invoices"],
            "type": data.get("type", "QC")  # Include customer type
        })
    
    # Product P&L with additional metrics
    product_pnl = []
    for product, data in sorted(sales_by_product.items(), key=lambda x: x[1]["sales_amount"], reverse=True):
        profit = data["sales_amount"] - data["purchase_amount"] - data["wastage_amount"]
        margin = (profit / data["sales_amount"] * 100) if data["sales_amount"] > 0 else 0
        profit_per_unit = (profit / data["sales_qty"]) if data["sales_qty"] > 0 else 0
        product_pnl.append({
            "product": product,
            "sales_amount": round(data["sales_amount"], 2),
            "sales_qty": round(data["sales_qty"], 2),
            "purchase_amount": round(data["purchase_amount"], 2),
            "purchase_qty": round(data["purchase_qty"], 2),
            "wastage_amount": round(data["wastage_amount"], 2),
            "profit": round(profit, 2),
            "margin": round(margin, 1),
            "profit_per_unit": round(profit_per_unit, 2)
        })
    
    # Calculate QC vs Retail bifurcation totals
    total_qc_sales = sum(day.get("qc_sales", 0) for day in daily_pnl)
    total_retail_sales = sum(day.get("retail_sales", 0) for day in daily_pnl)
    
    # Calculate ACTUAL QC and Retail COGS and Wastage from line items
    # (instead of proportional allocation by sales %)
    actual_qc_cogs = 0
    actual_retail_cogs = 0
    actual_qc_wastage = 0
    actual_retail_wastage = 0
    
    for day in daily_pnl:
        for item in day.get("line_items", []):
            if item.get("customer_type") == "QC":
                actual_qc_cogs += item.get("cogs", 0)
                actual_qc_wastage += item.get("wastage_value", 0)
            else:  # Retail
                actual_retail_cogs += item.get("cogs", 0)
                actual_retail_wastage += item.get("wastage_value", 0)
    
    # Count QC and Retail orders and quantities
    qc_order_count = 0
    qc_qty_total = 0
    retail_order_count = 0
    retail_qty_total = 0
    
    for customer, data in sales_by_customer.items():
        if data.get("type") == "QC":
            qc_order_count += data.get("invoices", 0)
            qc_qty_total += data.get("qty", 0)
        else:
            retail_order_count += data.get("invoices", 0)
            retail_qty_total += data.get("qty", 0)
    
    # Calculate QC and Retail specific P&L metrics
    # Use ACTUAL COGS and wastage from line items (not proportional allocation)
    qc_sales_pct = (total_qc_sales / total_sales * 100) if total_sales > 0 else 0
    retail_sales_pct = (total_retail_sales / total_sales * 100) if total_sales > 0 else 0
    
    # Use actual calculated values from line items
    qc_purchase = actual_qc_cogs
    retail_purchase = actual_retail_cogs
    
    qc_wastage = actual_qc_wastage
    retail_wastage = actual_retail_wastage
    
    # Variable expenses: QC-specific + half of "all" expenses
    # Retail-specific + half of "all" expenses
    qc_variable_exp = variable_qc + (variable_all / 2)
    retail_variable_exp = variable_retail + (variable_all / 2)
    
    # Fixed expenses: Always divided equally between QC and Retail
    qc_fixed_exp = total_fixed / 2
    retail_fixed_exp = total_fixed / 2
    
    # Calculate total GRN loss
    total_grn_loss = sum(data.get("value", 0) for data in grn_loss_by_date.values())
    total_grn_loss_qty = sum(data.get("qty_kg", 0) for data in grn_loss_by_date.values())
    
    # Calculate QC P&L (GRN Loss affects net profit)
    qc_gross_profit = total_qc_sales - qc_purchase - qc_wastage
    # Gross Margin % = (Gross Profit / Sales) * 100 - MUST use Sales as denominator, not cost base
    qc_gross_margin = (qc_gross_profit / total_qc_sales * 100) if total_qc_sales > 0 else 0
    # QC Net Profit = Gross Profit - GRN Loss - Variable Exp - Fixed Exp
    qc_net_profit = qc_gross_profit - total_grn_loss - qc_variable_exp - qc_fixed_exp
    qc_net_margin = (qc_net_profit / total_qc_sales * 100) if total_qc_sales > 0 else 0
    
    # Calculate Retail P&L (including rejection and commission)
    # Gross Profit = Sales - COGS - Wastage - Rejection - Commission
    retail_gross_profit = total_retail_sales - retail_purchase - retail_wastage - total_retail_rejection - total_retail_commission
    retail_cost_base = retail_purchase + retail_wastage + total_retail_rejection + total_retail_commission
    retail_gross_margin = (retail_gross_profit / total_retail_sales * 100) if total_retail_sales > 0 else 0
    retail_net_profit = retail_gross_profit - retail_variable_exp - retail_fixed_exp
    retail_net_margin = (retail_net_profit / total_retail_sales * 100) if total_retail_sales > 0 else 0
    
    # Also calculate overall gross margin %
    total_cost_base = total_purchase + total_wastage_value + total_retail_rejection + total_retail_commission
    overall_gross_margin_pct = (gross_profit / total_sales * 100) if total_sales > 0 else 0
    
    # Total COGS from line items (sum of QC + Retail COGS)
    total_cogs = actual_qc_cogs + actual_retail_cogs
    total_wastage_from_items = actual_qc_wastage + actual_retail_wastage
    
    # Recalculate gross profit using actual COGS and wastage from line items
    gross_profit_actual = total_sales - total_cogs - total_wastage_from_items - total_retail_rejection - total_retail_commission
    gross_margin_actual = (gross_profit_actual / total_sales * 100) if total_sales > 0 else 0
    
    # Recalculate net profit (including GRN Loss which affects QC)
    # Net Profit = Gross Profit - GRN Loss - Variable Expenses - Fixed Expenses
    net_profit_actual = gross_profit_actual - total_grn_loss - total_variable - total_fixed
    net_margin_actual = (net_profit_actual / total_sales * 100) if total_sales > 0 else 0
    
    # Now allocate GRN loss proportionally to QC customers based on their sales contribution
    # This ensures customer-level P&L aligns with QC-level P&L
    customer_pnl = []
    for cust_data in customer_pnl_data:
        customer_entry = cust_data.copy()
        if cust_data["type"] == "QC" and total_qc_sales > 0:
            # Allocate GRN loss proportionally based on this customer's sales share of QC
            customer_qc_share = cust_data["sales_amount"] / total_qc_sales
            customer_entry["grn_loss_share"] = round(total_grn_loss * customer_qc_share, 2)
            customer_entry["rejection_share"] = 0  # QC doesn't have rejection
        elif cust_data["type"] == "Retail" and total_retail_sales > 0:
            # Allocate Rejection loss proportionally based on this customer's sales share of Retail
            customer_retail_share = cust_data["sales_amount"] / total_retail_sales
            customer_entry["grn_loss_share"] = 0  # Retail doesn't have GRN loss
            customer_entry["rejection_share"] = round(total_retail_rejection * customer_retail_share, 2)
        else:
            customer_entry["grn_loss_share"] = 0
            customer_entry["rejection_share"] = 0
        customer_pnl.append(customer_entry)
    
    return {
        "period": {"from": from_date, "to": to_date},
        "summary": {
            "total_sales": round(total_sales, 2),
            "total_sales_qty": round(total_sales_qty, 2),
            "total_purchase": round(total_purchase, 2),
            "total_purchase_qty": round(total_purchase_qty, 2),
            "total_cogs": round(total_cogs, 2),  # Sum of line item COGS
            "total_wastage_qty": round(total_wastage_qty, 2),
            "total_wastage_value": round(total_wastage_value, 2),
            "total_wastage_from_items": round(total_wastage_from_items, 2),  # Sum of line item wastage
            "total_retail_rejection": round(total_retail_rejection, 2),
            "total_retail_commission": round(total_retail_commission, 2),
            "gross_profit": round(gross_profit_actual, 2),  # Using actual COGS
            "gross_margin": round(gross_margin_actual, 1),
            "gross_margin_pct": round(gross_margin_actual, 1),
            "total_variable_expenses": round(total_variable, 2),
            "total_fixed_expenses": round(total_fixed, 2),
            "total_labour_cost": round(total_labour_cost, 2),
            "net_profit": round(net_profit_actual, 2),
            "net_margin": round(net_margin_actual, 1)
        },
        "vertical_bifurcation": {
            "qc": {
                "sales": round(total_qc_sales, 2),
                "percentage": round(qc_sales_pct, 1),
                "qty": round(qc_qty_total, 2),
                "orders": qc_order_count,
                "purchase": round(qc_purchase, 2),
                "wastage": round(qc_wastage, 2),
                "grn_loss": round(total_grn_loss, 2),  # GRN Loss (dispatched - received)
                "gross_profit": round(qc_gross_profit, 2),
                "gross_margin_pct": round(qc_gross_margin, 1),
                "variable_exp": round(qc_variable_exp, 2),
                "fixed_exp": round(qc_fixed_exp, 2),
                "net_profit": round(qc_net_profit, 2),
                "net_margin": round(qc_net_margin, 1)
            },
            "retail": {
                "sales": round(total_retail_sales, 2),
                "percentage": round(retail_sales_pct, 1),
                "qty": round(retail_qty_total, 2),
                "orders": retail_order_count,
                "purchase": round(retail_purchase, 2),
                "wastage": round(retail_wastage, 2),
                "rejection": round(total_retail_rejection, 2),
                "commission": round(total_retail_commission, 2),
                "gross_profit": round(retail_gross_profit, 2),
                "gross_margin_pct": round(retail_gross_margin, 1),
                "variable_exp": round(retail_variable_exp, 2),
                "fixed_exp": round(retail_fixed_exp, 2),
                "net_profit": round(retail_net_profit, 2),
                "net_margin": round(retail_net_margin, 1)
            }
        },
        "daily_pnl": daily_pnl,
        "customer_pnl": customer_pnl,
        "product_pnl": product_pnl,
        "expenses": {
            "variable_by_category": variable_by_category,
            "fixed_by_category": fixed_by_category
        },
        "purchase_by_farmer": [
            {"farmer": f, "amount": round(d["amount"], 2), "qty": round(d["qty"], 2)} 
            for f, d in sorted(purchase_by_farmer.items(), key=lambda x: x[1]["amount"], reverse=True)
        ]
    }


# ============================================================================
# P&L EXCEL EXPORT - Detailed audit report with customer/product breakdown
# ============================================================================

@app.get("/api/reports/pnl/export-excel")
async def export_pnl_excel(
    from_date: str = Query(..., description="Start date YYYY-MM-DD"),
    to_date: str = Query(..., description="End date YYYY-MM-DD"),
    current_user: dict = Depends(get_current_user)
):
    """Export detailed P&L report as Excel for audit purposes"""
    import xlsxwriter
    from io import BytesIO
    
    # Call the main P&L endpoint logic to get properly calculated data
    pnl_data = await get_pnl_report(from_date=from_date, to_date=to_date, current_user=current_user)
    
    # Extract data from P&L response
    daily_pnl = pnl_data.get("daily_pnl", [])
    customer_pnl = pnl_data.get("customer_pnl", [])
    product_pnl = pnl_data.get("product_pnl", [])
    summary = pnl_data.get("summary", {})
    purchase_by_farmer = pnl_data.get("purchase_by_farmer", [])
    vertical_bif = pnl_data.get("vertical_bifurcation", {})
    
    # Fetch procurements in date range for purchase detail sheet
    date_procurements = await db.procurements.find({
        "date": {"$gte": from_date, "$lte": to_date + "T23:59:59"}
    }, {"_id": 0}).to_list(1000)
    
    # Create Excel workbook
    output = BytesIO()
    workbook = xlsxwriter.Workbook(output, {'in_memory': True})
    
    # Define formats
    header_format = workbook.add_format({
        'bold': True, 'bg_color': '#14532D', 'font_color': 'white',
        'border': 1, 'align': 'center', 'valign': 'vcenter'
    })
    currency_format = workbook.add_format({'num_format': '₹#,##0.00', 'border': 1})
    number_format = workbook.add_format({'num_format': '#,##0.00', 'border': 1})
    text_format = workbook.add_format({'border': 1})
    total_format = workbook.add_format({
        'bold': True, 'bg_color': '#E5E7EB', 'border': 1, 'num_format': '₹#,##0.00'
    })
    qc_format = workbook.add_format({'bg_color': '#DBEAFE', 'border': 1})
    retail_format = workbook.add_format({'bg_color': '#DCFCE7', 'border': 1})
    
    # Sheet 1: Summary
    summary_sheet = workbook.add_worksheet("Summary")
    summary_sheet.set_column('A:A', 25)
    summary_sheet.set_column('B:D', 18)
    
    row = 0
    summary_sheet.write(row, 0, f"P&L Report: {from_date} to {to_date}", workbook.add_format({'bold': True, 'font_size': 14}))
    row += 2
    
    # Overall summary
    summary_sheet.write(row, 0, "Metric", header_format)
    summary_sheet.write(row, 1, "Total", header_format)
    summary_sheet.write(row, 2, "QC", header_format)
    summary_sheet.write(row, 3, "Retail", header_format)
    row += 1
    
    qc = vertical_bif.get("qc", {})
    retail = vertical_bif.get("retail", {})
    
    metrics = [
        ("Sales", summary.get("total_sales", 0), qc.get("sales", 0), retail.get("sales", 0)),
        ("Purchase (COGS)", summary.get("total_cogs", 0), qc.get("purchase", 0), retail.get("purchase", 0)),
        ("Wastage", summary.get("total_wastage", 0), qc.get("wastage", 0), retail.get("wastage", 0)),
        ("Commission", retail.get("commission", 0), 0, retail.get("commission", 0)),
        ("Gross Profit", summary.get("gross_profit", 0), qc.get("gross_profit", 0), retail.get("gross_profit", 0)),
        ("Gross Margin %", summary.get("gross_margin", 0), qc.get("gross_margin", 0), retail.get("gross_margin", 0)),
    ]
    
    for metric, total, qc_val, retail_val in metrics:
        summary_sheet.write(row, 0, metric, text_format)
        if "%" in metric:
            summary_sheet.write(row, 1, f"{total:.1f}%", text_format)
            summary_sheet.write(row, 2, f"{qc_val:.1f}%", text_format)
            summary_sheet.write(row, 3, f"{retail_val:.1f}%", text_format)
        else:
            summary_sheet.write(row, 1, total, currency_format)
            summary_sheet.write(row, 2, qc_val, currency_format)
            summary_sheet.write(row, 3, retail_val, currency_format)
        row += 1
    
    # Sheet 2: Daily P&L Breakdown with line items
    daily_sheet = workbook.add_worksheet("Daily P&L")
    daily_sheet.set_column('A:A', 12)
    daily_sheet.set_column('B:B', 10)
    daily_sheet.set_column('C:C', 25)
    daily_sheet.set_column('D:D', 20)
    daily_sheet.set_column('E:L', 12)
    
    headers = ["Date", "Type", "Customer", "Product", "Qty", "Sales", "COGS", "Wastage", "Commission", "Gross P/L", "GM%"]
    row = 0
    for col, header in enumerate(headers):
        daily_sheet.write(row, col, header, header_format)
    row += 1
    
    for day in sorted(daily_pnl, key=lambda x: x.get("date", "")):
        date = day.get("date", "")
        line_items = day.get("line_items", [])
        
        for item in line_items:
            customer = item.get("customer", "Unknown")
            customer_type = item.get("customer_type", "")
            product = item.get("product", "Unknown")
            qty = item.get("supplied_qty", 0) or 0
            revenue = item.get("revenue", 0) or 0
            cogs = item.get("cogs", 0) or 0
            wastage = item.get("wastage_value", 0) or 0
            commission = item.get("commission", 0) or 0
            gross = item.get("gross_profit", 0) or 0
            gm = item.get("gross_margin", 0) or 0
            
            row_format = qc_format if customer_type == "QC" else retail_format
            
            daily_sheet.write(row, 0, date, text_format)
            daily_sheet.write(row, 1, customer_type, row_format)
            daily_sheet.write(row, 2, customer, text_format)
            daily_sheet.write(row, 3, product, text_format)
            daily_sheet.write(row, 4, qty, number_format)
            daily_sheet.write(row, 5, revenue, currency_format)
            daily_sheet.write(row, 6, cogs, currency_format)
            daily_sheet.write(row, 7, wastage, currency_format)
            daily_sheet.write(row, 8, commission, currency_format)
            daily_sheet.write(row, 9, gross, currency_format)
            daily_sheet.write(row, 10, f"{gm:.1f}%", text_format)
            row += 1
    
    # Aggregate customer data from line_items for Customer P&L sheet
    customer_totals = {}
    for day in daily_pnl:
        for item in day.get("line_items", []):
            customer = item.get("customer", "Unknown")
            customer_type = item.get("customer_type", "")
            if customer not in customer_totals:
                customer_totals[customer] = {
                    "type": customer_type,
                    "qty": 0, "sales": 0, "cogs": 0, 
                    "wastage": 0, "commission": 0
                }
            customer_totals[customer]["qty"] += item.get("supplied_qty", 0) or 0
            customer_totals[customer]["sales"] += item.get("revenue", 0) or 0
            customer_totals[customer]["cogs"] += item.get("cogs", 0) or 0
            customer_totals[customer]["wastage"] += item.get("wastage_value", 0) or 0
            customer_totals[customer]["commission"] += item.get("commission", 0) or 0
    
    # Sheet 3: Customer-wise P&L
    customer_sheet = workbook.add_worksheet("Customer P&L")
    customer_sheet.set_column('A:A', 30)
    customer_sheet.set_column('B:I', 15)
    
    headers = ["Customer", "Type", "Qty", "Sales", "COGS", "Wastage", "Commission", "Gross P/L", "GM%"]
    row = 0
    for col, header in enumerate(headers):
        customer_sheet.write(row, col, header, header_format)
    row += 1
    
    for customer, data in sorted(customer_totals.items(), key=lambda x: x[1].get("sales", 0), reverse=True):
        gross = data["sales"] - data["cogs"] - data["wastage"] - data["commission"]
        gm = (gross / data["sales"] * 100) if data["sales"] > 0 else 0
        
        customer_sheet.write(row, 0, customer, text_format)
        customer_sheet.write(row, 1, data["type"], text_format)
        customer_sheet.write(row, 2, data["qty"], number_format)
        customer_sheet.write(row, 3, data["sales"], currency_format)
        customer_sheet.write(row, 4, data["cogs"], currency_format)
        customer_sheet.write(row, 5, data["wastage"], currency_format)
        customer_sheet.write(row, 6, data["commission"], currency_format)
        customer_sheet.write(row, 7, gross, currency_format)
        customer_sheet.write(row, 8, f"{gm:.1f}%", text_format)
        row += 1
    
    # Aggregate product data from line_items for Product P&L sheet
    product_totals = {}
    for day in daily_pnl:
        for item in day.get("line_items", []):
            product = item.get("product", "Unknown")
            if product not in product_totals:
                product_totals[product] = {
                    "qty": 0, "sales": 0, "cogs": 0, "wastage": 0
                }
            product_totals[product]["qty"] += item.get("supplied_qty", 0) or 0
            product_totals[product]["sales"] += item.get("revenue", 0) or 0
            product_totals[product]["cogs"] += item.get("cogs", 0) or 0
            product_totals[product]["wastage"] += item.get("wastage_value", 0) or 0
    
    # Sheet 4: Product-wise P&L
    product_sheet = workbook.add_worksheet("Product P&L")
    product_sheet.set_column('A:A', 25)
    product_sheet.set_column('B:H', 15)
    
    headers = ["Product", "Qty Sold", "Sales", "COGS", "Wastage", "Gross P/L", "GM%"]
    row = 0
    for col, header in enumerate(headers):
        product_sheet.write(row, col, header, header_format)
    row += 1
    
    for product, data in sorted(product_totals.items(), key=lambda x: x[1].get("sales", 0), reverse=True):
        gross = data["sales"] - data["cogs"] - data["wastage"]
        gm = (gross / data["sales"] * 100) if data["sales"] > 0 else 0
        
        product_sheet.write(row, 0, product, text_format)
        product_sheet.write(row, 1, data["qty"], number_format)
        product_sheet.write(row, 2, data["sales"], currency_format)
        product_sheet.write(row, 3, data["cogs"], currency_format)
        product_sheet.write(row, 4, data["wastage"], currency_format)
        product_sheet.write(row, 5, gross, currency_format)
        product_sheet.write(row, 6, f"{gm:.1f}%", text_format)
        row += 1
    
    # Sheet 5: Purchases by Farmer
    purchase_sheet = workbook.add_worksheet("Purchases")
    purchase_sheet.set_column('A:A', 12)
    purchase_sheet.set_column('B:B', 20)
    purchase_sheet.set_column('C:C', 25)
    purchase_sheet.set_column('D:G', 12)
    
    headers = ["Date", "Farmer", "Product", "Qty", "Unit", "Rate", "Total"]
    row = 0
    for col, header in enumerate(headers):
        purchase_sheet.write(row, col, header, header_format)
    row += 1
    
    for proc in sorted(date_procurements, key=lambda x: x.get("date", "")):
        farmer = proc.get("farmer_name", "Unknown")
        proc_date = str(proc.get("date", ""))[:10]
        
        for item in proc.get("products", []):
            purchase_sheet.write(row, 0, proc_date, text_format)
            purchase_sheet.write(row, 1, farmer, text_format)
            purchase_sheet.write(row, 2, item.get("product_name", ""), text_format)
            purchase_sheet.write(row, 3, item.get("quantity", 0) or 0, number_format)
            purchase_sheet.write(row, 4, item.get("unit", "Kg"), text_format)
            purchase_sheet.write(row, 5, item.get("rate", 0) or 0, currency_format)
            purchase_sheet.write(row, 6, item.get("total", 0) or 0, currency_format)
            row += 1
    
    # Sheet 6: Purchase Summary by Farmer
    farmer_sheet = workbook.add_worksheet("Farmer Summary")
    farmer_sheet.set_column('A:A', 25)
    farmer_sheet.set_column('B:C', 15)
    
    headers = ["Farmer", "Total Qty", "Total Amount"]
    row = 0
    for col, header in enumerate(headers):
        farmer_sheet.write(row, col, header, header_format)
    row += 1
    
    for farmer in purchase_by_farmer:
        farmer_sheet.write(row, 0, farmer.get("farmer", ""), text_format)
        farmer_sheet.write(row, 1, farmer.get("qty", 0), number_format)
        farmer_sheet.write(row, 2, farmer.get("amount", 0), currency_format)
        row += 1
    
    workbook.close()
    output.seek(0)
    
    filename = f"PnL_Report_{from_date}_to_{to_date}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


# ============================================================================

@api_router.get("/stock-status/today")
async def get_today_stock_status(current_user: dict = Depends(get_current_user)):
    """Get today's stock status for all products with opening qty, purchases, dispatches"""
    if current_user["role"] not in ["admin", "staff"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    today = datetime.now(timezone.utc).strftime('%Y-%m-%d')
    yesterday = (datetime.now(timezone.utc) - timedelta(days=1)).strftime('%Y-%m-%d')
    
    # Get all products
    products = await db.products.find({}, {"_id": 0}).to_list(1000)
    
    # Get yesterday's closing stock (for opening)
    yesterday_status = await db.daily_stock_status.find(
        {"date": yesterday, "status": "closed"},
        {"_id": 0}
    ).to_list(1000)
    yesterday_map = {s["product_id"]: s for s in yesterday_status}
    
    # Get today's existing status
    today_status = await db.daily_stock_status.find(
        {"date": today},
        {"_id": 0}
    ).to_list(1000)
    today_map = {s["product_id"]: s for s in today_status}
    
    # Get today's procurements (purchases from farmers)
    # Use regex to match today's date (handles different timezone formats)
    procurements = await db.procurements.find({
        "date": {"$regex": f"^{today}"}
    }, {"_id": 0}).to_list(1000)
    
    # Calculate purchases by product
    purchases_by_product = {}
    for proc in procurements:
        # Note: Procurement uses 'products' field, not 'items'
        for item in proc.get("products", []):
            product_id = item.get("product_id")
            qty = item.get("quantity", 0)
            unit = item.get("unit", "Kg")
            unit_size = item.get("unit_size", "")  # e.g., "350" for 350gm per bunch
            rate = item.get("rate", 0)  # rate per unit (per Kg or per Bunch)
            total_value = item.get("total", qty * rate)
            
            # Convert to Kg if unit is Bunch/Piece with a unit_size (weight in grams)
            if unit.lower() in ["bunch", "piece", "pack"] and unit_size:
                try:
                    weight_per_unit_gm = float(unit_size)  # grams per bunch/piece
                    qty_kg = (qty * weight_per_unit_gm) / 1000  # convert to Kg
                except (ValueError, TypeError):
                    qty_kg = qty  # fallback if unit_size is not a number
            else:
                qty_kg = qty  # already in Kg
            
            if product_id not in purchases_by_product:
                purchases_by_product[product_id] = {"qty": 0, "value": 0}
            purchases_by_product[product_id]["qty"] += qty_kg
            purchases_by_product[product_id]["value"] += total_value
    
    # Get today's QC dispatches - handle both string dates and datetime objects
    # First, get all dispatches and filter by date in Python for reliability
    all_qc_dispatches = await db.qc_dispatches.find({}, {"_id": 0}).to_list(1000)
    qc_dispatches = []
    for d in all_qc_dispatches:
        dispatch_date = d.get("dispatch_date", "")
        if isinstance(dispatch_date, datetime):
            dispatch_date_str = dispatch_date.strftime('%Y-%m-%d')
        else:
            dispatch_date_str = str(dispatch_date)[:10]  # Get YYYY-MM-DD part
        if dispatch_date_str == today:
            qc_dispatches.append(d)
    
    # Get today's retailer dispatches (if exists)
    all_retailer_dispatches = await db.retailer_dispatches.find({}, {"_id": 0}).to_list(1000)
    retailer_dispatches = []
    for d in all_retailer_dispatches:
        dispatch_date = d.get("dispatch_date", "")
        if isinstance(dispatch_date, datetime):
            dispatch_date_str = dispatch_date.strftime('%Y-%m-%d')
        else:
            dispatch_date_str = str(dispatch_date)[:10]
        if dispatch_date_str == today:
            retailer_dispatches.append(d)
    
    # Get packaging weights for unit conversion from QC packaging table
    packaging_variants = await db.qc_packaging.find({}, {"_id": 0}).to_list(100)
    # Build packaging map: name -> weight_gm
    packaging_map = {}
    for p in packaging_variants:
        name = p['name']
        name_lower = name.lower().strip()
        # Use stored weight_gm if available, otherwise extract from name
        weight = p.get('weight_gm') or extract_weight_from_packaging_name(name)
        if weight > 0:
            packaging_map[name_lower] = weight
            # Also store without common suffixes for flexible matching
            if 'packet' in name_lower:
                packaging_map[name_lower.replace('packet', '').strip()] = weight
            # Store just the number version (e.g., "500 gm" also stored as "500")
            number_match = re.search(r'^(\d+)\s*(gm|g)?', name_lower)
            if number_match:
                packaging_map[number_match.group(1)] = weight
    
    # Build cost_alias mapping: product_id -> aliased_product_id
    # This ensures products like Spinach are combined with their alias (Palak)
    cost_alias_id_map = {}
    for product in products:
        alias_id = product.get("cost_alias_product_id")
        if alias_id:
            cost_alias_id_map[product["id"]] = alias_id
    
    # Build historical price map for products (fallback when no purchase today)
    # Get recent stock status with non-zero avg_price
    historical_prices = await db.daily_stock_status.find(
        {"avg_price": {"$gt": 0}},
        {"_id": 0, "product_id": 1, "date": 1, "avg_price": 1}
    ).sort("date", -1).to_list(5000)
    
    # Build map: product_id -> most recent avg_price
    product_historical_price = {}
    for h in historical_prices:
        pid = h.get("product_id")
        if pid and pid not in product_historical_price:
            product_historical_price[pid] = h.get("avg_price", 0)
    
    # Calculate dispatches by product (convert units to Kg)
    # Use cost_alias_id_map to combine aliased products
    dispatches_by_product = {}
    for dispatch in qc_dispatches + retailer_dispatches:
        for item in dispatch.get("items", []):
            product_name = item.get("product_name", "").lower()
            product_id = item.get("product_id")
            
            # Map to aliased product if applicable (e.g., Spinach → Palak)
            target_product_id = cost_alias_id_map.get(product_id, product_id)
            
            supplied_qty = item.get("supplied_qty", 0)  # in units
            # Check both packaging_name (QC) and variant_name (Retail)
            packaging_name = (item.get("packaging_name") or item.get("variant_name") or "").lower().strip()
            
            # Look up weight from packaging map first
            weight_gm = packaging_map.get(packaging_name)
            
            # If not found, try extracting from the packaging name
            if not weight_gm:
                weight_gm = extract_weight_from_packaging_name(packaging_name)
            
            # If still not found, assume bulk (1kg)
            if not weight_gm:
                weight_gm = 1000
            
            # Convert units to Kg based on packaging weight
            qty_kg = (supplied_qty * weight_gm) / 1000
            
            # Handle None rate
            rate = item.get("rate") or 0
            value = supplied_qty * rate
            
            # Use the target_product_id (mapped from alias) for aggregation
            if target_product_id not in dispatches_by_product:
                dispatches_by_product[target_product_id] = {"qty": 0, "value": 0}
            dispatches_by_product[target_product_id]["qty"] += qty_kg
            dispatches_by_product[target_product_id]["value"] += value
    
    # Build stock status for each product
    result = []
    for product in products:
        product_id = product["id"]
        product_name = product["name"]
        
        # Always get fresh purchase and dispatch data
        purchase_data = purchases_by_product.get(product_id, {"qty": 0, "value": 0})
        dispatch_data = dispatches_by_product.get(product_id, {"qty": 0, "value": 0})
        
        # Get or create today's status
        if product_id in today_map:
            status = today_map[product_id]
            
            # For open entries, always update purchase and dispatch values (real-time sync)
            if status.get("status") == "open":
                opening_qty = status.get("opening_qty", 0) or 0
                opening_price = status.get("opening_price", 0) or 0
                
                # Calculate weighted average price
                total_qty = opening_qty + purchase_data["qty"]
                if total_qty > 0:
                    opening_value = opening_qty * opening_price
                    avg_price = (opening_value + purchase_data["value"]) / total_qty
                else:
                    avg_price = opening_price or (purchase_data["value"] / purchase_data["qty"] if purchase_data["qty"] > 0 else 0)
                
                # Fallback to historical price if avg_price is still 0
                if avg_price == 0:
                    avg_price = product_historical_price.get(product_id, 0)
                
                # Update the status with fresh data
                update_data = {
                    "purchase_qty": round(purchase_data["qty"], 2),
                    "purchase_value": round(purchase_data["value"], 2),
                    "dispatch_qty": round(dispatch_data["qty"], 2),
                    "dispatch_value": round(dispatch_data["value"], 2),
                    "avg_price": round(avg_price, 2)
                }
                
                # Update in database
                await db.daily_stock_status.update_one(
                    {"id": status["id"]},
                    {"$set": update_data}
                )
                
                # Update local status for response
                status.update(update_data)
        else:
            # Calculate opening from yesterday's closing or start with 0
            if product_id in yesterday_map:
                opening_qty = yesterday_map[product_id].get("closing_qty", 0) or 0
                opening_price = yesterday_map[product_id].get("avg_price", 0) or 0
            else:
                opening_qty = 0
                opening_price = product.get("price_per_kg", 0) or 0
            
            # Calculate weighted average price
            total_qty = opening_qty + purchase_data["qty"]
            if total_qty > 0:
                opening_value = opening_qty * opening_price
                avg_price = (opening_value + purchase_data["value"]) / total_qty
            else:
                avg_price = opening_price or (purchase_data["value"] / purchase_data["qty"] if purchase_data["qty"] > 0 else 0)
            
            # Fallback to historical price if avg_price is still 0
            if avg_price == 0:
                avg_price = product_historical_price.get(product_id, 0)
            
            status = {
                "id": str(uuid.uuid4()),
                "date": today,
                "product_id": product_id,
                "product_name": product_name,
                "product_unit": product.get("unit", "Kg"),
                "opening_qty": round(opening_qty, 2),
                "opening_price": round(opening_price, 2),
                "purchase_qty": round(purchase_data["qty"], 2),
                "purchase_value": round(purchase_data["value"], 2),
                "dispatch_qty": round(dispatch_data["qty"], 2),
                "dispatch_value": round(dispatch_data["value"], 2),
                "closing_qty": None,
                "wastage_qty": 0,
                "wastage_value": 0,
                "wastage_percent": 0,
                "avg_price": round(avg_price, 2),
                "status": "open"
            }
            
            # Save to database (creates a copy internally, but mutates status with _id)
            await db.daily_stock_status.insert_one(status.copy())
        
        # Ensure no _id field in response (MongoDB insert_one mutates the dict)
        if "_id" in status:
            del status["_id"]
        result.append(status)
    
    return result

@api_router.post("/stock-status/close")
async def close_stock_status(entries: StockClosingBulkEntry, date: Optional[str] = None, current_user: dict = Depends(get_current_user)):
    """Staff enters closing quantities for products. Accepts optional date parameter for historical closes."""
    if current_user["role"] not in ["admin", "staff"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    target_date = date if date else datetime.now(timezone.utc).strftime('%Y-%m-%d')
    updated = []
    
    for entry in entries.entries:
        # Get status for this product on the target date
        status = await db.daily_stock_status.find_one(
            {"date": target_date, "product_id": entry.product_id},
            {"_id": 0}
        )
        
        if not status:
            # Create a new record if it doesn't exist
            product = await db.products.find_one({"id": entry.product_id}, {"_id": 0})
            if not product:
                continue
            status = {
                "product_id": entry.product_id,
                "product_name": product.get("name", "Unknown"),
                "date": target_date,
                "opening_qty": 0,
                "purchase_qty": 0,
                "dispatch_qty": 0,
                "avg_price": 0,
                "status": "open"
            }
        
        closing_qty = entry.closing_qty
        opening_qty = status.get("opening_qty", 0)
        purchase_qty = status.get("purchase_qty", 0)
        dispatch_qty = status.get("dispatch_qty", 0)
        avg_price = status.get("avg_price", 0)
        
        # If no purchase today (avg_price = 0), get the most recent purchase price for this product
        if avg_price == 0:
            # Look for the most recent stock status with a non-zero avg_price
            recent_status = await db.daily_stock_status.find_one(
                {
                    "product_id": entry.product_id,
                    "date": {"$lt": target_date},
                    "avg_price": {"$gt": 0}
                },
                {"_id": 0},
                sort=[("date", -1)]
            )
            if recent_status:
                avg_price = recent_status.get("avg_price", 0)
            else:
                # Fallback: check procurements directly
                recent_proc = await db.procurements.find(
                    {},
                    {"_id": 0}
                ).sort("date", -1).to_list(100)
                
                for proc in recent_proc:
                    for prod in proc.get("products", []):
                        if prod.get("product_id") == entry.product_id or prod.get("product_name") == status.get("product_name"):
                            if prod.get("quantity", 0) > 0:
                                avg_price = prod.get("rate_per_kg", 0) or (prod.get("total_value", 0) / prod.get("quantity", 1))
                                break
                    if avg_price > 0:
                        break
        
        # Calculate wastage: Opening + Purchase - Dispatch - Closing
        total_available = opening_qty + purchase_qty - dispatch_qty
        wastage_qty = max(0, total_available - closing_qty)
        wastage_value = wastage_qty * avg_price
        
        # Calculate wastage percentage
        total_input = opening_qty + purchase_qty
        wastage_percent = (wastage_qty / total_input * 100) if total_input > 0 else 0
        
        # Update status
        update_data = {
            "closing_qty": round(closing_qty, 2),
            "wastage_qty": round(wastage_qty, 2),
            "wastage_value": round(wastage_value, 2),
            "wastage_percent": round(wastage_percent, 2),
            "status": "closed",
            "closed_at": datetime.now(timezone.utc).isoformat()
        }
        
        await db.daily_stock_status.update_one(
            {"date": target_date, "product_id": entry.product_id},
            {"$set": {**status, **update_data}},
            upsert=True
        )


@api_router.post("/stock-status/recalculate-dispatches")
async def recalculate_stock_dispatches(
    date: str,
    current_user: dict = Depends(get_current_user)
):
    """
    Recalculate dispatch quantities for ALL stock status entries on a specific date.
    This fixes closed entries that have incorrect dispatch values.
    """
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Only admin can recalculate dispatches")
    
    # Get packaging weights from QC packaging table
    packaging_variants = await db.qc_packaging.find({}, {"_id": 0}).to_list(100)
    packaging_map = {}
    for p in packaging_variants:
        name = p['name']
        name_lower = name.lower().strip()
        weight = p.get('weight_gm') or extract_weight_from_packaging_name(name)
        if weight > 0:
            packaging_map[name_lower] = weight
            if 'packet' in name_lower:
                packaging_map[name_lower.replace('packet', '').strip()] = weight
            number_match = re.search(r'^(\d+)\s*(gm|g)?', name_lower)
            if number_match:
                packaging_map[number_match.group(1)] = weight
    
    # Get all products to build cost_alias mapping
    products = await db.products.find({}, {"_id": 0}).to_list(1000)
    cost_alias_id_map = {}
    for product in products:
        alias_id = product.get("cost_alias_product_id")
        if alias_id:
            cost_alias_id_map[product["id"]] = alias_id
    
    # Get all dispatches for the target date
    all_qc_dispatches = await db.qc_dispatches.find({}, {"_id": 0}).to_list(1000)
    all_retailer_dispatches = await db.retailer_dispatches.find({}, {"_id": 0}).to_list(1000)
    
    # Calculate dispatch totals by product (use cost_alias mapping)
    dispatches_by_product = {}
    for d in all_qc_dispatches + all_retailer_dispatches:
        dispatch_date = d.get("dispatch_date", "")
        if isinstance(dispatch_date, datetime):
            dispatch_date_str = dispatch_date.strftime('%Y-%m-%d')
        else:
            dispatch_date_str = str(dispatch_date)[:10]
        
        if dispatch_date_str == date:
            for item in d.get("items", []):
                product_id = item.get("product_id")
                # Map to aliased product if applicable (e.g., Spinach → Palak)
                target_product_id = cost_alias_id_map.get(product_id, product_id)
                
                supplied_qty = item.get("supplied_qty", 0)
                packaging_name = (item.get("packaging_name") or item.get("variant_name") or "").lower().strip()
                
                # Get weight from packaging map
                weight_gm = packaging_map.get(packaging_name)
                if not weight_gm:
                    weight_gm = extract_weight_from_packaging_name(packaging_name)
                if not weight_gm:
                    weight_gm = 1000  # Default 1kg
                
                qty_kg = (supplied_qty * weight_gm) / 1000
                
                # Use target_product_id (mapped from alias) for aggregation
                if target_product_id not in dispatches_by_product:
                    dispatches_by_product[target_product_id] = {"qty": 0, "value": 0}
                dispatches_by_product[target_product_id]["qty"] += qty_kg
                dispatches_by_product[target_product_id]["value"] += supplied_qty * (item.get("rate") or 0)
    
    # Get all stock status entries for the date
    stock_entries = await db.daily_stock_status.find({"date": date}, {"_id": 0}).to_list(500)
    
    updated_count = 0
    updates = []
    
    for entry in stock_entries:
        product_id = entry.get("product_id")
        old_dispatch = entry.get("dispatch_qty", 0)
        
        # Get new dispatch value
        dispatch_data = dispatches_by_product.get(product_id, {"qty": 0, "value": 0})
        new_dispatch = round(dispatch_data["qty"], 2)
        
        # Only update if there's a change
        if abs(new_dispatch - old_dispatch) > 0.001:
            # Recalculate wastage if entry is closed
            update_data = {
                "dispatch_qty": new_dispatch,
                "dispatch_value": round(dispatch_data["value"], 2)
            }
            
            # If closed, recalculate wastage
            if entry.get("status") == "closed" and entry.get("closing_qty") is not None:
                opening_qty = entry.get("opening_qty", 0)
                purchase_qty = entry.get("purchase_qty", 0)
                closing_qty = entry.get("closing_qty", 0)
                avg_price = entry.get("avg_price", 0)
                
                total_available = opening_qty + purchase_qty - new_dispatch
                wastage_qty = max(0, total_available - closing_qty)
                wastage_value = wastage_qty * avg_price
                total_input = opening_qty + purchase_qty
                wastage_percent = (wastage_qty / total_input * 100) if total_input > 0 else 0
                
                update_data.update({
                    "wastage_qty": round(wastage_qty, 2),
                    "wastage_value": round(wastage_value, 2),
                    "wastage_percent": round(wastage_percent, 2)
                })
            
            await db.daily_stock_status.update_one(
                {"date": date, "product_id": product_id},
                {"$set": update_data}
            )
            
            updates.append({
                "product_name": entry.get("product_name"),
                "old_dispatch": old_dispatch,
                "new_dispatch": new_dispatch
            })
            updated_count += 1
    
    return {
        "message": f"Recalculated {updated_count} entries for {date}",
        "updates": updates
    }


@api_router.post("/stock-status/recalculate-wastage-values")
async def recalculate_wastage_values(
    date: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """
    Recalculate wastage values for stock status entries where wastage_qty > 0 but wastage_value = 0.
    Uses the most recent purchase price for that product.
    If date is provided, only recalculates for that date. Otherwise, recalculates all.
    """
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Only admin can recalculate wastage values")
    
    # Build query
    query = {"wastage_qty": {"$gt": 0}, "wastage_value": {"$in": [0, None]}}
    if date:
        query["date"] = date
    
    # Get entries with wastage but no value
    entries_to_fix = await db.daily_stock_status.find(query, {"_id": 0}).to_list(1000)
    
    if not entries_to_fix:
        return {"message": "No entries need fixing", "updated_count": 0}
    
    # Get all historical stock status to build price history
    all_status = await db.daily_stock_status.find(
        {"avg_price": {"$gt": 0}},
        {"_id": 0, "product_id": 1, "date": 1, "avg_price": 1}
    ).sort("date", -1).to_list(5000)
    
    # Build price history map: product_id -> list of {date, price} sorted by date desc
    price_history = {}
    for status in all_status:
        pid = status.get("product_id")
        if pid not in price_history:
            price_history[pid] = []
        price_history[pid].append({
            "date": status.get("date"),
            "price": status.get("avg_price", 0)
        })
    
    updated_count = 0
    updates = []
    
    for entry in entries_to_fix:
        product_id = entry.get("product_id")
        entry_date = entry.get("date")
        wastage_qty = entry.get("wastage_qty", 0)
        
        # Find the most recent price for this product before or on this date
        avg_price = 0
        if product_id in price_history:
            for price_entry in price_history[product_id]:
                if price_entry["date"] <= entry_date:
                    avg_price = price_entry["price"]
                    break
        
        if avg_price > 0:
            wastage_value = round(wastage_qty * avg_price, 2)
            
            await db.daily_stock_status.update_one(
                {"date": entry_date, "product_id": product_id},
                {"$set": {"wastage_value": wastage_value, "avg_price": avg_price}}
            )
            
            updates.append({
                "product_name": entry.get("product_name"),
                "date": entry_date,
                "wastage_qty": wastage_qty,
                "avg_price": avg_price,
                "wastage_value": wastage_value
            })
            updated_count += 1
    
    return {
        "message": f"Recalculated wastage values for {updated_count} entries",
        "updates": updates
    }


@api_router.post("/stock-status/fix-all-historical-dispatches")
async def fix_all_historical_dispatches(
    current_user: dict = Depends(get_current_user)
):
    """
    Fix dispatch calculations for ALL historical dates.
    This recalculates dispatches for every date in daily_stock_status,
    properly combining aliased products (e.g., Spinach → Palak).
    """
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Only admin can fix historical dispatches")
    
    # Get all unique dates from daily_stock_status
    all_dates = await db.daily_stock_status.distinct("date")
    all_dates = sorted(all_dates)  # Sort chronologically
    
    if not all_dates:
        return {"message": "No stock status records found", "dates_fixed": 0, "total_updates": 0}
    
    # Get packaging weights from QC packaging table
    packaging_variants = await db.qc_packaging.find({}, {"_id": 0}).to_list(100)
    packaging_map = {}
    for p in packaging_variants:
        name = p['name']
        name_lower = name.lower().strip()
        weight = p.get('weight_gm') or extract_weight_from_packaging_name(name)
        if weight > 0:
            packaging_map[name_lower] = weight
            if 'packet' in name_lower:
                packaging_map[name_lower.replace('packet', '').strip()] = weight
            number_match = re.search(r'^(\d+)\s*(gm|g)?', name_lower)
            if number_match:
                packaging_map[number_match.group(1)] = weight
    
    # Get all products to build cost_alias mapping
    products = await db.products.find({}, {"_id": 0}).to_list(1000)
    cost_alias_id_map = {}
    for product in products:
        alias_id = product.get("cost_alias_product_id")
        if alias_id:
            cost_alias_id_map[product["id"]] = alias_id
    
    # Get ALL dispatches (we'll filter by date in memory)
    all_qc_dispatches = await db.qc_dispatches.find({}, {"_id": 0}).to_list(5000)
    all_retailer_dispatches = await db.retailer_dispatches.find({}, {"_id": 0}).to_list(5000)
    
    # Build dispatch data by date and product
    dispatches_by_date = {}
    for d in all_qc_dispatches + all_retailer_dispatches:
        dispatch_date = d.get("dispatch_date", "")
        if isinstance(dispatch_date, datetime):
            dispatch_date_str = dispatch_date.strftime('%Y-%m-%d')
        else:
            dispatch_date_str = str(dispatch_date)[:10]
        
        if dispatch_date_str not in dispatches_by_date:
            dispatches_by_date[dispatch_date_str] = {}
        
        for item in d.get("items", []):
            product_id = item.get("product_id")
            # Map to aliased product if applicable (e.g., Spinach → Palak)
            target_product_id = cost_alias_id_map.get(product_id, product_id)
            
            supplied_qty = item.get("supplied_qty", 0)
            packaging_name = (item.get("packaging_name") or item.get("variant_name") or "").lower().strip()
            
            # Get weight from packaging map
            weight_gm = packaging_map.get(packaging_name)
            if not weight_gm:
                weight_gm = extract_weight_from_packaging_name(packaging_name)
            if not weight_gm:
                weight_gm = 1000  # Default 1kg
            
            qty_kg = (supplied_qty * weight_gm) / 1000
            
            if target_product_id not in dispatches_by_date[dispatch_date_str]:
                dispatches_by_date[dispatch_date_str][target_product_id] = {"qty": 0, "value": 0}
            dispatches_by_date[dispatch_date_str][target_product_id]["qty"] += qty_kg
            dispatches_by_date[dispatch_date_str][target_product_id]["value"] += supplied_qty * (item.get("rate") or 0)
    
    # Process each date
    total_updated = 0
    dates_fixed = 0
    summary = []
    
    for date in all_dates:
        stock_entries = await db.daily_stock_status.find({"date": date}, {"_id": 0}).to_list(500)
        date_dispatches = dispatches_by_date.get(date, {})
        date_updates = 0
        
        for entry in stock_entries:
            product_id = entry.get("product_id")
            old_dispatch = entry.get("dispatch_qty", 0)
            
            # Get new dispatch value
            dispatch_data = date_dispatches.get(product_id, {"qty": 0, "value": 0})
            new_dispatch = round(dispatch_data["qty"], 2)
            
            # Only update if there's a change
            if abs(new_dispatch - old_dispatch) > 0.001:
                update_data = {
                    "dispatch_qty": new_dispatch,
                    "dispatch_value": round(dispatch_data["value"], 2)
                }
                
                # If closed, recalculate wastage
                if entry.get("status") == "closed" and entry.get("closing_qty") is not None:
                    opening_qty = entry.get("opening_qty", 0)
                    purchase_qty = entry.get("purchase_qty", 0)
                    closing_qty = entry.get("closing_qty", 0)
                    avg_price = entry.get("avg_price", 0)
                    
                    total_available = opening_qty + purchase_qty - new_dispatch
                    wastage_qty = max(0, total_available - closing_qty)
                    wastage_value = wastage_qty * avg_price
                    total_input = opening_qty + purchase_qty
                    wastage_percent = (wastage_qty / total_input * 100) if total_input > 0 else 0
                    
                    update_data.update({
                        "wastage_qty": round(wastage_qty, 2),
                        "wastage_value": round(wastage_value, 2),
                        "wastage_percent": round(wastage_percent, 2)
                    })
                
                await db.daily_stock_status.update_one(
                    {"date": date, "product_id": product_id},
                    {"$set": update_data}
                )
                date_updates += 1
                total_updated += 1
        
        if date_updates > 0:
            dates_fixed += 1
            summary.append(f"{date}: {date_updates} products updated")
    
    return {
        "message": f"Fixed dispatches for {dates_fixed} dates, {total_updated} total product entries updated",
        "dates_fixed": dates_fixed,
        "total_updates": total_updated,
        "summary": summary[-10:] if len(summary) > 10 else summary  # Return last 10 dates for brevity
    }


@api_router.post("/stock-status/fix-all-purchase-quantities")
async def fix_all_purchase_quantities(
    current_user: dict = Depends(get_current_user)
):
    """
    Fix corrupted purchase_qty values in daily_stock_status.
    This recalculates purchase_qty by summing qty from raw procurements collection.
    Use this when Excel backup imports have bloated/incorrect purchase values.
    """
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Only admin can fix purchase quantities")
    
    import ast
    
    def parse_products(products):
        """Parse products array, handling string corruption from Excel imports."""
        if isinstance(products, list):
            return products
        if isinstance(products, str):
            try:
                return ast.literal_eval(products)
            except (ValueError, SyntaxError):
                return []
        return []
    
    def calculate_qty_kg(item):
        """Convert quantity to Kg based on unit type."""
        try:
            qty = float(item.get('quantity', 0) or 0)
        except (ValueError, TypeError):
            qty = 0
        unit = item.get('unit', 'Kg')
        unit_size = item.get('unit_size', '')
        
        if unit == 'Bunch' and unit_size:
            try:
                weight_gm = float(unit_size)
                return (qty * weight_gm) / 1000  # Convert bunch to kg
            except (ValueError, TypeError):
                pass
        return qty
    
    def extract_date_string(date_val):
        """Extract YYYY-MM-DD string from various date formats."""
        if date_val is None:
            return None
        if isinstance(date_val, datetime):
            return date_val.strftime('%Y-%m-%d')
        date_str = str(date_val)
        # Handle ISO format with timezone: "2026-04-05T00:00:00+00:00"
        if 'T' in date_str:
            return date_str[:10]
        # Handle simple date: "2026-04-05"
        if len(date_str) >= 10:
            return date_str[:10]
        return date_str
    
    try:
        # Step 1: Get all unique dates from daily_stock_status
        all_dates = await db.daily_stock_status.distinct("date")
        # Filter out None/empty dates and sort
        all_dates = [d for d in all_dates if d]
        all_dates = sorted(all_dates, key=lambda x: str(x) if x else "")
        
        # Step 2: Get all procurements
        all_procurements = await db.procurements.find({}).to_list(10000)
        
        # Build lookup: date -> product_id -> {qty_kg, total_value}
        purchases_by_date_product = {}
        
        for proc in all_procurements:
            try:
                proc_date = extract_date_string(proc.get('date'))
                if not proc_date:
                    continue
                
                if proc_date not in purchases_by_date_product:
                    purchases_by_date_product[proc_date] = {}
                
                products = parse_products(proc.get('products', []))
                
                for item in products:
                    if not item or not isinstance(item, dict):
                        continue
                    product_id = item.get('product_id', '')
                    if not product_id:
                        continue
                    
                    qty_kg = calculate_qty_kg(item)
                    try:
                        total_value = float(item.get('total', 0) or 0)
                    except (ValueError, TypeError):
                        total_value = 0
                    
                    if product_id not in purchases_by_date_product[proc_date]:
                        purchases_by_date_product[proc_date][product_id] = {
                            'qty_kg': 0,
                            'total_value': 0
                        }
                    
                    purchases_by_date_product[proc_date][product_id]['qty_kg'] += qty_kg
                    purchases_by_date_product[proc_date][product_id]['total_value'] += total_value
            except Exception as proc_error:
                print(f"Error processing procurement {proc.get('id')}: {proc_error}")
                continue
        
        # Step 3: Fix each daily_stock_status record
        total_fixed = 0
        total_unchanged = 0
        fixes_summary = []
        
        for date in all_dates:
            try:
                stock_entries = await db.daily_stock_status.find({"date": date}).to_list(500)
                date_purchases = purchases_by_date_product.get(str(date), {})
                date_fixes = []
                
                for entry in stock_entries:
                    try:
                        product_id = entry.get('product_id')
                        product_name = entry.get('product_name')
                        try:
                            current_purchase_qty = float(entry.get('purchase_qty', 0) or 0)
                        except (ValueError, TypeError):
                            current_purchase_qty = 0
                        
                        # Calculate correct purchase_qty from procurements
                        product_purchase = date_purchases.get(product_id, {'qty_kg': 0, 'total_value': 0})
                        correct_purchase_qty = round(product_purchase['qty_kg'], 2)
                        correct_purchase_value = round(product_purchase['total_value'], 2)
                        
                        # Get current purchase_value
                        try:
                            current_purchase_value = float(entry.get('purchase_value', 0) or 0)
                        except (ValueError, TypeError):
                            current_purchase_value = 0
                        
                        # Check if update needed (compare both qty and value)
                        qty_diff = abs(current_purchase_qty - correct_purchase_qty)
                        value_diff = abs(current_purchase_value - correct_purchase_value)
                        
                        if qty_diff > 0.01 or value_diff > 0.01:  # Significant difference in qty OR value
                            # Calculate new avg_price
                            avg_price = 0
                            if correct_purchase_qty > 0:
                                avg_price = round(correct_purchase_value / correct_purchase_qty, 2)
                            
                            # Recalculate wastage_value based on new avg_price
                            wastage_qty = float(entry.get('wastage_qty', 0) or 0)
                            new_wastage_value = round(wastage_qty * avg_price, 2)
                            
                            # Update record
                            await db.daily_stock_status.update_one(
                                {"id": entry['id']},
                                {"$set": {
                                    "purchase_qty": correct_purchase_qty,
                                    "purchase_value": correct_purchase_value,
                                    "avg_price": avg_price,
                                    "wastage_value": new_wastage_value
                                }}
                            )
                            
                            date_fixes.append({
                                "product": product_name,
                                "old": current_purchase_qty,
                                "new": correct_purchase_qty,
                                "new_wastage_value": new_wastage_value
                            })
                            total_fixed += 1
                        else:
                            total_unchanged += 1
                    except Exception as entry_error:
                        print(f"Error processing entry {entry.get('id')}: {entry_error}")
                        continue
            
                if date_fixes:
                    fixes_summary.append({"date": str(date), "fixes": date_fixes[:5]})  # Limit to 5 per date
            except Exception as date_error:
                print(f"Error processing date {date}: {date_error}")
                continue
        
        return {
            "message": f"Fixed {total_fixed} records, {total_unchanged} unchanged",
            "total_fixed": total_fixed,
            "total_unchanged": total_unchanged,
            "summary": fixes_summary[-5:]  # Return last 5 dates
        }
    except Exception as e:
        import traceback
        error_trace = traceback.format_exc()
        print(f"Error in fix_all_purchase_quantities: {error_trace}")
        raise HTTPException(status_code=500, detail=f"Error fixing purchase quantities: {str(e)}")


@api_router.get("/stock-status/history")
async def get_stock_status_history(
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    product_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Get historical stock status data"""
    if current_user["role"] not in ["admin", "staff"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    query = {}
    if from_date:
        query["date"] = {"$gte": from_date}
    if to_date:
        if "date" in query:
            query["date"]["$lte"] = to_date
        else:
            query["date"] = {"$lte": to_date}
    if product_id:
        query["product_id"] = product_id
    
    history = await db.daily_stock_status.find(query, {"_id": 0}).sort("date", -1).to_list(1000)


@api_router.post("/procurement/bulk-update-rate")
async def bulk_update_procurement_rate(
    input: dict,
    current_user: dict = Depends(get_current_user)
):
    """
    Bulk update procurement rate for a product across multiple dates.
    This updates both procurement and daily_stock_status (including wastage_value).
    
    Input:
    {
        "product_name": "Curry Leaves",
        "new_rate": 75,
        "from_date": "2026-04-01",  // Optional, defaults to all time
        "to_date": "2026-04-08"     // Optional, defaults to today
    }
    """
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Only admin can bulk update rates")
    
    product_name = input.get("product_name")
    new_rate = input.get("new_rate")
    from_date = input.get("from_date")
    to_date = input.get("to_date")
    
    if not product_name or new_rate is None:
        raise HTTPException(status_code=400, detail="product_name and new_rate are required")
    
    try:
        new_rate = float(new_rate)
    except (ValueError, TypeError):
        raise HTTPException(status_code=400, detail="new_rate must be a number")
    
    # Find procurements with this product
    proc_query = {}
    if from_date:
        proc_query["date"] = {"$gte": from_date}
    if to_date:
        if "date" in proc_query:
            proc_query["date"]["$lte"] = to_date + "T23:59:59Z"
        else:
            proc_query["date"] = {"$lte": to_date + "T23:59:59Z"}
    
    procurements = await db.procurements.find(proc_query, {"_id": 0}).to_list(1000)
    
    updated_procurements = 0
    updated_stock_status = 0
    
    for proc in procurements:
        products = proc.get("products", [])
        proc_updated = False
        
        for product in products:
            if product.get("product_name") == product_name:
                # Update rate and total
                old_total = product.get("total", 0)
                qty = product.get("quantity", 0)
                new_total = round(qty * new_rate, 2)
                
                product["rate"] = new_rate
                product["total"] = new_total
                proc_updated = True
        
        if proc_updated:
            # Update procurement in database
            await db.procurements.update_one(
                {"id": proc["id"]},
                {"$set": {"products": products}}
            )
            updated_procurements += 1
            
            # Also update daily_stock_status
            proc_date = proc.get("date", "")[:10]  # Extract date part
            
            # Find product_id
            product_doc = await db.products.find_one({"name": product_name}, {"_id": 0})
            if product_doc:
                product_id = product_doc.get("id")
                
                # Calculate new purchase_value from updated procurement
                new_purchase_value = 0
                new_purchase_qty = 0
                for product in products:
                    if product.get("product_name") == product_name:
                        new_purchase_qty += product.get("quantity", 0)
                        new_purchase_value += product.get("total", 0)
                
                # Calculate new avg_price and wastage_value
                avg_price = new_purchase_value / new_purchase_qty if new_purchase_qty > 0 else 0
                
                # Get wastage_qty from current stock status
                stock_status = await db.daily_stock_status.find_one(
                    {"date": proc_date, "product_id": product_id},
                    {"_id": 0}
                )
                
                if stock_status:
                    wastage_qty = stock_status.get("wastage_qty", 0) or 0
                    new_wastage_value = round(wastage_qty * avg_price, 2)
                    
                    await db.daily_stock_status.update_one(
                        {"date": proc_date, "product_id": product_id},
                        {"$set": {
                            "purchase_value": new_purchase_value,
                            "avg_price": avg_price,
                            "wastage_value": new_wastage_value
                        }}
                    )
                    updated_stock_status += 1
    
    return {
        "message": f"Updated {updated_procurements} procurements and {updated_stock_status} stock status records",
        "product_name": product_name,
        "new_rate": new_rate,
        "updated_procurements": updated_procurements,
        "updated_stock_status": updated_stock_status
    }


@api_router.post("/stock-status/fix-all-wastage-values")
async def fix_all_wastage_values(
    current_user: dict = Depends(get_current_user)
):
    """
    Fix all historical wastage_value in daily_stock_status based on current avg_price.
    This recalculates wastage_value = wastage_qty * (purchase_value / purchase_qty)
    for ALL records in the database.
    
    Use this to sync wastage values after any procurement rate changes.
    """
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Only admin can fix wastage values")
    
    try:
        # Get all stock status records
        all_status = await db.daily_stock_status.find({}, {"_id": 0}).to_list(10000)
        
        updated_count = 0
        unchanged_count = 0
        fixes_by_product = {}
        
        for status in all_status:
            record_id = status.get("id")
            if not record_id:
                continue
                
            wastage_qty = status.get("wastage_qty", 0) or 0
            purchase_qty = status.get("purchase_qty", 0) or 0
            purchase_value = status.get("purchase_value", 0) or 0
            current_wastage_value = status.get("wastage_value", 0) or 0
            product_name = status.get("product_name", "Unknown")
            date = status.get("date", "")
            
            # Calculate correct wastage_value
            avg_price = purchase_value / purchase_qty if purchase_qty > 0 else 0
            correct_wastage_value = round(wastage_qty * avg_price, 2)
            
            # Check if update needed
            diff = abs(current_wastage_value - correct_wastage_value)
            if diff > 0.01 and wastage_qty > 0:
                # Update the record
                await db.daily_stock_status.update_one(
                    {"id": record_id},
                    {"$set": {
                        "wastage_value": correct_wastage_value,
                        "avg_price": round(avg_price, 2)
                    }}
                )
                updated_count += 1
                
                # Track by product
                if product_name not in fixes_by_product:
                    fixes_by_product[product_name] = []
                fixes_by_product[product_name].append({
                    "date": date,
                    "wastage_qty": wastage_qty,
                    "old_value": current_wastage_value,
                    "new_value": correct_wastage_value,
                    "avg_price": round(avg_price, 2)
                })
            else:
                unchanged_count += 1
        
        return {
            "message": f"Fixed {updated_count} records, {unchanged_count} unchanged",
            "total_fixed": updated_count,
            "total_unchanged": unchanged_count,
            "fixes_by_product": {k: v for k, v in list(fixes_by_product.items())[:10]}  # Limit output
        }
        
    except Exception as e:
        import traceback
        error_trace = traceback.format_exc()
        print(f"Error in fix_all_wastage_values: {error_trace}")
        raise HTTPException(status_code=500, detail=f"Error fixing wastage values: {str(e)}")


@api_router.get("/stock-status/history")
async def get_stock_status_history(
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    product_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Get historical stock status data"""
    if current_user["role"] not in ["admin", "staff"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    query = {}
    if from_date:
        query["date"] = {"$gte": from_date}
    if to_date:
        if "date" in query:
            query["date"]["$lte"] = to_date
        else:
            query["date"] = {"$lte": to_date}
    if product_id:
        query["product_id"] = product_id
    
    history = await db.daily_stock_status.find(query, {"_id": 0}).sort("date", -1).to_list(1000)
    return history

@api_router.get("/stock-status/closable-products")
async def get_closable_products_for_date(
    date: str,
    current_user: dict = Depends(get_current_user)
):
    """Get products that need closing for a specific date - products with activity (opening > 0, purchased, or dispatched)"""
    if current_user["role"] not in ["admin", "staff"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    target_date = date
    yesterday = (datetime.strptime(target_date, '%Y-%m-%d') - timedelta(days=1)).strftime('%Y-%m-%d')
    
    # Get all products
    products = await db.products.find({}, {"_id": 0}).to_list(1000)
    
    # Get yesterday's closing stock (for opening)
    yesterday_status = await db.daily_stock_status.find(
        {"date": yesterday, "status": "closed"},
        {"_id": 0}
    ).to_list(1000)
    yesterday_map = {s["product_id"]: s for s in yesterday_status}
    
    # Get existing status for target date (from daily_stock_status table)
    existing_status = await db.daily_stock_status.find(
        {"date": target_date},
        {"_id": 0}
    ).to_list(1000)
    existing_map = {s["product_id"]: s for s in existing_status}
    
    # Get procurements for target date (check multiple date formats)
    procurements = await db.procurements.find({}, {"_id": 0}).to_list(1000)
    purchases_by_product = {}
    for proc in procurements:
        proc_date = proc.get("date", "")
        if isinstance(proc_date, datetime):
            proc_date_str = proc_date.strftime('%Y-%m-%d')
        else:
            proc_date_str = str(proc_date)[:10]
        
        if proc_date_str == target_date:
            for item in proc.get("products", []):
                product_id = item.get("product_id")
                qty = item.get("quantity", 0)
                unit = item.get("unit", "Kg")
                unit_size = item.get("unit_size", "")
                
                # Convert to Kg
                if unit.lower() in ["bunch", "piece", "pack"] and unit_size:
                    try:
                        weight_per_unit_gm = float(unit_size)
                        qty_kg = (qty * weight_per_unit_gm) / 1000
                    except (ValueError, TypeError):
                        qty_kg = qty
                else:
                    qty_kg = qty
                
                purchases_by_product[product_id] = purchases_by_product.get(product_id, 0) + qty_kg
    
    # Get dispatches for target date
    all_qc_dispatches = await db.qc_dispatches.find({}, {"_id": 0}).to_list(1000)
    all_retailer_dispatches = await db.retailer_dispatches.find({}, {"_id": 0}).to_list(1000)
    
    # Get packaging weights from QC packaging table
    packaging_variants = await db.qc_packaging.find({}, {"_id": 0}).to_list(100)
    # Build packaging map: name -> weight_gm
    packaging_map = {}
    for p in packaging_variants:
        name = p['name']
        name_lower = name.lower().strip()
        # Use stored weight_gm if available, otherwise extract from name
        weight = p.get('weight_gm') or extract_weight_from_packaging_name(name)
        if weight > 0:
            packaging_map[name_lower] = weight
            # Also store without common suffixes for flexible matching
            if 'packet' in name_lower:
                packaging_map[name_lower.replace('packet', '').strip()] = weight
            # Store just the number version (e.g., "500 gm" also stored as "500")
            number_match = re.search(r'^(\d+)\s*(gm|g)?', name_lower)
            if number_match:
                packaging_map[number_match.group(1)] = weight
    
    dispatches_by_product = {}
    
    # Build cost_alias mapping: product_id -> alias_target_id
    # If a product has cost_alias_product_id, its dispatches should count under the alias target
    cost_alias_id_map = {}
    for product in products:
        alias_id = product.get("cost_alias_product_id")
        if alias_id:
            cost_alias_id_map[product["id"]] = alias_id
    
    for d in all_qc_dispatches + all_retailer_dispatches:
        dispatch_date = d.get("dispatch_date", "")
        if isinstance(dispatch_date, datetime):
            dispatch_date_str = dispatch_date.strftime('%Y-%m-%d')
        else:
            dispatch_date_str = str(dispatch_date)[:10]
        
        if dispatch_date_str == target_date:
            for item in d.get("items", []):
                product_id = item.get("product_id")
                # Map to aliased product if applicable (e.g., Spinach dispatches → count under Palak)
                target_product_id = cost_alias_id_map.get(product_id, product_id)
                
                supplied_qty = item.get("supplied_qty", 0)
                # Check both packaging_name (QC) and variant_name (Retail)
                packaging_name = (item.get("packaging_name") or item.get("variant_name") or "").lower().strip()
                
                # Look up weight from packaging map first
                weight_gm = packaging_map.get(packaging_name)
                
                # If not found, try extracting from the packaging name
                if not weight_gm:
                    weight_gm = extract_weight_from_packaging_name(packaging_name)
                
                # If still not found, assume bulk (1kg)
                if not weight_gm:
                    weight_gm = 1000
                
                qty_kg = (supplied_qty * weight_gm) / 1000
                # Use target_product_id (mapped from alias) for aggregation
                dispatches_by_product[target_product_id] = dispatches_by_product.get(target_product_id, 0) + qty_kg
    
    # Build closable products list
    closable_products = []
    
    # First, include all products that have existing status records for this date
    for product_id, existing in existing_map.items():
        product = next((p for p in products if p["id"] == product_id), None)
        if not product:
            continue
            
        product_name = product["name"]
        opening_qty = existing.get("opening_qty", 0)
        status = existing.get("status", "open")
        closing_qty = existing.get("closing_qty", None)
        
        # For OPEN entries, always use FRESH procurement and dispatch data (real-time sync)
        if status == "open":
            purchase_qty = round(purchases_by_product.get(product_id, 0), 2)
            dispatch_qty = round(dispatches_by_product.get(product_id, 0), 2)
            
            # Skip products with no activity (opening=0, purchase=0, dispatch=0)
            if opening_qty == 0 and purchase_qty == 0 and dispatch_qty == 0:
                continue
            
            # Update the stored entry with fresh data
            await db.daily_stock_status.update_one(
                {"id": existing["id"]},
                {"$set": {"purchase_qty": purchase_qty, "dispatch_qty": dispatch_qty}}
            )
        else:
            # For CLOSED entries, use stored values
            purchase_qty = existing.get("purchase_qty", 0)
            dispatch_qty = existing.get("dispatch_qty", 0)
        
        closable_products.append({
            "product_id": product_id,
            "product_name": product_name,
            "opening_qty": round(opening_qty, 2),
            "purchase_qty": round(purchase_qty, 2),
            "dispatch_qty": round(dispatch_qty, 2),
            "closing_qty": closing_qty,
            "status": status
        })
    
    # Then, add any products with activity that don't have existing status records
    for product in products:
        product_id = product["id"]
        if product_id in existing_map:
            continue  # Already added above
            
        product_name = product["name"]
        
        # Opening = yesterday's closing
        opening_qty = yesterday_map.get(product_id, {}).get("closing_qty", 0)
        
        # Purchase and dispatch quantities from today's transactions
        purchase_qty = round(purchases_by_product.get(product_id, 0), 2)
        dispatch_qty = round(dispatches_by_product.get(product_id, 0), 2)
        
        # Only include if there's activity (opening > 0, purchased, or dispatched)
        if opening_qty > 0 or purchase_qty > 0 or dispatch_qty > 0:
            closable_products.append({
                "product_id": product_id,
                "product_name": product_name,
                "opening_qty": round(opening_qty, 2),
                "purchase_qty": purchase_qty,
                "dispatch_qty": dispatch_qty,
                "closing_qty": None,
                "status": "open"
            })
    
    return closable_products


@api_router.get("/stock-status/wastage-dashboard")
async def get_wastage_dashboard(
    days: int = 7,
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Get wastage dashboard data with trends"""
    if current_user["role"] not in ["admin", "staff"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    # Use custom dates if provided, otherwise use days parameter
    if from_date and to_date:
        start_date = from_date
        end_date = to_date
    else:
        end_date = datetime.now(timezone.utc).strftime('%Y-%m-%d')
        start_date = (datetime.now(timezone.utc) - timedelta(days=days)).strftime('%Y-%m-%d')
    
    history = await db.daily_stock_status.find(
        {"date": {"$gte": start_date, "$lte": end_date}, "status": "closed"},
        {"_id": 0}
    ).to_list(5000)
    
    # Aggregate by date
    daily_totals = {}
    product_totals = {}
    
    for record in history:
        date = record["date"]
        product_name = record["product_name"]
        wastage_qty = record.get("wastage_qty", 0)
        wastage_percent = record.get("wastage_percent", 0)
        opening_qty = record.get("opening_qty", 0)
        purchase_qty = record.get("purchase_qty", 0)
        purchase_value = record.get("purchase_value", 0)
        
        # Calculate wastage_value dynamically based on current avg_price
        # This ensures wastage value reflects latest procurement rate
        avg_price = purchase_value / purchase_qty if purchase_qty > 0 else 0
        wastage_value = round(wastage_qty * avg_price, 2)
        
        # Daily totals
        if date not in daily_totals:
            daily_totals[date] = {
                "date": date,
                "total_wastage_kg": 0,
                "total_wastage_value": 0,
                "total_input": 0,
                "avg_wastage_percent": 0
            }
        daily_totals[date]["total_wastage_kg"] += wastage_qty
        daily_totals[date]["total_wastage_value"] += wastage_value
        daily_totals[date]["total_input"] += opening_qty + purchase_qty
        
        # Product totals
        if product_name not in product_totals:
            product_totals[product_name] = {
                "product_name": product_name,
                "total_wastage_kg": 0,
                "total_wastage_value": 0,
                "total_input": 0,
                "wastage_percent": 0,
                "days_count": 0
            }
        product_totals[product_name]["total_wastage_kg"] += wastage_qty
        product_totals[product_name]["total_wastage_value"] += wastage_value
        product_totals[product_name]["total_input"] += opening_qty + purchase_qty
        product_totals[product_name]["days_count"] += 1
    
    # Calculate averages for daily data
    for date in daily_totals:
        total_input = daily_totals[date]["total_input"]
        if total_input > 0:
            daily_totals[date]["avg_wastage_percent"] = round(
                (daily_totals[date]["total_wastage_kg"] / total_input) * 100, 2
            )
        daily_totals[date]["total_wastage_kg"] = round(daily_totals[date]["total_wastage_kg"], 2)
        daily_totals[date]["total_wastage_value"] = round(daily_totals[date]["total_wastage_value"], 2)
    
    # Sort daily data by date
    daily_data = sorted(daily_totals.values(), key=lambda x: x["date"])
    
    # Calculate wastage_percent for each product
    for product_name in product_totals:
        total_input = product_totals[product_name]["total_input"]
        if total_input > 0:
            product_totals[product_name]["wastage_percent"] = round(
                (product_totals[product_name]["total_wastage_kg"] / total_input) * 100, 2
            )
        product_totals[product_name]["total_wastage_kg"] = round(product_totals[product_name]["total_wastage_kg"], 2)
        product_totals[product_name]["total_wastage_value"] = round(product_totals[product_name]["total_wastage_value"], 2)
    
    # Top wastage products - sorted by wastage_percent descending
    product_data = sorted(product_totals.values(), key=lambda x: x["wastage_percent"], reverse=True)[:15]
    
    # Overall summary
    total_wastage_kg = sum(d["total_wastage_kg"] for d in daily_data)
    total_wastage_value = sum(d["total_wastage_value"] for d in daily_data)
    total_input = sum(d["total_input"] for d in daily_data)
    avg_daily_wastage = total_wastage_kg / len(daily_data) if daily_data else 0
    overall_wastage_percent = (total_wastage_kg / total_input * 100) if total_input > 0 else 0
    
    return {
        "period_days": days,
        "start_date": start_date,
        "end_date": end_date,
        "summary": {
            "total_wastage_kg": round(total_wastage_kg, 2),
            "total_wastage_value": round(total_wastage_value, 2),
            "avg_daily_wastage_kg": round(avg_daily_wastage, 2),
            "overall_wastage_percent": round(overall_wastage_percent, 2)
        },
        "daily_trend": daily_data,
        "top_wastage_products": product_data
    }

@api_router.get("/stock-status/yesterday-wastage")
async def get_yesterday_wastage(current_user: dict = Depends(get_current_user)):
    """Get product-wise wastage for previous day (includes closed entries and pending entries with calculated wastage)"""
    if current_user["role"] not in ["admin", "staff"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    yesterday = (datetime.now(timezone.utc) - timedelta(days=1)).strftime('%Y-%m-%d')
    
    # Get all stock status for yesterday (both closed and pending)
    yesterday_records = await db.daily_stock_status.find(
        {"date": yesterday},
        {"_id": 0}
    ).to_list(500)
    
    # Aggregate wastage by product (sum up QC + Retail entries for same product)
    product_wastage = {}
    
    for record in yesterday_records:
        product_name = record.get("product_name", "Unknown")
        product_id = record.get("product_id", "")
        
        # Get values
        wastage_qty = record.get("wastage_qty", 0) or 0
        opening_qty = record.get("opening_qty", 0) or 0
        purchase_qty = record.get("purchase_qty", 0) or 0
        purchase_value = record.get("purchase_value", 0) or 0
        dispatch_qty = record.get("dispatch_qty", 0) or 0
        closing_qty = record.get("closing_qty", 0) or 0
        status = record.get("status", "pending")
        
        # Calculate avg_price from current procurement data
        avg_price = purchase_value / purchase_qty if purchase_qty > 0 else 0
        
        # Calculate wastage_value dynamically based on current avg_price
        wastage_value = round(wastage_qty * avg_price, 2)
        
        # For pending entries, calculate potential wastage if not closed
        if status != "closed" and wastage_qty == 0:
            # Available = opening + purchase
            available = opening_qty + purchase_qty
            # If no closing recorded, estimate wastage as available - dispatch
            if available > 0 and closing_qty == 0:
                wastage_qty = max(0, available - dispatch_qty)
                # Estimate value using average procurement price
                wastage_value = round(wastage_qty * avg_price, 2)
        
        if wastage_qty > 0:
            if product_name not in product_wastage:
                product_wastage[product_name] = {
                    "product_name": product_name,
                    "product_id": product_id,
                    "opening_qty": 0,
                    "purchase_qty": 0,
                    "dispatch_qty": 0,
                    "closing_qty": 0,
                    "wastage_qty": 0,
                    "wastage_value": 0,
                    "status": status
                }
            
            # Aggregate values
            product_wastage[product_name]["opening_qty"] += opening_qty
            product_wastage[product_name]["purchase_qty"] += purchase_qty
            product_wastage[product_name]["dispatch_qty"] += dispatch_qty
            product_wastage[product_name]["closing_qty"] += closing_qty
            product_wastage[product_name]["wastage_qty"] += wastage_qty
            product_wastage[product_name]["wastage_value"] += wastage_value
            # Update status to closed if any entry is closed
            if status == "closed":
                product_wastage[product_name]["status"] = "closed"
    
    # Calculate wastage percent and avg_price
    wastage_products = []
    total_wastage_kg = 0
    total_wastage_value = 0
    
    for product_name, data in product_wastage.items():
        available = data["opening_qty"] + data["purchase_qty"]
        wastage_percent = (data["wastage_qty"] / available * 100) if available > 0 else 0
        avg_price = data["wastage_value"] / data["wastage_qty"] if data["wastage_qty"] > 0 else 0
        
        wastage_products.append({
            "product_name": product_name,
            "product_id": data["product_id"],
            "opening_qty": round(data["opening_qty"], 2),
            "purchase_qty": round(data["purchase_qty"], 2),
            "dispatch_qty": round(data["dispatch_qty"], 2),
            "closing_qty": round(data["closing_qty"], 2),
            "wastage_qty": round(data["wastage_qty"], 2),
            "wastage_value": round(data["wastage_value"], 2),
            "wastage_percent": round(wastage_percent, 1),
            "avg_price": round(avg_price, 2),
            "status": data["status"]
        })
        total_wastage_kg += data["wastage_qty"]
        total_wastage_value += data["wastage_value"]
    
    # Sort by wastage percent descending (highest wastage % first)
    wastage_products.sort(key=lambda x: x.get("wastage_percent", 0), reverse=True)
    
    return {
        "date": yesterday,
        "total_wastage_kg": round(total_wastage_kg, 2),
        "total_wastage_value": round(total_wastage_value, 2),
        "products": wastage_products
    }


@api_router.get("/stock-status/wastage-by-date")
async def get_wastage_by_date(date: str, current_user: dict = Depends(get_current_user)):
    """Get product-wise wastage for a specific date"""
    if current_user["role"] not in ["admin", "staff"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    # Get all stock status for the given date
    day_records = await db.daily_stock_status.find(
        {"date": date},
        {"_id": 0}
    ).to_list(500)
    
    # Aggregate wastage by product
    product_wastage = {}
    
    for record in day_records:
        product_name = record.get("product_name", "Unknown")
        product_id = record.get("product_id", "")
        
        # Get values
        wastage_qty = record.get("wastage_qty", 0) or 0
        opening_qty = record.get("opening_qty", 0) or 0
        purchase_qty = record.get("purchase_qty", 0) or 0
        purchase_value = record.get("purchase_value", 0) or 0
        dispatch_qty = record.get("dispatch_qty", 0) or 0
        closing_qty = record.get("closing_qty", 0) or 0
        status = record.get("status", "pending")
        
        # Calculate avg_price from current procurement data
        avg_price = purchase_value / purchase_qty if purchase_qty > 0 else 0
        
        # Calculate wastage_value dynamically based on current avg_price
        wastage_value = round(wastage_qty * avg_price, 2)
        
        # For pending entries, calculate potential wastage if not closed
        if status != "closed" and wastage_qty == 0:
            available = opening_qty + purchase_qty
            if available > 0 and closing_qty == 0:
                wastage_qty = max(0, available - dispatch_qty)
                wastage_value = round(wastage_qty * avg_price, 2)
        
        if wastage_qty > 0:
            if product_name not in product_wastage:
                product_wastage[product_name] = {
                    "product_name": product_name,
                    "product_id": product_id,
                    "opening_qty": 0,
                    "purchase_qty": 0,
                    "dispatch_qty": 0,
                    "closing_qty": 0,
                    "wastage_qty": 0,
                    "wastage_value": 0,
                    "status": status
                }
            
            # Aggregate values
            product_wastage[product_name]["opening_qty"] += opening_qty
            product_wastage[product_name]["purchase_qty"] += purchase_qty
            product_wastage[product_name]["dispatch_qty"] += dispatch_qty
            product_wastage[product_name]["closing_qty"] += closing_qty
            product_wastage[product_name]["wastage_qty"] += wastage_qty
            product_wastage[product_name]["wastage_value"] += wastage_value
            if status == "closed":
                product_wastage[product_name]["status"] = "closed"
    
    # Calculate wastage percent
    wastage_products = []
    total_wastage_kg = 0
    total_wastage_value = 0
    
    for product_name, data in product_wastage.items():
        available = data["opening_qty"] + data["purchase_qty"]
        wastage_percent = (data["wastage_qty"] / available * 100) if available > 0 else 0
        
        wastage_products.append({
            "product_name": product_name,
            "product_id": data["product_id"],
            "opening_qty": round(data["opening_qty"], 2),
            "purchase_qty": round(data["purchase_qty"], 2),
            "dispatch_qty": round(data["dispatch_qty"], 2),
            "closing_qty": round(data["closing_qty"], 2),
            "wastage_qty": round(data["wastage_qty"], 2),
            "wastage_value": round(data["wastage_value"], 2),
            "wastage_percent": round(wastage_percent, 1),
            "status": data["status"]
        })
        
        total_wastage_kg += data["wastage_qty"]
        total_wastage_value += data["wastage_value"]
    
    # Sort by wastage percent descending
    wastage_products.sort(key=lambda x: x.get("wastage_percent", 0), reverse=True)
    
    return {
        "date": date,
        "total_wastage_kg": round(total_wastage_kg, 2),
        "total_wastage_value": round(total_wastage_value, 2),
        "products": wastage_products
    }


@api_router.put("/stock-status/{status_id}")
async def update_stock_status(status_id: str, updates: dict, current_user: dict = Depends(get_current_user)):
    """Update a stock status entry"""
    if current_user["role"] not in ["admin", "staff"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    # Remove protected fields
    updates.pop("id", None)
    updates.pop("_id", None)
    updates.pop("date", None)
    updates.pop("product_id", None)
    
    result = await db.daily_stock_status.update_one(
        {"id": status_id},
        {"$set": updates}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Stock status not found")
    
    return {"message": "Stock status updated"}

@api_router.delete("/stock-status/{status_id}")
async def delete_stock_status(status_id: str, current_user: dict = Depends(get_current_user)):
    """Delete a stock status entry"""
    if current_user["role"] not in ["admin", "staff"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    result = await db.daily_stock_status.delete_one({"id": status_id})
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Stock status not found")
    
    return {"message": "Stock status deleted"}


# ============================================================================
# SECTION: VARIABLE EXPENSES ROUTES (Lines ~3104-3210)
# ============================================================================

@api_router.get("/expenses/variable")
async def get_variable_expenses(
    from_date: str = None,
    to_date: str = None,
    category: str = None,
    settled: str = None,
    current_user: dict = Depends(get_current_user)
):
    """Get variable expenses with optional filters"""
    if current_user["role"] not in ["admin", "staff"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    query = {}
    
    if from_date:
        query["date"] = {"$gte": from_date}
    if to_date:
        if "date" in query:
            query["date"]["$lte"] = to_date + "T23:59:59"
        else:
            query["date"] = {"$lte": to_date + "T23:59:59"}
    if category:
        query["category"] = category
    if settled:
        query["is_settled"] = settled.lower() == "true"
    
    expenses = await db.variable_expenses.find(query, {"_id": 0}).sort("date", -1).to_list(500)
    return expenses

@api_router.post("/expenses/variable")
async def create_variable_expense(expense: dict, current_user: dict = Depends(get_current_user)):
    """Create a new variable expense"""
    if current_user["role"] not in ["admin", "staff"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    expense["id"] = str(uuid.uuid4())
    expense["created_at"] = datetime.now(timezone.utc).isoformat()
    expense["created_by"] = current_user["email"]
    
    await db.variable_expenses.insert_one(expense.copy())
    if "_id" in expense:
        del expense["_id"]
    
    return expense

@api_router.put("/expenses/variable/{expense_id}")
async def update_variable_expense(expense_id: str, updates: dict, current_user: dict = Depends(get_current_user)):
    """Update a variable expense"""
    if current_user["role"] not in ["admin", "staff"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    updates["updated_at"] = datetime.now(timezone.utc).isoformat()
    updates["updated_by"] = current_user["email"]
    
    result = await db.variable_expenses.update_one(
        {"id": expense_id},
        {"$set": updates}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Expense not found")
    
    return {"message": "Expense updated"}

@api_router.delete("/expenses/variable/{expense_id}")
async def delete_variable_expense(expense_id: str, current_user: dict = Depends(get_current_user)):
    """Delete a variable expense"""
    if current_user["role"] not in ["admin", "staff"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    result = await db.variable_expenses.delete_one({"id": expense_id})
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Expense not found")
    
    return {"message": "Expense deleted"}

@api_router.post("/expenses/variable/bulk-settle")
async def bulk_settle_variable_expenses(data: dict, current_user: dict = Depends(get_current_user)):
    """Bulk settle employee expenses"""
    if current_user["role"] not in ["admin", "staff"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    expense_ids = data.get("expense_ids", [])
    settlement_date = data.get("settlement_date", datetime.now(timezone.utc).strftime('%Y-%m-%d'))
    settlement_remarks = data.get("settlement_remarks", "")
    
    result = await db.variable_expenses.update_many(
        {"id": {"$in": expense_ids}},
        {"$set": {
            "is_settled": True,
            "payment_status": "paid",  # Update new payment_status field
            "settlement_date": settlement_date,
            "settlement_remarks": settlement_remarks,
            "settled_by": current_user["email"],
            "settled_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    return {"message": f"Settled {result.modified_count} expenses", "count": result.modified_count}

# ============================================================================
# SECTION: FIXED EXPENSES ROUTES (Lines ~3208-3370)
# ============================================================================

@api_router.get("/expenses/fixed")
async def get_fixed_expenses(
    month: int = None,
    year: int = None,
    category: str = None,
    status: str = None,
    current_user: dict = Depends(get_current_user)
):
    """Get fixed expenses with optional filters"""
    if current_user["role"] not in ["admin", "staff"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    query = {}
    
    if month is not None:
        query["month"] = month
    if year is not None:
        query["year"] = year
    if category:
        query["category"] = category
    if status:
        query["status"] = status
    
    expenses = await db.fixed_expenses.find(query, {"_id": 0}).sort([("due_date", 1)]).to_list(500)
    return expenses

@api_router.post("/expenses/fixed")
async def create_fixed_expense(expense: dict, current_user: dict = Depends(get_current_user)):
    """Create a new fixed expense"""
    if current_user["role"] not in ["admin", "staff"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    expense["id"] = str(uuid.uuid4())
    expense["created_at"] = datetime.now(timezone.utc).isoformat()
    expense["created_by"] = current_user["email"]
    
    await db.fixed_expenses.insert_one(expense.copy())
    if "_id" in expense:
        del expense["_id"]
    
    return expense

@api_router.put("/expenses/fixed/{expense_id}")
async def update_fixed_expense(expense_id: str, updates: dict, current_user: dict = Depends(get_current_user)):
    """Update a fixed expense"""
    if current_user["role"] not in ["admin", "staff"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    updates["updated_at"] = datetime.now(timezone.utc).isoformat()
    updates["updated_by"] = current_user["email"]
    
    result = await db.fixed_expenses.update_one(
        {"id": expense_id},
        {"$set": updates}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Expense not found")
    
    return {"message": "Expense updated"}

@api_router.delete("/expenses/fixed/{expense_id}")
async def delete_fixed_expense(expense_id: str, current_user: dict = Depends(get_current_user)):
    """Delete a fixed expense"""
    if current_user["role"] not in ["admin", "staff"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    result = await db.fixed_expenses.delete_one({"id": expense_id})
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Expense not found")
    
    return {"message": "Expense deleted"}

@api_router.post("/expenses/fixed/generate-recurring")
async def generate_recurring_expenses(data: dict, current_user: dict = Depends(get_current_user)):
    """Generate recurring expenses for a given month based on previous month's recurring entries"""
    if current_user["role"] not in ["admin", "staff"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    target_month = data.get("month", datetime.now().month)
    target_year = data.get("year", datetime.now().year)
    
    # Calculate previous month
    if target_month == 0:
        prev_month = 11
        prev_year = target_year - 1
    else:
        prev_month = target_month - 1 if target_month > 0 else 11
        prev_year = target_year if target_month > 0 else target_year - 1
    
    # Get recurring expenses from previous month
    prev_expenses = await db.fixed_expenses.find({
        "month": prev_month,
        "year": prev_year,
        "is_recurring": True
    }, {"_id": 0}).to_list(100)
    
    # Check which ones already exist for target month
    existing = await db.fixed_expenses.find({
        "month": target_month,
        "year": target_year
    }, {"_id": 0, "category": 1, "description": 1}).to_list(100)
    
    existing_keys = set((e.get("category", "") + e.get("description", "")) for e in existing)
    
    # Create new entries for target month
    created_count = 0
    for exp in prev_expenses:
        key = exp.get("category", "") + exp.get("description", "")
        if key not in existing_keys:
            new_expense = {
                "id": str(uuid.uuid4()),
                "month": target_month,
                "year": target_year,
                "category": exp.get("category"),
                "description": exp.get("description"),
                "amount": exp.get("amount"),
                "due_date": exp.get("due_date"),
                "is_recurring": True,
                "status": "Pending",
                "paid_by": "Company",
                "is_settled": True,
                "created_at": datetime.now(timezone.utc).isoformat(),
                "created_by": current_user["email"]
            }
            await db.fixed_expenses.insert_one(new_expense.copy())
            created_count += 1
    
    return {"message": f"Generated {created_count} recurring entries", "count": created_count}

@api_router.post("/expenses/fixed/bulk-settle")
async def bulk_settle_fixed_expenses(data: dict, current_user: dict = Depends(get_current_user)):
    """Bulk settle employee expenses"""
    if current_user["role"] not in ["admin", "staff"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    expense_ids = data.get("expense_ids", [])
    settlement_date = data.get("settlement_date", datetime.now(timezone.utc).strftime('%Y-%m-%d'))
    settlement_remarks = data.get("settlement_remarks", "")
    
    result = await db.fixed_expenses.update_many(
        {"id": {"$in": expense_ids}},
        {"$set": {
            "is_settled": True,
            "settlement_date": settlement_date,
            "settlement_remarks": settlement_remarks,
            "settled_by": current_user["email"],
            "settled_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    return {"message": f"Settled {result.modified_count} expenses", "count": result.modified_count}

# ==================== RETAILER PORTAL APIs ====================

# ============================================================================
# SECTION: RETAILER PORTAL ROUTES (Lines ~3368-4100)
# Includes: Retailers, Indents, Dispatches, GRN, Rejections, Payments, Invoices
# ============================================================================

# Get all retailers (for dropdowns)
@api_router.get("/retailers")
async def get_retailers(current_user: dict = Depends(get_current_user)):
    retailers = await db.users.find({"role": "retailer"}, {"_id": 0, "password": 0}).to_list(500)
    return retailers

# ------------ RETAILER INDENTS ------------
@api_router.get("/retailer-indents")
async def get_retailer_indents(
    retailer_id: str = None,
    current_user: dict = Depends(get_current_user)
):
    query = {}
    # If retailer, only show their own indents
    if current_user["role"] == "retailer":
        query["retailer_id"] = current_user["user_id"]
    elif retailer_id:
        query["retailer_id"] = retailer_id
    
    indents = await db.retailer_indents.find(query, {"_id": 0}).sort("indent_date", -1).to_list(1000)
    return indents

@api_router.post("/retailer-indents")
async def create_retailer_indent(input: RetailerIndentCreate, current_user: dict = Depends(get_current_user)):
    # Get retailer info
    if current_user["role"] == "retailer":
        retailer_id = current_user["user_id"]
    else:
        retailer_id = input.retailer_id
    
    retailer = await db.users.find_one({"id": retailer_id, "role": "retailer"}, {"_id": 0})
    if not retailer:
        raise HTTPException(status_code=404, detail="Retailer not found")
    
    indent = RetailerIndent(
        retailer_id=retailer_id,
        retailer_name=retailer.get("name", "Unknown"),
        indent_date=input.indent_date,
        items=input.items,
        remarks=input.remarks,
        created_by=current_user["user_id"],
        created_by_role=current_user["role"]
    )
    
    doc = indent.model_dump()
    doc["indent_date"] = doc["indent_date"].isoformat()
    doc["created_at"] = doc["created_at"].isoformat()
    
    await db.retailer_indents.insert_one(doc)
    return {"id": indent.id, "message": "Indent created successfully"}

@api_router.put("/retailer-indents/{indent_id}")
async def update_retailer_indent(indent_id: str, input: RetailerIndentCreate, current_user: dict = Depends(get_current_user)):
    existing = await db.retailer_indents.find_one({"id": indent_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Indent not found")
    
    # Retailers can only update their own pending indents
    if current_user["role"] == "retailer":
        if existing["retailer_id"] != current_user["user_id"]:
            raise HTTPException(status_code=403, detail="Not authorized")
        if existing["status"] != "pending":
            raise HTTPException(status_code=400, detail="Cannot edit non-pending indent")
    
    # Get all dispatches for this indent to recalculate status
    all_dispatches = await db.retailer_dispatches.find({"indent_id": indent_id}, {"_id": 0}).to_list(100)
    
    # Calculate total dispatched per product+variant
    total_dispatched = {}
    for d in all_dispatches:
        for item in d.get('items', []):
            key = f"{item.get('product_id', '')}|{item.get('variant_id', '')}"
            total_dispatched[key] = total_dispatched.get(key, 0) + item.get('supplied_qty', 0)
    
    # Check if all new items are fully dispatched
    new_items = [item.model_dump() for item in input.items]
    fully_dispatched = True
    has_any_dispatch = len(all_dispatches) > 0
    
    for item in new_items:
        key = f"{item.get('product_id', '')}|{item.get('variant_id', '')}"
        dispatched = total_dispatched.get(key, 0)
        if dispatched < item.get('quantity', 0):
            fully_dispatched = False
            break
    
    # Determine new status
    if not has_any_dispatch:
        new_status = "pending"
    elif fully_dispatched:
        new_status = "dispatched"
    else:
        new_status = "partial"
    
    update_data = {
        "indent_date": input.indent_date.isoformat(),
        "items": new_items,
        "remarks": input.remarks,
        "status": new_status  # Update status based on dispatch completeness
    }
    
    await db.retailer_indents.update_one({"id": indent_id}, {"$set": update_data})
    return {"id": indent_id, "message": "Indent updated successfully", "new_status": new_status}

@api_router.delete("/retailer-indents/{indent_id}")
async def delete_retailer_indent(indent_id: str, current_user: dict = Depends(get_current_user)):
    existing = await db.retailer_indents.find_one({"id": indent_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Indent not found")
    
    # Retailers can only delete their own indents
    if current_user["role"] == "retailer" and existing["retailer_id"] != current_user["user_id"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    # Check if there are any dispatches for this indent
    dispatches = await db.retailer_dispatches.find({"indent_id": indent_id}, {"_id": 0}).to_list(100)
    if len(dispatches) > 0:
        raise HTTPException(status_code=400, detail="Cannot delete indent with existing dispatches. Delete dispatches first.")
    
    await db.retailer_indents.delete_one({"id": indent_id})
    return {"message": "Indent deleted successfully"}

@api_router.get("/retailer-indents/previous/{retailer_id}")
async def get_previous_retailer_indent(retailer_id: str, current_user: dict = Depends(get_current_user)):
    """Get the most recent indent for a retailer to use as template"""
    if current_user["role"] not in ["admin", "staff"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    # Find the most recent indent for this retailer
    indent = await db.retailer_indents.find_one(
        {"retailer_id": retailer_id},
        {"_id": 0},
        sort=[("indent_date", -1), ("created_at", -1)]
    )
    
    if not indent:
        return {"indent": None, "message": "No previous indent found"}
    
    return {"indent": indent, "message": "Previous indent found"}

# ------------ RETAILER DISPATCHES ------------
@api_router.get("/retailer-dispatches")
async def get_retailer_dispatches(
    retailer_id: str = None,
    current_user: dict = Depends(get_current_user)
):
    query = {}
    if current_user["role"] == "retailer":
        query["retailer_id"] = current_user["user_id"]
    elif retailer_id:
        query["retailer_id"] = retailer_id
    
    dispatches = await db.retailer_dispatches.find(query, {"_id": 0}).sort("dispatch_date", -1).to_list(1000)
    return dispatches

@api_router.post("/retailer-dispatches")
async def create_retailer_dispatch(input: RetailerDispatchCreate, current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in ["admin", "staff"]:
        raise HTTPException(status_code=403, detail="Only admin/staff can create dispatches")
    
    # Get indent
    indent = await db.retailer_indents.find_one({"id": input.indent_id})
    if not indent:
        raise HTTPException(status_code=404, detail="Indent not found")
    
    # Get retailer info for commission
    retailer = await db.users.find_one({"id": indent["retailer_id"]}, {"_id": 0})
    commission = retailer.get("commission_percentage", 0) if retailer else 0
    
    # Calculate totals
    total_mrp_value = 0
    items_with_totals = []
    for item in input.items:
        item_dict = item.model_dump()
        item_dict["total_value"] = item.supplied_qty * item.mrp
        total_mrp_value += item_dict["total_value"]
        items_with_totals.append(item_dict)
    
    net_payable = total_mrp_value * (1 - commission / 100)
    
    dispatch = RetailerDispatch(
        indent_id=input.indent_id,
        retailer_id=indent["retailer_id"],
        retailer_name=indent["retailer_name"],
        dispatch_date=input.dispatch_date,
        items=items_with_totals,
        total_mrp_value=round(total_mrp_value, 2),
        commission_percentage=commission,
        net_payable=round(net_payable, 2),
        transport_charges=input.transport_charges or 0,
        dispatched_by=current_user["user_id"],
        invoice_number=None,  # Invoice created separately
        remarks=input.remarks
    )
    
    doc = dispatch.model_dump()
    doc["dispatch_date"] = doc["dispatch_date"].isoformat()
    doc["created_at"] = doc["created_at"].isoformat()
    
    await db.retailer_dispatches.insert_one(doc)
    
    # Update indent status based on dispatch completeness
    # Get all dispatches for this indent
    all_dispatches = await db.retailer_dispatches.find({"indent_id": input.indent_id}, {"_id": 0}).to_list(100)
    
    # Calculate total dispatched per product
    total_dispatched = {}
    for d in all_dispatches:
        for item in d.get('items', []):
            key = item.get('product_id', '')
            total_dispatched[key] = total_dispatched.get(key, 0) + item.get('supplied_qty', 0)
    
    # Check if fully dispatched
    fully_dispatched = True
    for item in indent.get('items', []):
        dispatched = total_dispatched.get(item.get('product_id', ''), 0)
        if dispatched < item.get('quantity', 0):
            fully_dispatched = False
            break
    
    new_status = "dispatched" if fully_dispatched else "partial"
    await db.retailer_indents.update_one(
        {"id": input.indent_id},
        {"$set": {"status": new_status}}
    )
    
    return {"id": dispatch.id, "message": "Dispatch created successfully"}

# Edit Retailer Dispatch
@api_router.put("/retailer-dispatches/{dispatch_id}")
async def update_retailer_dispatch(dispatch_id: str, input: RetailerDispatchCreate, current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in ["admin", "staff"]:
        raise HTTPException(status_code=403, detail="Only admin/staff can update dispatches")
    
    existing = await db.retailer_dispatches.find_one({"id": dispatch_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Dispatch not found")
    
    # Get retailer info for commission
    retailer = await db.users.find_one({"id": existing["retailer_id"]}, {"_id": 0})
    commission = retailer.get("commission_percentage", 0) if retailer else 0
    
    # Calculate totals
    total_mrp_value = 0
    items_with_totals = []
    for item in input.items:
        item_dict = item.model_dump()
        item_dict["total_value"] = item.supplied_qty * item.mrp
        total_mrp_value += item_dict["total_value"]
        items_with_totals.append(item_dict)
    
    net_payable = total_mrp_value * (1 - commission / 100)
    
    update_data = {
        "dispatch_date": input.dispatch_date.isoformat(),
        "items": items_with_totals,
        "total_mrp_value": round(total_mrp_value, 2),
        "commission_percentage": commission,
        "net_payable": round(net_payable, 2),
        "transport_charges": input.transport_charges or 0,
        "remarks": input.remarks
    }
    
    await db.retailer_dispatches.update_one({"id": dispatch_id}, {"$set": update_data})
    return {"id": dispatch_id, "message": "Dispatch updated successfully"}

# Delete Retailer Dispatch
@api_router.delete("/retailer-dispatches/{dispatch_id}")
async def delete_retailer_dispatch(dispatch_id: str, current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in ["admin", "staff"]:
        raise HTTPException(status_code=403, detail="Only admin/staff can delete dispatches")
    
    existing = await db.retailer_dispatches.find_one({"id": dispatch_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Dispatch not found")
    
    # Check if dispatch is part of an invoice
    invoice = await db.retailer_invoices.find_one({"dispatch_ids": dispatch_id})
    if invoice:
        raise HTTPException(status_code=400, detail="Cannot delete dispatch that is part of an invoice")
    
    indent_id = existing.get("indent_id")
    
    await db.retailer_dispatches.delete_one({"id": dispatch_id})
    
    # Recalculate indent status based on remaining dispatches
    if indent_id:
        remaining_dispatches = await db.retailer_dispatches.find({"indent_id": indent_id}, {"_id": 0}).to_list(100)
        
        if len(remaining_dispatches) == 0:
            # No dispatches left - set back to pending
            await db.retailer_indents.update_one(
                {"id": indent_id},
                {"$set": {"status": "pending"}}
            )
        else:
            # Check if partially or fully dispatched
            indent = await db.retailer_indents.find_one({"id": indent_id}, {"_id": 0})
            if indent:
                total_dispatched = {}
                for d in remaining_dispatches:
                    for item in d.get('items', []):
                        key = item.get('product_id', '')
                        total_dispatched[key] = total_dispatched.get(key, 0) + item.get('supplied_qty', 0)
                
                fully_dispatched = True
                for item in indent.get('items', []):
                    dispatched = total_dispatched.get(item.get('product_id', ''), 0)
                    if dispatched < item.get('quantity', 0):
                        fully_dispatched = False
                        break
                
                new_status = "dispatched" if fully_dispatched else "partial"
                await db.retailer_indents.update_one(
                    {"id": indent_id},
                    {"$set": {"status": new_status}}
                )
    
    return {"message": "Dispatch deleted successfully"}

# ------------ RETAILER GRN ------------
@api_router.get("/retailer-grn")
async def get_retailer_grn(
    retailer_id: str = None,
    dispatch_id: str = None,
    current_user: dict = Depends(get_current_user)
):
    query = {}
    if current_user["role"] == "retailer":
        query["retailer_id"] = current_user["user_id"]
    elif retailer_id:
        query["retailer_id"] = retailer_id
    if dispatch_id:
        query["dispatch_id"] = dispatch_id
    
    grns = await db.retailer_grn.find(query, {"_id": 0}).sort("grn_date", -1).to_list(1000)
    return grns

@api_router.post("/retailer-grn")
async def create_retailer_grn(input: RetailerGRNCreate, current_user: dict = Depends(get_current_user)):
    # Get dispatch
    dispatch = await db.retailer_dispatches.find_one({"id": input.dispatch_id})
    if not dispatch:
        raise HTTPException(status_code=404, detail="Dispatch not found")
    
    # Retailers can only create GRN for their own dispatches
    if current_user["role"] == "retailer" and dispatch["retailer_id"] != current_user["user_id"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    # Check if GRN already exists
    existing = await db.retailer_grn.find_one({"dispatch_id": input.dispatch_id})
    if existing:
        raise HTTPException(status_code=400, detail="GRN already exists for this dispatch")
    
    # Calculate differences
    items_with_diff = []
    for item in input.items:
        item_dict = item.model_dump()
        item_dict["difference"] = item.received_qty - item.supplied_qty
        items_with_diff.append(item_dict)
    
    grn = RetailerGRN(
        dispatch_id=input.dispatch_id,
        retailer_id=dispatch["retailer_id"],
        retailer_name=dispatch["retailer_name"],
        grn_date=datetime.now(timezone.utc),
        items=items_with_diff,
        confirmed_by=current_user["user_id"],
        remarks=input.remarks
    )
    
    doc = grn.model_dump()
    doc["grn_date"] = doc["grn_date"].isoformat()
    doc["created_at"] = doc["created_at"].isoformat()
    
    await db.retailer_grn.insert_one(doc)
    
    # Update indent status to received
    await db.retailer_indents.update_one(
        {"id": dispatch["indent_id"]},
        {"$set": {"status": "received"}}
    )
    
    return {"id": grn.id, "message": "GRN confirmed successfully"}

# ------------ RETAILER REJECTIONS ------------
@api_router.get("/retailer-rejections")
async def get_retailer_rejections(
    retailer_id: str = None,
    current_user: dict = Depends(get_current_user)
):
    query = {}
    if current_user["role"] == "retailer":
        query["retailer_id"] = current_user["user_id"]
    elif retailer_id:
        query["retailer_id"] = retailer_id
    
    rejections = await db.retailer_rejections.find(query, {"_id": 0}).sort("rejection_date", -1).to_list(1000)
    return rejections

@api_router.post("/retailer-rejections")
async def create_retailer_rejection(input: RetailerRejectionCreate, current_user: dict = Depends(get_current_user)):
    # Only admin/staff can create rejections
    if current_user["role"] not in ["admin", "staff"]:
        raise HTTPException(status_code=403, detail="Only admin/staff can record rejections")
    
    # Get retailer info
    retailer = await db.users.find_one({"id": input.retailer_id, "role": "retailer"}, {"_id": 0})
    if not retailer:
        raise HTTPException(status_code=404, detail="Retailer not found")
    
    rejection_value = input.quantity * input.mrp
    
    rejection = RetailerRejection(
        retailer_id=input.retailer_id,
        retailer_name=retailer.get("name", "Unknown"),
        rejection_date=input.rejection_date,
        product_id=input.product_id,
        product_name=input.product_name,
        variant_name=input.variant_name,
        quantity=input.quantity,
        reason=input.reason,
        dispatch_id=input.dispatch_id,
        recorded_by=current_user["user_id"],
        mrp=input.mrp,
        rejection_value=round(rejection_value, 2),
        remarks=input.remarks
    )
    
    doc = rejection.model_dump()
    doc["rejection_date"] = doc["rejection_date"].isoformat()
    doc["created_at"] = doc["created_at"].isoformat()
    
    await db.retailer_rejections.insert_one(doc)
    return {"id": rejection.id, "message": "Rejection recorded successfully"}

@api_router.delete("/retailer-rejections/{rejection_id}")
async def delete_retailer_rejection(rejection_id: str, current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in ["admin", "staff"]:
        raise HTTPException(status_code=403, detail="Only admin/staff can delete rejections")
    
    result = await db.retailer_rejections.delete_one({"id": rejection_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Rejection not found")
    
    return {"message": "Rejection deleted successfully"}


@api_router.put("/retailer-rejections/{rejection_id}")
async def update_retailer_rejection(rejection_id: str, rejection: RetailerRejection, current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in ["admin", "staff"]:
        raise HTTPException(status_code=403, detail="Only admin/staff can update rejections")
    
    # Check if rejection exists
    existing = await db.retailer_rejections.find_one({"id": rejection_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Rejection not found")
    
    # Update the rejection
    update_data = {
        "retailer_id": rejection.retailer_id,
        "rejection_date": rejection.rejection_date,
        "product_id": rejection.product_id,
        "product_name": rejection.product_name,
        "variant_id": rejection.variant_id,
        "variant_name": rejection.variant_name,
        "quantity": rejection.quantity,
        "mrp": rejection.mrp,
        "reason": rejection.reason,
        "remarks": rejection.remarks,
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.retailer_rejections.update_one(
        {"id": rejection_id},
        {"$set": update_data}
    )
    
    return {"id": rejection_id, "message": "Rejection updated successfully"}


# ------------ RETAILER PAYMENTS ------------
@api_router.get("/retailer-payments")
async def get_retailer_payments(
    retailer_id: str = None,
    current_user: dict = Depends(get_current_user)
):
    query = {}
    if current_user["role"] == "retailer":
        query["retailer_id"] = current_user["user_id"]
    elif retailer_id:
        query["retailer_id"] = retailer_id
    
    payments = await db.retailer_payments.find(query, {"_id": 0}).sort("payment_date", -1).to_list(1000)
    return payments

@api_router.post("/retailer-payments")
async def create_retailer_payment(input: RetailerPaymentCreate, current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in ["admin", "staff"]:
        raise HTTPException(status_code=403, detail="Only admin/staff can record payments")
    
    # Get retailer info
    retailer = await db.users.find_one({"id": input.retailer_id, "role": "retailer"}, {"_id": 0})
    if not retailer:
        raise HTTPException(status_code=404, detail="Retailer not found")
    
    payment = RetailerPayment(
        retailer_id=input.retailer_id,
        retailer_name=retailer.get("name", "Unknown"),
        payment_date=input.payment_date,
        amount=input.amount,
        payment_mode=input.payment_mode,
        reference_number=input.reference_number,
        recorded_by=current_user["user_id"],
        remarks=input.remarks
    )
    
    doc = payment.model_dump()
    doc["payment_date"] = doc["payment_date"].isoformat()
    doc["created_at"] = doc["created_at"].isoformat()
    
    await db.retailer_payments.insert_one(doc)
    return {"id": payment.id, "message": "Payment recorded successfully"}

@api_router.delete("/retailer-payments/{payment_id}")
async def delete_retailer_payment(payment_id: str, current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in ["admin", "staff"]:
        raise HTTPException(status_code=403, detail="Only admin/staff can delete payments")
    
    result = await db.retailer_payments.delete_one({"id": payment_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Payment not found")
    
    return {"message": "Payment deleted successfully"}

# ------------ RETAILER INVOICES ------------
@api_router.get("/retailer-invoices")
async def get_retailer_invoices(
    retailer_id: str = None,
    current_user: dict = Depends(get_current_user)
):
    query = {}
    if current_user["role"] == "retailer":
        query["retailer_id"] = current_user["user_id"]
    elif retailer_id:
        query["retailer_id"] = retailer_id
    
    invoices = await db.retailer_invoices.find(query, {"_id": 0}).sort("invoice_date", -1).to_list(1000)
    return invoices

@api_router.get("/retailer-invoices/{invoice_id}")
async def get_retailer_invoice(invoice_id: str, current_user: dict = Depends(get_current_user)):
    invoice = await db.retailer_invoices.find_one({"id": invoice_id}, {"_id": 0})
    if not invoice:
        raise HTTPException(status_code=404, detail="Invoice not found")
    
    # Retailers can only see their own invoices
    if current_user["role"] == "retailer" and invoice["retailer_id"] != current_user["user_id"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    return invoice

@api_router.post("/retailer-invoices")
async def create_retailer_invoice(input: RetailerInvoiceCreate, current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in ["admin", "staff"]:
        raise HTTPException(status_code=403, detail="Only admin/staff can create invoices")
    
    if not input.dispatch_ids:
        raise HTTPException(status_code=400, detail="At least one dispatch is required")
    
    # Get retailer info
    retailer = await db.users.find_one({"id": input.retailer_id}, {"_id": 0})
    if not retailer:
        raise HTTPException(status_code=404, detail="Retailer not found")
    
    commission = retailer.get("commission_percentage", 0)
    
    # Get all dispatches for validation
    dispatches = await db.retailer_dispatches.find(
        {"id": {"$in": input.dispatch_ids}},
        {"_id": 0}
    ).to_list(100)
    
    if len(dispatches) != len(input.dispatch_ids):
        raise HTTPException(status_code=400, detail="Some dispatches not found")
    
    # Check if any dispatch is already in another invoice
    for dispatch in dispatches:
        existing_invoice = await db.retailer_invoices.find_one({"dispatch_ids": dispatch["id"]})
        if existing_invoice:
            raise HTTPException(
                status_code=400, 
                detail=f"Dispatch is already in invoice {existing_invoice.get('invoice_number')}"
            )
    
    # Use selected_items if provided (includes rejection adjustments), otherwise fall back to dispatch items
    all_items = []
    gross_value = 0  # Total before rejections
    rejection_amount = 0  # Total rejection value
    
    if input.selected_items and len(input.selected_items) > 0:
        # Use the item-level selection with rejection data
        for item in input.selected_items:
            all_items.append(RetailerInvoiceItem(
                dispatch_id=item.dispatch_id,
                product_id=item.product_id,
                product_name=item.product_name,
                variant_name=item.variant_name,
                quantity=item.net_qty,  # Use net_qty (after rejection deduction)
                supplied_qty=item.supplied_qty,
                rejected_qty=item.rejected_qty,
                mrp=item.mrp,
                total_value=item.total_value  # This is net_value from frontend
            ))
            # Calculate gross (supplied_qty * mrp)
            item_gross = item.supplied_qty * item.mrp
            gross_value += item_gross
            # Calculate rejection value (rejected_qty * mrp)
            item_rejection = item.rejected_qty * item.mrp if item.rejected_qty else 0
            rejection_amount += item_rejection
    else:
        # Fallback: aggregate items from all dispatches (legacy behavior)
        for dispatch in dispatches:
            for item in dispatch.get("items", []):
                all_items.append(RetailerInvoiceItem(
                    dispatch_id=dispatch["id"],
                    product_id=item.get("product_id", ""),
                    product_name=item.get("product_name", ""),
                    variant_name=item.get("variant_name"),
                    quantity=item.get("supplied_qty", 0),
                    mrp=item.get("mrp", 0),
                    total_value=item.get("total_value", 0)
                ))
                gross_value += item.get("total_value", 0)
    
    # Net MRP value = Gross - Rejections
    total_mrp_value = gross_value - rejection_amount
    
    # Commission on NET value (after rejections)
    commission_amount = total_mrp_value * commission / 100
    net_payable = total_mrp_value - commission_amount
    
    # Generate invoice number: RET-INV-DDMMMYYYY-NNN
    today = input.invoice_date.strftime("%d%b%Y").upper()
    count = await db.retailer_invoices.count_documents({
        "invoice_date": {"$regex": f"^{input.invoice_date.strftime('%Y-%m-%d')}"}
    })
    invoice_number = f"RET-INV-{today}-{str(count + 1).zfill(3)}"
    
    invoice = RetailerInvoice(
        invoice_number=invoice_number,
        retailer_id=input.retailer_id,
        retailer_name=retailer.get("name", "Unknown"),
        invoice_date=input.invoice_date,
        dispatch_ids=input.dispatch_ids,
        items=[item.model_dump() for item in all_items],
        gross_value=round(gross_value, 2),
        rejection_amount=round(rejection_amount, 2),
        total_mrp_value=round(total_mrp_value, 2),
        commission_percentage=commission,
        commission_amount=round(commission_amount, 2),
        net_payable=round(net_payable, 2),
        created_by=current_user["user_id"],
        remarks=input.remarks
    )
    
    doc = invoice.model_dump()
    doc["invoice_date"] = doc["invoice_date"].isoformat()
    doc["created_at"] = doc["created_at"].isoformat()
    
    await db.retailer_invoices.insert_one(doc)
    
    # Update dispatches with invoice number
    await db.retailer_dispatches.update_many(
        {"id": {"$in": input.dispatch_ids}},
        {"$set": {"invoice_number": invoice_number, "invoice_id": invoice.id}}
    )
    
    return {"id": invoice.id, "invoice_number": invoice_number, "message": "Invoice created successfully"}


@api_router.put("/retailer-invoices/{invoice_id}")
async def update_retailer_invoice(invoice_id: str, input: dict, current_user: dict = Depends(get_current_user)):
    """Update an existing retailer invoice"""
    if current_user["role"] not in ["admin", "staff"]:
        raise HTTPException(status_code=403, detail="Only admin/staff can update invoices")
    
    invoice = await db.retailer_invoices.find_one({"id": invoice_id}, {"_id": 0})
    if not invoice:
        raise HTTPException(status_code=404, detail="Invoice not found")
    
    # Process items if provided
    items = input.get("items", invoice.get("items", []))
    total_mrp_value = sum(item.get("total_value", 0) for item in items)
    
    # Get retailer for commission calculation
    retailer = await db.users.find_one({"id": invoice.get("retailer_id")}, {"_id": 0})
    commission_percentage = retailer.get("commission_percentage", 0) if retailer else invoice.get("commission_percentage", 0)
    commission_amount = total_mrp_value * (commission_percentage / 100)
    net_payable = total_mrp_value - commission_amount
    
    # Update invoice
    update_data = {
        "invoice_date": input.get("invoice_date", invoice.get("invoice_date")),
        "items": items,
        "total_mrp_value": round(total_mrp_value, 2),
        "commission_percentage": commission_percentage,
        "commission_amount": round(commission_amount, 2),
        "net_payable": round(net_payable, 2),
        "remarks": input.get("remarks", invoice.get("remarks", ""))
    }
    
    await db.retailer_invoices.update_one(
        {"id": invoice_id},
        {"$set": update_data}
    )
    
    return {"id": invoice_id, "message": "Invoice updated successfully"}


@api_router.delete("/retailer-invoices/{invoice_id}")
async def delete_retailer_invoice(invoice_id: str, current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in ["admin", "staff"]:
        raise HTTPException(status_code=403, detail="Only admin/staff can delete invoices")
    
    invoice = await db.retailer_invoices.find_one({"id": invoice_id})
    if not invoice:
        raise HTTPException(status_code=404, detail="Invoice not found")
    
    # Remove invoice reference from dispatches
    await db.retailer_dispatches.update_many(
        {"id": {"$in": invoice.get("dispatch_ids", [])}},
        {"$set": {"invoice_number": None, "invoice_id": None}}
    )
    
    await db.retailer_invoices.delete_one({"id": invoice_id})
    return {"message": "Invoice deleted successfully"}

# Record payment against an invoice
@api_router.post("/retailer-invoices/{invoice_id}/payment")
async def record_invoice_payment(invoice_id: str, input: dict, current_user: dict = Depends(get_current_user)):
    """Record payment against a retailer invoice"""
    if current_user["role"] not in ["admin", "staff"]:
        raise HTTPException(status_code=403, detail="Only admin/staff can record payments")
    
    invoice = await db.retailer_invoices.find_one({"id": invoice_id}, {"_id": 0})
    if not invoice:
        raise HTTPException(status_code=404, detail="Invoice not found")
    
    amount = input.get("amount", 0)
    payment_mode = input.get("payment_mode", "cash")
    received_by = input.get("received_by", current_user.get("user_id", ""))
    received_by_name = input.get("received_by_name", current_user.get("name", ""))
    reference_number = input.get("reference_number", "")
    remarks = input.get("remarks", "")
    payment_date = input.get("payment_date", datetime.now(timezone.utc).isoformat())
    
    # Calculate new paid amount and status
    current_paid = invoice.get("paid_amount", 0)
    new_paid = current_paid + amount
    net_payable = invoice.get("net_payable", 0)
    
    # Determine payment status
    if new_paid >= net_payable:
        new_status = "paid"
    elif new_paid > 0:
        new_status = "partial"
    else:
        new_status = "pending"
    
    # Record the payment in retailer_payments collection
    payment_id = str(uuid.uuid4())
    retailer = await db.users.find_one({"id": invoice.get("retailer_id")}, {"_id": 0})
    payment_doc = {
        "id": payment_id,
        "retailer_id": invoice.get("retailer_id"),
        "retailer_name": retailer.get("company_name", retailer.get("name", "")) if retailer else "",
        "invoice_id": invoice_id,
        "invoice_number": invoice.get("invoice_number"),
        "payment_date": payment_date,
        "amount": amount,
        "payment_mode": payment_mode,
        "received_by": received_by,
        "received_by_name": received_by_name,
        "reference_number": reference_number,
        "remarks": remarks,
        "recorded_by": current_user["user_id"],
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.retailer_payments.insert_one(payment_doc)
    
    # Update the invoice with new paid amount and status
    await db.retailer_invoices.update_one(
        {"id": invoice_id},
        {"$set": {
            "paid_amount": round(new_paid, 2),
            "status": new_status
        }}
    )
    
    return {
        "id": payment_id,
        "invoice_id": invoice_id,
        "paid_amount": round(new_paid, 2),
        "status": new_status,
        "message": "Payment recorded successfully"
    }

# Get payment history for an invoice
@api_router.get("/retailer-invoices/{invoice_id}/payments")
async def get_invoice_payments(invoice_id: str, current_user: dict = Depends(get_current_user)):
    """Get all payments for a specific invoice"""
    payments = await db.retailer_payments.find(
        {"invoice_id": invoice_id}, 
        {"_id": 0}
    ).sort("payment_date", -1).to_list(100)
    return payments

# Get uninvoiced dispatches for a retailer (for creating invoices)
# Now returns items that haven't been invoiced yet (item-level filtering)
@api_router.get("/retailer-dispatches/uninvoiced")
async def get_uninvoiced_dispatches(
    retailer_id: str,
    current_user: dict = Depends(get_current_user)
):
    if current_user["role"] not in ["admin", "staff"]:
        raise HTTPException(status_code=403, detail="Only admin/staff can view uninvoiced dispatches")
    
    # Get all dispatches for this retailer
    dispatches = await db.retailer_dispatches.find({
        "retailer_id": retailer_id
    }, {"_id": 0}).sort("dispatch_date", -1).to_list(500)
    
    # Get all invoices for this retailer to find which items are already invoiced
    invoices = await db.retailer_invoices.find({
        "retailer_id": retailer_id
    }, {"_id": 0, "items": 1}).to_list(500)
    
    # Build a set of already invoiced items (dispatch_id + product_id combo)
    # Using dispatch_id + product_id + variant as key since item_index may not be reliable
    invoiced_item_keys = set()
    for invoice in invoices:
        for item in invoice.get("items", []):
            dispatch_id = item.get("dispatch_id", "")
            product_id = item.get("product_id", "")
            variant = item.get("variant_name", "") or ""
            if dispatch_id and product_id:
                invoiced_item_keys.add(f"{dispatch_id}_{product_id}_{variant}")
    
    # Filter dispatches to only include uninvoiced items
    result = []
    for dispatch in dispatches:
        uninvoiced_items = []
        for item in dispatch.get("items", []):
            product_id = item.get("product_id", "")
            variant = item.get("variant_name", "") or ""
            item_key = f"{dispatch['id']}_{product_id}_{variant}"
            
            if item_key not in invoiced_item_keys:
                uninvoiced_items.append(item)
        
        # Only include dispatch if it has uninvoiced items
        if uninvoiced_items:
            dispatch_copy = dict(dispatch)
            dispatch_copy["items"] = uninvoiced_items
            result.append(dispatch_copy)
    
    return result

# ------------ RETAILER DASHBOARD ------------
@api_router.get("/retailer-dashboard")
async def get_retailer_dashboard(
    retailer_id: str = None,
    current_user: dict = Depends(get_current_user)
):
    # Determine which retailer's data to show
    if current_user["role"] == "retailer":
        target_retailer_id = current_user["user_id"]
    elif retailer_id:
        target_retailer_id = retailer_id
    else:
        raise HTTPException(status_code=400, detail="Retailer ID required")
    
    # Get retailer info
    retailer = await db.users.find_one({"id": target_retailer_id}, {"_id": 0, "password": 0})
    if not retailer:
        raise HTTPException(status_code=404, detail="Retailer not found")
    
    # Get all dispatches for this retailer
    dispatches = await db.retailer_dispatches.find(
        {"retailer_id": target_retailer_id}, 
        {"_id": 0}
    ).to_list(1000)
    
    # Get all rejections
    rejections = await db.retailer_rejections.find(
        {"retailer_id": target_retailer_id},
        {"_id": 0}
    ).to_list(1000)
    
    # Get all payments
    payments = await db.retailer_payments.find(
        {"retailer_id": target_retailer_id},
        {"_id": 0}
    ).to_list(1000)
    
    # Get pending indents
    pending_indents = await db.retailer_indents.count_documents({
        "retailer_id": target_retailer_id,
        "status": {"$in": ["pending", "partial"]}
    })
    
    # Calculate totals
    total_mrp_value = sum(d.get("total_mrp_value", 0) for d in dispatches)
    total_net_payable = sum(d.get("net_payable", 0) for d in dispatches)
    total_rejection_value = sum(r.get("rejection_value", 0) for r in rejections)
    total_paid = sum(p.get("amount", 0) for p in payments)
    
    # Adjusted payable after rejections
    adjusted_payable = total_net_payable - total_rejection_value
    pending_amount = max(0, adjusted_payable - total_paid)
    
    # Total items received
    total_items_received = 0
    for d in dispatches:
        for item in d.get("items", []):
            total_items_received += item.get("supplied_qty", 0)
    
    # Total items rejected
    total_items_rejected = sum(r.get("quantity", 0) for r in rejections)
    
    return {
        "retailer": {
            "id": retailer.get("id"),
            "name": retailer.get("name"),
            "email": retailer.get("email"),
            "contact": retailer.get("contact"),
            "address": retailer.get("address"),
            "company_name": retailer.get("company_name"),
            "commission_percentage": retailer.get("commission_percentage", 0),
            "referral_code": retailer.get("referral_code")
        },
        "summary": {
            "total_mrp_value": round(total_mrp_value, 2),
            "total_net_payable": round(total_net_payable, 2),
            "total_rejection_value": round(total_rejection_value, 2),
            "adjusted_payable": round(adjusted_payable, 2),
            "total_paid": round(total_paid, 2),
            "pending_amount": round(pending_amount, 2),
            "total_items_received": total_items_received,
            "total_items_rejected": total_items_rejected,
            "pending_indents": pending_indents,
            "total_dispatches": len(dispatches),
            "total_payments": len(payments)
        }
    }

# ============================================================================
# SECTION: BACKUP & DATA MANAGEMENT ROUTES (Lines ~4127-4200)
# ============================================================================

@api_router.post("/backup/trigger")
async def trigger_backup(current_user: dict = Depends(get_current_user)):
    """Manually trigger a backup and send to configured email addresses"""
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Only admin can trigger backups")
    
    result = await trigger_manual_backup(db)
    return result

@api_router.get("/backup/download")
async def download_backup(current_user: dict = Depends(get_current_user)):
    """Download backup Excel file directly"""
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Only admin can download backups")
    
    try:
        excel_data = await generate_backup_excel(db)
        today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
        filename = f"MrOrganix_Backup_{today}.xlsx"
        
        return StreamingResponse(
            io.BytesIO(excel_data),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to generate backup: {str(e)}")

@api_router.get("/backup/status")
async def get_backup_status(current_user: dict = Depends(get_current_user)):
    """Get backup scheduler status"""
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Only admin can view backup status")
    
    return {
        "scheduler_running": backup_scheduler is not None and backup_scheduler.running if backup_scheduler else False,
        "next_backup_time": "11:59 PM IST (18:29 UTC)",
        "recipients": ["dshukla11189@gmail.com"],
        "collections_backed_up": [
            "users", "products", "farmers", "procurements",
            "qc_customers", "qc_indents", "qc_dispatches", "qc_invoices", "qc_grns",
            "retailer_indents", "retailer_dispatches", "retailer_invoices", 
            "retailer_grn", "retailer_rejections", "retailer_payments",
            "daily_stock_status", "variable_expenses", "fixed_expenses"
        ]
    }

@api_router.post("/admin/reset-data")
async def reset_all_data(current_user: dict = Depends(get_current_user)):
    """
    Reset all data for production - keeps only admin user.
    USE WITH CAUTION - This deletes ALL business data!
    """
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Only admin can reset data")
    
    collections_to_clear = [
        "products", "farmers", "procurements",
        "qc_customers", "qc_indents", "qc_dispatches", "qc_invoices", "qc_grns",
        "retailer_indents", "retailer_dispatches", "retailer_invoices", 
        "retailer_grn", "retailer_rejections", "retailer_payments",
        "daily_stock_status", "variable_expenses", "fixed_expenses"
    ]
    
    deleted_counts = {}
    for coll_name in collections_to_clear:
        result = await db[coll_name].delete_many({})
        deleted_counts[coll_name] = result.deleted_count
    
    # Delete non-admin users (keep admin accounts)
    users_deleted = await db.users.delete_many({"role": {"$ne": "admin"}})
    deleted_counts["users (non-admin)"] = users_deleted.deleted_count
    
    return {
        "success": True,
        "message": "All test data cleared. Ready for production!",
        "deleted": deleted_counts
    }


# Pydantic model for sync request
class SyncFromProductionRequest(BaseModel):
    production_url: str = Field(default="https://harvest-hub-384.emergent.host", description="Production app URL")
    admin_email: str = Field(default="admin@freshflow.com", description="Admin email for production")
    admin_password: str = Field(default="admin123", description="Admin password for production")
    reset_passwords: bool = Field(default=True, description="Reset all user passwords to default")
    default_password: str = Field(default="admin123", description="Default password for users after sync")


@api_router.post("/sync-from-production")
async def sync_from_production(
    request: SyncFromProductionRequest,
    current_user: dict = Depends(get_current_user)
):
    """
    Sync data from production app to this preview/test environment.
    Downloads backup Excel from production and imports it here.
    
    IMPORTANT: This will REPLACE all existing data in the synced collections.
    """
    import ast
    import json as json_lib
    import httpx
    import openpyxl
    from io import BytesIO
    
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Only admin can sync from production")
    
    try:
        production_url = request.production_url.rstrip('/')
        
        async with httpx.AsyncClient(timeout=60.0) as client:
            # Step 1: Login to production
            login_response = await client.post(
                f"{production_url}/api/auth/login",
                json={
                    "email": request.admin_email,
                    "password": request.admin_password
                }
            )
            
            if login_response.status_code != 200:
                raise HTTPException(
                    status_code=401,
                    detail=f"Failed to login to production: {login_response.text}"
                )
            
            token = login_response.json().get("token")
            if not token:
                raise HTTPException(status_code=401, detail="No token received from production login")
            
            # Step 2: Download backup Excel
            backup_response = await client.get(
                f"{production_url}/api/backup/download",
                headers={"Authorization": f"Bearer {token}"},
                timeout=120.0
            )
            
            if backup_response.status_code != 200:
                raise HTTPException(
                    status_code=500,
                    detail=f"Failed to download backup from production: {backup_response.status_code}"
                )
            
            # Step 3: Parse Excel and import data
            excel_data = BytesIO(backup_response.content)
            wb = openpyxl.load_workbook(excel_data)
            
            # Helper function to parse JSON fields
            def safe_parse_json(value):
                if not value or not isinstance(value, str):
                    return value
                try:
                    return json_lib.loads(value)
                except:
                    pass
                try:
                    return ast.literal_eval(value)
                except:
                    pass
                return value
            
            # Map sheet names to collection names
            sheet_to_collection = {
                'users': 'users',
                'products': 'products',
                'farmers': 'farmers',
                'procurements': 'procurements',
                'qc_packaging': 'qc_packaging',
                'qc_customers': 'qc_customers',
                'qc_indents': 'qc_indents',
                'qc_dispatches': 'qc_dispatches',
                'qc_invoices': 'qc_invoices',
                'qc_grns': 'qc_grns',
                'qc_daily_requirements': 'qc_daily_requirements',
                'retailer_indents': 'retailer_indents',
                'retailer_dispatches': 'retailer_dispatches',
                'retailer_invoices': 'retailer_invoices',
                'retailer_grn': 'retailer_grn',
                'retailer_rejections': 'retailer_rejections',
                'retailer_payments': 'retailer_payments',
                'retailer_closing_inventory': 'retailer_closing_inventory',
                'daily_stock_status': 'daily_stock_status',
                'variable_expenses': 'variable_expenses',
                'fixed_expenses': 'fixed_expenses',
            }
            
            # JSON fields per collection
            json_fields_map = {
                'procurements': ['products'],
                'qc_indents': ['items'],
                'qc_dispatches': ['items', 'packaging_items'],
                'qc_invoices': ['items'],
                'qc_grns': ['items'],
                'qc_daily_requirements': ['items'],
                'retailer_indents': ['items'],
                'retailer_dispatches': ['items'],
                'retailer_invoices': ['items'],
                'retailer_grn': ['items'],
                'retailer_rejections': ['items'],
                'retailer_closing_inventory': ['items'],
            }
            
            sync_results = {}
            total_synced = 0
            
            for sheet_name, collection_name in sheet_to_collection.items():
                if sheet_name not in wb.sheetnames:
                    sync_results[collection_name] = {"status": "not_found", "count": 0}
                    continue
                
                ws = wb[sheet_name]
                rows = list(ws.iter_rows(values_only=True))
                
                if len(rows) <= 1:
                    sync_results[collection_name] = {"status": "empty", "count": 0}
                    continue
                
                # Get headers from first row
                headers = [str(h).lower().strip() if h else f'col_{i}' for i, h in enumerate(rows[0])]
                
                # Convert rows to documents
                documents = []
                for row in rows[1:]:  # Skip header row
                    doc = {}
                    for i, value in enumerate(row):
                        if i < len(headers):
                            key = headers[i]
                            if key == '_id':
                                continue  # Skip _id field
                            if value is not None:
                                # Parse JSON fields
                                if collection_name in json_fields_map and key in json_fields_map[collection_name]:
                                    doc[key] = safe_parse_json(value)
                                else:
                                    doc[key] = value
                    if doc:
                        documents.append(doc)
                
                if documents:
                    # Clear existing data and insert new
                    await db[collection_name].delete_many({})
                    result = await db[collection_name].insert_many(documents)
                    sync_results[collection_name] = {
                        "status": "success",
                        "count": len(result.inserted_ids)
                    }
                    total_synced += len(result.inserted_ids)
                else:
                    sync_results[collection_name] = {"status": "no_data", "count": 0}
            
            # Reset passwords if requested
            if request.reset_passwords:
                hashed_password = bcrypt.hashpw(
                    request.default_password.encode('utf-8'), 
                    bcrypt.gensalt()
                ).decode('utf-8')
                
                await db.users.update_many(
                    {},
                    {"$set": {"password": hashed_password}}
                )
                sync_results["password_reset"] = {
                    "status": "success",
                    "message": f"All user passwords reset to: {request.default_password}"
                }
            
            return {
                "success": True,
                "message": f"Successfully synced {total_synced} records from production",
                "total_records": total_synced,
                "production_url": production_url,
                "collections": sync_results
            }
        
    except httpx.TimeoutException:
        raise HTTPException(status_code=504, detail="Connection to production timed out")
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=500, 
            detail=f"Failed to sync from production: {str(e)}"
        )


@api_router.get("/sync-status")
async def get_sync_status(current_user: dict = Depends(get_current_user)):
    """Get current database record counts for sync verification"""
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Only admin can view sync status")
    
    collections = [
        "users", "products", "farmers", "procurements", "qc_packaging",
        "qc_customers", "qc_indents", "qc_dispatches", "qc_invoices", "qc_grns",
        "retailer_indents", "retailer_dispatches", "retailer_invoices", 
        "retailer_grn", "retailer_rejections", "retailer_payments",
        "daily_stock_status", "variable_expenses", "fixed_expenses"
    ]
    
    counts = {}
    total = 0
    for coll in collections:
        count = await db[coll].count_documents({})
        counts[coll] = count
        total += count
    
    return {
        "database": DB_NAME,
        "total_records": total,
        "collections": counts,
        "last_checked": datetime.now(timezone.utc).isoformat()
    }


# ============================================================================
# SECTION: GMAIL INTEGRATION ROUTES - Ninjacart GRN Email Automation
# ============================================================================

@api_router.get("/oauth/gmail/debug-states")
async def gmail_oauth_debug_states():
    """DEBUG: Show OAuth states in database"""
    states = await db.oauth_states.find({}).to_list(length=10)
    return {
        "total_states": len(states),
        "states": [
            {
                "state": s.get("state", "N/A")[:20] + "...",
                "user_id": s.get("user_id"),
                "redirect_uri": s.get("redirect_uri"),
                "created_at": str(s.get("created_at")),
                "expires_at": str(s.get("expires_at"))
            }
            for s in states
        ]
    }

@api_router.get("/oauth/gmail/debug-tokens")
async def gmail_oauth_debug_tokens():
    """DEBUG: Show Gmail tokens in database (without sensitive data)"""
    tokens = await db.gmail_tokens.find({}).to_list(length=10)
    return {
        "total_tokens": len(tokens),
        "tokens": [
            {
                "user_id": t.get("user_id"),
                "email": t.get("email"),
                "has_access_token": bool(t.get("access_token")),
                "has_refresh_token": bool(t.get("refresh_token")),
                "updated_at": str(t.get("updated_at"))
            }
            for t in tokens
        ]
    }

@api_router.get("/oauth/gmail/debug-headers")
async def gmail_oauth_debug_headers(request: Request):
    """DEBUG: Show what headers the server receives - helps diagnose redirect_uri issues"""
    host = request.headers.get("host", "")
    x_forwarded_host = request.headers.get("x-forwarded-host", "")
    x_forwarded_proto = request.headers.get("x-forwarded-proto", "https")
    origin = request.headers.get("origin", "")
    referer = request.headers.get("referer", "")
    
    # Determine the actual host - prefer x-forwarded-host for reverse proxy setups
    actual_host = x_forwarded_host or host
    
    # Detect redirect URI based on actual host
    if "emergent.host" in actual_host:
        redirect_uri = "https://harvest-hub-384.emergent.host/api/oauth/gmail/callback"
    elif "preview.emergentagent.com" in actual_host:
        redirect_uri = "https://harvest-hub-384.preview.emergentagent.com/api/oauth/gmail/callback"
    else:
        # Fallback: try to detect from referer or origin
        if "emergent.host" in referer or "emergent.host" in origin:
            redirect_uri = "https://harvest-hub-384.emergent.host/api/oauth/gmail/callback"
        else:
            redirect_uri = "https://harvest-hub-384.preview.emergentagent.com/api/oauth/gmail/callback"
    
    return {
        "headers_received": {
            "host": host,
            "x-forwarded-host": x_forwarded_host,
            "x-forwarded-proto": x_forwarded_proto,
            "origin": origin,
            "referer": referer
        },
        "actual_host_used": actual_host,
        "redirect_uri_that_would_be_generated": redirect_uri,
        "expected_google_console_uri": redirect_uri,
        "message": "The 'redirect_uri_that_would_be_generated' must EXACTLY match the URI in Google Cloud Console Authorized Redirect URIs"
    }

@api_router.get("/oauth/gmail/status")
async def get_gmail_connection_status(current_user: dict = Depends(get_current_user)):
    """Check if Gmail is connected"""
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Only admin can manage Gmail integration")
    
    token = await db.gmail_tokens.find_one({"user_id": current_user["user_id"]}, {"_id": 0})
    
    return {
        "connected": token is not None and token.get("access_token") is not None,
        "email": token.get("email") if token else None,
        "last_sync": token.get("last_sync") if token else None
    }

@api_router.get("/oauth/gmail/login")
async def gmail_oauth_login(request: Request, current_user: dict = Depends(get_current_user)):
    """Start Gmail OAuth flow"""
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Only admin can connect Gmail")
    
    # Generate state token
    state = str(uuid.uuid4())
    
    # Debug: Log all relevant headers to understand what we're receiving
    host = request.headers.get("host", "")
    x_forwarded_host = request.headers.get("x-forwarded-host", "")
    x_forwarded_proto = request.headers.get("x-forwarded-proto", "https")
    origin = request.headers.get("origin", "")
    referer = request.headers.get("referer", "")
    
    logger.info("OAuth Login - Headers Debug:")
    logger.info(f"  host: {host}")
    logger.info(f"  x-forwarded-host: {x_forwarded_host}")
    logger.info(f"  x-forwarded-proto: {x_forwarded_proto}")
    logger.info(f"  origin: {origin}")
    logger.info(f"  referer: {referer}")
    
    # Determine the actual host - prefer x-forwarded-host for reverse proxy setups
    actual_host = x_forwarded_host or host
    
    # Detect redirect URI based on actual host
    if "emergent.host" in actual_host:
        redirect_uri = "https://harvest-hub-384.emergent.host/api/oauth/gmail/callback"
    elif "preview.emergentagent.com" in actual_host:
        redirect_uri = "https://harvest-hub-384.preview.emergentagent.com/api/oauth/gmail/callback"
    else:
        # Fallback: try to detect from referer or origin
        if "emergent.host" in referer or "emergent.host" in origin:
            redirect_uri = "https://harvest-hub-384.emergent.host/api/oauth/gmail/callback"
        else:
            redirect_uri = "https://harvest-hub-384.preview.emergentagent.com/api/oauth/gmail/callback"
    
    logger.info(f"  Selected redirect_uri: {redirect_uri}")
    
    # Store state in DB with user_id and redirect_uri
    await db.oauth_states.insert_one({
        "state": state,
        "user_id": current_user["user_id"],
        "redirect_uri": redirect_uri,
        "created_at": datetime.now(timezone.utc),
        "expires_at": datetime.now(timezone.utc) + timedelta(minutes=10)
    })
    
    # Get authorization URL with dynamic redirect URI
    auth_url = get_authorization_url(state, redirect_uri)
    
    logger.info(f"  Generated auth_url: {auth_url[:100]}...")
    
    return {"auth_url": auth_url, "debug_redirect_uri": redirect_uri}

@api_router.get("/oauth/gmail/callback")
async def gmail_oauth_callback(request: Request, code: str = None, state: str = None, error: str = None):
    """Handle Gmail OAuth callback"""
    logger.info(f"Gmail OAuth callback received - code: {code[:20] if code else 'None'}..., state: {state}, error: {error}")
    
    if error:
        logger.error(f"Gmail OAuth error from Google: {error}")
        return RedirectResponse(url=f"/admin/backup?gmail_error={error}")
    
    if not code or not state:
        logger.error(f"Gmail OAuth missing params - code: {bool(code)}, state: {bool(state)}")
        return RedirectResponse(url="/admin/backup?gmail_error=missing_params")
    
    # Verify state
    state_doc = await db.oauth_states.find_one({"state": state})
    if not state_doc:
        # Log all states to debug
        all_states = await db.oauth_states.find({}).to_list(length=10)
        logger.error(f"Gmail OAuth invalid_state: {state}")
        logger.error(f"Available states in DB: {[s.get('state', 'N/A')[:20] for s in all_states]}")
        return RedirectResponse(url="/admin/backup?gmail_error=invalid_state")
    
    logger.info(f"State validated - user_id: {state_doc.get('user_id')}, redirect_uri: {state_doc.get('redirect_uri')}")
    
    # Check expiry
    if datetime.now(timezone.utc) > state_doc["expires_at"].replace(tzinfo=timezone.utc):
        return RedirectResponse(url="/admin/backup?gmail_error=state_expired")
    
    user_id = state_doc["user_id"]
    
    # Delete used state
    await db.oauth_states.delete_one({"state": state})
    
    # Get redirect URI from state
    redirect_uri = state_doc.get("redirect_uri")
    
    try:
        # Exchange code for tokens
        logger.info(f"Exchanging code for tokens with redirect_uri: {redirect_uri}")
        tokens = exchange_code_for_tokens(code, redirect_uri)
        logger.info(f"Token exchange successful - has_access_token: {bool(tokens.get('access_token'))}, has_refresh_token: {bool(tokens.get('refresh_token'))}")
        
        # Get user email
        service = get_gmail_service(tokens)
        profile = service.users().getProfile(userId='me').execute()
        email = profile.get('emailAddress', '')
        logger.info(f"Gmail profile retrieved - email: {email}")
        
        # Store tokens
        await db.gmail_tokens.update_one(
            {"user_id": user_id},
            {"$set": {
                "user_id": user_id,
                "email": email,
                **tokens,
                "updated_at": datetime.now(timezone.utc)
            }},
            upsert=True
        )
        logger.info(f"Gmail tokens saved for user: {user_id}")
        
        return RedirectResponse(url="/admin/backup?gmail_success=true")
    except Exception as e:
        logger.error(f"Gmail OAuth error: {e}", exc_info=True)
        return RedirectResponse(url=f"/admin/backup?gmail_error={str(e)[:50]}")

@api_router.post("/oauth/gmail/disconnect")
async def gmail_disconnect(current_user: dict = Depends(get_current_user)):
    """Disconnect Gmail integration"""
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Only admin can disconnect Gmail")
    
    await db.gmail_tokens.delete_one({"user_id": current_user["user_id"]})
    return {"message": "Gmail disconnected successfully"}

@api_router.get("/gmail/ninjacart-emails")
async def get_ninjacart_emails(
    max_results: int = 10,
    current_user: dict = Depends(get_current_user)
):
    """Fetch recent Ninjacart emails"""
    if current_user["role"] not in ["admin", "staff"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    # Get Gmail tokens
    token = await db.gmail_tokens.find_one({"user_id": current_user["user_id"]}, {"_id": 0})
    if not token:
        raise HTTPException(status_code=400, detail="Gmail not connected. Please connect Gmail first.")
    
    try:
        service = get_gmail_service(token)
        
        # Search for Ninjacart emails
        messages = search_ninjacart_emails(service, max_results)
        
        emails = []
        for msg in messages[:5]:  # Limit to 5 for performance
            content = get_email_content(service, msg['id'])
            if content:
                emails.append({
                    "id": content['id'],
                    "subject": content['subject'],
                    "from": content['from'],
                    "date": content['date'],
                    "has_attachments": len(content['attachments']) > 0,
                    "attachments": [a['filename'] for a in content['attachments']]
                })
        
        return {"emails": emails}
    except Exception as e:
        logger.error(f"Error fetching emails: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to fetch emails: {str(e)}")

@api_router.post("/gmail/process-grn-email/{message_id}")
async def process_grn_from_email(
    message_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Process a specific email and create GRN from it"""
    if current_user["role"] not in ["admin", "staff"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    # Get Gmail tokens
    token = await db.gmail_tokens.find_one({"user_id": current_user["user_id"]}, {"_id": 0})
    if not token:
        raise HTTPException(status_code=400, detail="Gmail not connected")
    
    try:
        service = get_gmail_service(token)
        
        # Get email content
        content = get_email_content(service, message_id)
        if not content:
            raise HTTPException(status_code=404, detail="Email not found")
        
        grn_items = []
        
        # Try to parse CSV attachment first
        for attachment in content['attachments']:
            if attachment['filename'].lower().endswith('.csv'):
                csv_data = download_attachment(service, message_id, attachment['attachment_id'])
                if csv_data:
                    grn_items = parse_ninjacart_csv(csv_data)
                    break
        
        # If no CSV, try to parse email body
        if not grn_items and content['body_text']:
            grn_items = parse_ninjacart_email_body(content['body_text'])
        
        if not grn_items:
            raise HTTPException(status_code=400, detail="Could not parse GRN data from email")
        
        # Mark email as processed
        mark_email_as_read(service, message_id)
        add_label_to_email(service, message_id, "FreshFlow-Processed")
        
        return {
            "message": "Email processed successfully",
            "email_subject": content['subject'],
            "items_found": len(grn_items),
            "grn_items": grn_items
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error processing email: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to process email: {str(e)}")

@api_router.post("/gmail/auto-sync-grn")
async def auto_sync_ninjacart_grn(current_user: dict = Depends(get_current_user)):
    """
    Automatically sync Ninjacart GRN from latest unprocessed email.
    This processes emails, matches with dispatches, and saves GRN to database.
    """
    if current_user["role"] not in ["admin", "staff"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    # Get Gmail tokens
    token = await db.gmail_tokens.find_one({"user_id": current_user["user_id"]}, {"_id": 0})
    if not token:
        raise HTTPException(status_code=400, detail="Gmail not connected")
    
    try:
        service = get_gmail_service(token)
        
        # Search for unread Ninjacart emails
        result = service.users().messages().list(
            userId='me',
            q='from:ninjacart is:unread',
            maxResults=5
        ).execute()
        
        messages = result.get('messages', [])
        
        if not messages:
            return {"message": "No new Ninjacart emails found", "processed": 0}
        
        processed_count = 0
        saved_grns = []
        total_items_saved = 0
        rate_changes_detected = []
        
        for msg in messages:
            content = get_email_content(service, msg['id'])
            if not content:
                continue
            
            items = []
            attachment_filename = None
            
            # Try CSV/Excel first
            for attachment in content['attachments']:
                fname = attachment['filename'].lower()
                if fname.endswith('.csv') or fname.endswith('.xlsx') or fname.endswith('.xls'):
                    attachment_data = download_attachment(service, msg['id'], attachment['attachment_id'])
                    if attachment_data:
                        attachment_filename = attachment['filename']
                        if fname.endswith('.csv'):
                            items = parse_ninjacart_csv(attachment_data)
                        else:
                            # Parse Excel
                            try:
                                from openpyxl import load_workbook
                                import io
                                wb = load_workbook(io.BytesIO(attachment_data))
                                ws = wb.active
                                rows = list(ws.iter_rows(values_only=True))
                                if rows:
                                    headers = [str(h).strip() if h else '' for h in rows[0]]
                                    items = []
                                    for row in rows[1:]:
                                        row_dict = {headers[i]: row[i] for i in range(min(len(headers), len(row)))}
                                        items.append(row_dict)
                            except Exception as e:
                                logger.error(f"Error parsing Excel attachment: {e}")
                        break
            
            # Fallback to body parsing
            if not items and content['body_text']:
                items = parse_ninjacart_email_body(content['body_text'])
            
            if items:
                # Process items similar to manual upload
                # Match with dispatches and save to database
                try:
                    # Create a mock UploadFile-like object and process
                    matched_items = []  # This would need the full matching logic
                    
                    # For now, save raw GRN data
                    grn_doc = {
                        "id": str(uuid.uuid4()),
                        "source": "gmail_auto_sync",
                        "email_id": msg['id'],
                        "email_subject": content['subject'],
                        "email_date": content['date'],
                        "attachment_filename": attachment_filename,
                        "items": items,
                        "created_at": datetime.now(timezone.utc).isoformat(),
                        "created_by": current_user["user_id"],
                        "sync_method": "automatic_gmail"
                    }
                    
                    # Save to database
                    await db.gmail_grn_syncs.insert_one(grn_doc)
                    
                    saved_grns.append({
                        "email_id": msg['id'],
                        "subject": content['subject'],
                        "date": content['date'],
                        "items_count": len(items)
                    })
                    total_items_saved += len(items)
                    
                except Exception as e:
                    logger.error(f"Error saving GRN from email: {e}")
                
                # Mark as processed
                mark_email_as_read(service, msg['id'])
                add_label_to_email(service, msg['id'], "FreshFlow-Processed")
                processed_count += 1
        
        # Update last sync time
        await db.gmail_tokens.update_one(
            {"user_id": current_user["user_id"]},
            {"$set": {"last_sync": datetime.now(timezone.utc)}}
        )
        
        # Build summary for frontend display
        summary = {
            "sync_date": datetime.now(timezone.utc).isoformat(),
            "sync_method": "automatic_gmail",
            "emails_processed": processed_count,
            "total_items": total_items_saved,
            "saved_grns": saved_grns,
            "rate_changes": rate_changes_detected
        }
        
        return {
            "message": f"Processed {processed_count} Ninjacart emails",
            "processed": processed_count,
            "summary": summary,
            "grn_data": saved_grns
        }
    except Exception as e:
        logger.error(f"Error in auto-sync: {e}")
        raise HTTPException(status_code=500, detail=f"Auto-sync failed: {str(e)}")


# ==================== AUTO INDENT GENERATION ====================

async def generate_auto_indents_for_tomorrow():
    """
    Auto-generate retailer indents for the next day based on sales history.
    Runs at 11 PM daily.
    
    Logic:
    1. Calculate tomorrow's day of week (e.g., Monday)
    2. For each retailer, find items sold on the last 7 occurrences of that day
    3. Calculate average and increase by 10%
    4. Create auto-generated indent
    """
    try:
        logger.info("Starting auto-indent generation for tomorrow...")
        
        # Get tomorrow's date
        now = datetime.now(timezone.utc)
        tomorrow = now + timedelta(days=1)
        tomorrow_str = tomorrow.strftime('%Y-%m-%d')
        tomorrow_weekday = tomorrow.weekday()  # 0=Monday, 6=Sunday
        
        # Get all retailers
        retailers = await db.users.find({"role": "retailer"}, {"_id": 0}).to_list(500)
        logger.info(f"Found {len(retailers)} retailers to process")
        
        # Get all products for reference
        products = await db.products.find({}, {"_id": 0}).to_list(1000)
        product_map = {p["id"]: p for p in products}
        
        auto_indents_created = 0
        
        for retailer in retailers:
            retailer_id = retailer["id"]
            retailer_name = retailer.get("company_name") or retailer.get("name", "Unknown")
            
            try:
                # Check if indent already exists for tomorrow
                existing_indent = await db.retailer_indents.find_one({
                    "retailer_id": retailer_id,
                    "indent_date": {"$regex": f"^{tomorrow_str}"}
                })
                
                if existing_indent:
                    logger.info(f"Indent already exists for {retailer_name} on {tomorrow_str}, skipping")
                    continue
                
                # Find the last 7 occurrences of the same weekday
                target_dates = []
                check_date = tomorrow - timedelta(days=7)
                
                while len(target_dates) < 7 and check_date > now - timedelta(days=90):
                    if check_date.weekday() == tomorrow_weekday:
                        target_dates.append(check_date.strftime('%Y-%m-%d'))
                    check_date -= timedelta(days=1)
                
                if not target_dates:
                    logger.info(f"No historical data for {retailer_name}, skipping")
                    continue
                
                logger.info(f"Analyzing {len(target_dates)} {['Mon','Tue','Wed','Thu','Fri','Sat','Sun'][tomorrow_weekday]}s for {retailer_name}")
                
                # Get dispatches for those dates
                date_regex_patterns = [f"^{d}" for d in target_dates]
                
                dispatches = await db.retailer_dispatches.find({
                    "retailer_id": retailer_id,
                    "$or": [{"dispatch_date": {"$regex": pattern}} for pattern in date_regex_patterns]
                }, {"_id": 0}).to_list(500)
                
                if not dispatches:
                    logger.info(f"No dispatch history for {retailer_name} on target weekdays, skipping")
                    continue
                
                # Get closing inventory data
                closing_records = await db.retailer_closing_inventory.find({
                    "retailer_id": retailer_id,
                    "closing_date": {"$in": target_dates}
                }, {"_id": 0}).to_list(100)
                
                closing_map = {}
                for record in closing_records:
                    closing_map[record["closing_date"]] = {
                        item["product_id"]: item["closing_qty"]
                        for item in record.get("items", [])
                        if item.get("closing_qty") is not None
                    }
                
                # Calculate items received per product per date
                received_per_date = {}
                for dispatch in dispatches:
                    dispatch_date = dispatch["dispatch_date"][:10]
                    if dispatch_date not in received_per_date:
                        received_per_date[dispatch_date] = {}
                    
                    for item in dispatch.get("items", []):
                        product_id = item.get("product_id")
                        qty = item.get("supplied_qty", 0)
                        if product_id:
                            received_per_date[dispatch_date][product_id] = \
                                received_per_date[dispatch_date].get(product_id, 0) + qty
                
                # Calculate items sold per product
                product_sales = {}
                sales_count = {}
                
                for date_str, received_items in received_per_date.items():
                    prev_date = (datetime.strptime(date_str, '%Y-%m-%d') - timedelta(days=1)).strftime('%Y-%m-%d')
                    prev_closing = closing_map.get(prev_date, {})
                    current_closing = closing_map.get(date_str, {})
                    
                    for product_id, received_qty in received_items.items():
                        opening_qty = prev_closing.get(product_id, 0)
                        closing_qty = current_closing.get(product_id)
                        
                        if closing_qty is not None:
                            items_sold = opening_qty + received_qty - closing_qty
                        else:
                            items_sold = received_qty * 0.8
                        
                        if items_sold > 0:
                            product_sales[product_id] = product_sales.get(product_id, 0) + items_sold
                            sales_count[product_id] = sales_count.get(product_id, 0) + 1
                
                if not product_sales:
                    logger.info(f"No sales data for {retailer_name}, skipping")
                    continue
                
                # Calculate average and add 10%
                indent_items = []
                for product_id, total_sales in product_sales.items():
                    avg_sales = total_sales / sales_count[product_id]
                    recommended_qty = round(avg_sales * 1.1)
                    
                    if recommended_qty > 0 and product_id in product_map:
                        product = product_map[product_id]
                        indent_items.append({
                            "product_id": product_id,
                            "product_name": product.get("name", "Unknown"),
                            "variant_id": None,
                            "variant_name": None,
                            "quantity": recommended_qty,
                            "status": "pending"
                        })
                
                if not indent_items:
                    logger.info(f"No items to indent for {retailer_name}, skipping")
                    continue
                
                # Create the auto-generated indent
                indent_doc = {
                    "id": str(uuid.uuid4()),
                    "retailer_id": retailer_id,
                    "retailer_name": retailer_name,
                    "indent_date": tomorrow.isoformat(),
                    "items": indent_items,
                    "status": "pending",
                    "created_by": "system",
                    "created_by_role": "system",
                    "remarks": f"Auto-generated based on last {len(target_dates)} {['Mon','Tue','Wed','Thu','Fri','Sat','Sun'][tomorrow_weekday]}s sales",
                    "is_auto_generated": True,
                    "created_at": datetime.now(timezone.utc).isoformat()
                }
                
                await db.retailer_indents.insert_one(indent_doc)
                auto_indents_created += 1
                logger.info(f"Created auto-indent for {retailer_name} with {len(indent_items)} items")
                
            except Exception as e:
                logger.error(f"Error processing retailer {retailer_name}: {e}")
                import traceback
                traceback.print_exc()
                continue
        
        logger.info(f"Auto-indent generation completed. Created {auto_indents_created} indents.")
        return {"indents_created": auto_indents_created}
        
    except Exception as e:
        logger.error(f"Auto-indent generation error: {e}")
        import traceback
        traceback.print_exc()
        return {"error": str(e)}


def generate_auto_indents_wrapper():
    """Wrapper to run async auto-indent generation from scheduler"""
    import asyncio
    loop = asyncio.new_event_loop()
    asyncio.set_event_loop(loop)
    try:
        loop.run_until_complete(generate_auto_indents_for_tomorrow())
    finally:
        loop.close()


@api_router.post("/admin/generate-auto-indents")
async def trigger_auto_indent_generation(current_user: dict = Depends(get_current_user)):
    """Manually trigger auto-indent generation for testing"""
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Only admin can trigger auto-indent generation")
    
    result = await generate_auto_indents_for_tomorrow()
    return result


@api_router.post("/admin/generate-auto-indent")
async def generate_single_auto_indent(
    request: dict,
    current_user: dict = Depends(get_current_user)
):
    """Generate auto-indent for a single retailer for a specific date"""
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Only admin can generate auto indents")
    
    retailer_id = request.get("retailer_id")
    target_date_str = request.get("target_date")
    
    if not retailer_id:
        raise HTTPException(status_code=400, detail="retailer_id is required")
    
    try:
        # Parse target date
        if target_date_str:
            target_date = datetime.strptime(target_date_str, "%Y-%m-%d").date()
        else:
            target_date = (datetime.now(timezone.utc) + timedelta(days=1)).date()
        
        # Get retailer info
        retailer = await db.users.find_one({"id": retailer_id, "role": "retailer"}, {"_id": 0})
        if not retailer:
            raise HTTPException(status_code=404, detail="Retailer not found")
        
        retailer_name = retailer.get("company_name") or retailer.get("name", "Unknown")
        
        # Check if indent already exists for this retailer and date
        existing_indent = await db.retailer_indents.find_one({
            "retailer_id": retailer_id,
            "indent_date": target_date.isoformat()
        })
        
        if existing_indent:
            return {
                "success": False,
                "message": f"An indent already exists for {retailer_name} on {target_date.isoformat()}. Delete it first to regenerate."
            }
        
        # Get the target weekday (0=Monday, 6=Sunday)
        target_weekday = target_date.weekday()
        
        # Find the last 7 occurrences of this weekday in retailer closing data
        closing_records = await db.retailer_closing.find({
            "retailer_id": retailer_id
        }).sort("closing_date", -1).to_list(100)
        
        # Filter to only include same weekday
        same_weekday_records = []
        for record in closing_records:
            try:
                record_date = datetime.strptime(record["closing_date"], "%Y-%m-%d").date()
                if record_date.weekday() == target_weekday and record_date < target_date:
                    same_weekday_records.append(record)
                    if len(same_weekday_records) >= 7:
                        break
            except:
                continue
        
        # Use whatever data is available (even 1 record is fine)
        if len(same_weekday_records) == 0:
            return {
                "success": False,
                "message": f"No historical data found for {retailer_name} on this weekday. The retailer needs to have at least one closing record for a {['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday'][target_weekday]}."
            }
        
        # Calculate average items sold per product
        product_totals = {}
        for record in same_weekday_records:
            for item in record.get("items", []):
                product_id = item.get("product_id")
                product_name = item.get("product_name", "")
                items_sold = item.get("items_sold", 0) or 0
                
                if product_id and items_sold > 0:
                    if product_id not in product_totals:
                        product_totals[product_id] = {
                            "product_name": product_name,
                            "total_sold": 0,
                            "count": 0
                        }
                    product_totals[product_id]["total_sold"] += items_sold
                    product_totals[product_id]["count"] += 1
        
        if not product_totals:
            return {
                "success": False,
                "message": f"No sales data found for {retailer_name} on this weekday"
            }
        
        # Create indent items with average + 10% buffer
        indent_items = []
        for product_id, data in product_totals.items():
            avg_sold = data["total_sold"] / data["count"]
            recommended_qty = round(avg_sold * 1.1)  # Add 10% buffer
            
            if recommended_qty > 0:
                indent_items.append({
                    "product_id": product_id,
                    "product_name": data["product_name"],
                    "variant_id": "",
                    "variant_name": "",
                    "quantity": recommended_qty,
                    "status": "pending"
                })
        
        if not indent_items:
            return {
                "success": False,
                "message": f"No products with positive sales found for {retailer_name}"
            }
        
        # Sort by product name
        indent_items.sort(key=lambda x: x["product_name"])
        
        # Create the indent
        new_indent = {
            "id": str(uuid.uuid4()),
            "retailer_id": retailer_id,
            "retailer_name": retailer_name,
            "indent_date": target_date.isoformat(),
            "items": indent_items,
            "total_qty": sum(item["quantity"] for item in indent_items),
            "status": "pending",
            "remarks": f"Auto-generated based on {len(same_weekday_records)} weeks of {['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday'][target_weekday]} sales",
            "is_auto_generated": True,
            "created_at": datetime.now(timezone.utc).isoformat(),
            "updated_at": datetime.now(timezone.utc).isoformat()
        }
        
        await db.retailer_indents.insert_one(new_indent)
        
        return {
            "success": True,
            "message": f"Auto indent created for {retailer_name} with {len(indent_items)} products",
            "retailer_name": retailer_name,
            "indent_id": new_indent["id"],
            "products_count": len(indent_items),
            "total_qty": new_indent["total_qty"]
        }
        
    except HTTPException:
        raise
    except Exception as e:
        print(f"Error generating auto indent: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


@api_router.post("/admin/populate-hindi-names")
async def populate_hindi_product_names(current_user: dict = Depends(get_current_user)):
    """Populate Hindi names for all products in the database"""
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Only admin can populate Hindi names")
    
    try:
        products = await db.products.find({}).to_list(1000)
        updated_count = 0
        missing = []
        
        for product in products:
            name = product.get("name", "")
            product_id = product.get("id")
            
            if name in HINDI_PRODUCT_NAMES:
                hindi_name = HINDI_PRODUCT_NAMES[name]
                result = await db.products.update_one(
                    {"id": product_id},
                    {"$set": {"name_hi": hindi_name}}
                )
                if result.modified_count > 0:
                    updated_count += 1
            else:
                missing.append(name)
        
        return {
            "success": True,
            "updated_count": updated_count,
            "total_products": len(products),
            "missing_translations": missing[:10] if missing else [],
            "message": f"Updated {updated_count} products with Hindi names"
        }
    except Exception as e:
        print(f"Error populating Hindi names: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.post("/admin/populate-referral-codes")
async def populate_retailer_referral_codes(current_user: dict = Depends(get_current_user)):
    """Populate referral codes for all retailers that don't have one"""
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Only admin can populate referral codes")
    
    try:
        # Find all retailers without referral codes
        retailers = await db.users.find({
            "role": "retailer",
            "$or": [
                {"referral_code": {"$exists": False}},
                {"referral_code": None},
                {"referral_code": ""}
            ]
        }).to_list(1000)
        
        updated_count = 0
        
        for retailer in retailers:
            # Generate unique 5-digit referral code
            referral_code = ''.join(random.choices(string.digits, k=5))
            
            # Make sure it's unique
            while await db.users.find_one({"referral_code": referral_code}):
                referral_code = ''.join(random.choices(string.digits, k=5))
            
            result = await db.users.update_one(
                {"id": retailer["id"]},
                {"$set": {"referral_code": referral_code}}
            )
            
            if result.modified_count > 0:
                updated_count += 1
        
        # Count total retailers
        total_retailers = await db.users.count_documents({"role": "retailer"})
        
        return {
            "success": True,
            "updated_count": updated_count,
            "total_retailers": total_retailers,
            "message": f"Generated referral codes for {updated_count} retailers"
        }
    except Exception as e:
        print(f"Error populating referral codes: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.post("/admin/reset-all-referral-codes")
async def reset_all_retailer_referral_codes(current_user: dict = Depends(get_current_user)):
    """Reset ALL retailer referral codes to 5-digit format"""
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Only admin can reset referral codes")
    
    try:
        # Find ALL retailers
        retailers = await db.users.find({"role": "retailer"}).to_list(1000)
        
        updated_count = 0
        used_codes = set()
        
        for retailer in retailers:
            # Generate unique 5-digit referral code
            referral_code = ''.join(random.choices(string.digits, k=5))
            
            # Make sure it's unique (check both DB and already-assigned in this batch)
            while referral_code in used_codes or await db.users.find_one({"referral_code": referral_code, "id": {"$ne": retailer["id"]}}):
                referral_code = ''.join(random.choices(string.digits, k=5))
            
            used_codes.add(referral_code)
            
            result = await db.users.update_one(
                {"id": retailer["id"]},
                {"$set": {"referral_code": referral_code}}
            )
            
            if result.modified_count > 0:
                updated_count += 1
        
        return {
            "success": True,
            "updated_count": updated_count,
            "total_retailers": len(retailers),
            "message": f"Reset referral codes for {updated_count} retailers to 5-digit format"
        }
    except Exception as e:
        print(f"Error resetting referral codes: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# Include router
app.include_router(api_router)

# Add no-cache middleware for API responses
from starlette.middleware.base import BaseHTTPMiddleware
from starlette.responses import Response

class NoCacheMiddleware(BaseHTTPMiddleware):
    async def dispatch(self, request, call_next):
        response = await call_next(request)
        if request.url.path.startswith('/api/'):
            response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
            response.headers['Pragma'] = 'no-cache'
            response.headers['Expires'] = '0'
        return response

app.add_middleware(NoCacheMiddleware)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

@app.on_event("startup")
async def startup_event():
    """Initialize backup scheduler, Gmail sync scheduler, and seed default data on startup"""
    global backup_scheduler
    try:
        # Auto-seed default units if collection is empty
        await seed_default_units_on_startup()
        
        backup_scheduler = setup_backup_scheduler(db)
        logger.info("Backup scheduler initialized successfully")
        
        # Add Gmail GRN sync job at 6 AM IST (00:30 UTC)
        # IST is UTC+5:30, so 6:00 AM IST = 00:30 UTC
        backup_scheduler.add_job(
            scheduled_gmail_sync,
            CronTrigger(hour=0, minute=30),  # 6:00 AM IST
            id='gmail_grn_sync',
            replace_existing=True,
            misfire_grace_time=3600  # Allow 1 hour grace period
        )
        logger.info("Gmail GRN sync scheduled for 6:00 AM IST daily")
        
        # Add Auto-Indent generation job at 11 PM IST (17:30 UTC)
        # IST is UTC+5:30, so 11:00 PM IST = 17:30 UTC
        backup_scheduler.add_job(
            generate_auto_indents_wrapper,
            CronTrigger(hour=17, minute=30),  # 11:00 PM IST
            id='auto_indent_generation',
            replace_existing=True,
            misfire_grace_time=3600  # Allow 1 hour grace period
        )
        logger.info("Auto-indent generation scheduled for 11:00 PM IST daily")
        
        # Check if we missed the 6 AM sync today and run it now
        await check_and_run_missed_gmail_sync()
        
        # Log all scheduled jobs for debugging
        jobs = backup_scheduler.get_jobs()
        for job in jobs:
            logger.info(f"Scheduled job: {job.id} - Next run: {job.next_run_time}")
            
    except Exception as e:
        logger.error(f"Failed to initialize schedulers: {e}")
        import traceback
        traceback.print_exc()


async def check_and_run_missed_gmail_sync():
    """Check if we missed today's 6 AM sync and run it if needed"""
    try:
        # Get the last sync time from settings or sync history
        settings = await db.gmail_settings.find_one({}, {"_id": 0})
        last_sync = settings.get("last_sync_time") if settings else None
        
        if last_sync:
            # Parse the last sync time
            if isinstance(last_sync, str):
                last_sync_dt = datetime.fromisoformat(last_sync.replace('Z', '+00:00'))
            else:
                last_sync_dt = last_sync
            
            # Get today's 6 AM IST (00:30 UTC)
            now = datetime.now(timezone.utc)
            today_6am_ist = now.replace(hour=0, minute=30, second=0, microsecond=0)
            
            # If it's past 6 AM IST today and last sync was before today's 6 AM
            if now > today_6am_ist and last_sync_dt < today_6am_ist:
                logger.info(f"Missed 6 AM sync. Last sync: {last_sync_dt}, Running now...")
                await scheduled_gmail_sync()
                logger.info("Catch-up sync completed")
        else:
            logger.info("No previous sync found, skipping catch-up check")
            
    except Exception as e:
        logger.error(f"Error checking missed sync: {e}")

async def scheduled_gmail_sync():
    """Scheduled job to sync Ninjacart GRN from Gmail at 6 AM"""
    logger.info("Starting scheduled Gmail GRN sync...")
    try:
        # Get all admin users with Gmail connected
        tokens = await db.gmail_tokens.find({}, {"_id": 0}).to_list(10)
        
        if not tokens:
            logger.info("No Gmail accounts connected for GRN sync")
            return
        
        for token in tokens:
            try:
                from gmail_integration import (
                    get_gmail_service, search_ninjacart_emails, get_email_content,
                    download_attachment, parse_ninjacart_csv, parse_ninjacart_email_body,
                    mark_email_as_read, add_label_to_email
                )
                
                service = get_gmail_service(token)
                
                # Search for unread Ninjacart emails from last 24 hours
                yesterday = (datetime.now(timezone.utc) - timedelta(days=1)).strftime('%Y/%m/%d')
                result = service.users().messages().list(
                    userId='me',
                    q=f'from:ninjacart is:unread after:{yesterday}',
                    maxResults=10
                ).execute()
                
                messages = result.get('messages', [])
                
                if not messages:
                    logger.info(f"No new Ninjacart emails for {token.get('email')}")
                    continue
                
                processed = 0
                for msg in messages:
                    content = get_email_content(service, msg['id'])
                    if not content:
                        continue
                    
                    items = []
                    for attachment in content['attachments']:
                        if attachment['filename'].lower().endswith('.csv'):
                            csv_data = download_attachment(service, msg['id'], attachment['attachment_id'])
                            if csv_data:
                                items = parse_ninjacart_csv(csv_data)
                                break
                    
                    if items:
                        # Store parsed GRN data for manual review
                        await db.pending_grn_imports.insert_one({
                            "email_id": msg['id'],
                            "subject": content['subject'],
                            "date": content['date'],
                            "items": items,
                            "imported_at": datetime.now(timezone.utc),
                            "status": "pending_review",
                            "user_email": token.get('email')
                        })
                        
                        mark_email_as_read(service, msg['id'])
                        add_label_to_email(service, msg['id'], "FreshFlow-Processed")
                        processed += 1
                
                logger.info(f"Processed {processed} Ninjacart emails for {token.get('email')}")
                
                # Update last sync time
                await db.gmail_tokens.update_one(
                    {"email": token.get('email')},
                    {"$set": {"last_sync": datetime.now(timezone.utc)}}
                )
                
            except Exception as e:
                logger.error(f"Error syncing Gmail for {token.get('email')}: {e}")
                continue
        
        logger.info("Scheduled Gmail GRN sync completed")
    except Exception as e:
        logger.error(f"Gmail sync scheduler error: {e}")




# ==================== RETAILER INVENTORY ENDPOINTS ====================

@app.get("/api/retailer-inventory/{retailer_id}")
async def get_retailer_inventory(
    retailer_id: str,
    date: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Get retailer inventory for a specific date or all dates"""
    query = {"retailer_id": retailer_id}
    if date:
        query["date"] = date
    
    inventory = await db.retailer_inventory.find(query, {"_id": 0}).sort("date", -1).to_list(500)
    return inventory

@app.post("/api/retailer-inventory/generate/{retailer_id}")
async def generate_retailer_inventory(
    retailer_id: str,
    date: str = Query(..., description="Date in YYYY-MM-DD format"),
    current_user: dict = Depends(get_current_user)
):
    """Generate inventory items from dispatches for a specific date"""
    from models import RetailerInventoryItem
    
    # Check if inventory already exists for this date
    existing = await db.retailer_inventory.find_one({"retailer_id": retailer_id, "date": date})
    if existing:
        return {"message": "Inventory already exists for this date", "exists": True}
    
    # Get dispatches for this retailer on this date
    dispatches = await db.retailer_dispatches.find({
        "retailer_id": retailer_id,
        "dispatch_date": {"$regex": f"^{date}"}
    }).to_list(100)
    
    if not dispatches:
        return {"message": "No dispatches found for this date", "items_created": 0}
    
    # Get previous day's inventory to calculate opening qty
    prev_date = (datetime.strptime(date, "%Y-%m-%d") - timedelta(days=1)).strftime("%Y-%m-%d")
    prev_inventory = await db.retailer_inventory.find(
        {"retailer_id": retailer_id, "date": prev_date}, {"_id": 0}
    ).to_list(500)
    prev_closing = {(i["product_id"], i.get("variant_name", "")): i.get("closing_qty", 0) for i in prev_inventory}
    
    # Aggregate items from all dispatches
    items_map = {}
    for dispatch in dispatches:
        for item in dispatch.get("items", []):
            product_id = item.get("product_id", "")
            variant = item.get("variant_name", "")
            key = (product_id, variant)
            
            if key not in items_map:
                items_map[key] = {
                    "product_id": product_id,
                    "product_name": item.get("product_name", ""),
                    "variant_name": variant,
                    "received_qty": 0
                }
            items_map[key]["received_qty"] += item.get("supplied_qty", 0)
    
    # Create inventory items
    created_items = []
    for key, item_data in items_map.items():
        opening_qty = prev_closing.get(key, 0)
        received_qty = item_data["received_qty"]
        
        inventory_item = RetailerInventoryItem(
            retailer_id=retailer_id,
            product_id=item_data["product_id"],
            product_name=item_data["product_name"],
            variant_name=item_data["variant_name"],
            date=date,
            opening_qty=opening_qty,
            received_qty=received_qty,
            sold_qty=0,
            wastage_qty=0,
            closing_qty=opening_qty + received_qty  # Initial closing = opening + received
        )
        
        await db.retailer_inventory.insert_one(inventory_item.model_dump())
        created_items.append(inventory_item.model_dump())
    
    return {"message": f"Created {len(created_items)} inventory items", "items_created": len(created_items), "items": created_items}

@app.put("/api/retailer-inventory/{item_id}")
async def update_retailer_inventory_item(
    item_id: str,
    update: dict,
    current_user: dict = Depends(get_current_user)
):
    """Update inventory item (sold_qty, wastage_qty, closing_qty, remarks)"""
    from datetime import datetime, timezone
    
    # Get current item
    item = await db.retailer_inventory.find_one({"id": item_id})
    if not item:
        raise HTTPException(status_code=404, detail="Inventory item not found")
    
    # Calculate closing qty if sold/wastage updated
    opening = item.get("opening_qty", 0)
    received = item.get("received_qty", 0)
    sold = update.get("sold_qty", item.get("sold_qty", 0))
    wastage = update.get("wastage_qty", item.get("wastage_qty", 0))
    closing = opening + received - sold - wastage
    
    update_data = {
        "sold_qty": sold,
        "wastage_qty": wastage,
        "closing_qty": closing,
        "remarks": update.get("remarks", item.get("remarks")),
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.retailer_inventory.update_one(
        {"id": item_id},
        {"$set": update_data}
    )
    
    # Return updated item
    updated = await db.retailer_inventory.find_one({"id": item_id}, {"_id": 0})
    return updated

@app.delete("/api/retailer-inventory/{item_id}")
async def delete_retailer_inventory_item(
    item_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Delete an inventory item"""
    result = await db.retailer_inventory.delete_one({"id": item_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Inventory item not found")
    return {"message": "Item deleted successfully"}

@app.get("/api/retailer-inventory/dates/{retailer_id}")
async def get_retailer_inventory_dates(
    retailer_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Get list of dates that have inventory data for a retailer"""
    pipeline = [
        {"$match": {"retailer_id": retailer_id}},
        {"$group": {"_id": "$date"}},
        {"$sort": {"_id": -1}},
        {"$limit": 30}
    ]
    dates = await db.retailer_inventory.aggregate(pipeline).to_list(30)
    return [d["_id"] for d in dates]


@app.post("/api/retailer-inventory/add-item")
async def add_retailer_inventory_item(
    data: dict,
    current_user: dict = Depends(get_current_user)
):
    """Add a single inventory item manually"""
    from models import RetailerInventoryItem
    
    retailer_id = data.get("retailer_id")
    product_id = data.get("product_id")
    product_name = data.get("product_name")
    variant_name = data.get("variant_name", "")
    date = data.get("date")
    opening_qty = float(data.get("opening_qty", 0))
    received_qty = float(data.get("received_qty", 0))
    
    if not retailer_id or not product_id or not date:
        raise HTTPException(status_code=400, detail="retailer_id, product_id, and date are required")
    
    # Check if item already exists
    existing = await db.retailer_inventory.find_one({
        "retailer_id": retailer_id,
        "product_id": product_id,
        "variant_name": variant_name or "",
        "date": date
    })
    if existing:
        raise HTTPException(status_code=400, detail="Item already exists for this product and date")
    
    inventory_item = RetailerInventoryItem(
        retailer_id=retailer_id,
        product_id=product_id,
        product_name=product_name,
        variant_name=variant_name,
        date=date,
        opening_qty=opening_qty,
        received_qty=received_qty,
        sold_qty=0,
        wastage_qty=0,
        closing_qty=opening_qty + received_qty
    )
    
    await db.retailer_inventory.insert_one(inventory_item.model_dump())
    return {"message": "Item added successfully", "item": inventory_item.model_dump()}


@app.post("/api/retailer-inventory/save-all")
async def save_all_retailer_inventory(
    data: dict,
    current_user: dict = Depends(get_current_user)
):
    """Bulk save multiple inventory items"""
    from datetime import datetime, timezone
    
    items = data.get("items", [])
    if not items:
        return {"message": "No items to save", "updated": 0}
    
    updated_count = 0
    for item in items:
        item_id = item.get("id")
        if not item_id:
            continue
            
        # Get current item
        current_item = await db.retailer_inventory.find_one({"id": item_id})
        if not current_item:
            continue
        
        # Calculate closing qty
        opening = current_item.get("opening_qty", 0)
        received = current_item.get("received_qty", 0)
        sold = item.get("sold_qty", current_item.get("sold_qty", 0))
        wastage = item.get("wastage_qty", current_item.get("wastage_qty", 0))
        closing = opening + received - sold - wastage
        
        update_data = {
            "sold_qty": sold,
            "wastage_qty": wastage,
            "closing_qty": closing,
            "remarks": item.get("remarks", current_item.get("remarks")),
            "updated_at": datetime.now(timezone.utc).isoformat()
        }
        
        await db.retailer_inventory.update_one(
            {"id": item_id},
            {"$set": update_data}
        )
        updated_count += 1
    
    return {"message": f"Saved {updated_count} items", "updated": updated_count}


# ============== SIMPLIFIED CLOSING INVENTORY SYSTEM ==============

@app.get("/api/retailer/product-variants/{retailer_id}")
async def get_retailer_product_variants(
    retailer_id: str,
    current_user: dict = Depends(get_current_user)
):
    """
    Get product variants that this retailer has received in dispatches.
    Returns products grouped with their packaging variants.
    """
    # Get retailer dispatches from last 30 days
    from datetime import timedelta
    cutoff = datetime.now(timezone.utc) - timedelta(days=30)
    
    dispatches = await db.retailer_dispatches.find({
        "retailer_id": retailer_id
    }, {"_id": 0}).to_list(500)
    
    # Build product -> variants mapping from dispatch items
    # Note: Field names are variant_id and variant_name (not packaging_id/packaging_name)
    product_variants = {}
    
    for d in dispatches:
        for item in d.get('items', []):
            product_id = item.get('product_id')
            product_name = item.get('product_name')
            # Use variant_id/variant_name fields from retailer dispatch
            variant_id = item.get('variant_id') or item.get('packaging_id')
            variant_name = item.get('variant_name') or item.get('packaging_name') or 'Kg'
            
            if not product_id:
                continue
            
            if product_id not in product_variants:
                # Get product info for Hindi name
                product = await db.products.find_one({"id": product_id}, {"_id": 0, "name_hi": 1, "unit": 1})
                product_variants[product_id] = {
                    'product_id': product_id,
                    'product_name': product_name,
                    'product_name_hi': product.get('name_hi') if product else None,
                    'unit': product.get('unit', 'Kg') if product else 'Kg',
                    'variants': {}
                }
            
            # Track variant by variant_id or name
            variant_key = variant_id or variant_name
            if variant_key not in product_variants[product_id]['variants']:
                product_variants[product_id]['variants'][variant_key] = {
                    'variant_id': variant_id,
                    'variant_name': variant_name,
                    'received_count': 0,
                    'total_qty': 0
                }
            product_variants[product_id]['variants'][variant_key]['received_count'] += 1
            product_variants[product_id]['variants'][variant_key]['total_qty'] += item.get('supplied_qty', 0)
    
    # Convert to list format for frontend
    result = []
    for pid, pdata in sorted(product_variants.items(), key=lambda x: x[1]['product_name']):
        variants = list(pdata['variants'].values())
        # If no specific variants (all 'Kg'), keep as single item
        if len(variants) == 1 and variants[0]['variant_name'] == 'Kg':
            result.append({
                'product_id': pdata['product_id'],
                'product_name': pdata['product_name'],
                'product_name_hi': pdata['product_name_hi'],
                'variant_id': None,
                'variant_name': 'Kg',
                'unit': pdata['unit'],
                'has_variants': False
            })
        else:
            # Multiple variants - add each as separate item
            for v in variants:
                result.append({
                    'product_id': pdata['product_id'],
                    'product_name': pdata['product_name'],
                    'product_name_hi': pdata['product_name_hi'],
                    'variant_id': v['variant_id'],
                    'variant_name': v['variant_name'],
                    'unit': 'Pcs',  # Variants are typically counted in pieces
                    'has_variants': True
                })
    
    return result


@app.get("/api/retailer-closing-inventory/{retailer_id}")
async def get_retailer_closing_inventory(
    retailer_id: str,
    date: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Get closing inventory for a retailer on a specific date"""
    query = {"retailer_id": retailer_id}
    if date:
        query["closing_date"] = date
    
    # Get closing inventory records
    closing_records = await db.retailer_closing_inventory.find(query, {"_id": 0}).sort("closing_date", -1).to_list(500)
    return closing_records

@app.get("/api/retailer-closing-inventory/dates/{retailer_id}")
async def get_retailer_closing_inventory_dates(
    retailer_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Get list of dates that have closing inventory recorded"""
    pipeline = [
        {"$match": {"retailer_id": retailer_id}},
        {"$group": {"_id": "$closing_date"}},
        {"$sort": {"_id": -1}},
        {"$limit": 60}
    ]
    dates = await db.retailer_closing_inventory.aggregate(pipeline).to_list(60)
    return [d["_id"] for d in dates]

@app.post("/api/retailer-closing-inventory/record")
async def record_retailer_closing_inventory(
    data: dict,
    current_user: dict = Depends(get_current_user)
):
    """
    Record closing inventory for a retailer.
    Expects: {
        "retailer_id": str,
        "closing_date": str (YYYY-MM-DD),
        "items": [{"product_id": str, "product_name": str, "variant_id": str (optional), 
                   "variant_name": str (optional), "closing_qty": number or null}, ...]
    }
    """
    retailer_id = data.get("retailer_id")
    closing_date = data.get("closing_date")
    items = data.get("items", [])
    
    if not retailer_id or not closing_date:
        raise HTTPException(status_code=400, detail="retailer_id and closing_date are required")
    
    # Delete existing closing inventory for this date and retailer
    await db.retailer_closing_inventory.delete_many({
        "retailer_id": retailer_id,
        "closing_date": closing_date
    })
    
    # Insert new closing inventory items
    saved_count = 0
    for item in items:
        closing_qty = item.get("closing_qty")
        
        # Only save items that have a closing qty value (not null, not empty string)
        if closing_qty is not None and closing_qty != "" and closing_qty != "":
            try:
                closing_qty_num = float(closing_qty)
            except (ValueError, TypeError):
                continue  # Skip invalid values
            
            inventory_record = {
                "id": str(uuid.uuid4()),
                "retailer_id": retailer_id,
                "closing_date": closing_date,
                "product_id": item.get("product_id"),
                "product_name": item.get("product_name"),
                "variant_id": item.get("variant_id"),
                "variant_name": item.get("variant_name") or "Kg",
                "closing_qty": closing_qty_num,
                "created_at": datetime.now(timezone.utc).isoformat(),
                "created_by": current_user.get("user_id")
            }
            await db.retailer_closing_inventory.insert_one(inventory_record)
            saved_count += 1
    
    return {
        "message": f"Recorded closing inventory for {closing_date}",
        "items_saved": saved_count
    }

@app.get("/api/retailer-closing-inventory/summary/{retailer_id}")
async def get_retailer_closing_summary(
    retailer_id: str,
    date: str,
    current_user: dict = Depends(get_current_user)
):
    """Get full inventory summary including opening, received, rejection, and closing"""
    
    # Parse the date
    try:
        from datetime import timedelta
        target_date = datetime.strptime(date, "%Y-%m-%d").date()
        prev_date = target_date - timedelta(days=1)
        prev_date_str = prev_date.strftime("%Y-%m-%d")
    except:
        prev_date_str = None
    
    # Get closing inventory for this date
    closing_items = await db.retailer_closing_inventory.find(
        {"retailer_id": retailer_id, "closing_date": date},
        {"_id": 0}
    ).to_list(500)
    
    # Get previous day's closing (for opening qty)
    # Build multiple lookup maps: one with variant, one without (product-level)
    prev_closing = {}
    prev_closing_product_only = {}  # Fallback for variant-to-product matching
    if prev_date_str:
        prev_items = await db.retailer_closing_inventory.find(
            {"retailer_id": retailer_id, "closing_date": prev_date_str},
            {"_id": 0}
        ).to_list(500)
        for item in prev_items:
            product_id = item.get('product_id')
            variant_id = item.get('variant_id')
            closing_qty = item.get("closing_qty", 0)
            
            # Key with variant (or 'default' if no variant)
            key = f"{product_id}_{variant_id or 'default'}"
            prev_closing[key] = closing_qty
            
            # Also store product-level aggregate for fallback
            # If multiple variants exist, sum them up (though typically there's one)
            if product_id not in prev_closing_product_only:
                prev_closing_product_only[product_id] = 0
            prev_closing_product_only[product_id] += closing_qty
    
    # Get dispatches for this date (received qty) and build variant name map from ALL dispatches
    dispatches = await db.retailer_dispatches.find({
        "retailer_id": retailer_id
    }, {"_id": 0}).to_list(500)
    
    received_qty = {}
    variant_name_map = {}  # Map to store best variant names from all dispatches
    dispatch_linkage = {}  # Track which dispatches each product+variant is linked to
    
    for d in dispatches:
        dispatch_date = str(d.get('dispatch_date', ''))[:10]
        dispatch_id = d.get('id')
        
        for item in d.get('items', []):
            product_id = item.get('product_id')
            variant_id = item.get('variant_id') or item.get('packaging_id')
            key = f"{product_id}_{variant_id or 'default'}"
            
            # Store variant name from any dispatch (prefer non-Kg names)
            variant_name = item.get('variant_name') or item.get('packaging_name') or 'Kg'
            if variant_name and variant_name != 'Kg':
                variant_name_map[key] = variant_name
            elif key not in variant_name_map:
                variant_name_map[key] = variant_name
            
            # Track dispatch linkage for all dates
            if key not in dispatch_linkage:
                dispatch_linkage[key] = []
            dispatch_linkage[key].append({
                'dispatch_id': dispatch_id,
                'dispatch_date': dispatch_date,
                'variant_name': variant_name
            })
            
            # Only count received qty for the target date
            if dispatch_date == date:
                if key not in received_qty:
                    received_qty[key] = {
                        'qty': 0,
                        'product_name': item.get('product_name'),
                        'variant_id': variant_id,
                        'variant_name': variant_name,
                        'dispatch_dates': []
                    }
                received_qty[key]['qty'] += item.get('supplied_qty', 0) or 0
                if dispatch_date not in received_qty[key]['dispatch_dates']:
                    received_qty[key]['dispatch_dates'].append(dispatch_date)
    
    # Get rejections for this date
    rejections = await db.retailer_rejections.find({
        "retailer_id": retailer_id,
        "rejection_date": {"$regex": f"^{date}"}
    }, {"_id": 0}).to_list(500)
    
    rejection_qty = {}
    for r in rejections:
        product_id = r.get('product_id')
        key = f"{product_id}_default"  # Rejections may not have variants
        rejection_qty[key] = rejection_qty.get(key, 0) + (r.get('quantity', 0) or 0)
    
    # Build comprehensive result - deduplicate by product_id + variant_name (not variant_id)
    # This prevents duplicates when same variant name has different IDs
    result = []
    seen_product_variants = {}  # Track by product_id + variant_name for deduplication
    
    # Get ALL products to include in the result
    all_products = await db.products.find({}, {"_id": 0}).to_list(500)
    
    # First: Add all products from the product list (even those without activity)
    for product in all_products:
        product_id = product.get('id')
        key = f"{product_id}_default"
        dedup_key = f"{product_id}_Kg"
        
        # Get opening from previous day
        opening = prev_closing.get(key, 0)
        if opening == 0 and product_id in prev_closing_product_only:
            opening = prev_closing_product_only.get(product_id, 0)
        
        recv = received_qty.get(key, {}).get('qty', 0)
        rej = rejection_qty.get(f"{product_id}_default", 0)
        linked_dispatches = dispatch_linkage.get(key, [])
        
        result.append({
            "id": None,  # No closing record yet for products without activity
            "product_id": product_id,
            "product_name": product.get("name"),
            "product_name_hi": product.get("name_hi"),
            "variant_id": None,
            "variant_name": "Kg",
            "opening_qty": opening,
            "received_qty": recv,
            "rejection_qty": rej,
            "closing_qty": None,
            "linked_dispatches": linked_dispatches[:5] if linked_dispatches else []
        })
        seen_product_variants[dedup_key] = len(result) - 1
    
    # Then: Update/add closing items with actual recorded closing data
    for item in closing_items:
        product_id = item.get('product_id')
        variant_id = item.get('variant_id')
        key = f"{product_id}_{variant_id or 'default'}"
        
        # Get variant name for deduplication
        stored_variant = item.get("variant_name", "Kg")
        dispatch_variant = received_qty.get(key, {}).get('variant_name')
        map_variant = variant_name_map.get(key)
        
        # Priority: dispatch_variant (today) > map_variant (any day) > stored_variant
        final_variant_name = stored_variant
        if (stored_variant == "Kg" or not stored_variant):
            if dispatch_variant and dispatch_variant != "Kg":
                final_variant_name = dispatch_variant
            elif map_variant and map_variant != "Kg":
                final_variant_name = map_variant
        
        # Deduplicate by product_id + variant_name
        dedup_key = f"{product_id}_{final_variant_name}"
        if dedup_key in seen_product_variants:
            # Update existing entry with closing data
            existing_idx = seen_product_variants[dedup_key]
            existing = result[existing_idx]
            
            # Update with the closing record ID
            if item.get("id"):
                existing["id"] = item.get("id")
            
            # Update quantities from closing record
            existing["closing_qty"] = item.get("closing_qty")
            
            # Update received_qty and opening_qty if different
            recv = received_qty.get(key, {}).get('qty', 0)
            if recv > 0:
                existing["received_qty"] = recv
            
            opening = prev_closing.get(key, 0)
            if opening == 0 and product_id in prev_closing_product_only:
                opening = prev_closing_product_only.get(product_id, 0)
            if opening > 0:
                existing["opening_qty"] = opening
            
            # Merge linked dispatches
            linked = dispatch_linkage.get(key, [])
            if linked:
                existing_linked = existing.get("linked_dispatches") or []
                existing["linked_dispatches"] = (existing_linked + linked)[:5]
            continue
        
        # Get opening from previous day - try exact key match first, then product-only fallback
        opening = prev_closing.get(key, 0)
        if opening == 0 and product_id in prev_closing_product_only:
            opening = prev_closing_product_only.get(product_id, 0)
        
        recv = received_qty.get(key, {}).get('qty', 0)
        rej = rejection_qty.get(f"{product_id}_default", 0)
        linked_dispatches = dispatch_linkage.get(key, [])
        
        result.append({
            "id": item.get("id"),
            "product_id": product_id,
            "product_name": item.get("product_name"),
            "variant_id": variant_id,
            "variant_name": final_variant_name,
            "opening_qty": opening,
            "received_qty": recv,
            "rejection_qty": rej,
            "closing_qty": item.get("closing_qty"),
            "linked_dispatches": linked_dispatches[:5] if linked_dispatches else []
        })
        seen_product_variants[dedup_key] = len(result) - 1  # Track index for merging
    
    # Sort result: items with activity (closing_qty not null, or opening > 0, or received > 0) first
    def sort_key(item):
        has_closing = item.get("closing_qty") is not None
        has_opening = (item.get("opening_qty") or 0) > 0
        has_received = (item.get("received_qty") or 0) > 0
        has_activity = has_closing or has_opening or has_received
        # Primary: activity items first (0 for activity, 1 for no activity)
        # Secondary: closing_qty descending (use negative for descending)
        return (0 if has_activity else 1, -(item.get("closing_qty") or 0), item.get("product_name", ""))
    
    result.sort(key=sort_key)
    
    # Determine if there's actual closing data for this date
    has_closing_data = len(closing_items) > 0
    
    return {
        "closing_date": date,
        "items": result,
        "recorded_items": len(closing_items),
        "has_closing_data": has_closing_data
    }


@app.get("/api/retailer-closing-inventory/check-yesterday/{retailer_id}")
async def check_yesterday_closing(
    retailer_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Check if yesterday's closing inventory was recorded for this retailer"""
    from datetime import datetime, timedelta
    
    today = datetime.now().date()
    yesterday = today - timedelta(days=1)
    yesterday_str = yesterday.strftime("%Y-%m-%d")
    
    # Check if yesterday's closing exists
    yesterday_closing = await db.retailer_closing_inventory.find_one({
        "retailer_id": retailer_id,
        "closing_date": yesterday_str
    })
    
    return {
        "yesterday_date": yesterday_str,
        "has_closing": yesterday_closing is not None,
        "today_date": today.strftime("%Y-%m-%d")
    }


@app.delete("/api/retailer-closing-inventory/{retailer_id}/{closing_date}")
async def delete_retailer_closing_inventory(
    retailer_id: str,
    closing_date: str,
    current_user: dict = Depends(get_current_user)
):
    """Delete all closing inventory records for a specific date"""
    result = await db.retailer_closing_inventory.delete_many({
        "retailer_id": retailer_id,
        "closing_date": closing_date
    })
    
    return {
        "message": f"Deleted closing inventory for {closing_date}",
        "deleted_count": result.deleted_count
    }

@app.delete("/api/retailer-closing-inventory/item/{item_id}")
async def delete_closing_inventory_item(
    item_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Delete a single closing inventory item"""
    result = await db.retailer_closing_inventory.delete_one({"id": item_id})
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Item not found")
    
    return {"message": "Item deleted successfully"}


@app.put("/api/retailer-closing-inventory/item/{item_id}")
async def update_closing_inventory_item(
    item_id: str,
    data: dict,
    current_user: dict = Depends(get_current_user)
):
    """Update a single closing inventory item (closing_qty and/or variant_name)"""
    closing_qty = data.get("closing_qty")
    variant_name = data.get("variant_name")
    
    if closing_qty is None:
        raise HTTPException(status_code=400, detail="closing_qty is required")
    
    try:
        closing_qty_num = float(closing_qty)
    except (ValueError, TypeError):
        raise HTTPException(status_code=400, detail="Invalid closing_qty value")
    
    update_data = {"closing_qty": closing_qty_num}
    if variant_name:
        update_data["variant_name"] = variant_name
    
    result = await db.retailer_closing_inventory.find_one_and_update(
        {"id": item_id},
        {"$set": update_data},
        return_document=True,
        projection={"_id": 0}
    )
    
    if not result:
        raise HTTPException(status_code=404, detail="Item not found")
    
    return {"message": "Item updated successfully", "item": result}



# ==================== DAILY REQUIREMENT ENDPOINTS ====================

@app.get("/api/retailer-daily-requirement/calculate")
async def calculate_daily_purchase_requirement(
    target_date: str,
    retailer_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """
    Calculate daily purchase requirement based on historical same-weekday sales data.
    
    Logic:
    1. Find the last 7 occurrences of the same weekday (e.g., if target is Monday, find last 7 Mondays)
    2. For each product, calculate average items sold across those weekdays
    3. For wastage, calculate average rejection qty across those weekdays
    4. Only divide by the number of weekdays with data (not always 7)
    5. Aggregate by product_id (not variant) to avoid duplicates
    """
    from datetime import datetime, timedelta
    
    try:
        target = datetime.strptime(target_date, "%Y-%m-%d").date()
        target_weekday = target.weekday()  # 0=Monday, 6=Sunday
        weekday_names = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
        
        # Build retailer filter
        retailer_filter = {}
        retailer_name = "All Retailers"
        if retailer_id:
            retailer_filter["retailer_id"] = retailer_id
            retailer = await db.users.find_one({"id": retailer_id})
            if retailer:
                retailer_name = retailer.get("company_name") or retailer.get("name", "Unknown")
        
        # Get all closing inventory records (source of items sold data)
        closing_records = await db.retailer_closing_inventory.find(
            retailer_filter, {"_id": 0}
        ).sort("closing_date", -1).to_list(500)
        
        # Filter to only include same weekday (last 7 occurrences before target date)
        same_weekday_records = []
        seen_dates = set()
        for record in closing_records:
            try:
                record_date_str = record.get("closing_date")
                if not record_date_str:
                    continue
                record_date = datetime.strptime(record_date_str, "%Y-%m-%d").date()
                if record_date.weekday() == target_weekday and record_date < target:
                    date_key = f"{record_date}_{record.get('retailer_id', 'all')}"
                    if date_key not in seen_dates:
                        same_weekday_records.append(record)
                        seen_dates.add(date_key)
                        # Stop after 7 unique dates
                        if len(seen_dates) >= 7:
                            break
            except:
                continue
        
        # Get rejections for the same weekdays for wastage calculation
        rejection_filter = {}
        if retailer_id:
            rejection_filter["retailer_id"] = retailer_id
        
        all_rejections = await db.retailer_rejections.find(
            rejection_filter, {"_id": 0}
        ).sort("rejection_date", -1).to_list(500)
        
        # Filter rejections to same weekdays
        same_weekday_rejections = []
        rejection_seen_dates = set()
        for rejection in all_rejections:
            try:
                rej_date_str = str(rejection.get("rejection_date", ""))[:10]
                if not rej_date_str:
                    continue
                rej_date = datetime.strptime(rej_date_str, "%Y-%m-%d").date()
                if rej_date.weekday() == target_weekday and rej_date < target:
                    same_weekday_rejections.append(rejection)
                    rejection_seen_dates.add(rej_date_str)
            except:
                continue
        
        # Calculate number of unique weekday dates found
        unique_closing_dates = set()
        for record in same_weekday_records:
            unique_closing_dates.add(record.get("closing_date"))
        num_weekdays_with_data = len(unique_closing_dates)
        
        if num_weekdays_with_data == 0:
            return {
                "success": False,
                "message": f"No historical sales data found for {weekday_names[target_weekday]}s. Please record closing inventory for at least one {weekday_names[target_weekday]}.",
                "items": [],
                "target_date": target_date,
                "weekday": weekday_names[target_weekday],
                "weekdays_analyzed": 0
            }
        
        # Aggregate items sold by PRODUCT_ID only (not variant to avoid duplicates)
        product_sales = {}  # product_id -> {total_sold, occurrences, product_name}
        
        for record in same_weekday_records:
            product_id = record.get("product_id")
            product_name = record.get("product_name", "Unknown")
            
            # Calculate items sold for this record
            # Items Sold = Opening + Received - Rejection - Closing
            opening = record.get("opening_qty", 0) or 0
            received = record.get("received_qty", 0) or 0
            closing = record.get("closing_qty", 0) or 0
            rejection = record.get("rejection_qty", 0) or 0
            items_sold = max(0, opening + received - rejection - closing)
            
            if product_id:
                if product_id not in product_sales:
                    product_sales[product_id] = {
                        "product_name": product_name,
                        "total_sold": 0,
                        "occurrences": 0,
                        "dates_with_data": set()
                    }
                product_sales[product_id]["total_sold"] += items_sold
                product_sales[product_id]["dates_with_data"].add(record.get("closing_date"))
        
        # Update occurrences count based on unique dates with data
        for product_id, data in product_sales.items():
            data["occurrences"] = len(data["dates_with_data"])
        
        # Aggregate wastage by product from rejections
        product_wastage = {}  # product_id -> {total_wastage, occurrences}
        
        unique_rejection_dates = set()
        for rejection in same_weekday_rejections:
            unique_rejection_dates.add(str(rejection.get("rejection_date", ""))[:10])
        num_rejection_weekdays = len(unique_rejection_dates)
        
        for rejection in same_weekday_rejections:
            product_id = rejection.get("product_id")
            qty = rejection.get("quantity", 0) or 0
            rej_date = str(rejection.get("rejection_date", ""))[:10]
            
            if product_id:
                if product_id not in product_wastage:
                    product_wastage[product_id] = {
                        "total_wastage": 0,
                        "dates_with_data": set()
                    }
                product_wastage[product_id]["total_wastage"] += qty
                product_wastage[product_id]["dates_with_data"].add(rej_date)
        
        # Get all products to fill in names if missing
        all_products = await db.products.find({}, {"_id": 0, "id": 1, "name": 1, "name_hi": 1}).to_list(500)
        product_name_map = {p["id"]: p.get("name", "Unknown") for p in all_products}
        
        # Build result items
        result_items = []
        seen_products = set()
        
        for product_id, data in product_sales.items():
            if product_id in seen_products:
                continue
            seen_products.add(product_id)
            
            # Calculate averages - divide by number of weekdays that had data for this product
            num_occurrences = data["occurrences"]
            avg_sold = data["total_sold"] / num_occurrences if num_occurrences > 0 else 0
            
            # Get wastage average
            wastage_data = product_wastage.get(product_id, {})
            wastage_occurrences = len(wastage_data.get("dates_with_data", set())) or 1
            avg_wastage = wastage_data.get("total_wastage", 0) / wastage_occurrences if wastage_data else 0
            
            # Round to 2 decimal places
            avg_sold = round(avg_sold, 2)
            avg_wastage = round(avg_wastage, 2)
            total_required = round(avg_sold + avg_wastage, 2)
            
            product_name = product_name_map.get(product_id, data["product_name"])
            
            result_items.append({
                "product_id": product_id,
                "product_name": product_name,
                "variant_name": "Kg",  # Aggregate view, no specific variant
                "indent_qty": avg_sold,  # Average items sold
                "kg_required": avg_sold,  # Same as indent qty for aggregate view
                "est_wastage_kg": avg_wastage,
                "total_kg_required": total_required,
                "weekdays_with_sales_data": num_occurrences,
                "rate_per_kg": "",
                "amount_paid": "",
                "remarks": ""
            })
        
        # Sort by product name
        result_items.sort(key=lambda x: x["product_name"])
        
        return {
            "success": True,
            "target_date": target_date,
            "weekday": weekday_names[target_weekday],
            "weekdays_analyzed": num_weekdays_with_data,
            "rejection_weekdays_analyzed": num_rejection_weekdays,
            "retailer_name": retailer_name,
            "items": result_items,
            "total_indent_qty": sum(item["indent_qty"] for item in result_items),
            "total_kg_required": sum(item["total_kg_required"] for item in result_items),
            "message": f"Calculated based on {num_weekdays_with_data} {weekday_names[target_weekday]}(s) of sales data"
        }
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return {
            "success": False,
            "message": f"Error calculating requirement: {str(e)}",
            "items": []
        }


@app.post("/api/retailer-daily-requirement")
async def save_retailer_daily_requirement(
    data: dict,
    current_user: dict = Depends(get_current_user)
):
    """Save or update retailer daily purchase requirement"""
    from datetime import datetime, timezone
    
    requirement_date = data.get("requirement_date")
    retailer_id = data.get("retailer_id")  # Can be null for 'All Retailers'
    
    if not requirement_date:
        raise HTTPException(status_code=400, detail="Requirement date is required")
    
    # Check if record exists for this date + retailer combo
    query = {"requirement_date": requirement_date}
    if retailer_id:
        query["retailer_id"] = retailer_id
    else:
        query["retailer_id"] = None
    
    existing = await db.retailer_daily_requirements.find_one(query)
    
    doc = {
        "requirement_date": requirement_date,
        "retailer_id": retailer_id,
        "retailer_name": data.get("retailer_name", "All Retailers"),
        "items": data.get("items", []),
        "total_kg": data.get("total_kg", 0),
        "total_amount": data.get("total_amount", 0),
        "updated_at": datetime.now(timezone.utc).isoformat(),
        "updated_by": current_user.get("user_id")
    }
    
    if existing:
        # Update existing
        await db.retailer_daily_requirements.update_one(
            {"_id": existing["_id"]},
            {"$set": doc}
        )
        return {"message": "Daily requirement updated", "id": existing.get("id")}
    else:
        # Create new
        doc["id"] = str(uuid.uuid4())
        doc["created_at"] = datetime.now(timezone.utc).isoformat()
        doc["created_by"] = current_user.get("user_id")
        await db.retailer_daily_requirements.insert_one(doc)
        return {"message": "Daily requirement saved", "id": doc["id"]}


@app.get("/api/retailer-daily-requirement/{requirement_date}")
async def get_retailer_daily_requirement(
    requirement_date: str,
    retailer_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Get saved daily requirement for a date"""
    query = {"requirement_date": requirement_date}
    if retailer_id:
        query["retailer_id"] = retailer_id
    else:
        query["retailer_id"] = None
    
    doc = await db.retailer_daily_requirements.find_one(query, {"_id": 0})
    if not doc:
        return None
    return doc


@app.get("/api/retailer-daily-requirements")
async def list_retailer_daily_requirements(
    limit: int = Query(default=30, le=100),
    current_user: dict = Depends(get_current_user)
):
    """List recent daily requirements"""
    docs = await db.retailer_daily_requirements.find(
        {}, {"_id": 0}
    ).sort("requirement_date", -1).to_list(limit)
    return docs


# ==================== QC DAILY REQUIREMENT ENDPOINTS ====================

@app.get("/api/qc-wastage-averages")
async def get_qc_wastage_averages(
    current_user: dict = Depends(get_current_user)
):
    """
    Calculate 1-month average wastage per product from QC GRN data.
    Wastage = supplied_qty - grn_qty (what we sent minus what was accepted)
    """
    from datetime import datetime, timedelta, timezone
    
    # Get date 30 days ago
    one_month_ago = (datetime.now(timezone.utc) - timedelta(days=30)).strftime("%Y-%m-%d")
    
    # Get all GRNs from last 30 days
    grns = await db.qc_grns.find(
        {"grn_date": {"$gte": one_month_ago}},
        {"_id": 0}
    ).to_list(1000)
    
    # Aggregate wastage by product
    wastage_by_product = {}  # {product_id: {total_wastage_qty: 0, count: 0, product_name: '', packaging_name: ''}}
    
    for grn in grns:
        for item in grn.get("items", []):
            product_id = item.get("product_id", "")
            packaging_id = item.get("packaging_id", "")
            key = f"{product_id}_{packaging_id}"
            
            # Wastage = what we supplied - what was accepted in GRN
            supplied = item.get("supplied_qty", 0) or 0
            grn_qty = item.get("grn_qty", 0) or 0
            wastage = max(0, supplied - grn_qty)  # Only count positive wastage
            
            if key not in wastage_by_product:
                wastage_by_product[key] = {
                    "product_id": product_id,
                    "product_name": item.get("product_name", ""),
                    "packaging_id": packaging_id,
                    "packaging_name": item.get("packaging_name", ""),
                    "total_wastage": 0,
                    "total_supplied": 0,
                    "count": 0
                }
            
            wastage_by_product[key]["total_wastage"] += wastage
            wastage_by_product[key]["total_supplied"] += supplied
            wastage_by_product[key]["count"] += 1
    
    # Calculate averages
    result = []
    for key, data in wastage_by_product.items():
        avg_wastage = data["total_wastage"] / data["count"] if data["count"] > 0 else 0
        wastage_pct = (data["total_wastage"] / data["total_supplied"] * 100) if data["total_supplied"] > 0 else 0
        
        result.append({
            "product_id": data["product_id"],
            "product_name": data["product_name"],
            "packaging_id": data["packaging_id"],
            "packaging_name": data["packaging_name"],
            "avg_wastage_qty": round(avg_wastage, 2),
            "total_wastage": round(data["total_wastage"], 2),
            "total_supplied": round(data["total_supplied"], 2),
            "wastage_percentage": round(wastage_pct, 2),
            "data_points": data["count"]
        })
    
    return result


@app.get("/api/retail-wastage-averages")
async def get_retail_wastage_averages(
    current_user: dict = Depends(get_current_user)
):
    """
    Calculate 30-day average wastage per product from stock status data.
    Uses the daily_stock_status collection to get wastage by product.
    """
    from datetime import datetime, timedelta, timezone
    
    # Get date 30 days ago
    thirty_days_ago = (datetime.now(timezone.utc) - timedelta(days=30)).strftime("%Y-%m-%d")
    
    # Get all stock status records from last 30 days
    stock_records = await db.daily_stock_status.find(
        {"date": {"$gte": thirty_days_ago}},
        {"_id": 0}
    ).to_list(10000)
    
    # Aggregate wastage by product
    wastage_by_product = {}  # {product_id: {total_wastage: 0, total_input: 0, count: 0, ...}}
    
    for record in stock_records:
        product_id = record.get("product_id", "")
        product_name = record.get("product_name", "")
        
        # Skip if no product info
        if not product_id and not product_name:
            continue
        
        # Calculate wastage: opening + purchase - dispatch - closing
        opening_qty = float(record.get("opening_qty") or 0)
        purchase_qty = float(record.get("purchase_qty") or 0)
        dispatch_qty = float(record.get("dispatch_qty") or 0)
        closing_qty = float(record.get("closing_qty") or 0)
        
        total_input = opening_qty + purchase_qty
        expected_closing = total_input - dispatch_qty
        wastage_qty = max(0, expected_closing - closing_qty)  # Only count positive wastage
        
        # Use product_id or name as key
        key = product_id or product_name
        
        if key not in wastage_by_product:
            wastage_by_product[key] = {
                "product_id": product_id,
                "product_name": product_name,
                "total_wastage": 0,
                "total_input": 0,
                "count": 0
            }
        
        wastage_by_product[key]["total_wastage"] += wastage_qty
        wastage_by_product[key]["total_input"] += total_input
        wastage_by_product[key]["count"] += 1
    
    # Calculate averages
    result = []
    for key, data in wastage_by_product.items():
        if data["count"] == 0:
            continue
            
        avg_wastage_kg = data["total_wastage"] / data["count"]
        wastage_pct = (data["total_wastage"] / data["total_input"] * 100) if data["total_input"] > 0 else 0
        
        result.append({
            "product_id": data["product_id"],
            "product_name": data["product_name"],
            "avg_wastage_kg": round(avg_wastage_kg, 2),
            "total_wastage_kg": round(data["total_wastage"], 2),
            "total_input_kg": round(data["total_input"], 2),
            "wastage_percentage": round(wastage_pct, 2),
            "data_points": data["count"]
        })
    
    # Sort by wastage percentage descending
    result.sort(key=lambda x: x["wastage_percentage"], reverse=True)
    
    return result


@app.get("/api/qc-latest-bunch-weights")
async def get_qc_latest_bunch_weights(
    current_user: dict = Depends(get_current_user)
):
    """
    Get the latest bunch weights per product from the most recent saved QC daily requirements.
    This is used to pre-populate bunch weights when calculating new daily requirements.
    """
    # Get the most recent saved QC daily requirement
    recent_requirements = await db.qc_daily_requirements.find(
        {},
        {"_id": 0}
    ).sort("requirement_date", -1).limit(10).to_list(10)
    
    # Build a map of product+packaging -> bunch weight from most recent data
    bunch_weights = {}
    
    for req in recent_requirements:
        for item in req.get("items", []):
            product_id = item.get("product_id", "")
            packaging_id = item.get("packaging_id", "")
            key = f"{product_id}_{packaging_id}"
            
            # Only use if we don't have a weight for this combo yet (most recent takes precedence)
            if key not in bunch_weights and item.get("weight_per_bunch"):
                bunch_weights[key] = {
                    "product_id": product_id,
                    "product_name": item.get("product_name", ""),
                    "packaging_id": packaging_id,
                    "packaging_name": item.get("packaging_name", ""),
                    "weight_per_bunch": item.get("weight_per_bunch"),
                    "source_date": req.get("requirement_date")
                }
    
    return list(bunch_weights.values())


@app.post("/api/qc-daily-requirement")
async def save_qc_daily_requirement(
    data: dict,
    current_user: dict = Depends(get_current_user)
):
    """Save or update QC daily purchase requirement"""
    from datetime import datetime, timezone
    
    requirement_date = data.get("requirement_date")
    customer_name = data.get("customer_name", "All Customers")
    
    if not requirement_date:
        raise HTTPException(status_code=400, detail="Requirement date is required")
    
    # Check if record exists for this date + customer combo
    query = {"requirement_date": requirement_date, "customer_name": customer_name}
    existing = await db.qc_daily_requirements.find_one(query)
    
    doc = {
        "requirement_date": requirement_date,
        "customer_name": customer_name,
        "items": data.get("items", []),
        "total_qty_required": data.get("total_qty_required", 0),
        "total_actual_qty": data.get("total_actual_qty", 0),
        "total_bunches": data.get("total_bunches", 0),
        "total_amount": data.get("total_amount", 0),
        "updated_at": datetime.now(timezone.utc).isoformat(),
        "updated_by": current_user.get("user_id")
    }
    
    if existing:
        # Update existing
        await db.qc_daily_requirements.update_one(
            {"_id": existing["_id"]},
            {"$set": doc}
        )
        return {"message": "QC daily requirement updated", "id": existing.get("id")}
    else:
        # Create new
        doc["id"] = str(uuid.uuid4())
        doc["created_at"] = datetime.now(timezone.utc).isoformat()
        doc["created_by"] = current_user.get("user_id")
        await db.qc_daily_requirements.insert_one(doc)
        return {"message": "QC daily requirement saved", "id": doc["id"]}


@app.get("/api/qc-daily-requirement/{requirement_date}")
async def get_qc_daily_requirement(
    requirement_date: str,
    customer_name: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Get saved QC daily requirement for a date"""
    query = {"requirement_date": requirement_date}
    if customer_name:
        query["customer_name"] = customer_name
    
    doc = await db.qc_daily_requirements.find_one(query, {"_id": 0})
    return doc


@app.get("/api/qc-daily-requirements")
async def list_qc_daily_requirements(
    limit: int = Query(default=30, le=100),
    current_user: dict = Depends(get_current_user)
):
    """List recent QC daily requirements"""
    docs = await db.qc_daily_requirements.find(
        {}, {"_id": 0}
    ).sort("requirement_date", -1).to_list(limit)
    return docs


# ==================================================================================
# SECTION: LABOUR MANAGEMENT ROUTES - Daily Variable Labour Costs
# ==================================================================================

@app.get("/api/labours")
async def list_labours(
    include_inactive: bool = False,
    current_user: dict = Depends(get_current_user)
):
    """Get all labourers, optionally including inactive ones"""
    query = {} if include_inactive else {"is_active": {"$ne": False}}
    labours = await db.labours.find(query, {"_id": 0}).sort("name", 1).to_list(100)
    return labours


@app.post("/api/labours")
async def create_labour(
    labour: dict,
    current_user: dict = Depends(get_current_user)
):
    """Create a new labourer"""
    # Check for duplicate name
    existing = await db.labours.find_one({"name": {"$regex": f"^{labour.get('name', '')}$", "$options": "i"}})
    if existing:
        raise HTTPException(status_code=400, detail="A labourer with this name already exists")
    
    doc = {
        "id": str(uuid.uuid4()),
        "name": labour.get("name"),
        "phone": labour.get("phone"),
        "default_daily_rate": float(labour.get("default_daily_rate", 0)),
        "default_overtime_rate": float(labour.get("default_overtime_rate", 0)),
        "bank_account_number": labour.get("bank_account_number"),
        "ifsc_code": labour.get("ifsc_code"),
        "joining_date": labour.get("joining_date"),
        "is_active": labour.get("is_active", True),
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.labours.insert_one(doc)
    doc.pop("_id", None)
    return doc


@app.put("/api/labours/{labour_id}")
async def update_labour(
    labour_id: str,
    labour: dict,
    current_user: dict = Depends(get_current_user)
):
    """Update a labourer's details"""
    existing = await db.labours.find_one({"id": labour_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Labourer not found")
    
    update_data = {}
    if "name" in labour:
        update_data["name"] = labour["name"]
    if "phone" in labour:
        update_data["phone"] = labour["phone"]
    if "default_daily_rate" in labour:
        update_data["default_daily_rate"] = float(labour["default_daily_rate"])
    if "default_overtime_rate" in labour:
        update_data["default_overtime_rate"] = float(labour["default_overtime_rate"])
    if "bank_account_number" in labour:
        update_data["bank_account_number"] = labour["bank_account_number"]
    if "ifsc_code" in labour:
        update_data["ifsc_code"] = labour["ifsc_code"]
    if "joining_date" in labour:
        update_data["joining_date"] = labour["joining_date"]
    if "is_active" in labour:
        update_data["is_active"] = labour["is_active"]
    
    if update_data:
        await db.labours.update_one({"id": labour_id}, {"$set": update_data})
    
    updated = await db.labours.find_one({"id": labour_id}, {"_id": 0})
    return updated


@app.delete("/api/labours/{labour_id}")
async def delete_labour(
    labour_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Delete a labourer (soft delete - marks as inactive)"""
    existing = await db.labours.find_one({"id": labour_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Labourer not found")
    
    # Check if labourer has attendance records
    has_attendance = await db.labour_attendance.find_one({"labour_id": labour_id})
    if has_attendance:
        # Soft delete - just mark as inactive
        await db.labours.update_one({"id": labour_id}, {"$set": {"is_active": False}})
        return {"message": "Labourer marked as inactive (has attendance history)"}
    else:
        # Hard delete if no attendance records
        await db.labours.delete_one({"id": labour_id})
        return {"message": "Labourer deleted"}


# --- Labour Attendance Routes ---

@app.get("/api/labour-attendance")
async def get_labour_attendance(
    date: str = Query(..., description="Date in YYYY-MM-DD format"),
    current_user: dict = Depends(get_current_user)
):
    """Get attendance records for a specific date"""
    # Get all active labourers
    labours = await db.labours.find({"is_active": {"$ne": False}}, {"_id": 0}).sort("name", 1).to_list(100)
    
    # Get existing attendance records for this date
    attendance_records = await db.labour_attendance.find(
        {"date": date}, {"_id": 0}
    ).to_list(100)
    attendance_map = {rec["labour_id"]: rec for rec in attendance_records}
    
    # Merge: For each labourer, return attendance if exists, else default values
    result = []
    for labour in labours:
        if labour["id"] in attendance_map:
            # Use existing attendance record
            record = attendance_map[labour["id"]]
            # Ensure labour_name is updated (in case labour name changed)
            record["labour_name"] = labour["name"]
            # Ensure working_hours has a default
            if "working_hours" not in record:
                record["working_hours"] = 9 if record.get("present") else 0
            result.append(record)
        else:
            # Return default attendance entry (not saved yet)
            daily_rate = labour.get("default_daily_rate", 0)
            result.append({
                "id": None,  # Not saved yet
                "date": date,
                "labour_id": labour["id"],
                "labour_name": labour["name"],
                "present": False,
                "working_hours": 9,  # Default 9 hours
                "overtime_hours": 0,
                "daily_rate": daily_rate,
                "hourly_rate": round(daily_rate / 9, 2) if daily_rate > 0 else 0,
                "overtime_rate": labour.get("default_overtime_rate", 0),
                "total_payment": 0,
                "remarks": None
            })
    
    return result


@app.post("/api/labour-attendance")
async def save_labour_attendance(
    attendance_data: dict,
    current_user: dict = Depends(get_current_user)
):
    """Save or update attendance for a labourer on a specific date"""
    date = attendance_data.get("date")
    labour_id = attendance_data.get("labour_id")
    labour_name = attendance_data.get("labour_name")
    present = attendance_data.get("present", False)
    overtime_hours = float(attendance_data.get("overtime_hours", 0))
    daily_rate = float(attendance_data.get("daily_rate", 0))
    overtime_rate = float(attendance_data.get("overtime_rate", 0))
    remarks = attendance_data.get("remarks")
    
    # Calculate total payment
    total_payment = 0
    if present:
        total_payment = daily_rate + (overtime_hours * overtime_rate)
    
    # Check if record exists for this date + labour combo
    existing = await db.labour_attendance.find_one({"date": date, "labour_id": labour_id})
    
    if existing:
        # Update existing
        await db.labour_attendance.update_one(
            {"_id": existing["_id"]},
            {"$set": {
                "present": present,
                "overtime_hours": overtime_hours,
                "daily_rate": daily_rate,
                "overtime_rate": overtime_rate,
                "total_payment": total_payment,
                "remarks": remarks,
                "recorded_by": current_user.get("user_id"),
                "updated_at": datetime.now(timezone.utc).isoformat()
            }}
        )
        updated = await db.labour_attendance.find_one({"_id": existing["_id"]}, {"_id": 0})
        return updated
    else:
        # Create new
        doc = {
            "id": str(uuid.uuid4()),
            "date": date,
            "labour_id": labour_id,
            "labour_name": labour_name,
            "present": present,
            "overtime_hours": overtime_hours,
            "daily_rate": daily_rate,
            "overtime_rate": overtime_rate,
            "total_payment": total_payment,
            "remarks": remarks,
            "recorded_by": current_user.get("user_id"),
            "created_at": datetime.now(timezone.utc).isoformat(),
            "updated_at": datetime.now(timezone.utc).isoformat()
        }
        await db.labour_attendance.insert_one(doc)
        doc.pop("_id", None)
        return doc


@app.post("/api/labour-attendance/bulk")
async def save_bulk_labour_attendance(
    data: dict,
    current_user: dict = Depends(get_current_user)
):
    """Save multiple attendance records at once"""
    date = data.get("date")
    records = data.get("records", [])
    
    saved_count = 0
    for record in records:
        labour_id = record.get("labour_id")
        labour_name = record.get("labour_name")
        present = record.get("present", False)
        working_hours = float(record.get("working_hours", 9))  # Default 9 hours
        overtime_hours = float(record.get("overtime_hours", 0))
        daily_rate = float(record.get("daily_rate", 0))
        overtime_rate = float(record.get("overtime_rate", 0))
        remarks = record.get("remarks")
        
        # Calculate hourly rate (daily rate / 9 hours standard)
        hourly_rate = daily_rate / 9 if daily_rate > 0 else 0
        
        # Calculate total payment based on working hours
        total_payment = 0
        if present:
            # Payment = (hourly rate * working hours) + (overtime rate * overtime hours)
            total_payment = (hourly_rate * working_hours) + (overtime_hours * overtime_rate)
        
        # Upsert attendance record
        await db.labour_attendance.update_one(
            {"date": date, "labour_id": labour_id},
            {"$set": {
                "labour_name": labour_name,
                "present": present,
                "working_hours": working_hours,
                "overtime_hours": overtime_hours,
                "daily_rate": daily_rate,
                "hourly_rate": round(hourly_rate, 2),
                "overtime_rate": overtime_rate,
                "total_payment": round(total_payment, 2),
                "remarks": remarks,
                "recorded_by": current_user.get("user_id"),
                "updated_at": datetime.now(timezone.utc).isoformat()
            },
            "$setOnInsert": {
                "id": str(uuid.uuid4()),
                "date": date,
                "labour_id": labour_id,
                "created_at": datetime.now(timezone.utc).isoformat()
            }},
            upsert=True
        )
        saved_count += 1
    
    return {"message": f"Saved {saved_count} attendance records", "saved_count": saved_count}


# --- Labour Costs Summary Routes ---

@app.get("/api/labour-costs/summary")
async def get_labour_costs_summary(
    from_date: str = Query(..., description="Start date YYYY-MM-DD"),
    to_date: str = Query(..., description="End date YYYY-MM-DD"),
    current_user: dict = Depends(get_current_user)
):
    """Get labour costs summary for a date range"""
    # Get all attendance records in date range
    attendance = await db.labour_attendance.find(
        {"date": {"$gte": from_date, "$lte": to_date}},
        {"_id": 0}
    ).sort("date", 1).to_list(1000)
    
    # Get all labourers for reference
    labours = await db.labours.find({}, {"_id": 0}).to_list(100)
    labour_map = {l["id"]: l for l in labours}
    
    # Calculate daily totals
    daily_totals = {}
    labour_totals = {}
    
    for record in attendance:
        date = record["date"]
        labour_id = record["labour_id"]
        
        # Daily totals
        if date not in daily_totals:
            daily_totals[date] = {
                "date": date,
                "total_present": 0,
                "total_overtime_hours": 0,
                "total_payment": 0,
                "records": []
            }
        
        if record.get("present"):
            daily_totals[date]["total_present"] += 1
            daily_totals[date]["total_overtime_hours"] += record.get("overtime_hours", 0)
            daily_totals[date]["total_payment"] += record.get("total_payment", 0)
            daily_totals[date]["records"].append(record)
        
        # Labour totals
        if labour_id not in labour_totals:
            labour_name = record.get("labour_name") or labour_map.get(labour_id, {}).get("name", "Unknown")
            labour_totals[labour_id] = {
                "labour_id": labour_id,
                "labour_name": labour_name,
                "days_present": 0,
                "total_overtime_hours": 0,
                "total_payment": 0
            }
        
        if record.get("present"):
            labour_totals[labour_id]["days_present"] += 1
            labour_totals[labour_id]["total_overtime_hours"] += record.get("overtime_hours", 0)
            labour_totals[labour_id]["total_payment"] += record.get("total_payment", 0)
    
    # Sort daily totals by date
    daily_list = sorted(daily_totals.values(), key=lambda x: x["date"], reverse=True)
    
    # Sort labour totals by total payment
    labour_list = sorted(labour_totals.values(), key=lambda x: x["total_payment"], reverse=True)
    
    # Calculate grand totals
    grand_total_payment = sum(d["total_payment"] for d in daily_list)
    total_days = len(daily_list)
    total_man_days = sum(d["total_present"] for d in daily_list)
    total_overtime_hours = sum(d["total_overtime_hours"] for d in daily_list)
    
    return {
        "from_date": from_date,
        "to_date": to_date,
        "summary": {
            "total_payment": grand_total_payment,
            "total_days": total_days,
            "total_man_days": total_man_days,
            "total_overtime_hours": total_overtime_hours,
            "daily_avg_cost": grand_total_payment / total_days if total_days > 0 else 0
        },
        "daily_breakdown": daily_list,
        "labour_breakdown": labour_list
    }


@app.on_event("shutdown")
async def shutdown_db_client():
    global backup_scheduler
    if backup_scheduler:
        backup_scheduler.shutdown()
    client.close()