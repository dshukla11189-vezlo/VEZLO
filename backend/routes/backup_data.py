"""
Backup & Data Management Routes
===============================
Extracted from server.py for modular organization.
Handles backup management, data sync, data reset, and admin utilities.
"""
from fastapi import APIRouter, HTTPException, Depends
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field
from datetime import datetime, timezone
from typing import Optional, List
import io
import os
import httpx

from dependencies import (
    db,
    get_current_user,
    hash_password,
    logger,
)
from backup_system import trigger_manual_backup, generate_backup_excel

# Note: backup_scheduler is a global variable managed in server.py

router = APIRouter(tags=["backup_data"])

# Global reference to backup scheduler (set from server.py)
backup_scheduler = None

# SECTION: BACKUP & DATA MANAGEMENT ROUTES (Lines ~4127-4200)
# ============================================================================

@router.post("/backup/trigger")
async def trigger_backup(current_user: dict = Depends(get_current_user)):
    """Manually trigger a backup and send to configured email addresses"""
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Only admin can trigger backups")
    
    result = await trigger_manual_backup(db)
    return result

@router.get("/backup/download")
async def download_backup(current_user: dict = Depends(get_current_user)):
    """Download backup Excel file directly"""
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Only admin can download backups")
    
    try:
        excel_data = await generate_backup_excel(db)
        today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
        filename = f"MrOrganix_Backup_{today}.xlsx"
        
        return StreamingResponse(
            io.BytesIO(excel_data),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to generate backup: {str(e)}")

@router.get("/backup/status")
async def get_backup_status(current_user: dict = Depends(get_current_user)):
    """Get backup scheduler status"""
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Only admin can view backup status")
    
    return {
        "scheduler_running": backup_scheduler is not None and backup_scheduler.running if backup_scheduler else False,
        "next_backup_time": "11:59 PM IST (18:29 UTC)",
        "recipients": ["dshukla11189@gmail.com"],
        "collections_backed_up": [
            # Core
            "users", "products", "units",
            # Procurement
            "farmers", "procurements", "procurement_templates",
            # Quick Commerce
            "qc_packaging", "qc_customers", "qc_indents", "qc_dispatches", 
            "qc_invoices", "qc_grns", "qc_daily_requirements",
            # Retailer
            "retailers", "retailer_indents", "retailer_dispatches", "retailer_invoices", 
            "retailer_grn", "retailer_rejections", "retailer_payments",
            "retailer_closing_inventory", "retailer_inventory", "retailer_daily_requirements",
            "retail_plans",
            # Stock & Expenses
            "daily_stock_status", "variable_expenses", "fixed_expenses",
            # Labour
            "labours", "labour_attendance",
            # Payments
            "payments"
        ]
    }

@router.post("/admin/reset-data")
async def reset_all_data(current_user: dict = Depends(get_current_user)):
    """
    Reset all data for production - keeps only admin user.
    USE WITH CAUTION - This deletes ALL business data!
    """
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Only admin can reset data")
    
    collections_to_clear = [
        # Core (keep users separately handled)
        "products", "units",
        # Procurement
        "farmers", "procurements", "procurement_templates",
        # Quick Commerce
        "qc_packaging", "qc_customers", "qc_indents", "qc_dispatches", 
        "qc_invoices", "qc_grns", "qc_daily_requirements",
        # Retailer
        "retailers", "retailer_indents", "retailer_dispatches", "retailer_invoices", 
        "retailer_grn", "retailer_rejections", "retailer_payments",
        "retailer_closing_inventory", "retailer_inventory", "retailer_daily_requirements",
        # Stock & Expenses
        "daily_stock_status", "variable_expenses", "fixed_expenses",
        # Labour
        "labours", "labour_attendance",
        # Payments
        "payments"
    ]
    
    deleted_counts = {}
    for coll_name in collections_to_clear:
        result = await db[coll_name].delete_many({})
        deleted_counts[coll_name] = result.deleted_count
    
    # Delete non-admin users (keep admin accounts)
    users_deleted = await db.users.delete_many({"role": {"$ne": "admin"}})
    deleted_counts["users (non-admin)"] = users_deleted.deleted_count
    
    return {
        "success": True,
        "message": "All test data cleared. Ready for production!",
        "deleted": deleted_counts
    }


# Pydantic model for sync request
class SyncFromProductionRequest(BaseModel):
    production_url: str = Field(default="https://harvest-hub-384.emergent.host", description="Production app URL")
    admin_email: str = Field(default="admin@freshflow.com", description="Admin email for production")
    admin_password: str = Field(default="admin123", description="Admin password for production")
    reset_passwords: bool = Field(default=True, description="Reset all user passwords to default")
    default_password: str = Field(default="admin123", description="Default password for users after sync")


@router.post("/sync-from-production")
async def sync_from_production(
    request: SyncFromProductionRequest,
    current_user: dict = Depends(get_current_user)
):
    """
    Sync data from production app to this preview/test environment.
    Downloads backup Excel from production and imports it here.
    
    IMPORTANT: This will REPLACE all existing data in the synced collections.
    """
    import ast
    import json as json_lib
    import httpx
    import openpyxl
    from io import BytesIO
    
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Only admin can sync from production")
    
    try:
        production_url = request.production_url.rstrip('/')
        
        async with httpx.AsyncClient(timeout=60.0) as client:
            # Step 1: Login to production
            login_response = await client.post(
                f"{production_url}/api/auth/login",
                json={
                    "identifier": request.admin_email,
                    "password": request.admin_password
                }
            )
            
            if login_response.status_code != 200:
                raise HTTPException(
                    status_code=401,
                    detail=f"Failed to login to production: {login_response.text}"
                )
            
            token = login_response.json().get("token")
            if not token:
                raise HTTPException(status_code=401, detail="No token received from production login")
            
            # Step 2: Download backup Excel
            backup_response = await client.get(
                f"{production_url}/api/backup/download",
                headers={"Authorization": f"Bearer {token}"},
                timeout=120.0
            )
            
            if backup_response.status_code != 200:
                raise HTTPException(
                    status_code=500,
                    detail=f"Failed to download backup from production: {backup_response.status_code}"
                )
            
            # Step 3: Parse Excel and import data
            excel_data = BytesIO(backup_response.content)
            wb = openpyxl.load_workbook(excel_data)
            
            # Helper function to parse JSON fields
            def safe_parse_json(value):
                if not value or not isinstance(value, str):
                    return value
                try:
                    return json_lib.loads(value)
                except:
                    pass
                try:
                    return ast.literal_eval(value)
                except:
                    pass
                return value
            
            # Map sheet names to collection names - COMPREHENSIVE list
            sheet_to_collection = {
                # Core
                'users': 'users',
                'products': 'products',
                'product_types': 'product_types',
                'product_categories': 'product_categories',
                'units': 'units',
                # Procurement
                'farmers': 'farmers',
                'procurements': 'procurements',
                'procurement_templates': 'procurement_templates',
                'farmer_payments': 'farmer_payments',
                # Quick Commerce
                'qc_packaging': 'qc_packaging',
                'qc_customers': 'qc_customers',
                'qc_indents': 'qc_indents',
                'qc_dispatches': 'qc_dispatches',
                'qc_invoices': 'qc_invoices',
                'qc_grns': 'qc_grns',
                'qc_daily_requirements': 'qc_daily_requirements',
                'customer_product_settings': 'customer_product_settings',
                # Retailer
                'retailers': 'retailers',
                'retailer_catalogue': 'retailer_catalogue',
                'retailer_indents': 'retailer_indents',
                'retailer_dispatches': 'retailer_dispatches',
                'retailer_invoices': 'retailer_invoices',
                'retailer_grn': 'retailer_grn',
                'retail_plans': 'retail_plans',
                'retailer_rejections': 'retailer_rejections',
                'retailer_payments': 'retailer_payments',
                'retailer_closing_inventory': 'retailer_closing_inventory',
                'retailer_inventory': 'retailer_inventory',
                'retailer_daily_requirements': 'retailer_daily_requirements',
                # Stock & Expenses
                'daily_stock_status': 'daily_stock_status',
                'variable_expenses': 'variable_expenses',
                'fixed_expenses': 'fixed_expenses',
                # Labour
                'labours': 'labours',
                'labour_attendance': 'labour_attendance',
                # Payments
                'payments': 'payments',
            }
            
            # JSON fields per collection (fields that contain lists/objects)
            json_fields_map = {
                'procurements': ['products'],
                'qc_indents': ['items'],
                'qc_dispatches': ['items', 'packaging_items'],
                'qc_invoices': ['items'],
                'qc_grns': ['items'],
                'qc_daily_requirements': ['items'],
                'retailer_indents': ['items'],
                'retailer_dispatches': ['items'],
                'retailer_invoices': ['items'],
                'retailer_grn': ['items'],
                'retailer_rejections': ['items'],
                'retailer_closing_inventory': ['items'],
                'retailer_inventory': ['items'],
                'retailer_daily_requirements': ['items'],
                'procurement_templates': ['products'],
                'farmers': ['materials_supplied'],
                'retail_plans': ['products'],
            }
            
            sync_results = {}
            total_synced = 0
            
            for sheet_name, collection_name in sheet_to_collection.items():
                if sheet_name not in wb.sheetnames:
                    sync_results[collection_name] = {"status": "not_found", "count": 0}
                    continue
                
                ws = wb[sheet_name]
                rows = list(ws.iter_rows(values_only=True))
                
                if len(rows) <= 1:
                    sync_results[collection_name] = {"status": "empty", "count": 0}
                    continue
                
                # Get headers from first row
                headers = [str(h).lower().strip() if h else f'col_{i}' for i, h in enumerate(rows[0])]
                
                # Convert rows to documents
                documents = []
                for row in rows[1:]:  # Skip header row
                    doc = {}
                    for i, value in enumerate(row):
                        if i < len(headers):
                            key = headers[i]
                            if key == '_id':
                                continue  # Skip _id field
                            if value is not None:
                                # Parse JSON fields
                                if collection_name in json_fields_map and key in json_fields_map[collection_name]:
                                    doc[key] = safe_parse_json(value)
                                else:
                                    doc[key] = value
                    if doc:
                        documents.append(doc)
                
                if documents:
                    # Clear existing data and insert new
                    await db[collection_name].delete_many({})
                    result = await db[collection_name].insert_many(documents)
                    sync_results[collection_name] = {
                        "status": "success",
                        "count": len(result.inserted_ids)
                    }
                    total_synced += len(result.inserted_ids)
                else:
                    sync_results[collection_name] = {"status": "no_data", "count": 0}
            
            # Reset passwords if requested
            if request.reset_passwords:
                hashed_password = bcrypt.hashpw(
                    request.default_password.encode('utf-8'), 
                    bcrypt.gensalt()
                ).decode('utf-8')
                
                await db.users.update_many(
                    {},
                    {"$set": {"password": hashed_password}}
                )
                sync_results["password_reset"] = {
                    "status": "success",
                    "message": f"All user passwords reset to: {request.default_password}"
                }
            
            return {
                "success": True,
                "message": f"Successfully synced {total_synced} records from production",
                "total_records": total_synced,
                "production_url": production_url,
                "collections": sync_results
            }
        
    except httpx.TimeoutException:
        raise HTTPException(status_code=504, detail="Connection to production timed out")
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=500, 
            detail=f"Failed to sync from production: {str(e)}"
        )


@router.post("/sync-from-production-direct")
async def sync_from_production_direct(
    request: SyncFromProductionRequest,
    current_user: dict = Depends(get_current_user)
):
    """
    Direct API sync from production - bypasses Excel backup.
    Pulls data directly from production API endpoints for 100% data sync.
    Use this when Excel backup doesn't include all collections.
    """
    import httpx
    
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Only admin can sync from production")
    
    try:
        production_url = request.production_url.rstrip('/')
        
        async with httpx.AsyncClient(timeout=120.0) as client:
            # Step 1: Login to production
            login_response = await client.post(
                f"{production_url}/api/auth/login",
                json={
                    "identifier": request.admin_email,
                    "password": request.admin_password
                }
            )
            
            if login_response.status_code != 200:
                raise HTTPException(
                    status_code=401,
                    detail=f"Failed to login to production: {login_response.text}"
                )
            
            token = login_response.json().get("token")
            if not token:
                raise HTTPException(status_code=401, detail="No token received from production login")
            
            headers = {"Authorization": f"Bearer {token}"}
            
            # Define API endpoints for each collection
            api_endpoints = {
                'qc_packaging': '/api/qc-packaging',
                'products': '/api/products?include_images=false',
                'product_types': '/api/product-types',
                'product_categories': '/api/product-categories',
                'units': '/api/units',
                'farmers': '/api/farmers',
                'labours': '/api/labours',
                'qc_customers': '/api/qc-customers',
                'retailers': '/api/retailers',
                'retailer_catalogue': '/api/retailer-catalogue',
                'retailer_indents': '/api/retailer-indents',
                'retailer_dispatches': '/api/retailer-dispatches',
                'retailer_invoices': '/api/retailer-invoices',
                'retailer_rejections': '/api/retailer-rejections',
                'retailer_payments': '/api/retailer-payments',
                'qc_indents': '/api/qc-indents',
                'qc_dispatches': '/api/qc-dispatches',
                'qc_grns': '/api/qc-grns',
                'qc_invoices': '/api/qc-invoices',
                'procurements': '/api/procurement',
                'farmer_payments': '/api/farmer-payments',
                'customer_product_settings': '/api/customer-product-settings',
                'retail_plans': '/api/retail-plans',
            }
            
            # Endpoints that need date range params (fetch last 6 months)
            from_date = (datetime.now(timezone.utc) - timedelta(days=180)).strftime('%Y-%m-%d')
            to_date = datetime.now(timezone.utc).strftime('%Y-%m-%d')
            
            date_range_endpoints = {
                'variable_expenses': f'/api/expenses/variable?from_date={from_date}&to_date={to_date}',
                'fixed_expenses': f'/api/expenses/fixed?from_date={from_date}&to_date={to_date}',
                # Note: labour_attendance requires per-day queries - handled by Excel backup sync
            }
            
            sync_results = {}
            total_synced = 0
            
            for collection_name, endpoint in api_endpoints.items():
                try:
                    # Fetch data from production API
                    response = await client.get(
                        f"{production_url}{endpoint}",
                        headers=headers,
                        timeout=60.0
                    )
                    
                    if response.status_code != 200:
                        sync_results[collection_name] = {
                            "status": "error",
                            "error": f"API returned {response.status_code}",
                            "count": 0
                        }
                        continue
                    
                    data = response.json()
                    
                    # Handle different response formats
                    if isinstance(data, list):
                        records = data
                    elif isinstance(data, dict) and 'items' in data:
                        records = data['items']
                    elif isinstance(data, dict) and collection_name in data:
                        records = data[collection_name]
                    else:
                        records = [data] if data else []
                    
                    if not records:
                        sync_results[collection_name] = {
                            "status": "empty",
                            "count": 0
                        }
                        continue
                    
                    # Clear existing and insert new
                    await db[collection_name].delete_many({})
                    
                    # Clean records (remove any _id fields)
                    cleaned_records = []
                    for record in records:
                        if isinstance(record, dict):
                            clean_record = {k: v for k, v in record.items() if k != '_id'}
                            cleaned_records.append(clean_record)
                    
                    if cleaned_records:
                        await db[collection_name].insert_many(cleaned_records)
                    
                    sync_results[collection_name] = {
                        "status": "synced",
                        "count": len(cleaned_records)
                    }
                    total_synced += len(cleaned_records)
                    
                except Exception as e:
                    sync_results[collection_name] = {
                        "status": "error",
                        "error": str(e),
                        "count": 0
                    }
            
            # Also sync collections that need date range params
            for collection_name, endpoint in date_range_endpoints.items():
                try:
                    response = await client.get(
                        f"{production_url}{endpoint}",
                        headers=headers,
                        timeout=60.0
                    )
                    
                    if response.status_code != 200:
                        sync_results[collection_name] = {
                            "status": "error",
                            "error": f"API returned {response.status_code}",
                            "count": 0
                        }
                        continue
                    
                    data = response.json()
                    
                    if isinstance(data, list):
                        records = data
                    elif isinstance(data, dict) and 'records' in data:
                        records = data['records']
                    elif isinstance(data, dict) and 'items' in data:
                        records = data['items']
                    else:
                        records = [data] if data else []
                    
                    if not records:
                        sync_results[collection_name] = {
                            "status": "empty",
                            "count": 0
                        }
                        continue
                    
                    # Clear existing and insert new
                    await db[collection_name].delete_many({})
                    
                    cleaned_records = []
                    for record in records:
                        if isinstance(record, dict):
                            clean_record = {k: v for k, v in record.items() if k != '_id'}
                            cleaned_records.append(clean_record)
                    
                    if cleaned_records:
                        await db[collection_name].insert_many(cleaned_records)
                    
                    sync_results[collection_name] = {
                        "status": "synced",
                        "count": len(cleaned_records)
                    }
                    total_synced += len(cleaned_records)
                    
                except Exception as e:
                    sync_results[collection_name] = {
                        "status": "error",
                        "error": str(e),
                        "count": 0
                    }
            
            # Reset passwords if requested
            if request.reset_passwords:
                await db.users.update_many(
                    {},
                    {"$set": {"password": pwd_context.hash(request.default_password)}}
                )
            
            return {
                "message": f"Direct API sync completed. {total_synced} records synced.",
                "method": "direct_api",
                "production_url": production_url,
                "collections": sync_results
            }
        
    except httpx.TimeoutException:
        raise HTTPException(status_code=504, detail="Connection to production timed out")
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=500, 
            detail=f"Failed to sync from production: {str(e)}"
        )


class FullSyncRequest(BaseModel):
    production_url: str
    admin_email: str
    admin_password: str
    include_images: bool = True
    sync_mode: str = "merge"  # "merge" or "replace"

@router.post("/sync-from-production-full")
async def sync_from_production_full(
    request: FullSyncRequest,
    current_user: dict = Depends(get_current_user)
):
    """
    Full sync from production including images.
    Supports both merge (update existing, add new) and replace (clear and import fresh) modes.
    """
    import httpx
    
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Only admin can sync from production")
    
    try:
        production_url = request.production_url.rstrip('/')
        
        async with httpx.AsyncClient(timeout=300.0) as client:
            # Step 1: Login to production
            login_response = await client.post(
                f"{production_url}/api/auth/login",
                json={
                    "identifier": request.admin_email,
                    "password": request.admin_password
                }
            )
            
            if login_response.status_code != 200:
                raise HTTPException(
                    status_code=401,
                    detail=f"Failed to login to production: {login_response.text}"
                )
            
            token = login_response.json().get("token")
            if not token:
                raise HTTPException(status_code=401, detail="No token received from production login")
            
            headers = {"Authorization": f"Bearer {token}"}
            
            # Define all API endpoints - WITH images for products
            api_endpoints = {
                'qc_packaging': '/api/qc-packaging',
                'products': '/api/products?include_images=true' if request.include_images else '/api/products?include_images=false',
                'product_types': '/api/product-types',
                'product_categories': '/api/product-categories',
                'units': '/api/units',
                'farmers': '/api/farmers',
                'labours': '/api/labours',
                'qc_customers': '/api/qc-customers',
                'retailers': '/api/retailers',
                'retailer_catalogue': '/api/retailer-catalogue',
                'retailer_indents': '/api/retailer-indents',
                'retailer_dispatches': '/api/retailer-dispatches',
                'retailer_invoices': '/api/retailer-invoices',
                'retailer_rejections': '/api/retailer-rejections',
                'retailer_payments': '/api/retailer-payments',
                'retailer_credit_notes': '/api/retailer-credit-notes',
                'qc_indents': '/api/qc-indents',
                'qc_dispatches': '/api/qc-dispatches',
                'qc_grns': '/api/qc-grns',
                'qc_invoices': '/api/qc-invoices',
                'procurements': '/api/procurement',
                'farmer_payments': '/api/farmer-payments',
                'customer_product_settings': '/api/customer-product-settings',
                'daily_mrp': '/api/daily-mrp',
                'daily_stock_status': '/api/stock-status',
                'users': '/api/users',
                'blinkit_prices': '/api/blinkit-prices',
                'retail_plans': '/api/retail-plans',
            }
            
            # Date range endpoints
            from_date = (datetime.now(timezone.utc) - timedelta(days=365)).strftime('%Y-%m-%d')
            to_date = datetime.now(timezone.utc).strftime('%Y-%m-%d')
            
            date_range_endpoints = {
                'variable_expenses': f'/api/expenses/variable?from_date={from_date}&to_date={to_date}',
                'fixed_expenses': f'/api/expenses/fixed?from_date={from_date}&to_date={to_date}',
                'labour_attendance': f'/api/labour-attendance?from_date={from_date}&to_date={to_date}',
            }
            
            sync_results = {}
            total_synced = 0
            images_synced = 0
            
            # Sync main collections
            for collection_name, endpoint in {**api_endpoints, **date_range_endpoints}.items():
                try:
                    response = await client.get(
                        f"{production_url}{endpoint}",
                        headers=headers,
                        timeout=120.0
                    )
                    
                    if response.status_code != 200:
                        sync_results[collection_name] = {
                            "status": "error",
                            "error": f"API returned {response.status_code}",
                            "count": 0
                        }
                        continue
                    
                    data = response.json()
                    
                    # Handle different response formats
                    if isinstance(data, list):
                        records = data
                    elif isinstance(data, dict) and 'items' in data:
                        records = data['items']
                    elif isinstance(data, dict) and 'records' in data:
                        records = data['records']
                    elif isinstance(data, dict) and collection_name in data:
                        records = data[collection_name]
                    else:
                        records = [data] if data else []
                    
                    if not records:
                        sync_results[collection_name] = {
                            "status": "empty",
                            "count": 0
                        }
                        continue
                    
                    # Clean records (remove any _id fields)
                    cleaned_records = []
                    for record in records:
                        if isinstance(record, dict):
                            clean_record = {k: v for k, v in record.items() if k != '_id'}
                            cleaned_records.append(clean_record)
                            # Count images
                            if collection_name == 'products' and clean_record.get('image'):
                                images_synced += 1
                    
                    if request.sync_mode == "replace":
                        # Clear existing and insert new
                        await db[collection_name].delete_many({})
                        if cleaned_records:
                            await db[collection_name].insert_many(cleaned_records)
                    else:
                        # Merge mode: update existing, insert new
                        for record in cleaned_records:
                            record_id = record.get('id')
                            if record_id:
                                await db[collection_name].update_one(
                                    {"id": record_id},
                                    {"$set": record},
                                    upsert=True
                                )
                            else:
                                # No id field - insert as new
                                await db[collection_name].insert_one(record)
                    
                    sync_results[collection_name] = {
                        "status": "synced",
                        "count": len(cleaned_records),
                        "mode": request.sync_mode
                    }
                    total_synced += len(cleaned_records)
                    
                except Exception as e:
                    sync_results[collection_name] = {
                        "status": "error",
                        "error": str(e),
                        "count": 0
                    }
            
            # Also sync retailer catalogue images separately if include_images is true
            if request.include_images:
                try:
                    # Fetch retailer catalogue with images
                    cat_response = await client.get(
                        f"{production_url}/api/retailer-catalogue",
                        headers=headers,
                        timeout=120.0
                    )
                    
                    if cat_response.status_code == 200:
                        catalogue_data = cat_response.json()
                        if isinstance(catalogue_data, list):
                            for item in catalogue_data:
                                if item.get('image_url'):
                                    images_synced += 1
                                    # Update catalogue entry with image
                                    await db.retailer_catalogue.update_one(
                                        {"id": item.get('id')},
                                        {"$set": {"image_url": item.get('image_url')}},
                                        upsert=False
                                    )
                except Exception as e:
                    sync_results['retailer_catalogue_images'] = {
                        "status": "error",
                        "error": str(e)
                    }
            
            return {
                "message": f"Full sync completed. {total_synced} records synced. {images_synced} images included.",
                "method": "full_sync",
                "sync_mode": request.sync_mode,
                "include_images": request.include_images,
                "production_url": production_url,
                "images_synced": images_synced,
                "collections": sync_results
            }
        
    except httpx.TimeoutException:
        raise HTTPException(status_code=504, detail="Connection to production timed out")
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=500, 
            detail=f"Full sync failed: {str(e)}"
        )


@router.get("/sync-status")
async def get_sync_status(current_user: dict = Depends(get_current_user)):
    """Get current database record counts for sync verification"""
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Only admin can view sync status")
    
    collections = [
        # Core
        "users", "products", "units",
        # Procurement
        "farmers", "procurements", "procurement_templates",
        # Quick Commerce
        "qc_packaging", "qc_customers", "qc_indents", "qc_dispatches", 
        "qc_invoices", "qc_grns", "qc_daily_requirements",
        # Retailer
        "retailers", "retailer_indents", "retailer_dispatches", "retailer_invoices", 
        "retailer_grn", "retailer_rejections", "retailer_payments",
        "retailer_closing_inventory", "retailer_inventory", "retailer_daily_requirements",
        "retail_plans",
        # Stock & Expenses
        "daily_stock_status", "variable_expenses", "fixed_expenses",
        # Labour
        "labours", "labour_attendance",
        # Payments
        "payments"
    ]
    
    counts = {}
    total = 0
    for coll in collections:
        count = await db[coll].count_documents({})
        counts[coll] = count
        total += count
    
    return {
        "database": os.environ.get('DB_NAME', 'unknown'),
        "total_records": total,
        "collections": counts,
        "last_checked": datetime.now(timezone.utc).isoformat()
    }


# ============================================================================
